"""
DataLineage IQ — Backend v4.1
Complete implementation:
  1. ERWIN DI  → Live API → Local Excel/JSON folder → Mock
  2. Data Dict → Live API → Local Excel/JSON folder → Mock
  3. Code Repo → GitHub API → Local folder scan → Mock
  4. Tachyon   → Internal LLM: business logic summary + executive insights
  5. Export    → 4-sheet Excel (color-coded) + structured JSON + PDF per field

NEW in v4.1 (per Business Requirements):
  6. Multi-field search  — comma-separated fields (customer_id, acct_num, transaction_ts)
  7. Equivalent fields   — identify same field across different SOR naming conventions
  8. System-of-Record    — show which SORs carry each field; user can pick authoritative SOR
  9. Confidentiality     — expose confidentiality classification per field
  10. Multi-hop lineage  — trace full Hogan→FDR→CCHUB→EIW/ADH/ICODE/CODI paths
  11. Similarity search  — fuzzy/alias matching for field names across systems
  12. Input file upload  — accept Business Requirements xlsx + metadata INPUTS folder
  13. Ingestion patterns — detect batch vs real-time fields
  14. /api/multi-search  — new endpoint for multi-field batch lookup
  15. /api/sor-select    — new endpoint for setting preferred SOR per field
  16. /api/upload-inputs — new endpoint for uploading BR xlsx / INPUTS zip
  17. /api/equivalent-fields — find equivalent/alias fields across systems
"""
from __future__ import annotations
import asyncio, base64, glob, io, json, os, re, difflib, zipfile, tempfile
from datetime import datetime
from pathlib import Path
from typing import Any, Dict, List, Optional

import httpx
from fastapi import FastAPI, HTTPException, UploadFile, File, Form
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import FileResponse, StreamingResponse
from pydantic import BaseModel

app = FastAPI(title="DataLineage IQ", version="4.1.0")
app.add_middleware(CORSMiddleware, allow_origins=["*"],
                   allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

# ─── CONFIG ──────────────────────────────────────────────────────────────────
CFG = {
    "ERWIN_BASE_URL":  os.getenv("ERWIN_BASE_URL",  "https://lpvra97a0104.wellsfargo.net:4512/erwinDISuite"),
    "ERWIN_API_KEY":   os.getenv("ERWIN_API_KEY",   ""),       # api-key value from ERWIN DI
    "ERWIN_SYSTEM":    os.getenv("ERWIN_SYSTEM",    ""),       # e.g. FNISH
    "ERWIN_ENV":       os.getenv("ERWIN_ENV",       ""),       # e.g. CDS_DCBTDISPUTE
    "ERWIN_PROJECT":   os.getenv("ERWIN_PROJECT",   ""),       # optional project id filter
    "DATA_DICT_URL":   os.getenv("DATA_DICT_URL",   "http://data-dict-host/api"),
    "LOCAL_META_DIR":  os.getenv("LOCAL_META_DIR",  "./metadata_files"),
    "LOCAL_DICT_DIR":  os.getenv("LOCAL_DICT_DIR",  "./dictionary_files"),
    "LOCAL_CODE_DIR":  os.getenv("LOCAL_CODE_DIR",  "./code_repo_samples"),
    "INPUTS_DIR":      os.getenv("INPUTS_DIR",      "./inputs"),          # NEW: Business Req + metadata INPUTS folder
    "GITHUB_TOKEN":    os.getenv("GITHUB_TOKEN",    ""),
    "GITHUB_REPO":     os.getenv("GITHUB_REPO",     ""),
    "TACHYON_URL":     os.getenv("TACHYON_URL",     ""),
    "TACHYON_API_KEY": os.getenv("TACHYON_API_KEY", ""),
    "TACHYON_MODEL":   os.getenv("TACHYON_MODEL",   "tachyon-llm"),
}
CODE_EXTS = [".sql",".hql",".py",".java",".scala",".ksh",".sh",".xml",".yaml",".yml"]
LANG_MAP  = {".sql":"SQL",".hql":"HQL",".py":"Python",".java":"Java",".scala":"Scala",
             ".ksh":"Shell",".sh":"Shell",".xml":"XML",".yaml":"YAML",".yml":"YAML"}

# ── NEW: Known Credit Services Systems of Record (SOR) ───────────────────────
# Based on business requirement: Hogan, FDR, CCHUB, EIW, ADH, ICODE, CODI, etc.
KNOWN_SORS = ["HOGAN","FDR","CCHUB","EIW","ADH","ICODE","CODI","BINK","BMG",
              "ICRDT","CBODS","ACDV","FICO","STRATA","LPS","CHEX","EXPERIAN"]

# ── NEW: In-memory SOR preference store (field_upper → preferred_sor) ─────────
_sor_preferences: Dict[str, str] = {}

# ── NEW: Known field aliases / equivalents across systems ──────────────────────
# customer_id appears as ECN_PRIMARY_ID, ECN, CUSTOMER_ID, ACCT_NUM, etc.
FIELD_ALIASES: Dict[str, List[str]] = {
    "CUSTOMER_ID":    ["ECN_PRIMARY_ID","ECN","CUST_ID","CUSTOMER_NUMBER","ACCT_HOLDER_ID"],
    "ECN_PRIMARY_ID": ["CUSTOMER_ID","ECN","CUST_ID"],
    "ACCT_NUM":       ["ACCOUNT_NUMBER","ACCT_ID","ACCOUNT_ID","ACCT_NUMBER"],
    "TRANSACTION_TS": ["TXN_TIMESTAMP","TRANS_DATE","TRANSACTION_DATE","TXN_DT","TRANS_TS"],
    "REVENUE_AMT":    ["AMOUNT","TXN_AMT","TRANSACTION_AMOUNT","SALE_AMOUNT","REVENUE"],
    "STATUS":         ["STATUS_CD","ACCT_STATUS","ACCT_STATUS_CD","STATUS_CODE"],
    "FICO_SCORE":     ["FICOSCOREV9SCORE","FICO_V9","CREDIT_SCORE","RISK_SCORE"],
    "RTD_CHECK_MTD_DT": ["CHECK_DATE","MTD_CHECK_DT","MONTHLY_CHECK_DATE"],
    "MAC_DELQ":       ["DELINQUENCY_STATUS","DELQ_STATUS","MAC_DELINQUENCY"],
}

# ── NEW: Confidentiality classification map ────────────────────────────────────
CONFIDENTIALITY_MAP: Dict[str, str] = {
    "CUSTOMER_ID":"PII-Restricted","ECN_PRIMARY_ID":"PII-Restricted",
    "ACCT_NUM":"PII-Restricted","SSN":"PII-Confidential","DOB":"PII-Confidential",
    "FICO_SCORE":"Confidential","FICOSCOREV9SCORE":"Confidential",
    "REVENUE_AMT":"Confidential","MAC_DELQ":"Confidential","RTD_CHECK_MTD_DT":"Internal",
    "STATUS":"Internal","REGION":"Public","UPDATED_BY":"Internal","CREATED_DATE":"Internal",
}
DEFAULT_CONFIDENTIALITY = "Internal"

# ─── REQUEST / RESPONSE ──────────────────────────────────────────────────────
class SearchRequest(BaseModel):
    query: str
    use_erwin: bool = True
    use_data_dict: bool = True
    use_code_repo: bool = True
    use_tachyon: bool = True
    github_token: Optional[str] = None
    github_repo:  Optional[str] = None
    local_code_dir: Optional[str] = None
    local_meta_dir: Optional[str] = None
    local_dict_dir: Optional[str] = None
    tachyon_url: Optional[str] = None
    tachyon_key: Optional[str] = None
    # NEW v4.1 fields
    preferred_sor: Optional[str] = None
    include_equivalent_fields: bool = True
    ingestion_context: Optional[str] = None
    # ERWIN DI connection — all user-configurable, override START_APP.bat values
    erwin_base_url:     Optional[str] = None  # Override ERWIN_BASE_URL
    erwin_api_key:      Optional[str] = None  # Override ERWIN_API_KEY
    erwin_system:       Optional[str] = None  # systemName param
    erwin_env:          Optional[str] = None  # environmentName param
    erwin_project:      Optional[str] = None  # projectIds param
    erwin_lineage_type: Optional[str] = None  # DUAL / SOURCE / TARGET
    # Per-endpoint overrides — applied only to the matching endpoint
    erwin_ep_tablename: Optional[str] = None  # override tableName param
    erwin_ep_colname:   Optional[str] = None  # override columnName param
    erwin_ep_objtype:   Optional[str] = None  # override type param for /api/objects/search
    erwin_ep_mapname:   Optional[str] = None  # override name param for /api/mappings
    erwin_ep_glosscat:  Optional[str] = None  # override category for /api/glossary/terms  # DUAL / SOURCE / TARGET
    # Per-endpoint overrides — applied only to the matching endpoint
    erwin_ep_tablename: Optional[str] = None  # override tableName param
    erwin_ep_colname:   Optional[str] = None  # override columnName param
    erwin_ep_objtype:   Optional[str] = None  # override type param for /api/objects/search
    erwin_ep_mapname:   Optional[str] = None  # override name param for /api/mappings
    erwin_ep_glosscat:  Optional[str] = None  # override category for /api/glossary/terms

class MultiSearchRequest(BaseModel):
    """NEW: Search for multiple fields at once — e.g. customer_id, acct_num, transaction_ts"""
    fields: List[str]
    use_erwin: bool = True
    use_data_dict: bool = True
    use_code_repo: bool = True
    use_tachyon: bool = True
    local_meta_dir: Optional[str] = None
    local_dict_dir: Optional[str] = None
    local_code_dir: Optional[str] = None
    tachyon_url: Optional[str] = None
    tachyon_key: Optional[str] = None
    erwin_base_url:     Optional[str] = None
    erwin_api_key:      Optional[str] = None
    erwin_system:       Optional[str] = None
    erwin_env:          Optional[str] = None
    erwin_project:      Optional[str] = None
    erwin_lineage_type: Optional[str] = None
    erwin_ep_tablename: Optional[str] = None
    erwin_ep_colname:   Optional[str] = None
    erwin_ep_objtype:   Optional[str] = None
    erwin_ep_mapname:   Optional[str] = None
    erwin_ep_glosscat:  Optional[str] = None

class SORSelectRequest(BaseModel):
    """NEW: User sets preferred System of Record for a field"""
    field_name: str
    preferred_sor: str

# ═══════════════════════════════════════════════════════════════════════════════
# ERWIN DI  —  Real API  →  Local Excel/JSON  →  Mock
# ═══════════════════════════════════════════════════════════════════════════════
# ═══════════════════════════════════════════════════════════════════════════════
# ERWIN DI — GENERIC ENDPOINT ENGINE
# ═══════════════════════════════════════════════════════════════════════════════
#
# How it works:
#  1. On first call, hits /api/swagger.json (or /v2/api-docs) to DISCOVER all
#     available ERWIN endpoints dynamically.
#  2. Builds a ranked call-plan: matches endpoints whose path/params contain
#     the query terms (tableName, columnName, systemName, etc.)
#  3. Calls each matched endpoint in priority order until data is returned.
#  4. Falls back to local Excel/JSON, then mock.
#
# This means ANY new ERWIN endpoint is automatically supported — no code change.
# ─────────────────────────────────────────────────────────────────────────────

# Cache discovered ERWIN endpoints (refreshed on server restart)
_erwin_swagger_cache: Optional[Dict] = None
_erwin_endpoints_cache: Optional[List[Dict]] = None

# Known ERWIN endpoint registry — used when Swagger discovery fails
# Format: {path, method, params[], priority, description, result_key}
ERWIN_KNOWN_ENDPOINTS: List[Dict] = [
    # ── Lineage endpoints ────────────────────────────────────────────────────
    {"path": "/api/datalineage/table",         "method": "GET", "priority": 1,
     "description": "Table lineage (DUAL/SOURCE/TARGET)",
     "params": ["lineageType","tableName","systemName","environmentName","projectIds"],
     "required": ["tableName"], "result_parser": "table_lineage"},

    {"path": "/api/datalineage/column",        "method": "GET", "priority": 2,
     "description": "Column/field lineage",
     "params": ["lineageType","tableName","columnName","systemName","environmentName","projectIds"],
     "required": ["tableName","columnName"], "result_parser": "column_lineage"},

    {"path": "/api/datalineage/table/impact",  "method": "GET", "priority": 3,
     "description": "Impact analysis for a table",
     "params": ["tableName","systemName","environmentName"],
     "required": ["tableName"], "result_parser": "table_lineage"},

    {"path": "/api/datalineage/column/impact", "method": "GET", "priority": 4,
     "description": "Impact analysis for a column",
     "params": ["tableName","columnName","systemName","environmentName"],
     "required": ["tableName","columnName"], "result_parser": "column_lineage"},

    {"path": "/api/datalineage/system",        "method": "GET", "priority": 1,
     "description": "System-level lineage — all tables and flows within a system",
     "params": ["lineageType","systemName","environmentName","projectIds"],
     "required": ["systemName"], "result_parser": "system_lineage"},

    # ── Object / metadata endpoints ──────────────────────────────────────────
    {"path": "/api/objects/search",            "method": "GET", "priority": 5,
     "description": "Search objects by name and type",
     "params": ["name","type","systemName","environmentName","projectId"],
     "required": ["name"], "result_parser": "objects"},

    {"path": "/api/objects",                   "method": "GET", "priority": 6,
     "description": "List all objects",
     "params": ["systemName","environmentName","type","page","pageSize"],
     "required": [], "result_parser": "objects"},

    {"path": "/api/tables",                    "method": "GET", "priority": 6,
     "description": "List tables",
     "params": ["systemName","environmentName","name","projectId"],
     "required": [], "result_parser": "objects"},

    {"path": "/api/columns",                   "method": "GET", "priority": 7,
     "description": "List columns",
     "params": ["tableName","systemName","environmentName","name"],
     "required": [], "result_parser": "objects"},

    # ── Mapping / ETL endpoints ──────────────────────────────────────────────
    {"path": "/api/mappings",                  "method": "GET", "priority": 8,
     "description": "ETL mapping documents",
     "params": ["tableName","systemName","environmentName","name","projectId"],
     "required": [], "result_parser": "mappings"},

    {"path": "/api/mappings/search",           "method": "GET", "priority": 8,
     "description": "Search mappings by name",
     "params": ["name","systemName","environmentName"],
     "required": ["name"], "result_parser": "mappings"},

    # ── System / environment discovery ───────────────────────────────────────
    {"path": "/api/systems",                   "method": "GET", "priority": 9,
     "description": "List all systems",
     "params": [], "required": [], "result_parser": "systems"},

    {"path": "/api/environments",              "method": "GET", "priority": 9,
     "description": "List environments",
     "params": ["systemName","systemId"],
     "required": [], "result_parser": "systems"},

    {"path": "/api/subject-areas",             "method": "GET", "priority": 10,
     "description": "Subject areas / data domains",
     "params": ["projectId","systemName"],
     "required": [], "result_parser": "objects"},

    # ── Data quality / profiling ─────────────────────────────────────────────
    {"path": "/api/data-quality/rules",        "method": "GET", "priority": 11,
     "description": "Data quality rules",
     "params": ["tableName","columnName","systemName"],
     "required": [], "result_parser": "objects"},

    {"path": "/api/data-quality/scores",       "method": "GET", "priority": 11,
     "description": "Data quality scores",
     "params": ["tableName","systemName","environmentName"],
     "required": [], "result_parser": "objects"},

    # ── Business glossary ────────────────────────────────────────────────────
    {"path": "/api/glossary/terms",            "method": "GET", "priority": 12,
     "description": "Business glossary terms",
     "params": ["name","category"],
     "required": [], "result_parser": "objects"},

    {"path": "/api/glossary/terms/search",     "method": "GET", "priority": 12,
     "description": "Search glossary terms",
     "params": ["name","q"],
     "required": ["name"], "result_parser": "objects"},

    # ── Projects ─────────────────────────────────────────────────────────────
    {"path": "/api/projects",                  "method": "GET", "priority": 13,
     "description": "List projects",
     "params": [], "required": [], "result_parser": "systems"},
]


async def _discover_erwin_endpoints(base: str, hdrs: Dict) -> List[Dict]:
    """
    Dynamically discover ERWIN endpoints via Swagger/OpenAPI spec.
    Falls back to ERWIN_KNOWN_ENDPOINTS if discovery fails.
    Returns list of endpoint dicts ready to call.
    """
    global _erwin_swagger_cache, _erwin_endpoints_cache
    if _erwin_endpoints_cache is not None:
        return _erwin_endpoints_cache

    swagger_urls = [
        f"{base}/v3/api-docs",
        f"{base}/v2/api-docs",
        f"{base}/api/swagger.json",
        f"{base}/swagger.json",
        f"{base}/openapi.json",
    ]
    endpoints: List[Dict] = []

    async with httpx.AsyncClient(timeout=8.0, verify=False) as c:
        for url in swagger_urls:
            try:
                r = await c.get(url, headers=hdrs)
                if r.status_code != 200: continue
                spec = r.json()
                _erwin_swagger_cache = spec
                # Parse OpenAPI paths
                for path, methods in spec.get("paths", {}).items():
                    for method, op in methods.items():
                        if method.upper() not in ("GET", "POST"): continue
                        params = [p.get("name","") for p in op.get("parameters", [])
                                  if p.get("in") in ("query","path")]
                        required = [p.get("name","") for p in op.get("parameters", [])
                                    if p.get("required", False)]
                        desc = op.get("summary","") or op.get("description","") or path
                        endpoints.append({
                            "path": path, "method": method.upper(),
                            "params": params, "required": required,
                            "description": desc, "result_parser": "generic",
                            "priority": _score_endpoint(path, params),
                            "from_swagger": True,
                        })
                if endpoints:
                    # Merge with known endpoints (known take priority for parsers)
                    known_paths = {e["path"] for e in ERWIN_KNOWN_ENDPOINTS}
                    for e in endpoints:
                        if e["path"] not in known_paths:
                            ERWIN_KNOWN_ENDPOINTS.append(e)
                    _erwin_endpoints_cache = sorted(ERWIN_KNOWN_ENDPOINTS, key=lambda x: x["priority"])
                    return _erwin_endpoints_cache
            except Exception:
                pass

    # Swagger not available — use known registry
    _erwin_endpoints_cache = sorted(ERWIN_KNOWN_ENDPOINTS, key=lambda x: x["priority"])
    return _erwin_endpoints_cache


def _score_endpoint(path: str, params: List[str]) -> int:
    """Score an endpoint by how useful it likely is for lineage discovery."""
    score = 50
    path_l = path.lower()
    if "lineage"    in path_l: score -= 40
    if "datalineage"in path_l: score -= 45
    if "impact"     in path_l: score -= 35
    if "mapping"    in path_l: score -= 30
    if "search"     in path_l: score -= 20
    if "table"      in path_l: score -= 15
    if "column"     in path_l: score -= 12
    if "quality"    in path_l: score -= 5
    if "tableName"  in params: score -= 10
    if "columnName" in params: score -= 8
    if "systemName" in params: score -= 5
    return max(1, score)


def _build_call_params(endpoint: Dict, table: str, column: Optional[str],
                        system: str, env: str, project: str,
                        lin_type: str, query: str,
                        ep_overrides: Optional[Dict] = None) -> Optional[Dict]:
    """
    Build the query params dict for an endpoint.
    Returns None if required params cannot be satisfied.
    """
    required = endpoint.get("required", [])
    available = endpoint.get("params", [])

    # Check required params are satisfiable
    val_map = {
        "tableName":       table,
        "table_name":      table,
        "name":            table,
        "q":               query,
        "columnName":      column or "",
        "column_name":     column or "",
        "systemName":      system,
        "system_name":     system,
        "environmentName": env,
        "environment_name":env,
        "projectIds":      project,
        "projectId":       project,
        "lineageType":     lin_type,
        "lineage_type":    lin_type,
        "type":            "TABLE",
    }
    # Apply per-endpoint overrides based on endpoint path
    ep_path = endpoint.get("path", "").lower()
    if ep_overrides:
        # tableName override for table/system lineage endpoints
        if "tableName" in endpoint.get("params", []) and ep_overrides.get("tablename"):
            val_map["tableName"] = ep_overrides["tablename"]
            val_map["name"]      = ep_overrides["tablename"]
        # columnName override for column lineage endpoints
        if "columnName" in endpoint.get("params", []) and ep_overrides.get("colname"):
            val_map["columnName"] = ep_overrides["colname"]
        # type override for object search
        if "type" in endpoint.get("params", []) and ep_overrides.get("objtype"):
            val_map["type"] = ep_overrides["objtype"]
        # name override for mappings/glossary
        if "name" in endpoint.get("params", []):
            if "mapping" in ep_path and ep_overrides.get("mapname"):
                val_map["name"] = ep_overrides["mapname"]
            elif "glossary" in ep_path and ep_overrides.get("glosscat"):
                val_map["category"] = ep_overrides["glosscat"]

    for req in required:
        if not val_map.get(req,""):
            return None  # Can't satisfy this required param

    # Build params dict with only what this endpoint accepts
    params: Dict = {}
    for p in available:
        v = val_map.get(p,"")
        if v:
            params[p] = v

    # Always include lineageType for lineage endpoints
    if "lineageType" in available and lin_type:
        params["lineageType"] = lin_type

    return params


async def fetch_erwin(
    query: str,
    local_dir: str,
    erwin_base_url: Optional[str] = None,
    erwin_api_key:  Optional[str] = None,
    erwin_system:   Optional[str] = None,
    erwin_env:      Optional[str] = None,
    erwin_project:  Optional[str] = None,
    erwin_lineage_type: Optional[str] = None,
    erwin_ep_tablename: Optional[str] = None,
    erwin_ep_colname:   Optional[str] = None,
    erwin_ep_objtype:   Optional[str] = None,
    erwin_ep_mapname:   Optional[str] = None,
    erwin_ep_glosscat:  Optional[str] = None,
) -> Dict:
    """
    Generic ERWIN DI integration engine.
    Discovers endpoints via Swagger, then tries each one in priority order.
    Any ERWIN endpoint is automatically supported — no code change needed.
    """
    # ── Parse query ───────────────────────────────────────────────────────────
    parts     = query.upper().split(".")
    table     = parts[-1]
    column    = parts[1] if len(parts) >= 3 else (parts[1] if len(parts)==2 and "_" not in parts[0][:3] else None)
    maybe_sys = parts[0] if len(parts) >= 2 else ""

    # ── Resolve connection params (UI > START_APP.bat > env > defaults) ───────
    base     = (erwin_base_url or CFG["ERWIN_BASE_URL"]).rstrip("/")
    api_key  = erwin_api_key  or CFG["ERWIN_API_KEY"]
    system   = erwin_system   or CFG["ERWIN_SYSTEM"] or maybe_sys or ""
    env      = erwin_env      or CFG["ERWIN_ENV"] or ""
    project  = erwin_project  or CFG["ERWIN_PROJECT"] or ""
    lin_type = erwin_lineage_type or "DUAL"

    if not api_key:
        result = _scan_local_meta(query, local_dir or CFG["LOCAL_META_DIR"])
        return result if result else _mock_erwin(query)

    hdrs = {
        "accept":        "application/json",
        "Authorization": api_key,  # raw api-key, no "Bearer" prefix
    }

    # Build per-endpoint override dict from request params
    ep_overrides = {
        "tablename": erwin_ep_tablename or "",
        "colname":   erwin_ep_colname   or "",
        "objtype":   erwin_ep_objtype   or "TABLE",
        "mapname":   erwin_ep_mapname   or "",
        "glosscat":  erwin_ep_glosscat  or "",
    }

    # ── Discover all endpoints ────────────────────────────────────────────────
    all_endpoints = await _discover_erwin_endpoints(base, hdrs)

    # ── Call each endpoint in priority order ─────────────────────────────────
    tried: List[str] = []
    async with httpx.AsyncClient(timeout=10.0, verify=False) as c:
        for ep in all_endpoints:
            params = _build_call_params(ep, table, column, system, env,
                                         project, lin_type, query,
                                         ep_overrides=ep_overrides)
            if params is None:
                continue  # Required params not satisfied

            url = base + ep["path"]
            tried.append(ep["path"])

            try:
                if ep["method"] == "GET":
                    r = await c.get(url, headers=hdrs, params=params)
                else:
                    r = await c.post(url, headers=hdrs, json=params)

                if r.status_code not in (200, 201):
                    continue
                if not r.text.strip() or r.text.strip() in ("[]", "{}", "null"):
                    continue

                raw = r.json()

                # ── Parse based on endpoint type ──────────────────────────────
                parser = ep.get("result_parser", "generic")
                if parser == "system_lineage":
                    result = _parse_erwin_system_lineage(raw, query)
                elif parser == "table_lineage":
                    result = _parse_erwin_table_lineage(raw, query)
                elif parser == "column_lineage":
                    result = _parse_erwin_column_lineage(raw, query)
                elif parser == "mappings":
                    result = _parse_erwin_mappings(raw, query)
                elif parser == "systems":
                    result = _parse_erwin_systems(raw, query)
                else:
                    # Generic parser — handles any response shape
                    result = _parse_erwin_generic(raw, query)

                if result.get("objects") or result.get("subject_areas"):
                    return {
                        **result,
                        "status": "live",
                        "erwin_endpoint": ep["path"],
                        "erwin_description": ep.get("description",""),
                        "endpoints_tried": tried,
                        "source_file": None,
                        "raw_response_keys": list(raw.keys()) if isinstance(raw, dict) else ["list:"+str(len(raw))],
                    }
            except Exception:
                continue

    # ── Fallback: local files → mock ──────────────────────────────────────────
    result = _scan_local_meta(query, local_dir or CFG["LOCAL_META_DIR"])
    if result:
        return {**result, "endpoints_tried": tried}
    return {**_mock_erwin(query), "endpoints_tried": tried}


def _parse_erwin_systems(raw: Any, query: str) -> Dict:
    """Parse /api/systems or /api/environments response."""
    objects = []
    items = raw if isinstance(raw, list) else (
        raw.get("systems") or raw.get("environments") or raw.get("items") or [])
    for item in (items if isinstance(items, list) else []):
        name = item.get("name") or item.get("systemName") or item.get("environmentName") or ""
        if not name: continue
        objects.append({
            "id": f"sys_{name.upper().replace(' ','_')}",
            "name": name, "type": "source_table",
            "database": "", "schema": "",
            "system": item.get("description","") or name,
            "fields": [], "logic": "",
        })
    return {"objects": objects, "edges": _infer_edges(objects), "source_file": None}


def _parse_erwin_generic(raw: Any, query: str) -> Dict:
    """
    Generic parser for any ERWIN endpoint response.
    Tries all known response shapes and extracts objects/edges.
    """
    # Try known parsers first
    for parser in [_parse_erwin_system_lineage, _parse_erwin_table_lineage,
                   _parse_erwin_column_lineage, _parse_erwin_mappings,
                   _parse_erwin_systems]:
        try:
            result = parser(raw, query)
            if result.get("objects"):
                return result
        except Exception:
            pass

    # Last resort: _parse_erwin_api (original v1/v2 handler)
    try:
        return _parse_erwin_api(raw, query, "live")
    except Exception:
        pass

    return {"objects": [], "edges": [], "source_file": None}


# ── NEW API: Expose ERWIN endpoint catalog to frontend ────────────────────────
@app.get("/api/erwin/endpoints")
async def erwin_endpoints():
    """Return the full catalog of known + discovered ERWIN endpoints."""
    base    = CFG["ERWIN_BASE_URL"].rstrip("/")
    api_key = CFG["ERWIN_API_KEY"]
    hdrs    = {"accept":"application/json","Authorization":api_key} if api_key else {}
    eps = await _discover_erwin_endpoints(base, hdrs) if api_key else ERWIN_KNOWN_ENDPOINTS
    return {
        "total": len(eps),
        "swagger_discovered": any(e.get("from_swagger") for e in eps),
        "base_url": base,
        "endpoints": [{"path":e["path"],"method":e["method"],
                        "description":e.get("description",""),
                        "params":e.get("params",[]),
                        "required":e.get("required",[]),
                        "priority":e.get("priority",99)} for e in eps],
    }


@app.post("/api/erwin/call")
async def erwin_call_direct(payload: Dict[str, Any]):
    """
    Call any ERWIN endpoint directly by path + params.
    Allows the UI to call any endpoint ad-hoc.
    POST { "path": "/api/systems", "params": {}, "method": "GET" }
    """
    base    = (payload.get("base_url") or CFG["ERWIN_BASE_URL"]).rstrip("/")
    api_key = payload.get("api_key") or CFG["ERWIN_API_KEY"]
    if not api_key:
        raise HTTPException(400, "ERWIN_API_KEY not configured")

    path   = payload.get("path","")
    method = payload.get("method","GET").upper()
    params = payload.get("params", {})
    if not path:
        raise HTTPException(400, "path is required")

    hdrs = {"accept":"application/json","Authorization": api_key}
    try:
        async with httpx.AsyncClient(timeout=15.0, verify=False) as c:
            if method == "GET":
                r = await c.get(base + path, headers=hdrs, params=params)
            else:
                r = await c.post(base + path, headers=hdrs, json=params)
        return {
            "status_code": r.status_code,
            "endpoint":    base + path,
            "params_sent": params,
            "response":    r.json() if r.text.strip() else {},
            "raw_size":    len(r.content),
        }
    except Exception as e:
        raise HTTPException(502, f"ERWIN call failed: {e}")


@app.get("/api/erwin/swagger")
async def erwin_swagger():
    """Return cached ERWIN Swagger spec (fetched from /v3/api-docs etc.)"""
    if _erwin_swagger_cache:
        return _erwin_swagger_cache
    base    = CFG["ERWIN_BASE_URL"].rstrip("/")
    api_key = CFG["ERWIN_API_KEY"]
    if not api_key:
        raise HTTPException(400,"ERWIN_API_KEY not set")
    hdrs = {"accept":"application/json","Authorization":api_key}
    for url in [f"{base}/v3/api-docs", f"{base}/v2/api-docs", f"{base}/api/swagger.json"]:
        try:
            async with httpx.AsyncClient(timeout=8.0,verify=False) as c:
                r = await c.get(url,headers=hdrs)
                if r.status_code==200:
                    return r.json()
        except Exception:
            pass
    raise HTTPException(404,"Swagger spec not found at any known URL")


# ── ERWIN RESPONSE PARSERS ────────────────────────────────────────────────────

def _parse_erwin_table_lineage(raw: Any, query: str) -> Dict:
    """
    Parse GET /api/datalineage/table response.

    Handles ALL known ERWIN DI response shapes:
      Shape A: { source: {...}, target: {...} }          ← YOUR ERWIN (confirmed from screenshots)
      Shape B: { sourceObjects: [...], targetObjects: [...] }
      Shape C: [ { source: {...}, target: {...} }, ... ] ← list of A
      Shape D: flat list [ { tableName, endPointType }, ... ]
      Shape E: { sourceTable: {...}, targetTable: {...} }
    """
    objects, edges = [], []
    seen = set()

    def add_obj(item: Dict, obj_type: str):
        if not isinstance(item, dict): return
        name = (item.get("tableName") or item.get("name") or item.get("objectName") or "")
        if not name or name.upper() in seen: return
        seen.add(name.upper())
        cols = item.get("columns") or item.get("fields") or item.get("attributes") or []
        # systemEnvironmentName is YOUR ERWIN's system field (confirmed from response)
        system = (item.get("systemEnvironmentName") or
                  item.get("systemName") or
                  item.get("applicationName") or
                  item.get("systemEnvName") or
                  CFG["ERWIN_ENV"] or CFG["ERWIN_SYSTEM"] or "")
        database = item.get("databaseName") or item.get("database") or ""
        schema   = item.get("schemaName")   or item.get("schema")   or ""
        # endPointType tells us source vs target directly
        ep_type = (item.get("endPointType") or "").upper()
        if ep_type == "SOURCE":   obj_type = "source_table"
        elif ep_type == "TARGET": obj_type = "target_table"
        objects.append({
            "id":       f"ew_{name.upper().replace(' ','_').replace('.','_')}",
            "name":     name,
            "type":     obj_type,
            "database": database,
            "schema":   schema,
            "system":   system,
            "fields":   [_norm_field(c) for c in cols],
            "logic":    item.get("transformationLogic") or item.get("mappingLogic") or "",
            # Extra ERWIN-specific fields preserved for display
            "node_type":    item.get("nodeType") or "",
            "object_type_id": item.get("objectTypeId") or 0,
            "sdi_classifications": item.get("sdiClassifications") or [],
        })

    def process_item(item: Dict):
        """Process one ERWIN record — may contain source/target keys or be a flat table."""
        if not isinstance(item, dict): return
        ep = (item.get("endPointType") or "").upper()
        if ep in ("SOURCE","TARGET"):
            # Flat record with endPointType — Shape D
            add_obj(item, "source_table" if ep=="SOURCE" else "target_table")
            return
        # Shape A: { source: {...}, target: {...} }
        src = item.get("source")
        tgt = item.get("target")
        if src and isinstance(src, dict): add_obj(src, "source_table")
        if tgt and isinstance(tgt, dict): add_obj(tgt, "target_table")
        # Extra transforms
        for trf in (item.get("transformations") or item.get("mappings") or []):
            add_obj(trf, "transformation")
        # Edges from this record
        if src and tgt:
            sid = f"ew_{(src.get('tableName','') or '').upper().replace(' ','_').replace('.','_')}"
            tid = f"ew_{(tgt.get('tableName','') or '').upper().replace(' ','_').replace('.','_')}"
            if sid and tid and sid != tid:
                edges.append({"from": sid, "to": tid, "type": "lineage"})

    if isinstance(raw, list):
        # Shape C or D — list of records
        for item in raw:
            process_item(item)
    elif isinstance(raw, dict):
        # Shape A — single { source, target }
        if "source" in raw or "target" in raw:
            process_item(raw)

        # Shape B — { sourceObjects: [...], targetObjects: [...] }
        for src in (raw.get("sourceObjects") or raw.get("sources") or
                    raw.get("sourceTable") or []):
            add_obj(src if isinstance(src, dict) else {}, "source_table")
        for tgt in (raw.get("targetObjects") or raw.get("targets") or
                    raw.get("targetTable") or []):
            add_obj(tgt if isinstance(tgt, dict) else {}, "target_table")
        for trf in (raw.get("transformations") or raw.get("mappings") or []):
            add_obj(trf, "transformation")

        # Shape E — top-level single table record
        if not objects and (raw.get("tableName") or raw.get("name")):
            add_obj(raw, _guess_type(raw.get("tableName") or raw.get("name","")))

        # Explicit relationship edges
        for rel in (raw.get("relationships") or raw.get("lineageEdges") or
                    raw.get("edges") or raw.get("lineage") or []):
            src_id = rel.get("sourceId") or rel.get("sourceTableId") or rel.get("from","")
            tgt_id = rel.get("targetId") or rel.get("targetTableId") or rel.get("to","")
            if src_id and tgt_id:
                edges.append({"from": src_id, "to": tgt_id,
                              "type": rel.get("type") or rel.get("relationshipType","flow")})

    # Deduplicate edges
    seen_edges: set = set()
    unique_edges = []
    for e in edges:
        k = f"{e['from']}|{e['to']}"
        if k not in seen_edges:
            seen_edges.add(k)
            unique_edges.append(e)

    # Auto-generate edges if none found from response
    if not unique_edges:
        unique_edges = _infer_edges(objects)

    return {"objects": objects, "edges": unique_edges, "source_file": None}


def _parse_erwin_system_lineage(raw: Any, query: str) -> Dict:
    """
    Parse GET /api/datalineage/system response.
    Returns all tables in a system and the data flows between them.
    ERWIN returns a list of systemObjects, each containing tables
    with their source/target relationships.
    """
    objects, edges = [], []
    seen = set()

    def add_obj(item: Dict, obj_type: str):
        name = (item.get("tableName") or item.get("objectName") or
                item.get("name") or "")
        if not name or name.upper() in seen:
            return
        seen.add(name.upper())
        cols = (item.get("columns") or item.get("fields") or
                item.get("attributes") or [])
        objects.append({
            "id":       f"ew_{name.upper().replace(' ','_')}",
            "name":     name,
            "type":     obj_type,
            "database": item.get("databaseName") or item.get("database") or "",
            "schema":   item.get("schemaName")   or item.get("schema")   or "",
            "system":   (item.get("systemName")  or item.get("applicationName")
                         or query or ""),
            "fields":   [_norm_field(c) for c in cols],
            "logic":    item.get("transformationLogic") or item.get("mappingLogic") or "",
        })

    # ── Handle all known response shapes ──────────────────────────────────────
    data = raw if isinstance(raw, dict) else {}

    # Shape 1: { systemObjects: [ { tables: [...] } ] }
    for sys_obj in (data.get("systemObjects") or data.get("systems") or []):
        if isinstance(sys_obj, dict):
            for tbl in (sys_obj.get("tables") or sys_obj.get("objects") or []):
                if isinstance(tbl, dict):
                    nm = tbl.get("tableName") or tbl.get("name") or ""
                    typ = _guess_type(nm)
                    add_obj(tbl, typ)

    # Shape 2: { sourceObjects: [...], targetObjects: [...] } — same as table lineage
    for src in (data.get("sourceObjects") or data.get("sources") or []):
        if isinstance(src, dict): add_obj(src, "source_table")
    for tgt in (data.get("targetObjects") or data.get("targets") or []):
        if isinstance(tgt, dict): add_obj(tgt, "target_table")
    for trf in (data.get("transformations") or data.get("mappings") or []):
        if isinstance(trf, dict): add_obj(trf, "transformation")

    # Shape 3: flat list of objects  [ { tableName, type, ... }, ... ]
    if not objects and isinstance(raw, list):
        for item in raw:
            if isinstance(item, dict):
                nm = item.get("tableName") or item.get("name") or ""
                add_obj(item, _guess_type(nm))

    # Shape 4: { lineageObjects: [...] }
    if not objects:
        for item in (data.get("lineageObjects") or data.get("items") or []):
            if isinstance(item, dict):
                nm = item.get("tableName") or item.get("name") or ""
                add_obj(item, _guess_type(nm))

    # ── Extract edges ──────────────────────────────────────────────────────────
    for rel in (data.get("relationships") or data.get("lineageEdges") or
                data.get("edges") or data.get("lineage") or []):
        src_id = (rel.get("sourceId") or rel.get("sourceTableId") or
                  rel.get("from") or "")
        tgt_id = (rel.get("targetId") or rel.get("targetTableId") or
                  rel.get("to") or "")
        if src_id and tgt_id:
            edges.append({"from": src_id, "to": tgt_id,
                          "type": rel.get("type") or "flow"})

    if not edges:
        edges = _infer_edges(objects)

    return {"objects": objects, "edges": edges, "source_file": None}


def _parse_erwin_column_lineage(raw: Any, query: str) -> Dict:
    """
    Parse GET /api/datalineage/column response.
    Returns column-level lineage with source/target column mappings.
    """
    objects, edges = [], []
    seen = set()

    def add_from_col(col: Dict, obj_type: str):
        tname = (col.get("tableName") or col.get("parentTable") or
                 col.get("objectName") or "UNKNOWN_TABLE")
        if tname.upper() not in seen:
            seen.add(tname.upper())
            objects.append({
                "id": f"ew_{tname.upper().replace(' ','_')}",
                "name": tname, "type": obj_type,
                "database": col.get("databaseName",""),
                "schema": col.get("schemaName",""),
                "system": col.get("systemName","") or CFG["ERWIN_SYSTEM"] or "",
                "fields": [_norm_field(col)], "logic": "",
            })
        else:
            # Add field to existing object
            for o in objects:
                if o["name"].upper() == tname.upper():
                    existing_names = [f["name"] for f in o["fields"]]
                    fn = col.get("columnName") or col.get("name","")
                    if fn and fn not in existing_names:
                        o["fields"].append(_norm_field(col))

    data = raw if isinstance(raw, dict) else {}
    for col in (data.get("sourceColumns") or data.get("sourceFields") or []):
        add_from_col(col, "source_table")
    for col in (data.get("targetColumns") or data.get("targetFields") or []):
        add_from_col(col, "target_table")
    # Mapping logic
    for m in (data.get("mappings") or data.get("transformations") or []):
        add_from_col({"tableName": m.get("name","TRANSFORM"), "columnName": "",
                      "transformationLogic": m.get("logic","")}, "transformation")
    if not edges:
        edges = _infer_edges(objects)

    return {"objects": objects, "edges": edges, "source_file": None}


def _parse_erwin_objects(raw: Any, query: str) -> Dict:
    """Parse GET /api/objects/search response — generic object search."""
    return _parse_erwin_api(raw, query, "live")


def _parse_erwin_mappings(raw: Any, query: str) -> Dict:
    """
    Parse GET /api/mappings response.
    Returns ETL mapping documents with source/target table info.
    """
    objects, edges = [], []
    items = raw if isinstance(raw, list) else (raw.get("mappings") or raw.get("items") or [])
    seen = set()
    for m in items:
        src_tbl = m.get("sourceTableName") or m.get("sourceName") or ""
        tgt_tbl = m.get("targetTableName") or m.get("targetName") or ""
        trf_name = m.get("mappingName") or m.get("name") or "ETL_TRANSFORM"
        cols = m.get("columns") or m.get("fields") or []
        for name, typ in [(src_tbl,"source_table"), (trf_name,"transformation"), (tgt_tbl,"target_table")]:
            if name and name.upper() not in seen:
                seen.add(name.upper())
                objects.append({
                    "id": f"ew_{name.upper().replace(' ','_')}",
                    "name": name, "type": typ,
                    "database": m.get("targetDatabaseName",""),
                    "schema": m.get("targetSchemaName",""),
                    "system": m.get("systemName","") or CFG["ERWIN_SYSTEM"] or "",
                    "fields": [_norm_field(c) for c in cols] if typ=="source_table" else [],
                    "logic": m.get("transformationLogic","") if typ=="transformation" else "",
                })
    if not edges:
        edges = _infer_edges(objects)
    return {"objects": objects, "edges": edges, "source_file": None}


def _parse_erwin_api(raw: Any, query: str, src: str) -> Dict:
    """
    ERWIN DI returns objects with columns/fields and relationships.
    Handles both v1 and v2 response schemas.
    """
    items = (raw.get("items") or raw.get("objects") or
             raw.get("data", {}).get("lineageObjects") or [])
    objects, edges = [], []

    for item in items:
        cols = (item.get("columns") or item.get("fields") or
                item.get("attributes") or [])
        fields = [_norm_field(f) for f in cols]

        name = item.get("name") or item.get("objectName") or ""
        nm = name.upper()
        obj_type = ("source_table" if any(k in nm for k in ["STG","SRC","RAW","LAND"])
                    else "target_table" if any(k in nm for k in ["DW","FACT","DIM","MART","AGG"])
                    else item.get("objectType","table").lower())
        if obj_type not in ("source_table","target_table","transformation","code_repo"):
            obj_type = "source_table"

        objects.append({
            "id":       item.get("id") or item.get("objectId") or f"ew_{name}",
            "name":     name,
            "type":     obj_type,
            "database": item.get("databaseName") or item.get("database") or "",
            "schema":   item.get("schemaName") or item.get("schema") or "",
            "system":   item.get("systemName") or item.get("applicationName") or "",
            "fields":   fields,
            "logic":    item.get("transformation") or item.get("mappingLogic") or "",
        })

    for rel in (raw.get("relationships") or raw.get("lineage") or raw.get("edges") or []):
        edges.append({
            "from": rel.get("sourceId") or rel.get("from") or "",
            "to":   rel.get("targetId") or rel.get("to") or "",
            "type": rel.get("type") or rel.get("relationshipType") or "flow",
        })

    return {"status": src, "objects": objects, "edges": edges,
            "source_file": None}


def _norm_field(f: Dict) -> Dict:
    if not isinstance(f, dict): return {"name":"","type":"VARCHAR","description":"","nullable":True,"pk":False,"risk":"Low"}
    # ERWIN DI field keys — covers both v1 and v2 and your specific ERWIN response
    name = (f.get("name") or f.get("columnName") or f.get("attributeName") or
            f.get("fieldName") or f.get("physicalColumnName") or "")
    dtype = (f.get("dataType") or f.get("type") or f.get("columnDataType") or
             f.get("datatype") or "VARCHAR")
    desc = (f.get("description") or f.get("businessDescription") or
            f.get("columnDefinition") or f.get("columnDescription") or "")
    # SDI classifications from ERWIN → map to risk
    sdi = f.get("sdiClassificationName") or f.get("sdiClassifications") or ""
    sdi_str = str(sdi).upper() if sdi else ""
    risk = f.get("riskLevel") or f.get("risk") or "Low"
    if any(k in sdi_str for k in ["CONFIDENTIAL","RESTRICTED","SENSITIVE","PII","HIGH"]): risk = "High"
    elif any(k in sdi_str for k in ["INTERNAL","MEDIUM"]): risk = "Medium"
    return {
        "name":        name,
        "type":        dtype,
        "description": desc,
        "nullable":    f.get("nullable", True),
        "pk":          f.get("isPrimaryKey") or f.get("primaryKey") or f.get("isKey") or False,
        "risk":        risk,
        "sdi":         sdi_str,
        "stm_id":      f.get("stmId") or 0,
    }


# ── Excel / JSON local scan ────────────────────────────────────────────────────
def _scan_local_meta(query: str, directory: str) -> Optional[Dict]:
    """
    Scan local directory for Excel (.xlsx/.xls) and JSON files containing
    metadata about the queried table.  Handles:
      - ERWIN export format (objects / columns)
      - Mapping document (source_table, target_table, transform)
      - Generic flat tables with header detection
    """
    table = query.upper().split(".")[0]
    for d in [directory]:
        if not d or not os.path.exists(d):
            continue
        # JSON first
        for fp in glob.glob(os.path.join(d, "**", "*.json"), recursive=True):
            try:
                data = json.loads(Path(fp).read_text(encoding="utf-8", errors="ignore"))
                if table not in json.dumps(data).upper():
                    continue
                objs = _extract_objects_from_json(data, query)
                if objs:
                    edges = _infer_edges(objs)
                    return {"status": "local_json", "objects": objs,
                            "edges": edges, "source_file": fp}
            except Exception:
                pass
        # Excel
        for fp in glob.glob(os.path.join(d, "**", "*.xlsx"), recursive=True):
            r = _parse_excel_meta(fp, query)
            if r:
                r["source_file"] = fp
                return r
        for fp in glob.glob(os.path.join(d, "**", "*.xls"), recursive=True):
            r = _parse_excel_meta(fp, query)
            if r:
                r["source_file"] = fp
                return r
    return None


def _extract_objects_from_json(data: Any, query: str) -> List[Dict]:
    table = query.upper().split(".")[0]
    objects = []
    # ERWIN-style
    items = data if isinstance(data, list) else (
        data.get("objects") or data.get("items") or data.get("tables") or [])
    for item in (items if isinstance(items, list) else []):
        if not isinstance(item, dict): continue
        nm = str(item.get("name") or item.get("tableName") or "")
        if table not in nm.upper(): continue
        cols = item.get("columns") or item.get("fields") or item.get("attributes") or []
        objects.append({
            "id": f"jn_{nm}", "name": nm,
            "type": _guess_type(nm),
            "database": item.get("database") or item.get("databaseName") or "",
            "schema":   item.get("schema") or item.get("schemaName") or "",
            "system":   item.get("system") or item.get("systemName") or "",
            "fields":   [_norm_field(c) for c in cols],
            "logic":    item.get("logic") or item.get("transformation") or "",
        })
    return objects


def _parse_excel_meta(fp: str, query: str) -> Optional[Dict]:
    """
    Parse Excel metadata/mapping document.  Detects header row automatically.
    Supports both:
      - Metadata sheet: TABLE, COLUMN, DATA_TYPE, DESCRIPTION, RISK, OWNER…
      - Mapping sheet:  SOURCE_TABLE, SOURCE_COLUMN, TARGET_TABLE, TARGET_COLUMN, TRANSFORM_LOGIC
    """
    try:
        import openpyxl
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        table = query.upper().split(".")[0]
        all_objects: Dict[str, Dict] = {}
        mapping_edges: List[Dict] = []

        for ws in wb.worksheets:
            rows = [r for r in ws.iter_rows(values_only=True) if any(r)]
            if not rows: continue

            # Find header row (first row containing recognisable keywords)
            hdr_idx, hdr = None, None
            for i, row in enumerate(rows[:8]):
                vals = [str(v).strip().upper() if v else "" for v in row]
                joined = " ".join(vals)
                if any(k in joined for k in ["TABLE","COLUMN","FIELD","SOURCE","TARGET","ATTRIBUTE"]):
                    hdr_idx, hdr = i, vals
                    break
            if hdr is None: continue

            # Map column indices
            def ci(*keys):
                for k in keys:
                    for j, h in enumerate(hdr):
                        if k in h: return j
                return None

            ti   = ci("TABLE_NAME","TABLE","OBJECT_NAME","OBJECT")
            fi   = ci("COLUMN","FIELD","ATTRIBUTE","COLUMN_NAME")
            dti  = ci("DATA_TYPE","DATATYPE","TYPE","DTYPE")
            dsi  = ci("DESCRIPTION","DEFINITION","DESC","COMMENT","MEANING")
            rki  = ci("RISK","RISK_LEVEL","SENSITIVITY")
            owi  = ci("OWNER","DATA_OWNER","STEWARD")
            sri  = ci("SOURCE_TABLE","SRC_TABLE","FROM_TABLE")
            sti  = ci("TARGET_TABLE","TGT_TABLE","TO_TABLE")
            sfi  = ci("SOURCE_COLUMN","SOURCE_FIELD","SRC_COLUMN","SRC_FIELD","FROM_COLUMN","FROM_FIELD")
            tfi  = ci("TARGET_COLUMN","TARGET_FIELD","TGT_COLUMN","TGT_FIELD","TO_COLUMN","TO_FIELD")
            tli  = ci("TRANSFORM","LOGIC","RULE","MAPPING_RULE","BUSINESS_RULE","EXPRESSION")
            syi  = ci("SYSTEM","SOURCE_SYSTEM","APPLICATION","APP")
            dbi  = ci("DATABASE","DB","DATABASE_NAME")
            sci  = ci("SCHEMA","SCHEMA_NAME")

            def cv(row, idx):
                if idx is None or idx >= len(row): return ""
                return str(row[idx]).strip() if row[idx] is not None else ""

            for row in rows[hdr_idx + 1:]:
                if not any(row): continue

                tname = cv(row, ti) or cv(row, sri) or cv(row, sti)
                if not tname: continue
                if table not in tname.upper(): continue

                fname   = cv(row, fi) or cv(row, sfi) or cv(row, tfi)
                dtype   = cv(row, dti) or "VARCHAR"
                desc    = cv(row, dsi)
                risk    = cv(row, rki) or "Low"
                owner   = cv(row, owi)
                src_tbl = cv(row, sri)
                tgt_tbl = cv(row, sti)
                logic   = cv(row, tli)
                system  = cv(row, syi)
                database= cv(row, dbi)
                schema  = cv(row, sci)

                # Build / update object entry
                obj_name = tname
                if obj_name not in all_objects:
                    all_objects[obj_name] = {
                        "id": f"xl_{obj_name.replace(' ','_')}",
                        "name": obj_name,
                        "type": _guess_type(obj_name),
                        "database": database, "schema": schema,
                        "system": system or ws.title,
                        "fields": [], "logic": "",
                    }
                if fname:
                    all_objects[obj_name]["fields"].append({
                        "name": fname, "type": dtype, "description": desc,
                        "nullable": True, "pk": False, "risk": risk,
                        "owner": owner, "logic": logic,
                        "source_table": src_tbl, "target_table": tgt_tbl,
                    })

                # Mapping edge
                if src_tbl and tgt_tbl:
                    mapping_edges.append({"src": src_tbl, "tgt": tgt_tbl})

        if not all_objects: return None

        objs = list(all_objects.values())
        edges = _infer_edges(objs)
        # Add edges from explicit mapping rows
        seen = {f"{e['from']}→{e['to']}" for e in edges}
        for me in mapping_edges:
            src_o = next((o for o in objs if o["name"] == me["src"]), None)
            tgt_o = next((o for o in objs if o["name"] == me["tgt"]), None)
            if src_o and tgt_o:
                k = f"{src_o['id']}→{tgt_o['id']}"
                if k not in seen:
                    seen.add(k)
                    edges.append({"from": src_o["id"], "to": tgt_o["id"], "type": "maps_to"})

        return {"status": "local_excel", "objects": objs, "edges": edges}
    except ImportError:
        return None
    except Exception:
        return None


def _guess_type(name: str) -> str:
    n = name.upper()
    # ERWIN target patterns — includes EDL/DSAS naming from your environment
    if any(k in n for k in ["EDL","DSAS","DW","FACT","DIM","MART","AGG","RPT","ODS","TGT","TARGET",
                              "FNISH_RAW","FNISH_CUR","FNISH_STD"]): return "target_table"
    # Source patterns
    if any(k in n for k in ["STG","SRC","RAW","LAND","STAGE","INBOUND","T_DCBMO","CDS_DCB",
                              "SOURCE","ORIG"]): return "source_table"
    # Transform patterns
    if any(k in n for k in ["ETL","TRANSFORM","PROC","JOB","MAP","INFORMATICA","ABT"]): return "transformation"
    return "source_table"


def _infer_edges(objects: List[Dict]) -> List[Dict]:
    edges = []
    seen: set = set()
    srcs = [o for o in objects if o["type"] == "source_table"]
    trfs = [o for o in objects if o["type"] == "transformation"]
    tgts = [o for o in objects if o["type"] == "target_table"]
    pairs = []
    if srcs and trfs:
        pairs += [(s["id"], t["id"], "reads") for s in srcs for t in trfs]
    if trfs and tgts:
        pairs += [(t["id"], g["id"], "writes") for t in trfs for g in tgts]
    if srcs and tgts and not trfs:
        pairs += [(s["id"], g["id"], "maps_to") for s in srcs for g in tgts]
    for frm, to, typ in pairs:
        k = f"{frm}→{to}"
        if k not in seen:
            seen.add(k)
            edges.append({"from": frm, "to": to, "type": typ})
    return edges


def _mock_erwin(query: str) -> Dict:
    q = query.upper()
    tbl = q.split(".")[0]
    fld = q.split(".")[1] if "." in q else None
    base = [
        {"name":"ID","type":"VARCHAR(50)","description":"Primary business key","nullable":False,"pk":True,"risk":"High"},
        {"name":"CREATED_DATE","type":"TIMESTAMP","description":"Record creation timestamp","nullable":True,"pk":False,"risk":"Medium"},
        {"name":"STATUS","type":"VARCHAR(20)","description":"Processing status: ACTIVE|INACTIVE|PENDING|DELETED","nullable":False,"pk":False,"risk":"High"},
        {"name":"UPDATED_BY","type":"VARCHAR(100)","description":"Username of last modifier","nullable":True,"pk":False,"risk":"Low"},
        {"name":"AMOUNT","type":"DECIMAL(18,4)","description":"Transaction amount","nullable":True,"pk":False,"risk":"High"},
        {"name":"REGION","type":"VARCHAR(50)","description":"Geographic region code","nullable":True,"pk":False,"risk":"Low"},
    ]
    if fld and not any(f["name"]==fld for f in base):
        base.append({"name":fld,"type":"DECIMAL(18,4)","description":f"Business field: {fld}","nullable":True,"pk":False,"risk":"High"})
    sk_fields = [{"name":"SK_"+f["name"],"type":f["type"],"description":f["description"],"nullable":f["nullable"],"pk":f["pk"],"risk":f["risk"]} for f in base]
    objs = [
        {"id":f"src_{tbl}_RAW","name":f"STG_{tbl}_RAW","type":"source_table","database":"STAGING_DB","schema":"STG","system":"Source System","fields":base,"logic":""},
        {"id":f"trf_{tbl}_ETL","name":f"ETL_{tbl}_TRANSFORM","type":"transformation","database":None,"schema":None,"system":"Informatica","fields":[],"logic":f"Normalize STATUS, cast dates, deduplicate on ID, apply SCD Type 2 for {tbl}"},
        {"id":f"tgt_{tbl}_DW","name":f"DW_{tbl}_FACT","type":"target_table","database":"DW_DB","schema":"FACTS","system":"Snowflake","fields":sk_fields,"logic":""},
    ]
    edges = [
        {"from":f"src_{tbl}_RAW","to":f"trf_{tbl}_ETL","type":"reads"},
        {"from":f"trf_{tbl}_ETL","to":f"tgt_{tbl}_DW","type":"writes"},
    ]
    return {"status":"mock","objects":objs,"edges":edges,"source_file":None}


# ═══════════════════════════════════════════════════════════════════════════════
# DATA DICTIONARY  —  Live API  →  Local Excel/JSON  →  Mock
# ═══════════════════════════════════════════════════════════════════════════════
async def fetch_dict(query: str, local_dir: str) -> Dict:
    # 1. Live API
    try:
        async with httpx.AsyncClient(timeout=5.0) as c:
            r = await c.get(f"{CFG['DATA_DICT_URL']}/search", params={"term": query, "q": query})
            if r.status_code == 200:
                return {"status":"live","entries":_parse_dict_api(r.json(), query),"source_file":None}
    except Exception:
        pass

    # 2. Local files
    result = _scan_local_dict(query, local_dir or CFG["LOCAL_DICT_DIR"])
    if result: return result

    # 3. Mock
    return _mock_dict(query)


def _parse_dict_api(raw: Any, query: str) -> List[Dict]:
    items = raw if isinstance(raw, list) else (raw.get("items") or raw.get("terms") or raw.get("results") or [])
    return [_norm_dict_entry(i) for i in items]


def _norm_dict_entry(i: Dict) -> Dict:
    return {
        "table":          i.get("tableName") or i.get("table") or "",
        "field":          i.get("columnName") or i.get("field") or i.get("term") or "",
        "data_type":      i.get("dataType") or i.get("type") or "VARCHAR",
        "description":    i.get("description") or i.get("definition") or i.get("businessDescription") or "",
        "risk_level":     i.get("riskLevel") or i.get("risk") or "Low",
        "impact":         i.get("impact") or i.get("impactAnalysis") or "",
        "owner":          i.get("owner") or i.get("dataOwner") or i.get("steward") or "",
        "tags":           i.get("tags") or i.get("categories") or [],
        "last_updated":   i.get("lastUpdated") or i.get("modifiedDate") or "",
        "business_rules": i.get("businessRules") or i.get("rules") or [],
    }


def _scan_local_dict(query: str, directory: str) -> Optional[Dict]:
    table = query.upper().split(".")[0]
    entries = []
    src_file = None
    for d in [directory]:
        if not d or not os.path.exists(d): continue
        for fp in glob.glob(os.path.join(d,"**","*.json"), recursive=True):
            try:
                data = json.loads(Path(fp).read_text(encoding="utf-8", errors="ignore"))
                if table not in json.dumps(data).upper(): continue
                items = data if isinstance(data,list) else (data.get("dictionary") or data.get("terms") or data.get("entries") or [])
                for item in (items if isinstance(items,list) else []):
                    if not isinstance(item,dict): continue
                    t = str(item.get("table") or item.get("tableName") or "")
                    f = str(item.get("field") or item.get("column") or item.get("term") or "")
                    if table in t.upper() or table in f.upper():
                        entries.append(_norm_dict_entry(item))
                if entries:
                    src_file = fp; break
            except Exception: pass
        if not entries:
            for fp in glob.glob(os.path.join(d,"**","*.xlsx"), recursive=True):
                res = _parse_excel_dict(fp, query)
                if res: entries.extend(res); src_file = fp; break
        if entries: break
    return {"status":"local_file","entries":entries,"source_file":src_file} if entries else None


def _parse_excel_dict(fp: str, query: str) -> List[Dict]:
    entries = []
    try:
        import openpyxl
        wb = openpyxl.load_workbook(fp, read_only=True, data_only=True)
        table = query.upper().split(".")[0]
        for ws in wb.worksheets:
            rows = [r for r in ws.iter_rows(values_only=True) if any(r)]
            if not rows: continue
            hdr = None
            for i, row in enumerate(rows[:6]):
                vals = [str(v).strip().upper() if v else "" for v in row]
                if any(k in " ".join(vals) for k in ["FIELD","COLUMN","TERM","ATTRIBUTE","DEFINITION"]):
                    hdr = {v:j for j,v in enumerate(vals) if v}
                    rows = rows[i+1:]
                    break
            if not hdr: continue
            def gi(*ks):
                for k in ks:
                    for hk,hv in hdr.items():
                        if k in hk: return hv
                return None
            ti=gi("TABLE"); fi=gi("FIELD","COLUMN","TERM","ATTRIBUTE"); dti=gi("DATA_TYPE","TYPE")
            dsi=gi("DESCRIPTION","DEFINITION","MEANING"); ri=gi("RISK"); ii=gi("IMPACT"); oi=gi("OWNER","STEWARD")
            for row in rows:
                if not any(row): continue
                def cv(idx): return str(row[idx]).strip() if idx is not None and idx<len(row) and row[idx] else ""
                tname=cv(ti); fname=cv(fi)
                if not fname: continue
                if table not in tname.upper() and table not in fname.upper(): continue
                entries.append({"table":tname,"field":fname,"data_type":cv(dti) or "VARCHAR",
                    "description":cv(dsi),"risk_level":cv(ri) or "Low","impact":cv(ii),
                    "owner":cv(oi),"tags":[],"last_updated":"","business_rules":[]})
    except Exception: pass
    return entries


def _mock_dict(query: str) -> Dict:
    tbl = query.upper().split(".")[0]
    fld = query.upper().split(".")[1] if "." in query.upper() else None
    entries = [
        {"table":f"STG_{tbl}_RAW","field":"ID","data_type":"VARCHAR(50)","description":"Unique surrogate key — must not be NULL or duplicate","risk_level":"High","impact":"All downstream joins and reports depend on this field","owner":"Data Governance","tags":["pk","certified","critical"],"last_updated":"2025-04-01","business_rules":["NOT NULL","UNIQUE","Must match LKP_ENTITY_ID"]},
        {"table":f"STG_{tbl}_RAW","field":"STATUS","data_type":"VARCHAR(20)","description":"Entity lifecycle status. Values: ACTIVE, INACTIVE, PENDING, DELETED","risk_level":"High","impact":"Controls workflow routing and all reporting segment filters","owner":"Business Analytics","tags":["operational","certified","filtered"],"last_updated":"2025-04-01","business_rules":["IN ('ACTIVE','INACTIVE','PENDING','DELETED')","NOT NULL"]},
        {"table":f"STG_{tbl}_RAW","field":"CREATED_DATE","data_type":"TIMESTAMP","description":"UTC timestamp when record was first inserted","risk_level":"Medium","impact":"Used in SLA metrics and data freshness checks","owner":"Data Engineering","tags":["audit","temporal"],"last_updated":"2025-03-15","business_rules":["NOT NULL","Must be > 2000-01-01","UTC timezone"]},
        {"table":f"STG_{tbl}_RAW","field":"AMOUNT","data_type":"DECIMAL(18,4)","description":"Transaction value in base currency (USD)","risk_level":"High","impact":"Financial reporting P&L — regulatory requirement","owner":"Finance Data Team","tags":["financial","kpi","certified"],"last_updated":"2025-04-10","business_rules":["NOT NULL","Must be >= 0","Precision: 18,4","Currency: USD"]},
        {"table":f"STG_{tbl}_RAW","field":"REGION","data_type":"VARCHAR(50)","description":"Geographic region code from master reference table","risk_level":"Low","impact":"Regional dashboard segmentation","owner":"MDM Team","tags":["geography","reference"],"last_updated":"2025-02-20","business_rules":["FK → REF_REGION(CODE)","Valid ISO codes only"]},
        {"table":f"STG_{tbl}_RAW","field":"UPDATED_BY","data_type":"VARCHAR(100)","description":"Username or SYSTEM_ID of the last modifier","risk_level":"Low","impact":"Audit trail only — no business logic dependency","owner":"Data Governance","tags":["audit"],"last_updated":"2025-01-10","business_rules":["Default: 'SYSTEM'"]},
    ]
    if fld and not any(e["field"]==fld for e in entries):
        entries.append({"table":f"STG_{tbl}_RAW","field":fld,"data_type":"DECIMAL(18,4)","description":f"Core KPI metric: {fld} — used in financial reporting and executive dashboards","risk_level":"High","impact":"Executive dashboards, regulatory filings, P&L reports","owner":"Finance Data Team","tags":["financial","kpi","certified","critical"],"last_updated":"2025-04-15","business_rules":["NOT NULL","Must be >= 0","Audited field — change requires JIRA ticket"]})
    return {"status":"mock","entries":entries,"source_file":None}


# ═══════════════════════════════════════════════════════════════════════════════
# CODE REPOSITORY  —  GitHub API  →  Local folder  →  Mock
# ═══════════════════════════════════════════════════════════════════════════════
async def fetch_code(query: str, gh_token: str, gh_repo: str, local_dir: str) -> List[Dict]:
    tbl = query.upper().split(".")[0]
    results = []

    # 1. GitHub Code Search
    token = gh_token or CFG["GITHUB_TOKEN"]
    repo  = gh_repo  or CFG["GITHUB_REPO"]
    if token and repo:
        try:
            hdrs = {"Authorization":f"token {token}","Accept":"application/vnd.github.v3+json"}
            async with httpx.AsyncClient(timeout=15.0) as c:
                for ext in ["sql","py","hql","java","scala","ksh"]:
                    url = f"https://api.github.com/search/code?q={tbl}+in:file+repo:{repo}+extension:{ext}"
                    r = await c.get(url, headers=hdrs)
                    if r.status_code != 200: continue
                    for item in r.json().get("items",[])[:2]:
                        fr = await c.get(item["url"], headers=hdrs)
                        if fr.status_code != 200: continue
                        content = base64.b64decode(fr.json().get("content","")).decode("utf-8","ignore")
                        logic = _best_line(content, tbl)
                        if logic:
                            results.append({
                                "file":item["name"],"path":item["path"],"url":item["html_url"],
                                "language":LANG_MAP.get("."+ext,ext.upper()),
                                "business_logic":logic,
                                "context":_context(content,tbl),
                                "source":"github",
                            })
        except Exception:
            pass

    # 2. Local folder
    if not results:
        dirs = [d for d in [local_dir, CFG["LOCAL_CODE_DIR"]] if d and os.path.exists(d)]
        for d in dirs:
            for ext in CODE_EXTS:
                for fp in glob.glob(os.path.join(d,"**",f"*{ext}"),recursive=True):
                    try:
                        content = Path(fp).read_text(encoding="utf-8",errors="ignore")
                        if tbl not in content.upper(): continue
                        logic = _best_line(content, tbl)
                        if logic:
                            results.append({
                                "file":os.path.basename(fp),"path":fp,"url":None,
                                "language":LANG_MAP.get(ext,ext[1:].upper()),
                                "business_logic":logic,
                                "context":_context(content,tbl),
                                "source":"local",
                            })
                    except Exception: pass
            if results: break

    # 3. Mock
    if not results:
        return _mock_code(query)

    return results


def _best_line(content: str, term: str) -> Optional[str]:
    best, best_score = None, 0
    for line in content.splitlines():
        if term not in line.upper(): continue
        clean = line.strip()
        if not clean or clean.startswith(("#","--","//","/*","*")): continue
        score = len(clean)
        kws = ["SELECT","INSERT","UPDATE","DELETE","MERGE","JOIN","FROM","WHERE","CREATE",
               "WITH","CASE","COALESCE","CAST","TRANSFORM","LOAD","WRITE","MAP","REDUCE"]
        for kw in kws:
            if kw in clean.upper(): score += 150
        if score > best_score: best_score, best = score, clean[:400]
    return best


def _context(content: str, term: str, ctx: int = 6) -> str:
    lines = content.splitlines()
    chunks, seen = [], set()
    for i, line in enumerate(lines):
        if term not in line.upper(): continue
        s, e = max(0,i-ctx), min(len(lines),i+ctx+1)
        chunk = "\n".join(lines[s:e])
        if chunk not in seen:
            seen.add(chunk)
            chunks.append(chunk)
        if len(chunks) >= 3: break
    return "\n---\n".join(chunks)


def _mock_code(query: str) -> List[Dict]:
    t = query.upper().split(".")[0]
    fld = query.upper().split(".")[1] if "." in query.upper() else "AMOUNT"
    return [
        {"file":f"stg_load_{t.lower()}.sql","path":f"etl/staging/{t.lower()}/stg_load.sql","url":None,"language":"SQL",
         "business_logic":f"INSERT /*+ APPEND */ INTO STG_{t}_RAW (ID,STATUS,CREATED_DATE,{fld},REGION)\nSELECT src_id, UPPER(TRIM(status)), SYSDATE, NVL({fld.lower()},0), region_cd\nFROM SRC_{t}\nWHERE status != 'DELETED' AND load_dt >= TRUNC(SYSDATE)-1",
         "context":f"-- Staging load for {t}\n-- Schedule: Daily 02:00 UTC\n-- Source: Oracle SRC_{t}\nINSERT /*+ APPEND */ INTO STG_{t}_RAW (ID,STATUS,CREATED_DATE,{fld},REGION)\nSELECT src_id, UPPER(TRIM(status)), SYSDATE, NVL({fld.lower()},0), region_cd\nFROM SRC_{t} WHERE status != 'DELETED'\nAND load_dt >= TRUNC(SYSDATE)-1\n-- Rows expected: ~500K/day","source":"mock"},
        {"file":f"transform_{t.lower()}.py","path":f"pipelines/transform_{t.lower()}.py","url":None,"language":"PySpark",
         "business_logic":f"df = spark.table('STG_{t}_RAW')\\\n  .withColumn('STATUS_CD', F.when(F.col('STATUS')=='ACTIVE','A').otherwise('I'))\\\n  .withColumn('DT', F.to_date('CREATED_DATE'))\\\n  .withColumn('{fld}_USD', F.col('{fld.lower()}')*F.lit(1.0))\\\n  .dropDuplicates(['ID'])",
         "context":f"# PySpark SCD Type 2 transform — {t}\nfrom pyspark.sql import functions as F\ndf = spark.table('STG_{t}_RAW')\\\n  .withColumn('STATUS_CD', F.when(F.col('STATUS')=='ACTIVE','A').otherwise('I'))\\\n  .withColumn('DT', F.to_date('CREATED_DATE'))\\\n  .withColumn('{fld}_USD', F.col('{fld.lower()}')*F.lit(1.0))\\\n  .dropDuplicates(['ID'])\n# Merge into DW\ndf.write.format('delta').mode('merge').save('/mnt/dw/{t.lower()}')","source":"mock"},
        {"file":f"DW{t}Load.java","path":f"src/main/java/etl/dw/DW{t}Load.java","url":None,"language":"Java",
         "business_logic":f"jdbcTemplate.execute(\"MERGE INTO DW_{t}_FACT T USING STG_{t}_RAW S ON(T.ID=S.ID) WHEN MATCHED THEN UPDATE SET T.STATUS_CD=UPPER(S.STATUS),T.{fld}=S.{fld.lower()} WHEN NOT MATCHED THEN INSERT(ID,STATUS_CD,CREATED_DT,{fld}) VALUES(S.src_id,UPPER(S.status),SYSDATE,NVL(S.{fld.lower()},0))\");",
         "context":f"// Java JDBC merge DW load — {t}\n// @Scheduled(cron = '0 3 * * *')\npublic void loadDW{t}() {{\n  jdbcTemplate.execute(\"MERGE INTO DW_{t}_FACT T USING STG_{t}_RAW S ON(T.ID=S.ID)\"\n    + \" WHEN MATCHED THEN UPDATE SET T.STATUS_CD=UPPER(S.STATUS),T.{fld}=S.{fld.lower()}\"\n    + \" WHEN NOT MATCHED THEN INSERT VALUES(S.src_id,UPPER(S.status),SYSDATE,NVL(S.{fld.lower()},0))\");\n}}","source":"mock"},
    ]


# ═══════════════════════════════════════════════════════════════════════════════
# TACHYON LLM
# ═══════════════════════════════════════════════════════════════════════════════
async def tachyon_call(prompt: str, tachyon_url: str, tachyon_key: str) -> Optional[str]:
    base = (tachyon_url or CFG["TACHYON_URL"] or "").rstrip("/")
    if not base: return None
    key  = tachyon_key or CFG["TACHYON_API_KEY"]
    hdrs = {"Content-Type":"application/json"}
    if key: hdrs["Authorization"] = f"Bearer {key}"
    endpoints = [base, f"{base}/chat/completions", f"{base}/completions", f"{base}/generate", f"{base}/chat"]
    payloads  = [
        {"model":CFG["TACHYON_MODEL"],"messages":[{"role":"user","content":prompt}],"max_tokens":400,"temperature":0.3},
        {"prompt":prompt,"max_tokens":400,"temperature":0.3},
        {"query":prompt,"model":CFG["TACHYON_MODEL"],"max_length":400},
    ]
    for ep in endpoints:
        for pl in payloads:
            try:
                async with httpx.AsyncClient(timeout=10.0) as c:
                    r = await c.post(ep, headers=hdrs, json=pl)
                    if r.status_code == 200:
                        d = r.json()
                        text = (d.get("choices",[{}])[0].get("message",{}).get("content")
                                or d.get("choices",[{}])[0].get("text")
                                or d.get("response") or d.get("output") or d.get("text")
                                or d.get("generated_text"))
                        if text and len(str(text).strip()) > 10:
                            return str(text).strip()
            except Exception: pass
    return None


async def tachyon_summarise_code(query: str, code: List[Dict], url: str, key: str) -> Optional[str]:
    if not code: return None
    snip = "\n\n".join(f"[{c['language']}] {c['file']}:\n{c.get('context',c['business_logic'])[:500]}" for c in code[:3])
    prompt = f"""You are a senior data engineer.
Analyse the following ETL code for the entity '{query}' and provide a JSON response:
{{
  "summary": "one sentence describing what this ETL does",
  "key_rules": ["rule1","rule2","rule3"],
  "risk": "one sentence on what could break if this changes",
  "quality_score": 7
}}
Only respond with the JSON object — no prose, no markdown fences.

Code:
{snip}"""
    resp = await tachyon_call(prompt, url, key)
    if not resp: return None
    try:
        j = json.loads(resp.strip().lstrip("```json").rstrip("```").strip())
        parts = [f"📌 {j.get('summary','')}"]
        if j.get("key_rules"): parts.append("Rules: " + " · ".join(j["key_rules"]))
        if j.get("risk"): parts.append(f"⚠️ {j['risk']}")
        return " | ".join(parts)
    except Exception:
        return resp[:300]


async def tachyon_insights(query: str, nodes: List[Dict], meta: List[Dict], url: str, key: str) -> Optional[str]:
    high = sum(1 for m in meta if m.get("risk_level")=="High")
    systems = list(set(n.get("system","") for n in nodes if n.get("system")))
    prompt = f"""You are a Chief Data Officer providing a governance briefing.
Entity: {query}
Systems involved: {systems}
Total fields: {len(meta)} ({high} HIGH risk)
Top fields: {[m['field_name'] for m in meta[:5]]}

Write 2 sentences ONLY:
1. What this data entity represents in the enterprise
2. The single most important governance action needed

Be specific and actionable. No headings, no bullets."""
    return await tachyon_call(prompt, url, key)


# ═══════════════════════════════════════════════════════════════════════════════
# GRAPH BUILDER
# ═══════════════════════════════════════════════════════════════════════════════
def build_graph(query: str, erwin: Dict, dd: Dict, code: List[Dict]):
    tbl  = query.upper().split(".")[0]
    fld  = query.upper().split(".")[1] if "." in query.upper() else None
    dd_map = {e["field"].upper(): e for e in dd.get("entries",[])}

    nodes, edges = [], []
    for obj in erwin.get("objects",[]):
        risk = "Low"
        for f in obj.get("fields",[]):
            de = dd_map.get(f["name"].upper(),{})
            r = de.get("risk_level", f.get("risk","Low"))
            if r=="High": risk="High"; break
            if r=="Medium" and risk=="Low": risk="Medium"
        nodes.append({
            "id":obj["id"],"label":obj["name"],"type":obj["type"],
            "system":obj.get("system",""),"database":obj.get("database",""),
            "schema_name":obj.get("schema",""),"description":obj.get("logic","") or obj["name"],
            "risk_level":risk,"source":erwin.get("status","erwin"),"fields":obj.get("fields",[]),
        })

    for e in erwin.get("edges",[]):
        edges.append({"id":f"e_{e['from']}_{e['to']}","source":e["from"],"target":e["to"],"label":e.get("type","flow"),"transformation":None})

    if code:
        cid = f"code_{tbl}"
        nodes.append({"id":cid,"label":f"Code Repo\n({len(code)} scripts)","type":"code_repo",
                       "system":"Code Repository","database":None,"schema_name":None,
                       "description":f"{len(code)} scripts","risk_level":"Low",
                       "source":code[0].get("source","local"),"code_results":code,"fields":[]})
        trf = f"trf_{tbl}_ETL"
        if any(n["id"]==trf for n in nodes):
            edges.append({"id":f"e_{cid}_{trf}","source":cid,"target":trf,"label":"implements","transformation":code[0]["business_logic"]})

    if fld:
        fid = f"field_{tbl}_{fld}"
        fm  = dd_map.get(fld.upper(),{})
        nodes.append({"id":fid,"label":f"{tbl}.{fld}","type":"field",
                       "system":"DW","field":fld,"table":tbl,
                       "description":fm.get("description",""),"risk_level":fm.get("risk_level","Medium"),
                       "data_type":fm.get("data_type","VARCHAR"),"source":"data_dict","fields":[]})
        tgt = f"tgt_{tbl}_DW"
        if any(n["id"]==tgt for n in nodes):
            edges.append({"id":f"e_{tgt}_{fid}","source":tgt,"target":fid,"label":"contains","transformation":None})

    return nodes, edges


def build_metadata(query: str, erwin: Dict, dd: Dict, code: List[Dict]) -> List[Dict]:
    dd_map = {e["field"].upper(): e for e in dd.get("entries",[])}
    seen   : set = set()
    meta   = []
    for obj in erwin.get("objects",[]):
        for f in obj.get("fields",[]):
            nm = f["name"].upper()
            if nm in seen: continue
            seen.add(nm)
            de = dd_map.get(nm,{})
            code_logic = next((c["business_logic"] for c in code if nm in c.get("business_logic","").upper()), None)
            ai_summary = next((c.get("ai_summary","") for c in code if c.get("ai_summary") and nm in c.get("ai_summary","").upper()), None)
            src_obj  = next((o for o in erwin.get("objects",[]) if o["type"]=="source_table"), None)
            tgt_obj  = next((o for o in erwin.get("objects",[]) if o["type"]=="target_table"), None)
            meta.append({
                "field_name":    f["name"],
                "table_name":    obj["name"],
                "data_type":     de.get("data_type", f.get("type","VARCHAR")),
                "description":   de.get("description", f.get("description","")),
                "business_logic":ai_summary or code_logic or (de.get("business_rules",[""])[0] if de.get("business_rules") else ""),
                "sources":       [src_obj["name"]] if src_obj else [],
                "targets":       [tgt_obj["name"]] if tgt_obj else [],
                "risk_level":    de.get("risk_level", f.get("risk","Low")),
                "impact":        de.get("impact",""),
                "owner":         de.get("owner", f.get("owner","")),
                "last_updated":  de.get("last_updated",""),
                "tags":          de.get("tags",[]),
                "is_pk":         f.get("pk",False),
                "is_nullable":   f.get("nullable",True),
                "business_rules":de.get("business_rules",[]),
                "source_system": obj.get("system",""),
            })
    return meta


# ═══════════════════════════════════════════════════════════════════════════════
# MAIN SEARCH ENDPOINT
# ═══════════════════════════════════════════════════════════════════════════════
@app.post("/api/search")
async def search(req: SearchRequest):
    t0 = datetime.now()
    sources_used, source_details = [], {}

    # Fan-out all source fetches concurrently
    erwin_task = fetch_erwin(
        req.query,
        req.local_meta_dir or "",
        erwin_base_url=req.erwin_base_url,
        erwin_api_key=req.erwin_api_key,
        erwin_system=req.erwin_system,
        erwin_env=req.erwin_env,
        erwin_project=req.erwin_project,
        erwin_lineage_type=req.erwin_lineage_type,
        erwin_ep_tablename=getattr(req,'erwin_ep_tablename',None),
        erwin_ep_colname=getattr(req,'erwin_ep_colname',None),
        erwin_ep_objtype=getattr(req,'erwin_ep_objtype',None),
        erwin_ep_mapname=getattr(req,'erwin_ep_mapname',None),
        erwin_ep_glosscat=getattr(req,'erwin_ep_glosscat',None),
    ) if req.use_erwin else None
    dict_task  = fetch_dict(req.query, req.local_dict_dir or "")  if req.use_data_dict else None
    code_task  = fetch_code(req.query,
                            req.github_token or "", req.github_repo or "",
                            req.local_code_dir or "")             if req.use_code_repo else None

    tasks  = [t for t in [erwin_task, dict_task, code_task] if t is not None]
    labels = ([("erwin",)  if req.use_erwin   else []] +
              [("dict",)   if req.use_data_dict else []] +
              [("code",)   if req.use_code_repo  else []])
    labels = [l[0] for l in labels if l]

    results = await asyncio.gather(*tasks, return_exceptions=True)
    erwin_data = {} ; dict_data = {} ; code_results = []

    for label, res in zip(labels, results):
        if isinstance(res, Exception): continue
        if label == "erwin":
            erwin_data = res
            stat = res.get("status","mock")
            tag  = {"live":"🟢 LIVE","local_excel":"📂 LOCAL EXCEL","local_json":"📂 LOCAL JSON","mock":"⚪ MOCK"}
            sources_used.append(f"ERWIN DI [{tag.get(stat,stat.upper())}]")
            source_details["erwin"] = {"status":stat,"objects":len(res.get("objects",[])),"source_file":res.get("source_file")}
        elif label == "dict":
            dict_data = res
            stat = res.get("status","mock")
            sources_used.append(f"Data Dictionary [{tag.get(stat,stat.upper())}]")
            source_details["dict"] = {"status":stat,"entries":len(res.get("entries",[])),"source_file":res.get("source_file")}
        elif label == "code":
            code_results = res
            src = res[0].get("source","local") if res else "mock"
            sources_used.append(f"Code Repo [{'🐙 GITHUB' if src=='github' else '📁 LOCAL' if src=='local' else '⚪ MOCK'}]")
            source_details["code"] = {"status":src,"files":len(res)}

    # Tachyon AI enrichment
    ai_insights = None
    t_url = req.tachyon_url or CFG["TACHYON_URL"]
    t_key = req.tachyon_key or CFG["TACHYON_API_KEY"]
    if req.use_tachyon and t_url:
        ai_sum = await tachyon_summarise_code(req.query, code_results, t_url, t_key)
        if ai_sum and code_results:
            code_results[0]["ai_summary"] = ai_sum

    nodes, edges = build_graph(req.query, erwin_data, dict_data, code_results)
    metadata     = build_metadata(req.query, erwin_data, dict_data, code_results)
    # NEW v4.1: enrich metadata with confidentiality, SOR, equivalent fields, multi-hop, ingestion pattern
    metadata     = _enrich_metadata_v41(metadata, nodes, edges, code_results)

    if req.use_tachyon and t_url and nodes:
        ai_insights = await tachyon_insights(req.query, nodes, metadata, t_url, t_key)
    if ai_insights:
        sources_used.append("Tachyon AI [🤖 LIVE]")
        source_details["tachyon"] = {"status":"live"}

    ms = int((datetime.now()-t0).total_seconds()*1000)
    return {
        "query":req.query, "nodes":nodes, "edges":edges,
        "metadata":metadata, "sources_used":sources_used,
        "source_details":source_details, "search_time_ms":ms,
        "ai_insights":ai_insights,
        "summary":f"{len(nodes)} objects · {len(edges)} edges · {len(metadata)} fields · {len(sources_used)} sources",
    }


# ═══════════════════════════════════════════════════════════════════════════════
# EXCEL EXPORT  —  4-sheet color-coded workbook
# ═══════════════════════════════════════════════════════════════════════════════
@app.post("/api/export/excel")
async def export_excel(payload: Dict[str,Any]):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter

        q        = payload.get("query","export")
        metadata = payload.get("metadata",[])
        nodes    = payload.get("nodes",[])
        edges    = payload.get("edges",[])

        # Palette
        BG_DARK  = "050C18"; BG_MID   = "0B1730"; BG_ALT   = "0F1E35"
        HDR_COL  = "818CF8"; HDR_BG   = "0F1E35"; TITLE_COL= "38BDF8"
        CELL_COL = "CBD5E1"; ALT_BG   = "091425"
        RISK_BG  = {"High":"3D0A0F","Medium":"2D1E06","Low":"061A10"}
        RISK_COL = {"High":"F43F5E","Medium":"F59E0B","Low":"10B981"}
        thin = Border(
            left=Side(style="thin",color="1A2F52"), right=Side(style="thin",color="1A2F52"),
            top=Side(style="thin",color="1A2F52"),  bottom=Side(style="thin",color="1A2F52"))

        def hdr_font(): return Font(name="Calibri",bold=True,color=HDR_COL,size=10)
        def cell_font(risk=None): return Font(name="Calibri",color=RISK_COL.get(risk,CELL_COL),size=9,bold=bool(risk))
        def hdr_fill(): return PatternFill("solid",fgColor=HDR_BG)
        def row_fill(i,risk=None):
            if risk: return PatternFill("solid",fgColor=RISK_BG[risk])
            return PatternFill("solid",fgColor=ALT_BG if i%2==0 else BG_MID)
        def title_style(cell,text):
            cell.value=text; cell.font=Font(name="Calibri",bold=True,color=TITLE_COL,size=13)
            cell.fill=PatternFill("solid",fgColor=BG_DARK)
            cell.alignment=Alignment(horizontal="center",vertical="center"); cell.border=thin
        def sh(cell,text):
            cell.value=text; cell.font=hdr_font(); cell.fill=hdr_fill()
            cell.alignment=Alignment(horizontal="center",vertical="center",wrap_text=True); cell.border=thin
        def sc(cell,val,i=0,risk=None):
            cell.value=val; cell.font=cell_font(risk); cell.fill=row_fill(i,risk)
            cell.alignment=Alignment(vertical="center",wrap_text=True); cell.border=thin

        wb = openpyxl.Workbook()

        # ── Sheet 1 Metadata ──────────────────────────────────────────────────
        ws1 = wb.active; ws1.title = "📋 Metadata"
        ws1.sheet_view.showGridLines = False; ws1.freeze_panes = "A3"
        ws1.row_dimensions[1].height = 28; ws1.row_dimensions[2].height = 20

        MCOLS = get_column_letter
        hdrs1 = ["Field","Table","Type","PK","Nullable","Risk","Description",
                 "Business Logic","Source System","Sources","Targets","Owner","Updated","Tags","Rules"]
        wids1 = [18,22,13,6,8,9,42,55,18,22,22,18,12,18,50]
        ws1.merge_cells(f"A1:{MCOLS(len(hdrs1))}1")
        title_style(ws1["A1"], f"DataLineage IQ — Metadata: {q.upper()}  ·  {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        for j,(h,w) in enumerate(zip(hdrs1,wids1),1):
            sh(ws1.cell(2,j), h)
            ws1.column_dimensions[MCOLS(j)].width = w

        for i,m in enumerate(metadata,1):
            risk = m.get("risk_level","Low")
            ws1.row_dimensions[i+2].height = 16
            vals = [m.get("field_name",""), m.get("table_name",""), m.get("data_type",""),
                    "✓" if m.get("is_pk") else "", "✓" if m.get("is_nullable") else "",
                    risk, m.get("description",""), m.get("business_logic",""),
                    m.get("source_system",""), ", ".join(m.get("sources",[])),
                    ", ".join(m.get("targets",[])), m.get("owner",""), m.get("last_updated",""),
                    ", ".join(m.get("tags",[])), "; ".join(m.get("business_rules",[]))]
            for j,v in enumerate(vals,1):
                sc(ws1.cell(i+2,j), v, i, risk if j==6 else None)

        # ── Sheet 2 Mapping Document ──────────────────────────────────────────
        ws2 = wb.create_sheet("🗺️ Mapping")
        ws2.sheet_view.showGridLines = False; ws2.freeze_panes = "A3"
        ws2.row_dimensions[1].height = 28; ws2.row_dimensions[2].height = 20
        hdrs2 = ["Map ID","Source Table","Source Field","Src Type",
                 "Transformation Logic","Target Table","Target Field","Tgt Type","Risk","Impact","Owner"]
        wids2 = [10,22,18,12,55,22,18,12,9,35,18]
        ws2.merge_cells(f"A1:{MCOLS(len(hdrs2))}1")
        title_style(ws2["A1"], f"DataLineage IQ — Mapping Document: {q.upper()}")
        for j,(h,w) in enumerate(zip(hdrs2,wids2),1):
            sh(ws2.cell(2,j), h); ws2.column_dimensions[MCOLS(j)].width = w
        src_nodes = [n for n in nodes if n.get("type")=="source_table"]
        tgt_nodes = [n for n in nodes if n.get("type")=="target_table"]
        for i,m in enumerate(metadata,1):
            risk=m.get("risk_level","Low"); ws2.row_dimensions[i+2].height = 16
            src=m.get("sources",["—"])[0] if m.get("sources") else "—"
            tgt=m.get("targets",["—"])[0] if m.get("targets") else "—"
            sn=next((n for n in src_nodes if n["label"]==src),{})
            tn=next((n for n in tgt_nodes if n["label"]==tgt),{})
            sf=next((f for f in sn.get("fields",[]) if f["name"]==m["field_name"]),{})
            tf=next((f for f in tn.get("fields",[]) if m["field_name"] in f["name"]),{})
            vals=[f"MAP-{i:04d}", src, m["field_name"], sf.get("type",m.get("data_type","")),
                  m.get("business_logic",""), tgt, tf.get("name",m["field_name"]),
                  tf.get("type",""), risk, m.get("impact",""), m.get("owner","")]
            for j,v in enumerate(vals,1):
                sc(ws2.cell(i+2,j), v, i, risk if j==9 else None)

        # ── Sheet 3 Lineage Graph ─────────────────────────────────────────────
        ws3 = wb.create_sheet("🕸️ Lineage Graph")
        ws3.sheet_view.showGridLines = False
        ws3.merge_cells(f"A1:F1"); ws3.row_dimensions[1].height = 28; ws3.row_dimensions[2].height = 20
        title_style(ws3["A1"], f"DataLineage IQ — Lineage Graph: {q.upper()}")
        for j,(h,w) in enumerate(zip(["Node ID","Name","Type","System","Database","Risk"],
                                     [20,26,18,22,16,9]),1):
            sh(ws3.cell(2,j),h); ws3.column_dimensions[MCOLS(j)].width=w
        for i,n in enumerate(nodes,1):
            ws3.row_dimensions[i+2].height=15; risk=n.get("risk_level","Low")
            for j,v in enumerate([n.get("id",""),n.get("label","").replace("\n"," "),n.get("type",""),n.get("system",""),n.get("database",""),risk],1):
                sc(ws3.cell(i+2,j),v,i,risk if j==6 else None)
        base_row=len(nodes)+4
        ws3.cell(base_row,1).value="EDGES"; ws3.cell(base_row,1).font=hdr_font()
        for j,(h,w) in enumerate(zip(["Edge ID","Source","Target","Label","Transform Logic"],
                                     [18,22,22,12,55]),1):
            sh(ws3.cell(base_row+1,j),h); ws3.column_dimensions[MCOLS(j)].width=w
        for i,e in enumerate(edges,1):
            for j,v in enumerate([e.get("id",""),e.get("source",""),e.get("target",""),e.get("label",""),e.get("transformation","") or ""],1):
                sc(ws3.cell(base_row+1+i,j),v,i)

        # ── Sheet 4 Summary ───────────────────────────────────────────────────
        ws4 = wb.create_sheet("📊 Summary")
        ws4.sheet_view.showGridLines = False
        ws4.column_dimensions["A"].width=28; ws4.column_dimensions["B"].width=42
        ws4.merge_cells("A1:B1"); ws4.row_dimensions[1].height=28
        title_style(ws4["A1"],"DataLineage IQ — Executive Summary")
        rows4=[("Query",q),("Generated",datetime.now().strftime("%Y-%m-%d %H:%M UTC")),
               ("Total Objects",len(nodes)),("Total Relationships",len(edges)),
               ("Total Fields",len(metadata)),
               ("High Risk Fields",sum(1 for m in metadata if m.get("risk_level")=="High")),
               ("Medium Risk Fields",sum(1 for m in metadata if m.get("risk_level")=="Medium")),
               ("Low Risk Fields",sum(1 for m in metadata if m.get("risk_level")=="Low")),
               ("Source Systems",len(set(n.get("system","") for n in nodes if n.get("type")=="source_table"))),
               ("Target Systems",len(set(n.get("system","") for n in nodes if n.get("type")=="target_table"))),
               ("AI Insights",payload.get("ai_insights","N/A") or "N/A")]
        for i,(k,v) in enumerate(rows4,2):
            ws4.row_dimensions[i].height=18
            kc=ws4.cell(i,1); kc.value=k; kc.font=hdr_font(); kc.fill=hdr_fill(); kc.border=thin
            vc=ws4.cell(i,2); vc.value=str(v); vc.font=cell_font(); vc.fill=row_fill(i); vc.border=thin

        buf=io.BytesIO(); wb.save(buf); buf.seek(0)
        fname=f"DataLineage_{q}_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
        return StreamingResponse(buf,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                  headers={"Content-Disposition":f"attachment; filename={fname}"})
    except ImportError:
        raise HTTPException(500,"openpyxl not installed")
    except Exception as e:
        raise HTTPException(500,str(e))


# ── JSON Export ───────────────────────────────────────────────────────────────
@app.post("/api/export/json")
async def export_json(payload: Dict[str,Any]):
    q=payload.get("query","export")
    doc={
        "export_info":{"tool":"DataLineage IQ","version":"4.0","query":q,"exported_at":datetime.now().isoformat()},
        "summary":{"nodes":len(payload.get("nodes",[])),"edges":len(payload.get("edges",[])),"fields":len(payload.get("metadata",[]))},
        "lineage_graph":{"nodes":payload.get("nodes",[]),"edges":payload.get("edges",[])},
        "metadata":payload.get("metadata",[]),
        "mapping_document":[{
            "mapping_id":f"MAP-{i+1:04d}",
            "source_table":m.get("sources",[""])[0],"source_field":m["field_name"],
            "target_table":m.get("targets",[""])[0],"target_field":m["field_name"],
            "data_type":m.get("data_type",""),"business_logic":m.get("business_logic",""),
            "risk_level":m.get("risk_level",""),"impact":m.get("impact",""),
            "owner":m.get("owner",""),"business_rules":m.get("business_rules",[]),
        } for i,m in enumerate(payload.get("metadata",[]))],
        "sources_used":payload.get("sources_used",[]),
        "ai_insights":payload.get("ai_insights"),
    }
    buf=io.BytesIO(json.dumps(doc,indent=2,ensure_ascii=False,default=str).encode())
    return StreamingResponse(buf,media_type="application/json",
                              headers={"Content-Disposition":f"attachment; filename=DataLineage_{q}_{datetime.now().strftime('%Y%m%d_%H%M')}.json"})


# ── PDF Field Report ──────────────────────────────────────────────────────────
@app.post("/api/generate-report")
async def generate_pdf(payload: Dict[str,Any]):
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.platypus import SimpleDocTemplate,Paragraph,Spacer,Table,TableStyle,HRFlowable
        from reportlab.lib.styles import getSampleStyleSheet,ParagraphStyle
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_CENTER,TA_LEFT

        f=payload.get("field_name","?"); tbl=payload.get("table_name","")
        risk=payload.get("risk_level","Low"); desc=payload.get("description","")
        logic=payload.get("business_logic",""); rules=payload.get("business_rules",[])
        sources=payload.get("sources",[]); targets=payload.get("targets",[])
        impact=payload.get("impact",""); owner=payload.get("owner","")
        dtype=payload.get("data_type","VARCHAR"); tags=payload.get("tags",[])
        ai=payload.get("ai_insights","")

        RC={"High":colors.HexColor("#F43F5E"),"Medium":colors.HexColor("#F59E0B"),"Low":colors.HexColor("#10B981")}
        rc=RC.get(risk,colors.grey)
        buf=io.BytesIO()
        doc=SimpleDocTemplate(buf,pagesize=A4,rightMargin=45,leftMargin=45,topMargin=50,bottomMargin=40)

        T=ParagraphStyle("T",fontSize=20,fontName="Helvetica-Bold",textColor=colors.HexColor("#0F172A"),alignment=TA_CENTER,spaceAfter=4)
        Sub=ParagraphStyle("Sub",fontSize=10,fontName="Helvetica",textColor=colors.HexColor("#6366F1"),alignment=TA_CENTER,spaceAfter=16)
        H2=ParagraphStyle("H2",fontSize=11,fontName="Helvetica-Bold",textColor=colors.HexColor("#1E293B"),spaceBefore=12,spaceAfter=5)
        Bd=ParagraphStyle("Bd",fontSize=9,fontName="Helvetica",textColor=colors.HexColor("#334155"),leading=13)
        Cd=ParagraphStyle("Cd",fontSize=8,fontName="Courier",textColor=colors.HexColor("#1E293B"),
                           backColor=colors.HexColor("#F1F5F9"),leading=12,leftIndent=8)
        Ft=ParagraphStyle("Ft",fontSize=7,fontName="Helvetica",textColor=colors.HexColor("#94A3B8"),alignment=TA_CENTER)
        thin_border=[("BOX",(0,0),(-1,-1),0.5,colors.HexColor("#E2E8F0")),("INNERGRID",(0,0),(-1,-1),0.4,colors.HexColor("#E2E8F0")),("TOPPADDING",(0,0),(-1,-1),5),("BOTTOMPADDING",(0,0),(-1,-1),5),("LEFTPADDING",(0,0),(-1,-1),8)]

        story=[Paragraph("DataLineage IQ",T),Paragraph("Field Lineage & Impact Report",Sub),
               HRFlowable(width="100%",thickness=2,color=colors.HexColor("#6366F1")),Spacer(1,10)]

        # Header badge
        badge=[[Paragraph(f"<b>Field:</b> {f}",Bd),Paragraph(f"<b>Table:</b> {tbl}",Bd),
                Paragraph(f"<b>Type:</b> {dtype}",Bd),Paragraph(f"<b>Risk:</b> ",Bd),
                Paragraph(f"<b>{risk}</b>",ParagraphStyle("R",fontSize=9,fontName="Helvetica-Bold",textColor=rc))]]
        bt=Table(badge,colWidths=[115,155,100,48,62])
        bt.setStyle(TableStyle(thin_border+[("BACKGROUND",(0,0),(-1,-1),colors.HexColor("#F8FAFC"))]))
        story+=[bt,Spacer(1,12)]

        # Overview
        story.append(Paragraph("Field Overview",H2))
        ov=[["Attribute","Value"],["Description",desc or "—"],["Business Owner",owner or "—"],
            ["Tags",", ".join(tags) if tags else "—"]]
        ot=Table(ov,colWidths=[130,360])
        ot.setStyle(TableStyle(thin_border+[("BACKGROUND",(0,0),(0,-1),colors.HexColor("#EEF2FF")),
            ("FONTNAME",(0,0),(0,-1),"Helvetica-Bold"),("FONTSIZE",(0,0),(-1,-1),8),
            ("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1E293B")),("TEXTCOLOR",(0,0),(-1,0),colors.white),
            ("FONTNAME",(0,0),(-1,0),"Helvetica-Bold")]))
        story+=[ot,Paragraph("Business Logic",H2),Paragraph(logic or "No logic defined.",Cd),Spacer(1,6)]

        if rules:
            story.append(Paragraph("Business Rules",H2))
            for r in rules: story.append(Paragraph(f"▸ {r}",Bd))
            story.append(Spacer(1,6))

        if ai:
            story.append(Paragraph("🤖 Tachyon AI Insights",H2))
            story.append(Paragraph(ai,ParagraphStyle("AI",fontSize=9,fontName="Helvetica",textColor=colors.HexColor("#7C3AED"),leading=13,backColor=colors.HexColor("#F5F3FF"),leftIndent=8,rightIndent=8)))
            story.append(Spacer(1,6))

        story.append(Paragraph("Data Lineage",H2))
        ld=[["Direction","Object(s)"],
            ["⬆ Sources","\n".join(sources) if sources else "—"],
            ["⬇ Targets","\n".join(targets) if targets else "—"]]
        lt=Table(ld,colWidths=[100,390])
        lt.setStyle(TableStyle(thin_border+[("BACKGROUND",(0,0),(-1,0),colors.HexColor("#6366F1")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),9),("BACKGROUND",(0,1),(0,1),colors.HexColor("#DCFCE7")),
            ("BACKGROUND",(0,2),(0,2),colors.HexColor("#FEE2E2"))]))
        story+=[lt,Paragraph("Impact Analysis",H2),Paragraph(impact or "No impact assessment available.",Bd),Spacer(1,8)]

        # Risk matrix
        story.append(Paragraph("Risk Assessment Matrix",H2))
        rm=[["Level","Meaning","Required Action"],
            ["🔴 High","Breaking change — regulatory/financial impact","Data Steward sign-off + CAB approval"],
            ["🟡 Medium","Moderate downstream impact","Peer review + full UAT regression"],
            ["🟢 Low","Minimal risk","Standard PR review"]]
        rmt=Table(rm,colWidths=[75,225,190])
        rmt.setStyle(TableStyle(thin_border+[("BACKGROUND",(0,0),(-1,0),colors.HexColor("#1E293B")),
            ("TEXTCOLOR",(0,0),(-1,0),colors.white),("FONTNAME",(0,0),(-1,0),"Helvetica-Bold"),
            ("FONTSIZE",(0,0),(-1,-1),8),
            ("ROWBACKGROUNDS",(0,1),(-1,-1),[colors.HexColor("#FEF2F2"),colors.HexColor("#FFFBEB"),colors.HexColor("#F0FDF4")])]))
        story+=[rmt,Spacer(1,16),
                HRFlowable(width="100%",thickness=0.5,color=colors.HexColor("#E2E8F0")),Spacer(1,4),
                Paragraph(f"DataLineage IQ v4.0  ·  {datetime.now().strftime('%B %d, %Y %H:%M UTC')}  ·  CONFIDENTIAL",Ft)]
        doc.build(story); buf.seek(0)
        return StreamingResponse(buf,media_type="application/pdf",
                                  headers={"Content-Disposition":f"attachment; filename=lineage_{f}_{datetime.now().strftime('%Y%m%d')}.pdf"})
    except ImportError:
        raise HTTPException(500,"reportlab not installed")
    except Exception as e:
        raise HTTPException(500,str(e))


# ═══════════════════════════════════════════════════════════════════════════════
# NEW v4.1 — HELPER FUNCTIONS (additive, no existing logic changed)
# ═══════════════════════════════════════════════════════════════════════════════

def _get_confidentiality(field_name: str) -> str:
    """Return confidentiality classification. Uses static map + keyword rules."""
    fu = field_name.upper()
    if fu in CONFIDENTIALITY_MAP:
        return CONFIDENTIALITY_MAP[fu]
    if any(k in fu for k in ["SSN","DOB","BIRTH","PASSPORT","LICENSE","NATIONAL_ID"]):
        return "PII-Confidential"
    if any(k in fu for k in ["CUSTOMER","CUST_ID","ECN","ACCT","ACCOUNT","EMAIL","PHONE","ADDRESS","NAME"]):
        return "PII-Restricted"
    if any(k in fu for k in ["SCORE","FICO","CREDIT","RISK","DELQ","DELINQ","BALANCE","AMOUNT","SALARY","INCOME","LIMIT"]):
        return "Confidential"
    return DEFAULT_CONFIDENTIALITY


def _get_equivalent_fields(field_name: str, threshold: float = 0.72) -> List[Dict]:
    """Find equivalent/alias field names across SORs using static map + fuzzy matching."""
    fu = field_name.upper()
    results: List[Dict] = []
    seen: set = set()
    for canonical, aliases in FIELD_ALIASES.items():
        all_names = [canonical] + aliases
        all_upper = [n.upper() for n in all_names]
        if fu in all_upper:
            for a in all_names:
                if a.upper() != fu and a.upper() not in seen:
                    seen.add(a.upper())
                    results.append({"field": a, "canonical": canonical,
                                    "similarity": 1.0, "alias_type": "known_equivalent"})
    all_known = list(FIELD_ALIASES.keys()) + [a for aliases in FIELD_ALIASES.values() for a in aliases]
    for candidate in all_known:
        cu = candidate.upper()
        if cu == fu or cu in seen: continue
        ratio = difflib.SequenceMatcher(None, fu, cu).ratio()
        if ratio >= threshold:
            seen.add(cu)
            results.append({"field": candidate, "canonical": candidate,
                             "similarity": round(ratio, 2), "alias_type": "fuzzy_match"})
    return sorted(results, key=lambda x: x["similarity"], reverse=True)[:8]


def _detect_ingestion_pattern(field_name: str, code_results: List[Dict]) -> str:
    """Detect batch vs real-time streaming based on code context and field name patterns."""
    fu = field_name.upper()
    rt_kw = ["KAFKA","STREAM","CDC","REAL_TIME","REALTIME","EVENT","TOPIC","KINESIS","FLINK"]
    batch_kw = ["BATCH","TRUNCATE","APPEND","DAILY","MONTHLY","SYSDATE","TRUNC","SCHEDULE"]
    for c in code_results:
        ctx = (c.get("context","") + c.get("business_logic","")).upper()
        if any(k in ctx for k in rt_kw): return "Real-Time Streaming"
        if any(k in ctx for k in batch_kw): return "Batch"
    if any(k in fu for k in ["TS","TIMESTAMP","EVENT","STREAM","FEED","REAL"]): return "Real-Time Streaming"
    if any(k in fu for k in ["MTD","YTD","MONTHLY","DAILY","BATCH","LOAD_DT"]): return "Batch"
    return "Batch"


def _identify_sors(field_name: str, nodes: List[Dict], metadata: List[Dict]) -> List[Dict]:
    """Identify all Systems of Record carrying this field."""
    fu = field_name.upper()
    pref = _sor_preferences.get(fu, "")
    sor_list: List[Dict] = []
    seen_sys: set = set()
    for m in metadata:
        if m.get("field_name","").upper() != fu: continue
        sys_name = m.get("source_system","")
        if sys_name and sys_name not in seen_sys:
            seen_sys.add(sys_name)
            sor_name = next((s for s in KNOWN_SORS if s in sys_name.upper()), sys_name)
            sor_list.append({
                "sor": sor_name, "system": sys_name,
                "database": m.get("sources",[""])[0] if m.get("sources") else "",
                "table": m.get("table_name",""), "field_name": m.get("field_name",""),
                "data_type": m.get("data_type",""),
                "is_authoritative": sor_name.upper()==pref.upper() if pref else False,
                "is_known_sor": any(s in sys_name.upper() for s in KNOWN_SORS),
            })
    for n in nodes:
        sys_name = n.get("system","")
        if not sys_name or sys_name in seen_sys: continue
        for f in n.get("fields",[]):
            if f.get("name","").upper() != fu: continue
            seen_sys.add(sys_name)
            sor_name = next((s for s in KNOWN_SORS if s in sys_name.upper()), sys_name)
            sor_list.append({
                "sor": sor_name, "system": sys_name,
                "database": n.get("database",""), "table": n.get("label",""),
                "field_name": f.get("name",""), "data_type": f.get("type",""),
                "is_authoritative": sor_name.upper()==pref.upper() if pref else False,
                "is_known_sor": any(s in sys_name.upper() for s in KNOWN_SORS),
            })
    if not sor_list:
        mock_sors = (["HOGAN","FDR","CCHUB"] if any(k in fu for k in ["ECN","CUSTOMER","CUST","ACCT"])
                     else ["FICO","FDR","EIW"] if any(k in fu for k in ["FICO","SCORE","RISK","DELQ"])
                     else ["FDR","HOGAN","ICODE"] if any(k in fu for k in ["TRANS","TXN","PAYMENT"])
                     else ["HOGAN","FDR"])
        for i, sor in enumerate(mock_sors):
            sor_list.append({
                "sor": sor, "system": sor, "database": f"{sor}_DB",
                "table": f"{sor}_{fu[:15]}", "field_name": field_name,
                "data_type": "VARCHAR",
                "is_authoritative": (sor.upper()==pref.upper()) if pref else i==0,
                "is_known_sor": True,
            })
    if pref:
        for s in sor_list: s["is_authoritative"] = s["sor"].upper()==pref.upper()
    elif not any(s["is_authoritative"] for s in sor_list) and sor_list:
        sor_list[0]["is_authoritative"] = True
    return sor_list


def _build_lineage_hops(field_name: str, nodes: List[Dict], edges: List[Dict]) -> List[Dict]:
    """Build ordered multi-hop lineage path: Source SOR -> Intermediate hops -> Target DW."""
    adj: Dict[str, List[str]] = {}
    for e in edges:
        adj.setdefault(e["source"],[]).append(e["target"])
    targets_set = {e["target"] for e in edges}
    sources = [n for n in nodes if n["id"] not in targets_set] or ([nodes[0]] if nodes else [])
    best_path: List[str] = []
    queue: List[tuple] = [(s["id"], [s["id"]]) for s in sources]
    while queue:
        nid, path = queue.pop(0)
        if len(path) > len(best_path): best_path = path
        for nxt in adj.get(nid, []):
            if nxt not in path: queue.append((nxt, path+[nxt]))
    hops = []
    for i, nid in enumerate(best_path):
        n = next((x for x in nodes if x["id"]==nid), None)
        if not n: continue
        hops.append({
            "hop": i+1, "node_id": nid, "system": n.get("system",""),
            "table": n.get("label","").replace("\n"," "), "type": n.get("type",""),
            "field_present": any(f.get("name","").upper()==field_name.upper() for f in n.get("fields",[])),
            "transformation": n.get("description","") if n.get("type")=="transformation" else "",
        })
    return hops


def _enrich_metadata_v41(meta: List[Dict], nodes: List[Dict],
                          edges: List[Dict], code_results: List[Dict]) -> List[Dict]:
    """Add v4.1 fields to existing metadata rows. Does NOT modify existing keys."""
    for m in meta:
        fn = m.get("field_name","")
        m.setdefault("confidentiality",   _get_confidentiality(fn))
        m.setdefault("equivalent_fields", _get_equivalent_fields(fn))
        m.setdefault("sors",              _identify_sors(fn, nodes, meta))
        m.setdefault("preferred_sor",     _sor_preferences.get(fn.upper(),
                                          m.get("sors",[{}])[0].get("sor","") if m.get("sors") else ""))
        m.setdefault("ingestion_pattern", _detect_ingestion_pattern(fn, code_results))
        m.setdefault("lineage_hops",      _build_lineage_hops(fn, nodes, edges))
    return meta


# ═══════════════════════════════════════════════════════════════════════════════
# NEW v4.1 — API ENDPOINTS
# ═══════════════════════════════════════════════════════════════════════════════

@app.post("/api/multi-search")
async def multi_search(req: MultiSearchRequest):
    """Search multiple fields at once — returns per-field lineage + combined graph."""
    t0 = datetime.now()
    results_by_field: Dict[str, Any] = {}
    all_nodes: List[Dict] = []
    all_edges: List[Dict] = []
    all_meta: List[Dict] = []
    seen_ids: set = set()
    for field in req.fields[:10]:
        field = field.strip()
        if not field: continue
        try:
            ed = await fetch_erwin(
                field, req.local_meta_dir or "",
                erwin_base_url=getattr(req,'erwin_base_url',None),
                erwin_api_key=getattr(req,'erwin_api_key',None),
                erwin_system=getattr(req,'erwin_system',None),
                erwin_env=getattr(req,'erwin_env',None),
                erwin_project=getattr(req,'erwin_project',None),
                erwin_lineage_type=getattr(req,'erwin_lineage_type',None),
                erwin_ep_tablename=getattr(req,'erwin_ep_tablename',None),
                erwin_ep_colname=getattr(req,'erwin_ep_colname',None),
                erwin_ep_objtype=getattr(req,'erwin_ep_objtype',None),
                erwin_ep_mapname=getattr(req,'erwin_ep_mapname',None),
                erwin_ep_glosscat=getattr(req,'erwin_ep_glosscat',None),
            )
            dd = await fetch_dict(field, req.local_dict_dir or "")
            cr = await fetch_code(field, "", "", req.local_code_dir or "")
            nodes, edges = build_graph(field, ed, dd, cr)
            meta = _enrich_metadata_v41(build_metadata(field, ed, dd, cr), nodes, edges, cr)
            results_by_field[field] = {"nodes":nodes,"edges":edges,"metadata":meta,
                                       "erwin_status":ed.get("status","mock"),"dict_status":dd.get("status","mock")}
            for n in nodes:
                if n["id"] not in seen_ids: seen_ids.add(n["id"]); all_nodes.append(n)
            for e in edges:
                if e not in all_edges: all_edges.append(e)
            all_meta.extend(meta)
        except Exception as ex:
            results_by_field[field] = {"error": str(ex)}
    ms = int((datetime.now()-t0).total_seconds()*1000)
    return {"fields_searched":req.fields,"results_by_field":results_by_field,
            "combined_graph":{"nodes":all_nodes,"edges":all_edges},
            "combined_metadata":all_meta,"search_time_ms":ms,
            "summary":f"{len(req.fields)} fields searched · {len(all_nodes)} nodes · {len(all_meta)} metadata rows"}


@app.post("/api/sor-select")
async def sor_select(req: SORSelectRequest):
    """Set preferred System of Record for a field. Persists for the session."""
    _sor_preferences[req.field_name.upper()] = req.preferred_sor.upper()
    return {"status":"ok","field":req.field_name,"preferred_sor":req.preferred_sor,
            "message":f"Authoritative SOR for '{req.field_name}' set to '{req.preferred_sor}'."}


@app.get("/api/sor-preferences")
async def get_sor_preferences():
    """Get all SOR preference selections."""
    return {"preferences":_sor_preferences,"count":len(_sor_preferences)}


@app.post("/api/equivalent-fields")
async def equivalent_fields_endpoint(payload: Dict[str,Any]):
    """Find equivalent/alias fields across SORs. e.g. CUSTOMER_ID -> ECN_PRIMARY_ID, CUST_ID"""
    field = payload.get("field","")
    threshold = float(payload.get("threshold", 0.72))
    return {"field":field,"confidentiality":_get_confidentiality(field),
            "preferred_sor":_sor_preferences.get(field.upper(),""),
            "equivalent_fields":_get_equivalent_fields(field,threshold),
            "note":"Includes known aliases and fuzzy-matched names across Credit Services SORs."}


@app.post("/api/upload-inputs")
async def upload_inputs(file: UploadFile = File(...)):
    """Upload Business Requirements xlsx or INPUTS zip file for metadata ingestion."""
    os.makedirs(CFG["INPUTS_DIR"], exist_ok=True)
    fname = file.filename or "upload"
    content = await file.read()
    summary: Dict[str,Any] = {"filename":fname,"size_bytes":len(content)}
    try:
        if fname.endswith(".zip"):
            with zipfile.ZipFile(io.BytesIO(content)) as zf:
                extracted = []
                for member in zf.namelist():
                    if any(member.endswith(ext) for ext in [".xlsx",".json",".sql",".py",".hql",".java",".scala",".txt",".csv"]):
                        target = os.path.join(CFG["INPUTS_DIR"], os.path.basename(member))
                        with open(target,"wb") as out: out.write(zf.read(member))
                        extracted.append(os.path.basename(member))
            summary.update({"status":"extracted","extracted_files":extracted,"inputs_dir":CFG["INPUTS_DIR"]})
        elif fname.endswith(".xlsx"):
            target = os.path.join(CFG["INPUTS_DIR"], fname)
            with open(target,"wb") as out: out.write(content)
            try:
                import openpyxl
                wb = openpyxl.load_workbook(io.BytesIO(content),read_only=True,data_only=True)
                fields_found = []
                for ws in wb.worksheets:
                    for row in ws.iter_rows(values_only=True):
                        for v in row:
                            sv = str(v).strip() if v else ""
                            if sv and len(sv)<=60 and "_" in sv and sv.upper()==sv:
                                if sv not in fields_found: fields_found.append(sv)
                summary.update({"status":"saved","fields_detected":fields_found[:50],
                                 "total_fields_detected":len(fields_found),"inputs_dir":CFG["INPUTS_DIR"]})
            except Exception:
                summary.update({"status":"saved","inputs_dir":CFG["INPUTS_DIR"]})
        else:
            raise HTTPException(400,"Unsupported file type. Upload .xlsx or .zip")
        return summary
    except HTTPException: raise
    except Exception as e:
        raise HTTPException(500,f"Upload failed: {e}")


@app.get("/api/known-sors")
async def known_sors():
    """Return known Credit Services Systems of Record and typical lineage flow."""
    return {"sors":KNOWN_SORS,
            "typical_flow":"Hogan -> FDR -> CCHUB -> EIW / ADH / ICODE / CODI / BINK",
            "example":"ECN_PRIMARY_ID (CUSTOMER_ID) exists in Hogan, FDR, CCHUB, EIW, ADH, ICODE, CODI. "
                       "Use /api/sor-select to set preferred authoritative source for your application."}


@app.get("/api/field-aliases")
async def field_aliases():
    """Return the complete known field alias/equivalent map."""
    return {"aliases":FIELD_ALIASES,"total_groups":len(FIELD_ALIASES)}


# ── HEALTH / CONFIG ───────────────────────────────────────────────────────────
@app.get("/api/health")
async def health():
    return {"status":"ok","version":"4.1.0","timestamp":datetime.now().isoformat(),
            "features":{"erwin":"api+excel+json+mock","data_dict":"api+excel+json+mock",
                         "code":"github+local+mock","tachyon":"summarise+insights","export":"excel+json+pdf",
                         "multi_search":"multi-field batch","sor_select":"preferred SOR","equivalent_fields":"alias matching",
                         "confidentiality":"PII classification","lineage_hops":"multi-hop path","upload_inputs":"BR xlsx + zip"}}

@app.get("/api/config")
async def get_config():
    return {k:v for k,v in CFG.items() if "KEY" not in k and "TOKEN" not in k}

# HTML embedded directly - no file dependency
_HTML_B64 = "PCFET0NUWVBFIGh0bWw+CjxodG1sIGxhbmc9ImVuIj4KPGhlYWQ+CjxtZXRhIGNoYXJzZXQ9IlVURi04Ii8+CjxtZXRhIG5hbWU9InZpZXdwb3J0IiBjb250ZW50PSJ3aWR0aD1kZXZpY2Utd2lkdGgsaW5pdGlhbC1zY2FsZT0xLjAiLz4KPG1ldGEgaHR0cC1lcXVpdj0iQ2FjaGUtQ29udHJvbCIgY29udGVudD0ibm8tY2FjaGUsIG5vLXN0b3JlLCBtdXN0LXJldmFsaWRhdGUiLz4KPG1ldGEgaHR0cC1lcXVpdj0iUHJhZ21hIiBjb250ZW50PSJuby1jYWNoZSIvPgo8bWV0YSBodHRwLWVxdWl2PSJFeHBpcmVzIiBjb250ZW50PSIwIi8+Cjx0aXRsZT5EYXRhTGluZWFnZSBJUSB2NC4xPC90aXRsZT4KPHNjcmlwdCBzcmM9Imh0dHBzOi8vY2RuanMuY2xvdWRmbGFyZS5jb20vYWpheC9saWJzL2QzLzcuOC41L2QzLm1pbi5qcyI+PC9zY3JpcHQ+CjxsaW5rIGhyZWY9Imh0dHBzOi8vZm9udHMuZ29vZ2xlYXBpcy5jb20vY3NzMj9mYW1pbHk9QmViYXMrTmV1ZSZmYW1pbHk9SmV0QnJhaW5zK01vbm86d2dodEA0MDA7NTAwOzYwMCZmYW1pbHk9RE0rU2Fuczp3Z2h0QDMwMDs0MDA7NTAwOzYwMDs3MDAmZGlzcGxheT1zd2FwIiByZWw9InN0eWxlc2hlZXQiLz4KPHN0eWxlPgovKiDilZDilZDilZAgREVTSUdOIFRPS0VOUyDilZDilZDilZAgKi8KOnJvb3R7CiAgLS1iZzojMDIwODEwOy0tYmcxOiMwNjBmMWU7LS1iZzI6IzBiMTczMDstLWJnMzojMTAxZjNjOwogIC0tcGFuZWw6cmdiYSgyNTUsMjU1LDI1NSwuMDQpOy0tcGFuZWwyOnJnYmEoMjU1LDI1NSwyNTUsLjA3KTstLXBhbmVsMzpyZ2JhKDI1NSwyNTUsMjU1LC4xMSk7CiAgLS1iZHI6cmdiYSgyNTUsMjU1LDI1NSwuMDcpOy0tYmRyMjpyZ2JhKDI1NSwyNTUsMjU1LC4xMik7CiAgLS1nb2xkOiNGNUE2MjM7LS1nb2xkMjojRkZEMTY2Oy0tZ29sZDM6I0I4NzgwQTsKICAtLXRlYWw6IzAwQzlCMTstLXRlYWwyOiM0REREQzg7LS10ZWFsMzojMDA4RjdFOwogIC0tc2t5OiMzQjlFRkY7LS1za3kyOiM3NEJGRkY7CiAgLS1saW1lOiM3RkQ5NjI7LS1saW1lMjojQThFRjhGOwogIC0tcm9zZTojRkY0RjZEOy0tcm9zZTI6I0ZGODA5ODsKICAtLXZpbzojOUI2REZGOy0tdmlvMjojQkZBMkZGOwogIC0tdDA6I0YyRjhGRjstLXQxOiNBOEM0RTA7LS10MjojNEU2RThBOy0tdDM6IzFFMzQ1MDsKICAtLXI6MTBweDstLXIyOjdweDstLXIzOjRweDsKICAtLXNoOjAgOHB4IDMycHggcmdiYSgwLDAsMCwuNTUpOwp9CiosKjo6YmVmb3JlLCo6OmFmdGVye2JveC1zaXppbmc6Ym9yZGVyLWJveDttYXJnaW46MDtwYWRkaW5nOjB9Cmh0bWx7Y29sb3Itc2NoZW1lOmRhcmt9CmJvZHl7Zm9udC1mYW1pbHk6J0RNIFNhbnMnLHNhbnMtc2VyaWY7YmFja2dyb3VuZDp2YXIoLS1iZyk7Y29sb3I6dmFyKC0tdDApO21pbi1oZWlnaHQ6MTAwdmg7b3ZlcmZsb3cteDpoaWRkZW59CmJvZHk6OmJlZm9yZXtjb250ZW50OicnO3Bvc2l0aW9uOmZpeGVkO2luc2V0OjA7cG9pbnRlci1ldmVudHM6bm9uZTt6LWluZGV4OjA7CiAgYmFja2dyb3VuZDpyYWRpYWwtZ3JhZGllbnQoZWxsaXBzZSAxMzAlIDgwJSBhdCAtNSUgMCUscmdiYSgyNDUsMTY2LDM1LC4wOSkgMCUsdHJhbnNwYXJlbnQgNTUlKSwKICAgIHJhZGlhbC1ncmFkaWVudChlbGxpcHNlIDgwJSAxMDAlIGF0IDEwNSUgMjAlLHJnYmEoMCwyMDEsMTc3LC4wNykgMCUsdHJhbnNwYXJlbnQgNTAlKSwKICAgIHJhZGlhbC1ncmFkaWVudChlbGxpcHNlIDYwJSA2MCUgYXQgNTAlIDExMCUscmdiYSg1OSwxNTgsMjU1LC4wNikgMCUsdHJhbnNwYXJlbnQgNTUlKX0KYm9keTo6YWZ0ZXJ7Y29udGVudDonJztwb3NpdGlvbjpmaXhlZDtpbnNldDowO3BvaW50ZXItZXZlbnRzOm5vbmU7ei1pbmRleDowOwogIGJhY2tncm91bmQtaW1hZ2U6bGluZWFyLWdyYWRpZW50KHJnYmEoMjQ1LDE2NiwzNSwuMDE2KSAxcHgsdHJhbnNwYXJlbnQgMXB4KSwKICAgIGxpbmVhci1ncmFkaWVudCg5MGRlZyxyZ2JhKDAsMjAxLDE3NywuMDE2KSAxcHgsdHJhbnNwYXJlbnQgMXB4KTsKICBiYWNrZ3JvdW5kLXNpemU6NjBweCA2MHB4fQoKLyog4pWQ4pWQ4pWQIFNDUk9MTEJBUiDilZDilZDilZAgKi8KOjotd2Via2l0LXNjcm9sbGJhcnt3aWR0aDo0cHg7aGVpZ2h0OjRweH0KOjotd2Via2l0LXNjcm9sbGJhci10cmFja3tiYWNrZ3JvdW5kOnRyYW5zcGFyZW50fQo6Oi13ZWJraXQtc2Nyb2xsYmFyLXRodW1ie2JhY2tncm91bmQ6dmFyKC0tYmRyMik7Ym9yZGVyLXJhZGl1czoycHh9CgovKiDilZDilZDilZAgSEVBREVSIOKVkOKVkOKVkCAqLwpoZWFkZXJ7cG9zaXRpb246c3RpY2t5O3RvcDowO3otaW5kZXg6MTAwO2hlaWdodDo1NHB4OwogIGJhY2tncm91bmQ6cmdiYSgyLDgsMTYsLjkyKTtiYWNrZHJvcC1maWx0ZXI6Ymx1cigyNHB4KTsKICBib3JkZXItYm90dG9tOjFweCBzb2xpZCB2YXIoLS1iZHIyKTsKICBkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2p1c3RpZnktY29udGVudDpzcGFjZS1iZXR3ZWVuO3BhZGRpbmc6MCAyMHB4fQoubG9nb3tkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDoxMXB4fQoubG9nby1tYXJre3dpZHRoOjMycHg7aGVpZ2h0OjMycHg7Ym9yZGVyLXJhZGl1czo4cHg7CiAgYmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoMTM1ZGVnLHZhcigtLWdvbGQpLHZhcigtLXRlYWwpKTsKICBkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2p1c3RpZnktY29udGVudDpjZW50ZXI7Zm9udC1zaXplOjE2cHg7CiAgYm94LXNoYWRvdzowIDAgMThweCByZ2JhKDI0NSwxNjYsMzUsLjM1KTthbmltYXRpb246Z2xvdyAzcyBlYXNlLWluLW91dCBpbmZpbml0ZSBhbHRlcm5hdGV9CkBrZXlmcmFtZXMgZ2xvd3swJXtib3gtc2hhZG93OjAgMCAxOHB4IHJnYmEoMjQ1LDE2NiwzNSwuMzUpfTEwMCV7Ym94LXNoYWRvdzowIDAgMjhweCByZ2JhKDAsMjAxLDE3NywuNCl9fQoubG9nby10aXRsZXtmb250LWZhbWlseTonQmViYXMgTmV1ZScsc2Fucy1zZXJpZjtmb250LXNpemU6MTlweDtsZXR0ZXItc3BhY2luZzoycHg7CiAgYmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoOTBkZWcsdmFyKC0tZ29sZDIpLHZhcigtLXRlYWwyKSk7CiAgLXdlYmtpdC1iYWNrZ3JvdW5kLWNsaXA6dGV4dDstd2Via2l0LXRleHQtZmlsbC1jb2xvcjp0cmFuc3BhcmVudH0KLmxvZ28tc3Vie2ZvbnQtc2l6ZTo4cHg7Y29sb3I6dmFyKC0tdDIpO2xldHRlci1zcGFjaW5nOjJweDtmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZX0KLmFwaS1iYWRnZXtkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo1cHg7cGFkZGluZzo0cHggMTBweDsKICBiYWNrZ3JvdW5kOnZhcigtLXBhbmVsKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJkcjIpO2JvcmRlci1yYWRpdXM6MjBweDsKICBmb250LXNpemU6MTBweDtjb2xvcjp2YXIoLS10Mik7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2V9Ci5saXZlLWRvdHt3aWR0aDo2cHg7aGVpZ2h0OjZweDtib3JkZXItcmFkaXVzOjUwJTtiYWNrZ3JvdW5kOnZhcigtLWxpbWUpOwogIGJveC1zaGFkb3c6MCAwIDhweCB2YXIoLS1saW1lKTthbmltYXRpb246cHVsc2UgMnMgaW5maW5pdGV9CkBrZXlmcmFtZXMgcHVsc2V7MCUsMTAwJXtvcGFjaXR5OjF9NTAle29wYWNpdHk6LjN9fQoKLyog4pWQ4pWQ4pWQIExBWU9VVCDilZDilZDilZAgKi8KLmFwcHtwb3NpdGlvbjpyZWxhdGl2ZTt6LWluZGV4OjE7ZGlzcGxheTpncmlkO2dyaWQtdGVtcGxhdGUtY29sdW1uczoyOTJweCAxZnI7bWluLWhlaWdodDpjYWxjKDEwMHZoIC0gNTRweCl9CgovKiDilZDilZDilZAgU0lERUJBUiDilZDilZDilZAgKi8KLnNie2JhY2tncm91bmQ6dmFyKC0tYmcxKTtib3JkZXItcmlnaHQ6MXB4IHNvbGlkIHZhcigtLWJkcjIpOwogIGhlaWdodDpjYWxjKDEwMHZoIC0gNTRweCk7cG9zaXRpb246c3RpY2t5O3RvcDo1NHB4OwogIG92ZXJmbG93LXk6YXV0bztvdmVyZmxvdy14OmhpZGRlbjtkaXNwbGF5OmZsZXg7ZmxleC1kaXJlY3Rpb246Y29sdW1ufQouc2Itc2Vje3BhZGRpbmc6MTNweCAxMnB4O2JvcmRlci1ib3R0b206MXB4IHNvbGlkIHZhcigtLWJkcil9Ci5zYi1sYmx7Zm9udC1zaXplOjlweDtmb250LXdlaWdodDo3MDA7bGV0dGVyLXNwYWNpbmc6MS41cHg7dGV4dC10cmFuc2Zvcm06dXBwZXJjYXNlOwogIGNvbG9yOnZhcigtLXQyKTttYXJnaW4tYm90dG9tOjlweDtkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo2cHg7CiAgZm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2V9Ci5zYi1sYmw6OmFmdGVye2NvbnRlbnQ6Jyc7ZmxleDoxO2hlaWdodDoxcHg7YmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoOTBkZWcsdmFyKC0tYmRyMiksdHJhbnNwYXJlbnQpfQoKLyogU2VhcmNoIGlucHV0ICovCi5zcmNoLXdyYXB7cG9zaXRpb246cmVsYXRpdmV9Ci5zcmNoLWljb3twb3NpdGlvbjphYnNvbHV0ZTtsZWZ0OjlweDt0b3A6NTAlO3RyYW5zZm9ybTp0cmFuc2xhdGVZKC01MCUpO2ZvbnQtc2l6ZToxM3B4O2NvbG9yOnZhcigtLWdvbGQpfQppbnB1dC5zcmNoe3dpZHRoOjEwMCU7aGVpZ2h0OjM3cHg7cGFkZGluZzowIDlweCAwIDI5cHg7CiAgYmFja2dyb3VuZDp2YXIoLS1wYW5lbCk7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1iZHIyKTtib3JkZXItcmFkaXVzOnZhcigtLXIyKTsKICBjb2xvcjp2YXIoLS10MCk7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2U7Zm9udC1zaXplOjEycHg7b3V0bGluZTpub25lO3RyYW5zaXRpb246Ym9yZGVyLWNvbG9yIC4yc30KaW5wdXQuc3JjaDpmb2N1c3tib3JkZXItY29sb3I6dmFyKC0tZ29sZCk7Ym94LXNoYWRvdzowIDAgMCAzcHggcmdiYSgyNDUsMTY2LDM1LC4xKX0KaW5wdXQuc3JjaDo6cGxhY2Vob2xkZXJ7Y29sb3I6dmFyKC0tdDIpfQouZ28tYnRue3dpZHRoOjEwMCU7aGVpZ2h0OjM0cHg7bWFyZ2luLXRvcDo3cHg7Ym9yZGVyOm5vbmU7Ym9yZGVyLXJhZGl1czp2YXIoLS1yMik7CiAgYmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoMTM1ZGVnLHZhcigtLWdvbGQzKSx2YXIoLS10ZWFsMykpOwogIGNvbG9yOiNmZmY7Zm9udC1mYW1pbHk6J0JlYmFzIE5ldWUnLHNhbnMtc2VyaWY7Zm9udC1zaXplOjE1cHg7bGV0dGVyLXNwYWNpbmc6MnB4OwogIGN1cnNvcjpwb2ludGVyO3Bvc2l0aW9uOnJlbGF0aXZlO292ZXJmbG93OmhpZGRlbjtib3gtc2hhZG93OjAgMCAxNHB4IHJnYmEoMjQ1LDE2NiwzNSwuMjIpOwogIHRyYW5zaXRpb246YWxsIC4yc30KLmdvLWJ0bjo6YmVmb3Jle2NvbnRlbnQ6Jyc7cG9zaXRpb246YWJzb2x1dGU7aW5zZXQ6MDsKICBiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCg5MGRlZyx0cmFuc3BhcmVudCxyZ2JhKDI1NSwyNTUsMjU1LC4xNSksdHJhbnNwYXJlbnQpOwogIHRyYW5zZm9ybTp0cmFuc2xhdGVYKC0xMDAlKTt0cmFuc2l0aW9uOnRyYW5zZm9ybSAuNDVzfQouZ28tYnRuOmhvdmVyOjpiZWZvcmV7dHJhbnNmb3JtOnRyYW5zbGF0ZVgoMTAwJSl9Ci5nby1idG46aG92ZXJ7Ym94LXNoYWRvdzowIDAgMjZweCByZ2JhKDI0NSwxNjYsMzUsLjQ1KTt0cmFuc2Zvcm06dHJhbnNsYXRlWSgtMXB4KX0KLmdvLWJ0bjpkaXNhYmxlZHtvcGFjaXR5Oi40O2N1cnNvcjpub3QtYWxsb3dlZDt0cmFuc2Zvcm06bm9uZX0KCi8qIFNvdXJjZSB0b2dnbGVzICovCi5zcmMtdG9ne2Rpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjdweDtwYWRkaW5nOjdweCA5cHg7Ym9yZGVyLXJhZGl1czp2YXIoLS1yMik7CiAgY3Vyc29yOnBvaW50ZXI7dHJhbnNpdGlvbjphbGwgLjEzczttYXJnaW4tYm90dG9tOjNweDtib3JkZXI6MXB4IHNvbGlkIHRyYW5zcGFyZW50O3VzZXItc2VsZWN0Om5vbmV9Ci5zcmMtdG9nLm9ue2JhY2tncm91bmQ6dmFyKC0tcGFuZWwpO2JvcmRlci1jb2xvcjp2YXIoLS1iZHIyKX0KLnNyYy10b2c6aG92ZXJ7YmFja2dyb3VuZDp2YXIoLS1wYW5lbDIpfQouc3QtaWNve2ZvbnQtc2l6ZToxMnB4O3dpZHRoOjE2cHg7dGV4dC1hbGlnbjpjZW50ZXI7ZmxleC1zaHJpbms6MH0KLnN0LW5hbWV7Zm9udC1zaXplOjExcHg7Zm9udC13ZWlnaHQ6NjAwO2NvbG9yOnZhcigtLXQwKTtmbGV4OjF9Ci5zdC1kZXNje2ZvbnQtc2l6ZTo5cHg7Y29sb3I6dmFyKC0tdDIpfQouc3QtZG90e3dpZHRoOjZweDtoZWlnaHQ6NnB4O2JvcmRlci1yYWRpdXM6NTAlO2JhY2tncm91bmQ6dmFyKC0tdDMpO21hcmdpbi1sZWZ0OmF1dG87CiAgbWFyZ2luLXJpZ2h0OjRweDtmbGV4LXNocmluazowO3RyYW5zaXRpb246YWxsIC4zc30KLnN0LWRvdC5saXZle2JhY2tncm91bmQ6dmFyKC0tbGltZSk7Ym94LXNoYWRvdzowIDAgNnB4IHZhcigtLWxpbWUpfQouc3QtZG90LmZpbGV7YmFja2dyb3VuZDp2YXIoLS1nb2xkKTtib3gtc2hhZG93OjAgMCA2cHggdmFyKC0tZ29sZCl9Ci5zdC1kb3QubW9ja3tiYWNrZ3JvdW5kOnZhcigtLXQyKX0KLnN0LWRvdC5haXtiYWNrZ3JvdW5kOnZhcigtLXZpbyk7Ym94LXNoYWRvdzowIDAgNnB4IHZhcigtLXZpbyl9Ci5jaGstYm94e3dpZHRoOjE0cHg7aGVpZ2h0OjE0cHg7Ym9yZGVyLXJhZGl1czozcHg7Ym9yZGVyOjEuNXB4IHNvbGlkIHZhcigtLWJkcjIpOwogIGRpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7anVzdGlmeS1jb250ZW50OmNlbnRlcjtmb250LXNpemU6OXB4OwogIGNvbG9yOnRyYW5zcGFyZW50O3RyYW5zaXRpb246YWxsIC4xM3M7ZmxleC1zaHJpbms6MH0KLnNyYy10b2cub24gLmNoay1ib3h7YmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoMTM1ZGVnLHZhcigtLWdvbGQpLHZhcigtLXRlYWwpKTsKICBib3JkZXItY29sb3I6dHJhbnNwYXJlbnQ7Y29sb3I6I2ZmZjtib3gtc2hhZG93OjAgMCA4cHggcmdiYSgyNDUsMTY2LDM1LC4zKX0KCi8qIENoaXBzICovCi5jaGlwc3tkaXNwbGF5OmZsZXg7ZmxleC13cmFwOndyYXA7Z2FwOjRweH0KLmNoaXB7YmFja2dyb3VuZDp2YXIoLS1wYW5lbCk7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1iZHIpO2JvcmRlci1yYWRpdXM6OXB4OwogIHBhZGRpbmc6MnB4IDhweDtmb250LXNpemU6OXB4O2ZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlO2NvbG9yOnZhcigtLXRlYWwyKTsKICBjdXJzb3I6cG9pbnRlcjt0cmFuc2l0aW9uOmFsbCAuMTFzO3VzZXItc2VsZWN0Om5vbmV9Ci5jaGlwOmhvdmVye2JhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDEzNWRlZyx2YXIoLS1nb2xkMyksdmFyKC0tdGVhbDMpKTsKICBib3JkZXItY29sb3I6dHJhbnNwYXJlbnQ7Y29sb3I6I2ZmZjt0cmFuc2Zvcm06dHJhbnNsYXRlWSgtMXB4KX0KCi8qIENvbmZpZyBpbnB1dHMgKi8KLmNmZy1sYmx7Zm9udC1zaXplOjlweDtjb2xvcjp2YXIoLS10Mik7ZGlzcGxheTpibG9jazttYXJnaW4tYm90dG9tOjNweDsKICBmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZX0KLmNmZy1pbnt3aWR0aDoxMDAlO2hlaWdodDoyNnB4O3BhZGRpbmc6MCA3cHg7bWFyZ2luLWJvdHRvbTo1cHg7CiAgYmFja2dyb3VuZDp2YXIoLS1wYW5lbCk7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1iZHIpO2JvcmRlci1yYWRpdXM6dmFyKC0tcjMpOwogIGNvbG9yOnZhcigtLXQwKTtmb250LXNpemU6MTBweDtmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZTtvdXRsaW5lOm5vbmU7dHJhbnNpdGlvbjpib3JkZXItY29sb3IgLjE1c30KLmNmZy1pbjpmb2N1c3tib3JkZXItY29sb3I6dmFyKC0tZ29sZCl9Ci5jZmctaW46OnBsYWNlaG9sZGVye2NvbG9yOnZhcigtLXQzKX0KLmNmZy1zZWN7Zm9udC1zaXplOjlweDtmb250LXdlaWdodDo3MDA7Y29sb3I6dmFyKC0tZ29sZDIpO21hcmdpbjo4cHggMCA1cHg7CiAgbGV0dGVyLXNwYWNpbmc6LjVweDtmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZX0KZGV0YWlscz5zdW1tYXJ5e2N1cnNvcjpwb2ludGVyO3VzZXItc2VsZWN0Om5vbmU7bGlzdC1zdHlsZTpub25lO2ZvbnQtc2l6ZToxMHB4OwogIGZvbnQtd2VpZ2h0OjcwMDtjb2xvcjp2YXIoLS1nb2xkMik7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2U7bGV0dGVyLXNwYWNpbmc6LjVweH0KZGV0YWlscz5zdW1tYXJ5Ojotd2Via2l0LWRldGFpbHMtbWFya2Vye2Rpc3BsYXk6bm9uZX0KCi8qIE11bHRpLWZpZWxkIHRleHRhcmVhICovCi5tdWx0aS10YXt3aWR0aDoxMDAlO2JhY2tncm91bmQ6dmFyKC0tcGFuZWwpO2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyMik7CiAgYm9yZGVyLXJhZGl1czp2YXIoLS1yMyk7Y29sb3I6dmFyKC0tdDApO2ZvbnQtc2l6ZToxMHB4OwogIGZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlO3BhZGRpbmc6N3B4IDhweDtyZXNpemU6bm9uZTtvdXRsaW5lOm5vbmU7CiAgbGluZS1oZWlnaHQ6MS41O3RyYW5zaXRpb246Ym9yZGVyLWNvbG9yIC4xNXN9Ci5tdWx0aS10YTpmb2N1c3tib3JkZXItY29sb3I6dmFyKC0tdGVhbCl9Ci5tdWx0aS1idG57d2lkdGg6MTAwJTtoZWlnaHQ6MjdweDttYXJnaW4tdG9wOjVweDtiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCgxMzVkZWcsdmFyKC0tdGVhbDMpLCMxZDZmYTYpOwogIGJvcmRlcjpub25lO2JvcmRlci1yYWRpdXM6dmFyKC0tcjMpO2NvbG9yOiNmZmY7Zm9udC1zaXplOjEwcHg7Zm9udC13ZWlnaHQ6NzAwOwogIGN1cnNvcjpwb2ludGVyO2xldHRlci1zcGFjaW5nOi41cHg7dHJhbnNpdGlvbjphbGwgLjE1c30KLm11bHRpLWJ0bjpob3Zlcntib3gtc2hhZG93OjAgMCAxNnB4IHJnYmEoMCwyMDEsMTc3LC4zNSk7dHJhbnNmb3JtOnRyYW5zbGF0ZVkoLTFweCl9CgovKiBVcGxvYWQgKi8KLnVwbG9hZC1idG57d2lkdGg6MTAwJTtoZWlnaHQ6MjdweDtiYWNrZ3JvdW5kOnZhcigtLXBhbmVsKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJkcjIpOwogIGJvcmRlci1yYWRpdXM6dmFyKC0tcjMpO2NvbG9yOnZhcigtLXQxKTtmb250LXNpemU6MTBweDtmb250LXdlaWdodDo2MDA7Y3Vyc29yOnBvaW50ZXI7dHJhbnNpdGlvbjphbGwgLjEzc30KLnVwbG9hZC1idG46aG92ZXJ7YmFja2dyb3VuZDp2YXIoLS1wYW5lbDIpfQojdXBsb2FkLXN0YXR1c3tmb250LXNpemU6OXB4O2NvbG9yOnZhcigtLXQyKTttYXJnaW4tdG9wOjRweDsKICBmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZTttaW4taGVpZ2h0OjE0cHg7bGluZS1oZWlnaHQ6MS41fQoKLyogQUkgcGFuZWwgKi8KLmFpLXBhbmVse21hcmdpbi10b3A6OHB4O3BhZGRpbmc6MTBweDsKICBiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCgxMzVkZWcscmdiYSgxNTUsMTA5LDI1NSwuMDgpLHJnYmEoMCwyMDEsMTc3LC4wNSkpOwogIGJvcmRlcjoxcHggc29saWQgcmdiYSgxNTUsMTA5LDI1NSwuMik7Ym9yZGVyLXJhZGl1czp2YXIoLS1yMik7CiAgYm9yZGVyLWxlZnQ6M3B4IHNvbGlkIHZhcigtLXZpbyl9Ci5haS1wYW5lbC10aXRsZXtmb250LXNpemU6OXB4O2ZvbnQtd2VpZ2h0OjcwMDtjb2xvcjp2YXIoLS12aW8yKTtsZXR0ZXItc3BhY2luZzouOHB4OwogIHRleHQtdHJhbnNmb3JtOnVwcGVyY2FzZTttYXJnaW4tYm90dG9tOjVweDtmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZX0KLmFpLXRleHR7Zm9udC1zaXplOjEwcHg7Y29sb3I6dmFyKC0tdDEpO2xpbmUtaGVpZ2h0OjEuNTU7Zm9udC1zdHlsZTppdGFsaWN9CgovKiBSZWNlbnQgKi8KLnJlY2VudC1pdGVte3BhZGRpbmc6NXB4IDdweDtib3JkZXItcmFkaXVzOnZhcigtLXIzKTtjdXJzb3I6cG9pbnRlcjsKICBmb250LXNpemU6MTBweDtjb2xvcjp2YXIoLS10MSk7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2U7CiAgdHJhbnNpdGlvbjphbGwgLjExcztib3JkZXI6MXB4IHNvbGlkIHRyYW5zcGFyZW50fQoucmVjZW50LWl0ZW06aG92ZXJ7YmFja2dyb3VuZDp2YXIoLS1wYW5lbCk7Ym9yZGVyLWNvbG9yOnZhcigtLWJkcil9Ci5yZWNlbnQtaXRlbTo6YmVmb3Jle2NvbnRlbnQ6J+KfsyAnO2NvbG9yOnZhcigtLXQzKX0KCi8qIOKVkOKVkOKVkCBNQUlOIOKVkOKVkOKVkCAqLwoubWFpbntkaXNwbGF5OmZsZXg7ZmxleC1kaXJlY3Rpb246Y29sdW1uO292ZXJmbG93OmhpZGRlbjttaW4taGVpZ2h0OjB9Ci50b3AtYmFye3BhZGRpbmc6MTJweCAxOHB4IDA7Ym9yZGVyLWJvdHRvbToxcHggc29saWQgdmFyKC0tYmRyMik7CiAgYmFja2dyb3VuZDpyZ2JhKDIsOCwxNiwuOCk7YmFja2Ryb3AtZmlsdGVyOmJsdXIoMTZweCk7ZmxleC1zaHJpbms6MH0KLnJlcy1iYXJ7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6N3B4O3BhZGRpbmctYm90dG9tOjEwcHg7ZmxleC13cmFwOndyYXB9Ci5yZXMtdGl0bGV7Zm9udC1mYW1pbHk6J0JlYmFzIE5ldWUnLHNhbnMtc2VyaWY7Zm9udC1zaXplOjIwcHg7bGV0dGVyLXNwYWNpbmc6MXB4OwogIGJhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDkwZGVnLHZhcigtLXQwKSx2YXIoLS1nb2xkMikpOwogIC13ZWJraXQtYmFja2dyb3VuZC1jbGlwOnRleHQ7LXdlYmtpdC10ZXh0LWZpbGwtY29sb3I6dHJhbnNwYXJlbnQ7ZmxleDoxfQouc3BpbGx7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6NHB4O3BhZGRpbmc6M3B4IDlweDtib3JkZXItcmFkaXVzOjIwcHg7CiAgZm9udC1zaXplOjlweDtmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJkcil9Ci5zZG90e3dpZHRoOjVweDtoZWlnaHQ6NXB4O2JvcmRlci1yYWRpdXM6NTAlfQouc3AtbGl2ZXtib3JkZXItY29sb3I6cmdiYSgxMjcsMjE3LDk4LC4zKTtjb2xvcjp2YXIoLS1saW1lMil9Ci5zcC1maWxle2JvcmRlci1jb2xvcjpyZ2JhKDI0NSwxNjYsMzUsLjMpO2NvbG9yOnZhcigtLWdvbGQyKX0KLnNwLW1vY2t7Ym9yZGVyLWNvbG9yOnJnYmEoNzgsMTEwLDEzOCwuMyk7Y29sb3I6dmFyKC0tdDEpfQouc3AtYWl7Ym9yZGVyLWNvbG9yOnJnYmEoMTU1LDEwOSwyNTUsLjMpO2NvbG9yOnZhcigtLXZpbzIpfQouZXhwLXJvd3tkaXNwbGF5OmZsZXg7Z2FwOjVweH0KLmV4cC1idG57ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6NHB4O3BhZGRpbmc6NHB4IDEwcHg7Ym9yZGVyOm5vbmU7CiAgYm9yZGVyLXJhZGl1czp2YXIoLS1yMyk7Zm9udC1zaXplOjEwcHg7Zm9udC13ZWlnaHQ6NzAwO2N1cnNvcjpwb2ludGVyO3RyYW5zaXRpb246YWxsIC4xNHN9Ci5leHAteGx7YmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoMTM1ZGVnLCMxNDUzMmQsIzE2YTM0YSk7Y29sb3I6I2ZmZn0KLmV4cC14bDpob3Zlcntib3gtc2hhZG93OjAgMCAyMHB4IHJnYmEoMjIsMTYzLDc0LC40KTt0cmFuc2Zvcm06dHJhbnNsYXRlWSgtMXB4KX0KLmV4cC1qc3tiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCgxMzVkZWcsIzFlM2E1ZiwjMjU2M2ViKTtjb2xvcjojZmZmfQouZXhwLWpzOmhvdmVye2JveC1zaGFkb3c6MCAwIDIwcHggcmdiYSgzNyw5OSwyMzUsLjQpO3RyYW5zZm9ybTp0cmFuc2xhdGVZKC0xcHgpfQoKLyogVGFicyAqLwoudGFic3tkaXNwbGF5OmZsZXg7Z2FwOjNweDtiYWNrZ3JvdW5kOnZhcigtLXBhbmVsKTsKICBib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJkcjIpO2JvcmRlci1yYWRpdXM6dmFyKC0tcjIpO3BhZGRpbmc6M3B4O3dpZHRoOmZpdC1jb250ZW50fQoudGFiLWJ0bntwYWRkaW5nOjVweCAxM3B4O2JvcmRlci1yYWRpdXM6NXB4O2ZvbnQtc2l6ZToxMXB4O2ZvbnQtd2VpZ2h0OjYwMDsKICBjdXJzb3I6cG9pbnRlcjtjb2xvcjp2YXIoLS10Mik7Ym9yZGVyOm5vbmU7YmFja2dyb3VuZDp0cmFuc3BhcmVudDt0cmFuc2l0aW9uOmFsbCAuMTNzfQoudGFiLWJ0bi5vbntiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCgxMzVkZWcsdmFyKC0tZ29sZDMpLHZhcigtLXRlYWwzKSk7Y29sb3I6I2ZmZjsKICBib3gtc2hhZG93OjAgMCAxMnB4IHJnYmEoMjQ1LDE2NiwzNSwuMjgpfQoudGFiLWJ0bjpob3Zlcjpub3QoLm9uKXtjb2xvcjp2YXIoLS10MCk7YmFja2dyb3VuZDp2YXIoLS1wYW5lbDIpfQoKLmNvbnRlbnR7cGFkZGluZzoxNHB4IDE4cHg7b3ZlcmZsb3cteTphdXRvO2ZsZXg6MX0KCi8qIOKVkOKVkOKVkCBFTVBUWSAvIExPQURJTkcg4pWQ4pWQ4pWQICovCi5lbXB0eXtkaXNwbGF5OmZsZXg7ZmxleC1kaXJlY3Rpb246Y29sdW1uO2FsaWduLWl0ZW1zOmNlbnRlcjtwYWRkaW5nOjgwcHggNDBweDt0ZXh0LWFsaWduOmNlbnRlcn0KLmVtcHR5LWljb257Zm9udC1zaXplOjU0cHg7bWFyZ2luLWJvdHRvbToxMnB4OwogIGZpbHRlcjpkcm9wLXNoYWRvdygwIDAgMThweCByZ2JhKDI0NSwxNjYsMzUsLjM1KSk7YW5pbWF0aW9uOmZsb2F0IDMuNXMgZWFzZS1pbi1vdXQgaW5maW5pdGV9CkBrZXlmcmFtZXMgZmxvYXR7MCUsMTAwJXt0cmFuc2Zvcm06dHJhbnNsYXRlWSgwKX01MCV7dHJhbnNmb3JtOnRyYW5zbGF0ZVkoLThweCl9fQouZW1wdHktdGl0bGV7Zm9udC1mYW1pbHk6J0JlYmFzIE5ldWUnLHNhbnMtc2VyaWY7Zm9udC1zaXplOjI0cHg7bGV0dGVyLXNwYWNpbmc6MnB4OwogIGJhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDkwZGVnLHZhcigtLWdvbGQyKSx2YXIoLS10ZWFsMikpOwogIC13ZWJraXQtYmFja2dyb3VuZC1jbGlwOnRleHQ7LXdlYmtpdC10ZXh0LWZpbGwtY29sb3I6dHJhbnNwYXJlbnQ7bWFyZ2luLWJvdHRvbTo4cHh9Ci5lbXB0eS1kZXNje2ZvbnQtc2l6ZToxMnB4O2NvbG9yOnZhcigtLXQyKTttYXgtd2lkdGg6MzQwcHg7bGluZS1oZWlnaHQ6MS43fQouY3RhLXJvd3tkaXNwbGF5OmZsZXg7Z2FwOjZweDtmbGV4LXdyYXA6d3JhcDtqdXN0aWZ5LWNvbnRlbnQ6Y2VudGVyO21hcmdpbi10b3A6MTZweH0KCi5sb2FkaW5ne2Rpc3BsYXk6bm9uZTtmbGV4LWRpcmVjdGlvbjpjb2x1bW47YWxpZ24taXRlbXM6Y2VudGVyO3BhZGRpbmc6ODBweDtnYXA6MTNweDt0ZXh0LWFsaWduOmNlbnRlcn0KLmxvYWRpbmcuc2hvd3tkaXNwbGF5OmZsZXh9Ci5yaW5ne3dpZHRoOjQycHg7aGVpZ2h0OjQycHg7cG9zaXRpb246cmVsYXRpdmV9Ci5yaW5nOjpiZWZvcmUsLnJpbmc6OmFmdGVye2NvbnRlbnQ6Jyc7cG9zaXRpb246YWJzb2x1dGU7Ym9yZGVyLXJhZGl1czo1MCU7Ym9yZGVyOjJweCBzb2xpZCB0cmFuc3BhcmVudH0KLnJpbmc6OmJlZm9yZXtpbnNldDowO2JvcmRlci10b3AtY29sb3I6dmFyKC0tZ29sZCk7YW5pbWF0aW9uOnNwaW4gLjc1cyBsaW5lYXIgaW5maW5pdGV9Ci5yaW5nOjphZnRlcntpbnNldDo1cHg7Ym9yZGVyLXRvcC1jb2xvcjp2YXIoLS10ZWFsKTthbmltYXRpb246c3BpbiAuNXMgbGluZWFyIGluZmluaXRlIHJldmVyc2V9CkBrZXlmcmFtZXMgc3Bpbnt0b3t0cmFuc2Zvcm06cm90YXRlKDM2MGRlZyl9fQoubGQtc3RlcHN7ZGlzcGxheTpmbGV4O2ZsZXgtZGlyZWN0aW9uOmNvbHVtbjtnYXA6NXB4O21hcmdpbi10b3A6OHB4fQoubGQtc3RlcHtkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo2cHg7Zm9udC1zaXplOjEwcHg7Y29sb3I6dmFyKC0tdDIpOwogIG9wYWNpdHk6LjI7dHJhbnNpdGlvbjpvcGFjaXR5IC4zcztmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZX0KLmxkLXN0ZXAuYWN0aXZle29wYWNpdHk6MTtjb2xvcjp2YXIoLS1nb2xkMil9Ci5sZC1zdGVwLmRvbmV7b3BhY2l0eTouNTU7Y29sb3I6dmFyKC0tbGltZSl9CgovKiDilZDilZDilZAgR1JBUEggUEFORSDilZDilZDilZAgKi8KLmN0cmwtcm93e2Rpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjdweDtmbGV4LXdyYXA6d3JhcDttYXJnaW4tYm90dG9tOjExcHh9Ci5sdi1iYXJ7ZGlzcGxheTpmbGV4O2JhY2tncm91bmQ6dmFyKC0tcGFuZWwpO2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyMik7Ym9yZGVyLXJhZGl1czp2YXIoLS1yMik7b3ZlcmZsb3c6aGlkZGVufQoubHYtYnRue3BhZGRpbmc6NXB4IDEycHg7Ym9yZGVyOm5vbmU7YmFja2dyb3VuZDp0cmFuc3BhcmVudDtmb250LXNpemU6MTFweDtmb250LXdlaWdodDo2MDA7CiAgY3Vyc29yOnBvaW50ZXI7Y29sb3I6dmFyKC0tdDIpO2JvcmRlci1yaWdodDoxcHggc29saWQgdmFyKC0tYmRyKTt0cmFuc2l0aW9uOmFsbCAuMTNzfQoubHYtYnRuOmxhc3QtY2hpbGR7Ym9yZGVyLXJpZ2h0Om5vbmV9Ci5sdi1idG4ub257YmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoMTM1ZGVnLHJnYmEoMjQ1LDE2NiwzNSwuMTgpLHJnYmEoMCwyMDEsMTc3LC4xOCkpO2NvbG9yOnZhcigtLWdvbGQyKX0KLmJje2Rpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjRweDtmb250LXNpemU6MTBweDtjb2xvcjp2YXIoLS10Mik7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2V9Ci5iYy1pdGVte2NvbG9yOnZhcigtLXRlYWwyKTtjdXJzb3I6cG9pbnRlcn0uYmMtaXRlbTpob3Zlcnt0ZXh0LWRlY29yYXRpb246dW5kZXJsaW5lfQouYmMtc2Vwe2NvbG9yOnZhcigtLXQzKX0uYmMtY3Vye2NvbG9yOnZhcigtLXQwKTtjdXJzb3I6ZGVmYXVsdH0KLnZpZXctdG9ne2Rpc3BsYXk6ZmxleDtiYWNrZ3JvdW5kOnZhcigtLXBhbmVsKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJkcjIpO2JvcmRlci1yYWRpdXM6dmFyKC0tcjIpO292ZXJmbG93OmhpZGRlbn0KLnZ0LWJ0bntwYWRkaW5nOjVweCAxMXB4O2JvcmRlcjpub25lO2JhY2tncm91bmQ6dHJhbnNwYXJlbnQ7Zm9udC1zaXplOjEwcHg7Zm9udC13ZWlnaHQ6NjAwOwogIGN1cnNvcjpwb2ludGVyO2NvbG9yOnZhcigtLXQyKTtib3JkZXItcmlnaHQ6MXB4IHNvbGlkIHZhcigtLWJkcik7dHJhbnNpdGlvbjphbGwgLjEzc30KLnZ0LWJ0bjpsYXN0LWNoaWxke2JvcmRlci1yaWdodDpub25lfQoudnQtYnRuLm9ue2JhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDEzNWRlZyxyZ2JhKDAsMjAxLDE3NywuMTUpLHJnYmEoNTksMTU4LDI1NSwuMTUpKTtjb2xvcjp2YXIoLS10ZWFsMil9CgouZ2JveHtiYWNrZ3JvdW5kOnZhcigtLWJnMik7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1iZHIyKTtib3JkZXItcmFkaXVzOnZhcigtLXIpO292ZXJmbG93OmhpZGRlbjtib3gtc2hhZG93OnZhcigtLXNoKX0KLmctaGRye3BhZGRpbmc6OXB4IDEzcHg7Ym9yZGVyLWJvdHRvbToxcHggc29saWQgdmFyKC0tYmRyMik7CiAgZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtqdXN0aWZ5LWNvbnRlbnQ6c3BhY2UtYmV0d2VlbjsKICBiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCg5MGRlZyxyZ2JhKDI0NSwxNjYsMzUsLjA1KSx0cmFuc3BhcmVudCl9Ci5nLXRpdGxle2ZvbnQtZmFtaWx5OidCZWJhcyBOZXVlJyxzYW5zLXNlcmlmO2ZvbnQtc2l6ZToxM3B4O2xldHRlci1zcGFjaW5nOjFweDtjb2xvcjp2YXIoLS10MCl9Ci5nLXN1Yntmb250LXNpemU6OXB4O2NvbG9yOnZhcigtLXQyKTtmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZTttYXJnaW4tdG9wOjFweH0KLmctdG9vbHN7ZGlzcGxheTpmbGV4O2dhcDozcHh9Ci5nLWJ0bnt3aWR0aDoyNXB4O2hlaWdodDoyNXB4O2JhY2tncm91bmQ6dmFyKC0tcGFuZWwpO2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyMik7CiAgYm9yZGVyLXJhZGl1czo0cHg7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtqdXN0aWZ5LWNvbnRlbnQ6Y2VudGVyOwogIGN1cnNvcjpwb2ludGVyO2NvbG9yOnZhcigtLXQxKTtmb250LXNpemU6MTFweDt0cmFuc2l0aW9uOmFsbCAuMTNzfQouZy1idG46aG92ZXJ7YmFja2dyb3VuZDp2YXIoLS1nb2xkMyk7Y29sb3I6I2ZmZjtib3JkZXItY29sb3I6dmFyKC0tZ29sZCk7Ym94LXNoYWRvdzowIDAgOXB4IHJnYmEoMjQ1LDE2NiwzNSwuMyl9Ci5nLWxlZ2VuZHtkaXNwbGF5OmZsZXg7Z2FwOjExcHg7cGFkZGluZzo2cHggMTNweDtib3JkZXItdG9wOjFweCBzb2xpZCB2YXIoLS1iZHIpOwogIGJhY2tncm91bmQ6cmdiYSgyLDgsMTYsLjUpO2ZsZXgtd3JhcDp3cmFwfQoubGVnLWl7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6NHB4O2ZvbnQtc2l6ZTo5cHg7Y29sb3I6dmFyKC0tdDIpO2ZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlfQoubGVnLWR7d2lkdGg6N3B4O2hlaWdodDo3cHg7Ym9yZGVyLXJhZGl1czo1MCV9CgovKiBFUkQgKi8KI2VyZC12aWV3e2hlaWdodDo0OTBweDtwb3NpdGlvbjpyZWxhdGl2ZTtvdmVyZmxvdzpoaWRkZW47Y3Vyc29yOmdyYWI7dXNlci1zZWxlY3Q6bm9uZX0KI2VyZC12aWV3OmFjdGl2ZXtjdXJzb3I6Z3JhYmJpbmd9CiNlcmQtd29ybGR7cG9zaXRpb246YWJzb2x1dGU7dHJhbnNmb3JtLW9yaWdpbjowIDB9CiNlcmQtcGFue3dpZHRoOjEwMCU7aGVpZ2h0OjEwMCU7cG9zaXRpb246YWJzb2x1dGU7dG9wOjA7bGVmdDowfQojZXJkLWhvc3R7cG9zaXRpb246YWJzb2x1dGU7dG9wOjA7bGVmdDowO3dpZHRoOjMwMDBweDtoZWlnaHQ6MjAwMHB4fQojZXJkLXN2Z3twb3NpdGlvbjphYnNvbHV0ZTt0b3A6MDtsZWZ0OjA7cG9pbnRlci1ldmVudHM6bm9uZTtvdmVyZmxvdzp2aXNpYmxlO3dpZHRoOjMwMDBweDtoZWlnaHQ6MjAwMHB4fQouZXJkLWNhcmR7cG9zaXRpb246YWJzb2x1dGU7Ym9yZGVyLXJhZGl1czo3cHg7b3ZlcmZsb3c6aGlkZGVuO21pbi13aWR0aDoxOTBweDsKICBib3gtc2hhZG93OjAgNHB4IDE2cHggcmdiYSgwLDAsMCwuNSk7YmFja2Ryb3AtZmlsdGVyOmJsdXIoNHB4KX0KLmVyZC1oZHJ7cGFkZGluZzo3cHggMTBweDtkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo2cHg7Ym9yZGVyLWJvdHRvbToxcHggc29saWR9Ci5lcmQtZHJhZ3tjdXJzb3I6bW92ZTtkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo2cHg7ZmxleDoxO292ZXJmbG93OmhpZGRlbn0KLmVyZC1pY297Zm9udC1zaXplOjExcHg7ZmxleC1zaHJpbms6MH0KLmVyZC1uYW1le2ZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlO2ZvbnQtc2l6ZToxMHB4O2ZvbnQtd2VpZ2h0OjYwMDsKICBvdmVyZmxvdzpoaWRkZW47dGV4dC1vdmVyZmxvdzplbGxpcHNpczt3aGl0ZS1zcGFjZTpub3dyYXB9Ci5lcmQtYmFkZ2V7Zm9udC1zaXplOjdweDtmb250LXdlaWdodDo4MDA7cGFkZGluZzoxcHggNXB4O2JvcmRlci1yYWRpdXM6M3B4O2xldHRlci1zcGFjaW5nOi41cHh9Ci5lcmQtZmN7Zm9udC1zaXplOjhweDtvcGFjaXR5Oi40O2ZsZXgtc2hyaW5rOjB9Ci5lcmQtZmllbGR7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtwYWRkaW5nOjNweCA5cHg7CiAgYm9yZGVyLWJvdHRvbToxcHggc29saWQgcmdiYSgyNTUsMjU1LDI1NSwuMDQpO2N1cnNvcjpwb2ludGVyO3RyYW5zaXRpb246YmFja2dyb3VuZCAuMDhzfQouZXJkLWZpZWxkOmxhc3QtY2hpbGR7Ym9yZGVyLWJvdHRvbTpub25lfQouZXJkLWZpZWxkOmhvdmVye2JhY2tncm91bmQ6cmdiYSgyNTUsMjU1LDI1NSwuMDcpfQouZXJkLWZpZWxkLmhpe2JhY2tncm91bmQ6cmdiYSgyNDUsMTY2LDM1LC4xKTtib3JkZXItbGVmdDoycHggc29saWQgdmFyKC0tZ29sZCl9Ci5lcmQtYW5je3dpZHRoOjhweDtoZWlnaHQ6OHB4O2JvcmRlci1yYWRpdXM6NTAlO2ZsZXgtc2hyaW5rOjA7Ym9yZGVyOjEuNXB4IHNvbGlkfQouZXJkLWZue2ZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlO2ZvbnQtc2l6ZTo5cHg7Zm9udC13ZWlnaHQ6NjAwO2ZsZXg6MTsKICBvdmVyZmxvdzpoaWRkZW47dGV4dC1vdmVyZmxvdzplbGxpcHNpczt3aGl0ZS1zcGFjZTpub3dyYXA7bWFyZ2luOjAgNXB4fQouZXJkLWZ0e2ZvbnQtc2l6ZTo4cHg7Y29sb3I6dmFyKC0tdDIpO2ZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlO2ZsZXgtc2hyaW5rOjB9Ci5lcmQtZnJpc2t7d2lkdGg6NXB4O2hlaWdodDo1cHg7Ym9yZGVyLXJhZGl1czo1MCU7bWFyZ2luLWxlZnQ6NHB4O2ZsZXgtc2hyaW5rOjB9Ci8qIENhcmQgdGhlbWVzICovCi5lYy1zcmN7Ym9yZGVyOjEuNXB4IHNvbGlkIHJnYmEoMTI3LDIxNyw5OCwuMzUpO2JhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDE2MGRlZyxyZ2JhKDEyNywyMTcsOTgsLjA3KSxyZ2JhKDYsMTUsMzAsLjk3KSl9Ci5lYy1zcmMgLmVyZC1oZHJ7YmFja2dyb3VuZDpyZ2JhKDEyNywyMTcsOTgsLjEpO2JvcmRlci1jb2xvcjpyZ2JhKDEyNywyMTcsOTgsLjIpfQouZWMtc3JjIC5lcmQtbmFtZXtjb2xvcjp2YXIoLS1saW1lMil9LmVjLXNyYyAuZXJkLWJhZGdle2JhY2tncm91bmQ6cmdiYSgxMjcsMjE3LDk4LC4xOCk7Y29sb3I6dmFyKC0tbGltZTIpfQouZWMtc3JjIC5lcmQtYW5je2JvcmRlci1jb2xvcjpyZ2JhKDEyNywyMTcsOTgsLjQ1KTtiYWNrZ3JvdW5kOnJnYmEoMTI3LDIxNyw5OCwuMSl9Ci5lYy10Z3R7Ym9yZGVyOjEuNXB4IHNvbGlkIHJnYmEoNTksMTU4LDI1NSwuMzUpO2JhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDE2MGRlZyxyZ2JhKDU5LDE1OCwyNTUsLjA3KSxyZ2JhKDYsMTUsMzAsLjk3KSl9Ci5lYy10Z3QgLmVyZC1oZHJ7YmFja2dyb3VuZDpyZ2JhKDU5LDE1OCwyNTUsLjEpO2JvcmRlci1jb2xvcjpyZ2JhKDU5LDE1OCwyNTUsLjIpfQouZWMtdGd0IC5lcmQtbmFtZXtjb2xvcjp2YXIoLS1za3kyKX0uZWMtdGd0IC5lcmQtYmFkZ2V7YmFja2dyb3VuZDpyZ2JhKDU5LDE1OCwyNTUsLjE4KTtjb2xvcjp2YXIoLS1za3kyKX0KLmVjLXRndCAuZXJkLWFuY3tib3JkZXItY29sb3I6cmdiYSg1OSwxNTgsMjU1LC40NSk7YmFja2dyb3VuZDpyZ2JhKDU5LDE1OCwyNTUsLjEpfQouZWMtdHJme2JvcmRlcjoxLjVweCBzb2xpZCByZ2JhKDI0NSwxNjYsMzUsLjM1KTtiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCgxNjBkZWcscmdiYSgyNDUsMTY2LDM1LC4wNykscmdiYSg2LDE1LDMwLC45NykpfQouZWMtdHJmIC5lcmQtaGRye2JhY2tncm91bmQ6cmdiYSgyNDUsMTY2LDM1LC4xKTtib3JkZXItY29sb3I6cmdiYSgyNDUsMTY2LDM1LC4yKX0KLmVjLXRyZiAuZXJkLW5hbWV7Y29sb3I6dmFyKC0tZ29sZDIpfS5lYy10cmYgLmVyZC1iYWRnZXtiYWNrZ3JvdW5kOnJnYmEoMjQ1LDE2NiwzNSwuMTgpO2NvbG9yOnZhcigtLWdvbGQyKX0KLmVjLXRyZiAuZXJkLWFuY3tib3JkZXItY29sb3I6cmdiYSgyNDUsMTY2LDM1LC40NSk7YmFja2dyb3VuZDpyZ2JhKDI0NSwxNjYsMzUsLjEpfQouZWMtY29kZXtib3JkZXI6MS41cHggc29saWQgcmdiYSgwLDIwMSwxNzcsLjM1KTtiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCgxNjBkZWcscmdiYSgwLDIwMSwxNzcsLjA3KSxyZ2JhKDYsMTUsMzAsLjk3KSl9Ci5lYy1jb2RlIC5lcmQtaGRye2JhY2tncm91bmQ6cmdiYSgwLDIwMSwxNzcsLjEpO2JvcmRlci1jb2xvcjpyZ2JhKDAsMjAxLDE3NywuMil9Ci5lYy1jb2RlIC5lcmQtbmFtZXtjb2xvcjp2YXIoLS10ZWFsMil9LmVjLWNvZGUgLmVyZC1iYWRnZXtiYWNrZ3JvdW5kOnJnYmEoMCwyMDEsMTc3LC4xOCk7Y29sb3I6dmFyKC0tdGVhbDIpfQouZWMtY29kZSAuZXJkLWFuY3tib3JkZXItY29sb3I6cmdiYSgwLDIwMSwxNzcsLjQ1KTtiYWNrZ3JvdW5kOnJnYmEoMCwyMDEsMTc3LC4xKX0KCi8qIExldmVsIGNhcmRzICovCi5sdi1jYXJkc3tkaXNwbGF5OmdyaWQ7Z3JpZC10ZW1wbGF0ZS1jb2x1bW5zOnJlcGVhdCgzLDFmcik7Z2FwOjlweDttYXJnaW4tdG9wOjExcHh9Ci5sdi1jYXJke2JhY2tncm91bmQ6dmFyKC0tcGFuZWwpO2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyKTtib3JkZXItcmFkaXVzOnZhcigtLXIpOwogIHBhZGRpbmc6MTFweDtjdXJzb3I6cG9pbnRlcjt0cmFuc2l0aW9uOmFsbCAuMTdzfQoubHYtY2FyZDpob3Zlcntib3JkZXItY29sb3I6dmFyKC0tZ29sZCk7dHJhbnNmb3JtOnRyYW5zbGF0ZVkoLTJweCk7Ym94LXNoYWRvdzowIDZweCAxOHB4IHJnYmEoMjQ1LDE2NiwzNSwuMTQpfQoubHYtY2FyZC1pY29ue2ZvbnQtc2l6ZToxOHB4O21hcmdpbi1ib3R0b206NHB4fQoubHYtY2FyZC1uYW1le2ZvbnQtd2VpZ2h0OjcwMDtmb250LXNpemU6MTFweDttYXJnaW4tYm90dG9tOjNweH0KLmx2LWNhcmQtZGVzY3tmb250LXNpemU6OXB4O2NvbG9yOnZhcigtLXQyKTtsaW5lLWhlaWdodDoxLjR9CgovKiDilZDilZDilZAgTUVUQURBVEEg4pWQ4pWQ4pWQICovCi5tZXRhLWdyaWR7ZGlzcGxheTpncmlkO2dyaWQtdGVtcGxhdGUtY29sdW1uczoxZnIgMzA4cHg7Z2FwOjEzcHg7YWxpZ24taXRlbXM6c3RhcnR9Ci5tZXRhLXRibHtiYWNrZ3JvdW5kOnZhcigtLWJnMik7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1iZHIyKTtib3JkZXItcmFkaXVzOnZhcigtLXIpOwogIG92ZXJmbG93OmhpZGRlbjtib3gtc2hhZG93OnZhcigtLXNoKX0KLm1ldGEtaGRye2Rpc3BsYXk6Z3JpZDtwYWRkaW5nOjhweCAxMnB4OwogIGJhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDkwZGVnLHJnYmEoMjQ1LDE2NiwzNSwuMDgpLHJnYmEoMCwyMDEsMTc3LC4wNikpOwogIGJvcmRlci1ib3R0b206MXB4IHNvbGlkIHZhcigtLWJkcjIpO2ZvbnQtc2l6ZTo5cHg7Zm9udC13ZWlnaHQ6NzAwO2NvbG9yOnZhcigtLXQyKTsKICB0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7bGV0dGVyLXNwYWNpbmc6LjVweDtmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZX0KLm1ldGEtcm93e2Rpc3BsYXk6Z3JpZDtwYWRkaW5nOjhweCAxMnB4O2JvcmRlci1ib3R0b206MXB4IHNvbGlkIHZhcigtLWJkcik7CiAgY3Vyc29yOnBvaW50ZXI7dHJhbnNpdGlvbjpiYWNrZ3JvdW5kIC4xczthbGlnbi1pdGVtczpjZW50ZXJ9Ci5tZXRhLXJvdzpob3ZlcntiYWNrZ3JvdW5kOnZhcigtLXBhbmVsMil9Ci5tZXRhLXJvdy5zZWx7YmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoOTBkZWcscmdiYSgyNDUsMTY2LDM1LC4wNykscmdiYSgwLDIwMSwxNzcsLjA1KSk7CiAgYm9yZGVyLWxlZnQ6MnB4IHNvbGlkIHZhcigtLWdvbGQpfQoubWV0YS1yb3c6bGFzdC1jaGlsZHtib3JkZXItYm90dG9tOm5vbmV9CgovKiBGaWVsZCBkZXRhaWwgKi8KLmZkZXRhaWx7YmFja2dyb3VuZDp2YXIoLS1iZzIpO2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyMik7Ym9yZGVyLXJhZGl1czp2YXIoLS1yKTsKICBvdmVyZmxvdzpoaWRkZW47cG9zaXRpb246c3RpY2t5O3RvcDoxM3B4O2JveC1zaGFkb3c6dmFyKC0tc2gpfQouZmQtaGVhZHtwYWRkaW5nOjEycHggMTNweDtiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCgxMzVkZWcscmdiYSgyNDUsMTY2LDM1LC4wOCkscmdiYSgwLDIwMSwxNzcsLjA2KSk7CiAgYm9yZGVyLWJvdHRvbToxcHggc29saWQgdmFyKC0tYmRyMil9Ci5mZC1uYW1le2ZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlO2ZvbnQtc2l6ZToxM3B4O2ZvbnQtd2VpZ2h0OjYwMDsKICBiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCg5MGRlZyx2YXIoLS1nb2xkMiksdmFyKC0tdGVhbDIpKTsKICAtd2Via2l0LWJhY2tncm91bmQtY2xpcDp0ZXh0Oy13ZWJraXQtdGV4dC1maWxsLWNvbG9yOnRyYW5zcGFyZW50fQouZmQtdGJse2ZvbnQtc2l6ZTo5cHg7Y29sb3I6dmFyKC0tdDIpO2ZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlO21hcmdpbi10b3A6MnB4fQouZmQtYm9keXtwYWRkaW5nOjEycHggMTNweDttYXgtaGVpZ2h0OmNhbGMoMTAwdmggLSAyNjBweCk7b3ZlcmZsb3cteTphdXRvfQouZmQtc2Vje21hcmdpbi1ib3R0b206MTJweH0KLmZkLXNlYy10e2ZvbnQtc2l6ZTo5cHg7Zm9udC13ZWlnaHQ6NzAwO2NvbG9yOnZhcigtLXQyKTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7CiAgbGV0dGVyLXNwYWNpbmc6LjZweDttYXJnaW4tYm90dG9tOjZweDtwYWRkaW5nLWJvdHRvbTozcHg7CiAgYm9yZGVyLWJvdHRvbToxcHggc29saWQgdmFyKC0tYmRyKTtmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZX0KLmZkLXJvd3tkaXNwbGF5OmZsZXg7Z2FwOjZweDttYXJnaW4tYm90dG9tOjRweDthbGlnbi1pdGVtczpmbGV4LXN0YXJ0fQouZmQta3tmb250LXNpemU6OXB4O2ZvbnQtd2VpZ2h0OjYwMDtjb2xvcjp2YXIoLS10Mik7bWluLXdpZHRoOjkwcHg7ZmxleC1zaHJpbms6MDsKICBtYXJnaW4tdG9wOjFweDtmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZX0KLmZkLXZ7Zm9udC1zaXplOjExcHg7Y29sb3I6dmFyKC0tdDApO2xpbmUtaGVpZ2h0OjEuNX0KLmZkLWNvZGV7YmFja2dyb3VuZDpyZ2JhKDAsMCwwLC40NSk7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1iZHIyKTtib3JkZXItcmFkaXVzOjRweDsKICBwYWRkaW5nOjZweCA4cHg7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2U7Zm9udC1zaXplOjlweDtjb2xvcjp2YXIoLS10ZWFsMik7CiAgbGluZS1oZWlnaHQ6MS42O292ZXJmbG93LXg6YXV0bzt3b3JkLWJyZWFrOmJyZWFrLWFsbDttYXJnaW4tdG9wOjNweDsKICBib3JkZXItbGVmdDoycHggc29saWQgdmFyKC0tZ29sZCl9Ci5saW5lYWdlLWZsb3d7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6NXB4O2ZsZXgtd3JhcDp3cmFwO21hcmdpbi10b3A6NHB4fQoubGYtbntiYWNrZ3JvdW5kOnZhcigtLXBhbmVsKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJkcjIpO2JvcmRlci1yYWRpdXM6M3B4OwogIHBhZGRpbmc6M3B4IDdweDtmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZTtmb250LXNpemU6OXB4fQoubGYtc3Jje2JvcmRlci1jb2xvcjpyZ2JhKDEyNywyMTcsOTgsLjMpO2NvbG9yOnZhcigtLWxpbWUyKX0KLmxmLXRndHtib3JkZXItY29sb3I6cmdiYSg1OSwxNTgsMjU1LC4zKTtjb2xvcjp2YXIoLS1za3kyKX0KLmxmLWFycntjb2xvcjp2YXIoLS10Myk7Zm9udC1zaXplOjEycHh9Ci5wZGYtYnRue3dpZHRoOjEwMCU7aGVpZ2h0OjI5cHg7YmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoMTM1ZGVnLHZhcigtLWdvbGQzKSx2YXIoLS10ZWFsMykpOwogIGJvcmRlcjpub25lO2JvcmRlci1yYWRpdXM6dmFyKC0tcjIpO2NvbG9yOiNmZmY7Zm9udC1mYW1pbHk6J0JlYmFzIE5ldWUnLHNhbnMtc2VyaWY7CiAgZm9udC1zaXplOjE0cHg7bGV0dGVyLXNwYWNpbmc6MnB4O2N1cnNvcjpwb2ludGVyO3RyYW5zaXRpb246YWxsIC4xOHM7CiAgYm94LXNoYWRvdzowIDAgMTJweCByZ2JhKDI0NSwxNjYsMzUsLjIyKX0KLnBkZi1idG46aG92ZXJ7Ym94LXNoYWRvdzowIDAgMjRweCByZ2JhKDI0NSwxNjYsMzUsLjQyKTt0cmFuc2Zvcm06dHJhbnNsYXRlWSgtMXB4KX0KLmZkLWZvb3R7cGFkZGluZzo5cHggMTNweDtib3JkZXItdG9wOjFweCBzb2xpZCB2YXIoLS1iZHIpO2JhY2tncm91bmQ6cmdiYSgyLDgsMTYsLjQ1KX0KCi8qIFNPUiBjYXJkcyAqLwouc29yLWNhcmR7YmFja2dyb3VuZDpyZ2JhKDAsMCwwLC4zKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJkcik7Ym9yZGVyLXJhZGl1czo1cHg7CiAgcGFkZGluZzo1cHggOHB4O21hcmdpbi1ib3R0b206M3B4O2Rpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjVweDtjdXJzb3I6cG9pbnRlcjt0cmFuc2l0aW9uOmFsbCAuMTFzfQouc29yLWNhcmQ6aG92ZXJ7YmFja2dyb3VuZDp2YXIoLS1wYW5lbDIpfQouc29yLWNhcmQuYXV0aHtib3JkZXItbGVmdDoycHggc29saWQgdmFyKC0tZ29sZCk7YmFja2dyb3VuZDpyZ2JhKDI0NSwxNjYsMzUsLjA1KX0KLnNvci1uYW1le2ZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlO2ZvbnQtc2l6ZToxMHB4O2ZvbnQtd2VpZ2h0OjcwMDtmbGV4OjF9Ci5zb3Itc3lze2ZvbnQtc2l6ZTo4cHg7Y29sb3I6dmFyKC0tdDIpfQouc29yLWF1dGh7bWFyZ2luLWxlZnQ6YXV0bztmb250LXNpemU6OHB4O2ZvbnQtd2VpZ2h0OjcwMDtjb2xvcjp2YXIoLS1nb2xkMik7CiAgYmFja2dyb3VuZDpyZ2JhKDI0NSwxNjYsMzUsLjEyKTtwYWRkaW5nOjFweCA1cHg7Ym9yZGVyLXJhZGl1czozcHg7ZmxleC1zaHJpbms6MH0KCi8qIEhvcCByb3dzICovCi5ob3Atcm93e2Rpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjVweDttYXJnaW4tYm90dG9tOjRweH0KLmhvcC1ue3dpZHRoOjE3cHg7aGVpZ2h0OjE3cHg7Ym9yZGVyLXJhZGl1czo1MCU7YmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoMTM1ZGVnLHZhcigtLWdvbGQzKSx2YXIoLS10ZWFsMykpOwogIGRpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7anVzdGlmeS1jb250ZW50OmNlbnRlcjtmb250LXNpemU6N3B4O2ZvbnQtd2VpZ2h0OjcwMDsKICBjb2xvcjojZmZmO2ZsZXgtc2hyaW5rOjB9Ci5ob3AtbGJse2ZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlO2ZvbnQtc2l6ZTo5cHg7Y29sb3I6dmFyKC0tdDEpO2ZsZXg6MX0KLmhvcC10eXBle2ZvbnQtc2l6ZTo4cHg7Y29sb3I6dmFyKC0tdDMpfQoKLyogRXF1aXYgZmllbGRzICovCi5lcS1pdGVte2Rpc3BsYXk6aW5saW5lLWJsb2NrO21hcmdpbjoycHg7cGFkZGluZzoycHggN3B4O2JhY2tncm91bmQ6dmFyKC0tcGFuZWwpOwogIGJvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyMik7Ym9yZGVyLXJhZGl1czo4cHg7Zm9udC1zaXplOjlweDsKICBmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZTtjb2xvcjp2YXIoLS10ZWFsMik7Y3Vyc29yOnBvaW50ZXI7dHJhbnNpdGlvbjphbGwgLjFzfQouZXEtaXRlbTpob3ZlcntiYWNrZ3JvdW5kOnZhcigtLXBhbmVsMil9Ci5lcS1leGFjdHtib3JkZXItY29sb3I6cmdiYSgxMjcsMjE3LDk4LC4zNSk7Y29sb3I6dmFyKC0tbGltZTIpfQoKLyogQmFkZ2VzICovCi5yYntwYWRkaW5nOjJweCA3cHg7Ym9yZGVyLXJhZGl1czoxMHB4O2ZvbnQtc2l6ZTo5cHg7Zm9udC13ZWlnaHQ6NzAwOwogIHRleHQtdHJhbnNmb3JtOnVwcGVyY2FzZTtsZXR0ZXItc3BhY2luZzouM3B4fQoucmItSHtiYWNrZ3JvdW5kOnJnYmEoMjU1LDc5LDEwOSwuMSk7Y29sb3I6dmFyKC0tcm9zZTIpO2JvcmRlcjoxcHggc29saWQgcmdiYSgyNTUsNzksMTA5LC4yNSl9Ci5yYi1Ne2JhY2tncm91bmQ6cmdiYSgyNDUsMTY2LDM1LC4xKTtjb2xvcjp2YXIoLS1nb2xkMik7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDI0NSwxNjYsMzUsLjI1KX0KLnJiLUx7YmFja2dyb3VuZDpyZ2JhKDEyNywyMTcsOTgsLjEpO2NvbG9yOnZhcigtLWxpbWUyKTtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMTI3LDIxNyw5OCwuMjUpfQovKiBVbmlmaWVkIHNlYXJjaCBzb3VyY2UgcGlsbHMgKi8KLnNyYy1waWxse2Rpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjRweDtwYWRkaW5nOjNweCA4cHg7Ym9yZGVyLXJhZGl1czoxMHB4OwogIGZvbnQtc2l6ZTo5cHg7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2U7Y29sb3I6dmFyKC0tdDIpOwogIGJhY2tncm91bmQ6dmFyKC0tcGFuZWwpO2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyKTtjdXJzb3I6ZGVmYXVsdDt0cmFuc2l0aW9uOmFsbCAuMnN9Ci5zcmMtcGlsbC5saXZle2JvcmRlci1jb2xvcjpyZ2JhKDEyNywyMTcsOTgsLjM1KTtjb2xvcjp2YXIoLS1saW1lMil9Ci5zcmMtcGlsbC5maWxle2JvcmRlci1jb2xvcjpyZ2JhKDI0NSwxNjYsMzUsLjM1KTtjb2xvcjp2YXIoLS1nb2xkMil9Ci5zcmMtcGlsbC5tb2Nre2JvcmRlci1jb2xvcjp2YXIoLS1iZHIpO2NvbG9yOnZhcigtLXQzKX0KLnNyYy1waWxsLmFjdGl2ZXtib3JkZXItY29sb3I6cmdiYSgwLDIwMSwxNzcsLjM1KTtjb2xvcjp2YXIoLS10ZWFsMil9Ci5zcC1kb3R7d2lkdGg6NXB4O2hlaWdodDo1cHg7Ym9yZGVyLXJhZGl1czo1MCU7YmFja2dyb3VuZDp2YXIoLS10Myk7ZmxleC1zaHJpbms6MDt0cmFuc2l0aW9uOmFsbCAuM3N9Ci5zcmMtcGlsbC5saXZlICAuc3AtZG90e2JhY2tncm91bmQ6dmFyKC0tbGltZSk7Ym94LXNoYWRvdzowIDAgNXB4IHZhcigtLWxpbWUpfQouc3JjLXBpbGwuZmlsZSAgLnNwLWRvdHtiYWNrZ3JvdW5kOnZhcigtLWdvbGQpO2JveC1zaGFkb3c6MCAwIDVweCB2YXIoLS1nb2xkKX0KLnNyYy1waWxsLmFjdGl2ZSAuc3AtZG90e2JhY2tncm91bmQ6dmFyKC0tdGVhbCk7Ym94LXNoYWRvdzowIDAgNXB4IHZhcigtLXRlYWwpfQovKiBBSSBTdW1tYXJ5IGJveCBpbiBtYWluIGNvbnRlbnQgKi8KLmFpLXN1bW1hcnktYm94e21hcmdpbi1ib3R0b206MTRweDtwYWRkaW5nOjE0cHggMTZweDsKICBiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCgxMzVkZWcscmdiYSgxNTUsMTA5LDI1NSwuMDgpLHJnYmEoMCwyMDEsMTc3LC4wNikpOwogIGJvcmRlcjoxcHggc29saWQgcmdiYSgxNTUsMTA5LDI1NSwuMjIpO2JvcmRlci1yYWRpdXM6dmFyKC0tcik7Ym9yZGVyLWxlZnQ6M3B4IHNvbGlkIHZhcigtLXZpbyk7CiAgYm94LXNoYWRvdzp2YXIoLS1zaCl9Ci5haS1zdW0taGVhZGVye2Rpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjhweDttYXJnaW4tYm90dG9tOjEwcHh9Ci5haS1zdW0tdGl0bGV7Zm9udC1mYW1pbHk6J0JlYmFzIE5ldWUnLHNhbnMtc2VyaWY7Zm9udC1zaXplOjE0cHg7bGV0dGVyLXNwYWNpbmc6MXB4OwogIGJhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDkwZGVnLHZhcigtLXZpbzIpLHZhcigtLXRlYWwyKSk7CiAgLXdlYmtpdC1iYWNrZ3JvdW5kLWNsaXA6dGV4dDstd2Via2l0LXRleHQtZmlsbC1jb2xvcjp0cmFuc3BhcmVudH0KLmFpLXN1bS1iYWRnZXtmb250LXNpemU6OHB4O2ZvbnQtd2VpZ2h0OjcwMDtwYWRkaW5nOjJweCA3cHg7Ym9yZGVyLXJhZGl1czo4cHg7CiAgYmFja2dyb3VuZDpyZ2JhKDE1NSwxMDksMjU1LC4xNSk7Y29sb3I6dmFyKC0tdmlvMik7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDE1NSwxMDksMjU1LC4yNSk7CiAgZm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2V9Ci5haS1zdW0tZ3JpZHtkaXNwbGF5OmdyaWQ7Z3JpZC10ZW1wbGF0ZS1jb2x1bW5zOjFmciAxZnI7Z2FwOjEwcHg7bWFyZ2luLWJvdHRvbToxMHB4fQouYWktc3VtLWNhcmR7YmFja2dyb3VuZDpyZ2JhKDAsMCwwLC4zKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJkcik7Ym9yZGVyLXJhZGl1czo2cHg7cGFkZGluZzo5cHggMTFweH0KLmFpLXN1bS1jYXJkLWxhYmVse2ZvbnQtc2l6ZTo4cHg7Zm9udC13ZWlnaHQ6NzAwO2NvbG9yOnZhcigtLXQyKTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7CiAgbGV0dGVyLXNwYWNpbmc6LjZweDttYXJnaW4tYm90dG9tOjVweDtmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZX0KLmFpLXN1bS1jYXJkLXZhbHtmb250LXNpemU6MTFweDtjb2xvcjp2YXIoLS10MCk7bGluZS1oZWlnaHQ6MS41NX0KLmFpLXN1bS1pbnNpZ2h0e2ZvbnQtc2l6ZToxMXB4O2NvbG9yOnZhcigtLXQxKTtsaW5lLWhlaWdodDoxLjY7Zm9udC1zdHlsZTppdGFsaWM7CiAgcGFkZGluZzo4cHggMTFweDtiYWNrZ3JvdW5kOnJnYmEoMTU1LDEwOSwyNTUsLjA1KTtib3JkZXItcmFkaXVzOjVweDtib3JkZXItbGVmdDoycHggc29saWQgdmFyKC0tdmlvKX0KQGtleWZyYW1lcyBzaGltbWVyezAle2xlZnQ6LTEwMCV9NjAlLDEwMCV7bGVmdDoyMDAlfX0KQGtleWZyYW1lcyBhaVB1bHNlezAlLDEwMCV7Ym94LXNoYWRvdzowIDAgMTZweCByZ2JhKDE1NSwxMDksMjU1LC4yKX01MCV7Ym94LXNoYWRvdzowIDAgMjhweCByZ2JhKDE1NSwxMDksMjU1LC40KX19CiNhaS10cmlnZ2VyLWJ0bjpob3Zlcntib3JkZXItY29sb3I6cmdiYSgxNTUsMTA5LDI1NSwuNSk7YmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoMTM1ZGVnLHJnYmEoMTU1LDEwOSwyNTUsLjE4KSxyZ2JhKDAsMjAxLDE3NywuMTIpKTt0cmFuc2Zvcm06dHJhbnNsYXRlWSgtMXB4KTtib3gtc2hhZG93OjAgNnB4IDI0cHggcmdiYSgxNTUsMTA5LDI1NSwuMil9CiNhaS10cmlnZ2VyLWJ0bjpob3ZlciAjYWktYnRuLWFycm93e3RyYW5zZm9ybTp0cmFuc2xhdGVYKDRweCl9CkBrZXlmcmFtZXMgdHlwZUlue2Zyb217b3BhY2l0eTowO3RyYW5zZm9ybTp0cmFuc2xhdGVZKDRweCl9dG97b3BhY2l0eToxO3RyYW5zZm9ybTp0cmFuc2xhdGVZKDApfX0KLmFpLXR5cGluZ3thbmltYXRpb246dHlwZUluIC40cyBlYXNlIGZvcndhcmRzfQouZHR5cGV7YmFja2dyb3VuZDpyZ2JhKDU5LDE1OCwyNTUsLjEpO2JvcmRlcjoxcHggc29saWQgcmdiYSg1OSwxNTgsMjU1LC4yKTsKICBib3JkZXItcmFkaXVzOjNweDtwYWRkaW5nOjFweCA1cHg7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2U7Zm9udC1zaXplOjlweDtjb2xvcjp2YXIoLS1za3kyKX0KLnRhZ3tiYWNrZ3JvdW5kOnZhcigtLXBhbmVsKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJkcik7Ym9yZGVyLXJhZGl1czo4cHg7CiAgcGFkZGluZzoxcHggNnB4O2ZvbnQtc2l6ZTo4cHg7Y29sb3I6dmFyKC0tdDEpO2Rpc3BsYXk6aW5saW5lLWJsb2NrO21hcmdpbjoxcHh9Ci5jb25mLXBpbGx7ZGlzcGxheTppbmxpbmUtYmxvY2s7cGFkZGluZzoycHggN3B4O2JvcmRlci1yYWRpdXM6OHB4O2ZvbnQtc2l6ZTo4cHg7CiAgZm9udC13ZWlnaHQ6NzAwO2xldHRlci1zcGFjaW5nOi40cHg7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2V9Ci5jb25mLXBpaS1ye2JhY2tncm91bmQ6cmdiYSgyNTUsNzksMTA5LC4xMik7Y29sb3I6I0ZGODA5ODtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMjU1LDc5LDEwOSwuMjUpfQouY29uZi1waWktY3tiYWNrZ3JvdW5kOnJnYmEoMjM5LDY4LDY4LC4xMik7Y29sb3I6I0Y4NzE3MTtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMjM5LDY4LDY4LC4yNSl9Ci5jb25mLWNvbmZ7YmFja2dyb3VuZDpyZ2JhKDI0NSwxNjYsMzUsLjEyKTtjb2xvcjp2YXIoLS1nb2xkMik7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDI0NSwxNjYsMzUsLjI1KX0KLmNvbmYtaW50e2JhY2tncm91bmQ6cmdiYSg1OSwxNTgsMjU1LC4wOCk7Y29sb3I6dmFyKC0tc2t5Mik7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDU5LDE1OCwyNTUsLjIpfQouY29uZi1wdWJ7YmFja2dyb3VuZDpyZ2JhKDEyNywyMTcsOTgsLjA4KTtjb2xvcjp2YXIoLS1saW1lMik7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDEyNywyMTcsOTgsLjIpfQouaW5nLXBpbGx7Ym9yZGVyLXJhZGl1czo4cHg7cGFkZGluZzoycHggN3B4O2ZvbnQtc2l6ZTo5cHg7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2U7ZGlzcGxheTppbmxpbmUtYmxvY2t9Ci5pbmctYmF0Y2h7YmFja2dyb3VuZDpyZ2JhKDU5LDE1OCwyNTUsLjA4KTtjb2xvcjp2YXIoLS1za3kyKTtib3JkZXI6MXB4IHNvbGlkIHJnYmEoNTksMTU4LDI1NSwuMil9Ci5pbmctcnR7YmFja2dyb3VuZDpyZ2JhKDI0NSwxNjYsMzUsLjEpO2NvbG9yOnZhcigtLWdvbGQyKTtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMjQ1LDE2NiwzNSwuMil9Ci5yaXNrLXRibHtiYWNrZ3JvdW5kOnJnYmEoMCwwLDAsLjMpO2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyKTtib3JkZXItcmFkaXVzOjRweDtvdmVyZmxvdzpoaWRkZW59Ci5yaXNrLXJvd3tkaXNwbGF5OmdyaWQ7Z3JpZC10ZW1wbGF0ZS1jb2x1bW5zOjcycHggMWZyO2JvcmRlci1ib3R0b206MXB4IHNvbGlkIHZhcigtLWJkcil9Ci5yaXNrLXJvdzpsYXN0LWNoaWxke2JvcmRlci1ib3R0b206bm9uZX0KLnJpc2stbHtwYWRkaW5nOjVweCA3cHg7ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6M3B4O2ZvbnQtc2l6ZTo5cHg7Zm9udC13ZWlnaHQ6NzAwfQoucmlzay1ye3BhZGRpbmc6NXB4IDdweDtib3JkZXItbGVmdDoxcHggc29saWQgdmFyKC0tYmRyKX0KLnJpc2stcmR7Zm9udC1zaXplOjlweDtjb2xvcjp2YXIoLS10MCl9LnJpc2stcmF7Zm9udC1zaXplOjhweDtjb2xvcjp2YXIoLS10Mik7bWFyZ2luLXRvcDoxcHh9CgovKiDilZDilZDilZAgU09VUkNFUyDilZDilZDilZAgKi8KLnNyYy1jYXJkc3tkaXNwbGF5OmdyaWQ7Z3JpZC10ZW1wbGF0ZS1jb2x1bW5zOnJlcGVhdChhdXRvLWZpdCxtaW5tYXgoMjc1cHgsMWZyKSk7Z2FwOjExcHh9Ci5zcmMtY2FyZHtiYWNrZ3JvdW5kOnZhcigtLWJnMik7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1iZHIyKTtib3JkZXItcmFkaXVzOnZhcigtLXIpOwogIHBhZGRpbmc6MTNweDtib3gtc2hhZG93OnZhcigtLXNoKTt0cmFuc2l0aW9uOmFsbCAuMTdzfQouc3JjLWNhcmQ6aG92ZXJ7Ym9yZGVyLWNvbG9yOnZhcigtLWJkcjIpO3RyYW5zZm9ybTp0cmFuc2xhdGVZKC0xcHgpfQouc2MtaGRye2Rpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjdweDttYXJnaW4tYm90dG9tOjEwcHh9Ci5zYy1pY297d2lkdGg6MjVweDtoZWlnaHQ6MjVweDtib3JkZXItcmFkaXVzOjVweDtkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2p1c3RpZnktY29udGVudDpjZW50ZXI7Zm9udC1zaXplOjEycHh9Ci5zYy1uYW1le2ZvbnQtc2l6ZToxMnB4O2ZvbnQtd2VpZ2h0OjcwMH0KLnNjLXN0YXR7Zm9udC1zaXplOjlweDtmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZTttYXJnaW4tdG9wOjFweH0KLnNjLWJhZGdle3BhZGRpbmc6MnB4IDZweDtib3JkZXItcmFkaXVzOjhweDtmb250LXNpemU6OHB4O2ZvbnQtd2VpZ2h0OjcwMDtsZXR0ZXItc3BhY2luZzouNHB4O21hcmdpbi1sZWZ0OmF1dG99Ci5iLWxpdmV7YmFja2dyb3VuZDpyZ2JhKDEyNywyMTcsOTgsLjEpO2NvbG9yOnZhcigtLWxpbWUyKTtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMTI3LDIxNyw5OCwuMil9Ci5iLWZpbGV7YmFja2dyb3VuZDpyZ2JhKDI0NSwxNjYsMzUsLjEpO2NvbG9yOnZhcigtLWdvbGQyKTtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMjQ1LDE2NiwzNSwuMil9Ci5iLW1vY2t7YmFja2dyb3VuZDpyZ2JhKDc4LDExMCwxMzgsLjEpO2NvbG9yOnZhcigtLXQxKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJkcil9Ci5iLWFpe2JhY2tncm91bmQ6cmdiYSgxNTUsMTA5LDI1NSwuMSk7Y29sb3I6dmFyKC0tdmlvMik7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDE1NSwxMDksMjU1LC4yKX0KLmNvZGUtYmxre2JhY2tncm91bmQ6cmdiYSgwLDAsMCwuNDUpO2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyKTtib3JkZXItcmFkaXVzOjRweDsKICBwYWRkaW5nOjZweCA4cHg7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2U7Zm9udC1zaXplOjlweDtjb2xvcjp2YXIoLS10ZWFsMik7CiAgbGluZS1oZWlnaHQ6MS42O292ZXJmbG93LXg6YXV0bzt3b3JkLWJyZWFrOmJyZWFrLWFsbDttYXJnaW46NHB4IDB9Ci5sYW5nLXBpbGx7ZGlzcGxheTppbmxpbmUtYmxvY2s7cGFkZGluZzoxcHggNnB4O2JvcmRlci1yYWRpdXM6M3B4O2ZvbnQtc2l6ZTo4cHg7Zm9udC13ZWlnaHQ6NzAwOwogIGZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlO21hcmdpbi1ib3R0b206M3B4fQoubHAtU1FMe2JhY2tncm91bmQ6IzBkMjAzNTtjb2xvcjojMzhCREY4fS5scC1QeXRob24sLmxwLVB5U3Bhcmt7YmFja2dyb3VuZDojMGQxZDBkO2NvbG9yOiMzNEQzOTl9Ci5scC1KYXZhLC5scC1TY2FsYXtiYWNrZ3JvdW5kOiMxZDBkMGQ7Y29sb3I6I0Y4NzE3MX0KLmxwLUhRTCwubHAtU2hlbGx7YmFja2dyb3VuZDojMTgwZDFkO2NvbG9yOiNDMDg0RkN9Ci5scC1KU09OLC5scC1ZQU1MLC5scC1YTUx7YmFja2dyb3VuZDojMWQxODA4O2NvbG9yOiNGQ0QzNER9Ci5haS1pbnNpZ2h0LWNhcmR7bWFyZ2luLXRvcDo2cHg7cGFkZGluZzo4cHggMTBweDsKICBiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCgxMzVkZWcscmdiYSgxNTUsMTA5LDI1NSwuMDgpLHJnYmEoMCwyMDEsMTc3LC4wNSkpOwogIGJvcmRlcjoxcHggc29saWQgcmdiYSgxNTUsMTA5LDI1NSwuMTgpO2JvcmRlci1yYWRpdXM6NHB4O2JvcmRlci1sZWZ0OjJweCBzb2xpZCB2YXIoLS12aW8pfQouYWktaW5zaWdodC10ZXh0e2ZvbnQtc2l6ZTo5cHg7Y29sb3I6dmFyKC0tdmlvMik7bGluZS1oZWlnaHQ6MS41NTtmb250LXN0eWxlOml0YWxpY30KCi8qIOKVkOKVkOKVkCBNVUxUSSDilZDilZDilZAgKi8KLm1yZXMtY2FyZHtiYWNrZ3JvdW5kOnZhcigtLWJnMik7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1iZHIyKTtib3JkZXItcmFkaXVzOnZhcigtLXIpOwogIHBhZGRpbmc6MTNweDttYXJnaW4tYm90dG9tOjEwcHg7Ym94LXNoYWRvdzp2YXIoLS1zaCl9Ci5tcmVzLWhkcntmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZTtmb250LXNpemU6MTJweDtmb250LXdlaWdodDo2MDA7CiAgYmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoOTBkZWcsdmFyKC0tZ29sZDIpLHZhcigtLXRlYWwyKSk7CiAgLXdlYmtpdC1iYWNrZ3JvdW5kLWNsaXA6dGV4dDstd2Via2l0LXRleHQtZmlsbC1jb2xvcjp0cmFuc3BhcmVudDttYXJnaW4tYm90dG9tOjhweH0KCi8qIOKVkOKVkOKVkCBUT0FTVCDilZDilZDilZAgKi8KLnRvYXN0e3Bvc2l0aW9uOmZpeGVkO2JvdHRvbToxNnB4O3JpZ2h0OjE2cHg7ei1pbmRleDo1MDA7CiAgYmFja2dyb3VuZDp2YXIoLS1iZzIpO2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyMik7Ym9yZGVyLXJhZGl1czp2YXIoLS1yMik7CiAgcGFkZGluZzo5cHggMTNweDtmb250LXNpemU6MTFweDtjb2xvcjp2YXIoLS10MCk7Ym94LXNoYWRvdzp2YXIoLS1zaCk7CiAgdHJhbnNmb3JtOnRyYW5zbGF0ZVkoNjBweCk7dHJhbnNpdGlvbjp0cmFuc2Zvcm0gLjIyczsKICBkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo2cHg7Ym9yZGVyLWxlZnQ6M3B4IHNvbGlkIHZhcigtLWdvbGQpO21heC13aWR0aDozMjBweH0KLnRvYXN0LnNob3d7dHJhbnNmb3JtOnRyYW5zbGF0ZVkoMCl9CgovKiDilZDilZDilZAgTU9EQUwg4pWQ4pWQ4pWQICovCi5vdmVybGF5e3Bvc2l0aW9uOmZpeGVkO2luc2V0OjA7YmFja2dyb3VuZDpyZ2JhKDIsOCwxNiwuODgpO2JhY2tkcm9wLWZpbHRlcjpibHVyKDE0cHgpOwogIHotaW5kZXg6NDAwO2Rpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7anVzdGlmeS1jb250ZW50OmNlbnRlcjtwYWRkaW5nOjIwcHg7CiAgb3BhY2l0eTowO3BvaW50ZXItZXZlbnRzOm5vbmU7dHJhbnNpdGlvbjpvcGFjaXR5IC4yc30KLm92ZXJsYXkub3BlbntvcGFjaXR5OjE7cG9pbnRlci1ldmVudHM6YWxsfQoubW9kYWx7YmFja2dyb3VuZDp2YXIoLS1iZzIpO2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyMik7Ym9yZGVyLXJhZGl1czp2YXIoLS1yKTsKICB3aWR0aDoxMDAlO21heC13aWR0aDo1MDBweDttYXgtaGVpZ2h0Ojg2dmg7b3ZlcmZsb3cteTphdXRvOwogIHRyYW5zZm9ybTp0cmFuc2xhdGVZKDE2cHgpO3RyYW5zaXRpb246dHJhbnNmb3JtIC4ycztib3gtc2hhZG93OjAgMjRweCA2NHB4IHJnYmEoMCwwLDAsLjgpfQoub3ZlcmxheS5vcGVuIC5tb2RhbHt0cmFuc2Zvcm06dHJhbnNsYXRlWSgwKX0KLm1vZGFsLWhkcntwYWRkaW5nOjEzcHggMTdweDtib3JkZXItYm90dG9tOjFweCBzb2xpZCB2YXIoLS1iZHIpOwogIGRpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7anVzdGlmeS1jb250ZW50OnNwYWNlLWJldHdlZW47CiAgYmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoOTBkZWcscmdiYSgyNDUsMTY2LDM1LC4wNiksdHJhbnNwYXJlbnQpO3Bvc2l0aW9uOnN0aWNreTt0b3A6MH0KLm1vZGFsLXRpdGxle2ZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlO2ZvbnQtc2l6ZToxMnB4O2ZvbnQtd2VpZ2h0OjYwMDtjb2xvcjp2YXIoLS10ZWFsMil9Ci5tb2RhbC14e3dpZHRoOjI0cHg7aGVpZ2h0OjI0cHg7Ym9yZGVyLXJhZGl1czo0cHg7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1iZHIyKTsKICBiYWNrZ3JvdW5kOnZhcigtLXBhbmVsKTtjb2xvcjp2YXIoLS10MSk7Y3Vyc29yOnBvaW50ZXI7Zm9udC1zaXplOjEycHg7CiAgZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtqdXN0aWZ5LWNvbnRlbnQ6Y2VudGVyO3RyYW5zaXRpb246YWxsIC4xMnN9Ci5tb2RhbC14OmhvdmVye2JhY2tncm91bmQ6dmFyKC0tcm9zZSk7Ym9yZGVyLWNvbG9yOnZhcigtLXJvc2UpO2NvbG9yOiNmZmZ9Ci5tb2RhbC1ib2R5e3BhZGRpbmc6MTdweH0KLm1vZGFsLWZvb3R7cGFkZGluZzo5cHggMTdweDtib3JkZXItdG9wOjFweCBzb2xpZCB2YXIoLS1iZHIpOwogIGRpc3BsYXk6ZmxleDtqdXN0aWZ5LWNvbnRlbnQ6ZmxleC1lbmQ7YmFja2dyb3VuZDpyZ2JhKDIsOCwxNiwuNDUpfQouYWN0aW9uLWJhci1idG57ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6NnB4O3BhZGRpbmc6OXB4IDE2cHg7Ym9yZGVyOm5vbmU7Ym9yZGVyLXJhZGl1czo3cHg7Y29sb3I6I2ZmZjtmb250LXNpemU6MTJweDtmb250LXdlaWdodDo3MDA7Y3Vyc29yOnBvaW50ZXI7dHJhbnNpdGlvbjphbGwgLjJzfQojbXVsdGktYWN0aW9uLWJhcntkaXNwbGF5Om5vbmU7cGFkZGluZzoxMHB4IDE4cHg7YmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoOTBkZWcscmdiYSgyNDUsMTY2LDM1LC4wOCkscmdiYSgwLDIwMSwxNzcsLjA2KSk7CiAgYm9yZGVyLWJvdHRvbToycHggc29saWQgcmdiYSgyNDUsMTY2LDM1LC4zKTtwb3NpdGlvbjpzdGlja3k7dG9wOjA7ei1pbmRleDo1MDsKICBmbGV4LXdyYXA6d3JhcDtnYXA6OHB4O2FsaWduLWl0ZW1zOmNlbnRlcjtiYWNrZHJvcC1maWx0ZXI6Ymx1cigxMnB4KX0KI211bHRpLWFjdGlvbi1iYXIgLm1hYi10aXRsZXtmb250LWZhbWlseTonQmViYXMgTmV1ZScsc2Fucy1zZXJpZjtmb250LXNpemU6MTRweDtsZXR0ZXItc3BhY2luZzoxcHg7CiAgYmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoOTBkZWcsdmFyKC0tZ29sZDIpLHZhcigtLXRlYWwyKSk7LXdlYmtpdC1iYWNrZ3JvdW5kLWNsaXA6dGV4dDstd2Via2l0LXRleHQtZmlsbC1jb2xvcjp0cmFuc3BhcmVudDttYXJnaW4tcmlnaHQ6NHB4fQojbXVsdGktYWN0aW9uLWJhciBidXR0b257ZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6NXB4O3BhZGRpbmc6OHB4IDE0cHg7CiAgYm9yZGVyOm5vbmU7Ym9yZGVyLXJhZGl1czo2cHg7Zm9udC1zaXplOjExcHg7Zm9udC13ZWlnaHQ6NzAwO2N1cnNvcjpwb2ludGVyO3RyYW5zaXRpb246YWxsIC4xNXN9CiNtdWx0aS1hY3Rpb24tYmFyIGJ1dHRvbjpob3Zlcnt0cmFuc2Zvcm06dHJhbnNsYXRlWSgtMXB4KTtmaWx0ZXI6YnJpZ2h0bmVzcygxLjE1KX0KLm1hYi1saW5lYWdle2JhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDEzNWRlZyx2YXIoLS1nb2xkMyksdmFyKC0tdGVhbDMpKTtjb2xvcjojZmZmO2JveC1zaGFkb3c6MCAwIDE0cHggcmdiYSgyNDUsMTY2LDM1LC4zKX0KLm1hYi1tZXRhe2JhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDEzNWRlZyx2YXIoLS10ZWFsMyksdmFyKC0tc2t5KSk7Y29sb3I6I2ZmZjtib3gtc2hhZG93OjAgMCAxNHB4IHJnYmEoMCwyMDEsMTc3LC4yNSl9Ci5tYWItc3Jje2JhY2tncm91bmQ6dmFyKC0tcGFuZWwyKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJkcjIpIWltcG9ydGFudDtjb2xvcjp2YXIoLS10MCl9Ci5hY3Rpb24tYmFyLWJ0bjpob3Zlcnt0cmFuc2Zvcm06dHJhbnNsYXRlWSgtMnB4KTtmaWx0ZXI6YnJpZ2h0bmVzcygxLjE1KX0KLmJ0bi1jbG9zZXtwYWRkaW5nOjVweCAxMnB4O2JhY2tncm91bmQ6dmFyKC0tcGFuZWwpO2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyMik7CiAgYm9yZGVyLXJhZGl1czp2YXIoLS1yMik7Y29sb3I6dmFyKC0tdDEpO2ZvbnQtc2l6ZToxMXB4O2ZvbnQtd2VpZ2h0OjYwMDtjdXJzb3I6cG9pbnRlcn0KPC9zdHlsZT4KPC9oZWFkPgo8Ym9keT4KCjxoZWFkZXI+CiAgPGRpdiBjbGFzcz0ibG9nbyI+CiAgICA8ZGl2IGNsYXNzPSJsb2dvLW1hcmsiPiYjeDFGNTE3OzwvZGl2PgogICAgPGRpdj4KICAgICAgPGRpdiBjbGFzcz0ibG9nby10aXRsZSI+RGF0YUxpbmVhZ2UgSVE8L2Rpdj4KICAgICAgPGRpdiBjbGFzcz0ibG9nby1zdWIiPkludGVsbGlnZW50IExpbmVhZ2UgJmJ1bGw7IHY0LjE8L2Rpdj4KICAgIDwvZGl2PgogIDwvZGl2PgogIDxkaXYgY2xhc3M9ImFwaS1iYWRnZSI+PGRpdiBjbGFzcz0ibGl2ZS1kb3QiPjwvZGl2PjxzcGFuIGlkPSJhcGktc3RhdHVzIj5Db25uZWN0aW5nLi4uPC9zcGFuPjwvZGl2Pgo8L2hlYWRlcj4KCjxkaXYgY2xhc3M9ImFwcCI+Cgo8IS0tIOKVkOKVkOKVkCBTSURFQkFSIOKVkOKVkOKVkCAtLT4KPGFzaWRlIGNsYXNzPSJzYiI+CgogIDwhLS0g4pSA4pSAIFNJTkdMRSBVTklGSUVEIFNFQVJDSCDilIDilIAgLS0+CiAgPGRpdiBjbGFzcz0ic2Itc2VjIj4KICAgIDxkaXYgY2xhc3M9InNiLWxibCI+JiN4MUY1MEQ7IFNlYXJjaDwvZGl2PgogICAgPCEtLSBIaW50OiBvbmUgZmllbGQgT1IgbXVsdGlwbGUgZmllbGRzIGNvbW1hL25ld2xpbmUgc2VwYXJhdGVkIC0tPgogICAgPGRpdiBzdHlsZT0iZm9udC1zaXplOjlweDtjb2xvcjp2YXIoLS10Mik7bWFyZ2luLWJvdHRvbTo1cHg7bGluZS1oZWlnaHQ6MS41Ij4KICAgICAgU2luZ2xlIGZpZWxkIDxzcGFuIHN0eWxlPSJjb2xvcjp2YXIoLS10MykiPm9yPC9zcGFuPiBtdWx0aXBsZSBmaWVsZHMgKGNvbW1hIC8gb25lIHBlciBsaW5lKQogICAgPC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJzcmNoLXdyYXAiPgogICAgICA8c3BhbiBjbGFzcz0ic3JjaC1pY28iPiYjeDIzMTU7PC9zcGFuPgogICAgICA8aW5wdXQgaWQ9InEiIGNsYXNzPSJzcmNoIiB0eXBlPSJ0ZXh0IiBwbGFjZWhvbGRlcj0iVEFCTEUgIC8gIEZJRUxEICAvICBmaWVsZDEsIGZpZWxkMiwgZmllbGQzIiBhdXRvY29tcGxldGU9Im9mZiIvPgogICAgPC9kaXY+CiAgICA8IS0tIE11bHRpLWxpbmUgbW9kZSB0b2dnbGUgLS0+CiAgICA8dGV4dGFyZWEgaWQ9Im11bHRpLXRhIiBjbGFzcz0ibXVsdGktdGEiIHJvd3M9IjMiCiAgICAgIHBsYWNlaG9sZGVyPSJjdXN0b21lcl9pZCYjMTA7YWNjdF9udW0mIzEwO3RyYW5zYWN0aW9uX3RzIgogICAgICBzdHlsZT0iZGlzcGxheTpub25lO21hcmdpbi10b3A6NXB4Ij48L3RleHRhcmVhPgogICAgPGRpdiBzdHlsZT0iZGlzcGxheTpmbGV4O2dhcDo1cHg7bWFyZ2luLXRvcDo2cHgiPgogICAgICA8YnV0dG9uIGNsYXNzPSJnby1idG4iIGlkPSJnby1idG4iIG9uY2xpY2s9ImRvVW5pZmllZFNlYXJjaCgpIiBzdHlsZT0iZmxleDoxIj4mI3gyNkExOyBESVNDT1ZFUiBMSU5FQUdFPC9idXR0b24+CiAgICAgIDxidXR0b24gaWQ9Im11bHRpLXRvZ2dsZS1idG4iIHRpdGxlPSJNdWx0aS1maWVsZCBtb2RlIgogICAgICAgIG9uY2xpY2s9InRvZ2dsZU11bHRpTW9kZSgpIgogICAgICAgIHN0eWxlPSJ3aWR0aDozMnB4O2hlaWdodDozNHB4O2JhY2tncm91bmQ6dmFyKC0tcGFuZWwpO2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyMik7CiAgICAgICAgICAgICAgIGJvcmRlci1yYWRpdXM6dmFyKC0tcjIpO2NvbG9yOnZhcigtLXQxKTtmb250LXNpemU6MTNweDtjdXJzb3I6cG9pbnRlcjsKICAgICAgICAgICAgICAgdHJhbnNpdGlvbjphbGwgLjE1cztmbGV4LXNocmluazowIgogICAgICAgID4mI3gxRjUwMDs8L2J1dHRvbj4KICAgIDwvZGl2PgogICAgPCEtLSBTb3VyY2Ugc3RhdHVzIHBpbGxzIOKAlCBhbHdheXMgdmlzaWJsZSBiZWxvdyBzZWFyY2ggLS0+CiAgICA8ZGl2IGlkPSJzcmMtc3RhdHVzLXJvdyIgc3R5bGU9ImRpc3BsYXk6ZmxleDtmbGV4LXdyYXA6d3JhcDtnYXA6NHB4O21hcmdpbi10b3A6OHB4Ij4KICAgICAgPGRpdiBjbGFzcz0ic3JjLXBpbGwiIGlkPSJwaWxsLWVyd2luIiB0aXRsZT0iRVJXSU4gREk6IEFQSSDihpIgTG9jYWwg4oaSIE1vY2siPgogICAgICAgIDxzcGFuIGNsYXNzPSJzcC1kb3QiIGlkPSJzZG90LWVyd2luIj48L3NwYW4+JiN4MUYzREI7IEVSV0lOIERJCiAgICAgIDwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJzcmMtcGlsbCIgaWQ9InBpbGwtZGljdCIgdGl0bGU9IkRhdGEgRGljdGlvbmFyeTogQVBJIOKGkiBMb2NhbCDihpIgTW9jayI+CiAgICAgICAgPHNwYW4gY2xhc3M9InNwLWRvdCIgaWQ9InNkb3QtZGljdCI+PC9zcGFuPiYjeDFGNEQ2OyBEaWN0CiAgICAgIDwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJzcmMtcGlsbCIgaWQ9InBpbGwtY29kZSIgdGl0bGU9IkNvZGUgUmVwbzogR2l0SHViIOKGkiBMb2NhbCDihpIgTW9jayI+CiAgICAgICAgPHNwYW4gY2xhc3M9InNwLWRvdCIgaWQ9InNkb3QtY29kZSI+PC9zcGFuPiYjeDFGNEJCOyBDb2RlCiAgICAgIDwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJzcmMtcGlsbCIgaWQ9InBpbGwtdGFjaHlvbiIgdGl0bGU9IlRhY2h5b24gQUk6IExMTSBzdW1tYXJ5IGlmIHJlYWNoYWJsZSI+CiAgICAgICAgPHNwYW4gY2xhc3M9InNwLWRvdCIgaWQ9InNkb3QtdGFjaHlvbiI+PC9zcGFuPiYjeDFGOTE2OyBBSQogICAgICA8L2Rpdj4KICAgIDwvZGl2PgogIDwvZGl2PgoKICA8IS0tIOKUgOKUgCBRVUlDSyBDSElQUyDilIDilIAgLS0+CiAgPGRpdiBjbGFzcz0ic2Itc2VjIj4KICAgIDxkaXYgY2xhc3M9InNiLWxibCI+JiN4MjZBMTsgUXVpY2s8L2Rpdj4KICAgIDxkaXYgY2xhc3M9ImNoaXBzIj4KICAgICAgPGRpdiBjbGFzcz0iY2hpcCIgb25jbGljaz0icXVpY2tTZWFyY2goJ0NVU1RPTUVSJykiPkNVU1RPTUVSPC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9ImNoaXAiIG9uY2xpY2s9InF1aWNrU2VhcmNoKCdPUkRFUl9GQUNUJykiPk9SREVSX0ZBQ1Q8L2Rpdj4KICAgICAgPGRpdiBjbGFzcz0iY2hpcCIgb25jbGljaz0icXVpY2tTZWFyY2goJ1NBTEVTJykiPlNBTEVTPC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9ImNoaXAiIG9uY2xpY2s9InF1aWNrU2VhcmNoKCdFTVBMT1lFRScpIj5FTVBMT1lFRTwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJjaGlwIiBvbmNsaWNrPSJxdWlja1NlYXJjaCgnUFJPRFVDVCcpIj5QUk9EVUNUPC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9ImNoaXAiIG9uY2xpY2s9InF1aWNrU2VhcmNoKCdTQUxFUy5SRVZFTlVFX0FNVCcpIj5TQUxFUy5SRVZFTlVFX0FNVDwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJjaGlwIiBvbmNsaWNrPSJxdWlja1NlYXJjaCgnRU1QTE9ZRUUuRU1QX0lEJykiPkVNUExPWUVFLkVNUF9JRDwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJjaGlwIiBvbmNsaWNrPSJxdWlja1NlYXJjaCgnRUNOX1BSSU1BUllfSUQnKSI+RUNOX1BSSU1BUllfSUQ8L2Rpdj4KICAgICAgPGRpdiBjbGFzcz0iY2hpcCIgb25jbGljaz0icXVpY2tTZWFyY2goJ0ZJQ09TQ09SRVY5U0NPUkUnKSI+RklDT1NDT1JFVjlTQ09SRTwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJjaGlwIiBvbmNsaWNrPSJxdWlja1NlYXJjaCgnUlREX0NIRUNLX01URF9EVCcpIj5SVERfQ0hFQ0tfTVREX0RUPC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9ImNoaXAiIG9uY2xpY2s9InF1aWNrU2VhcmNoKCdNQUNfREVMUScpIj5NQUNfREVMUTwvZGl2PgogICAgPC9kaXY+CiAgPC9kaXY+CgogIDwhLS0gSGlkZGVuIHRvZ2dsZXMga2VwdCBmb3IgSlMgY29tcGF0aWJpbGl0eSAtLT4KICA8ZGl2IHN0eWxlPSJkaXNwbGF5Om5vbmUiPgogICAgPGRpdiBpZD0idG9nLWVyd2luIiAgY2xhc3M9InNyYy10b2cgb24iPjwvZGl2PgogICAgPGRpdiBpZD0idG9nLWRpY3QiICAgY2xhc3M9InNyYy10b2cgb24iPjwvZGl2PgogICAgPGRpdiBpZD0idG9nLWNvZGUiICAgY2xhc3M9InNyYy10b2cgb24iPjwvZGl2PgogICAgPGRpdiBpZD0idG9nLXRhY2h5b24iIGNsYXNzPSJzcmMtdG9nIG9uIj48L2Rpdj4KICAgIDxkaXYgaWQ9ImNoay1lcndpbiI+JiN4MjcxMzs8L2Rpdj4KICAgIDxkaXYgaWQ9ImNoay1kaWN0Ij4mI3gyNzEzOzwvZGl2PgogICAgPGRpdiBpZD0iY2hrLWNvZGUiPiYjeDI3MTM7PC9kaXY+CiAgICA8ZGl2IGlkPSJjaGstdGFjaHlvbiI+JiN4MjcxMzs8L2Rpdj4KICAgIDxkaXYgaWQ9ImRvdC1lcndpbiI+PC9kaXY+CiAgICA8ZGl2IGlkPSJkb3QtZGljdCI+PC9kaXY+CiAgICA8ZGl2IGlkPSJkb3QtY29kZSI+PC9kaXY+CiAgICA8ZGl2IGlkPSJkb3QtdGFjaHlvbiI+PC9kaXY+CiAgICA8ZGl2IGlkPSJhaS1zZWMiPjwvZGl2PgogICAgPGRpdiBpZD0iYWktdGV4dCI+PC9kaXY+CiAgPC9kaXY+CgogIDxkaXYgY2xhc3M9InNiLXNlYyI+CiAgICA8ZGl2IGNsYXNzPSJzYi1sYmwiPiYjeDFGNEMyOyBVcGxvYWQgSW5wdXRzPC9kaXY+CiAgICA8ZGl2IHN0eWxlPSJmb250LXNpemU6OXB4O2NvbG9yOnZhcigtLXQyKTttYXJnaW4tYm90dG9tOjVweDtsaW5lLWhlaWdodDoxLjQiPlVwbG9hZCBCUiB4bHN4IG9yIElOUFVUUyB6aXA8L2Rpdj4KICAgIDxpbnB1dCB0eXBlPSJmaWxlIiBpZD0iZmlsZS1pbnB1dCIgYWNjZXB0PSIueGxzeCwuemlwIiBzdHlsZT0iZGlzcGxheTpub25lIiBvbmNoYW5nZT0iZG9VcGxvYWQodGhpcykiLz4KICAgIDxidXR0b24gY2xhc3M9InVwbG9hZC1idG4iIG9uY2xpY2s9ImRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdmaWxlLWlucHV0JykuY2xpY2soKSI+JiN4MUY0QzE7IENob29zZSBGaWxlICgueGxzeCAvIC56aXApPC9idXR0b24+CiAgICA8ZGl2IGlkPSJ1cGxvYWQtc3RhdHVzIj48L2Rpdj4KICA8L2Rpdj4KCiAgPGRpdiBjbGFzcz0ic2Itc2VjIj4KICAgIDxkZXRhaWxzPgogICAgICA8c3VtbWFyeT4mI3gyNjk5OyYjeEZFMEY7IENvbmZpZ3VyYXRpb248L3N1bW1hcnk+CiAgICAgIDxkaXYgc3R5bGU9Im1hcmdpbi10b3A6OXB4Ij4KICAgICAgICA8ZGl2IGNsYXNzPSJjZmctc2VjIj4mI3gxRjNEQjsgRVJXSU4gREkg4oCUIENvbm5lY3Rpb248L2Rpdj4KCiAgICAgICAgPCEtLSDilIDilIAgU2hhcmVkIGNvbm5lY3Rpb24gKGFwcGxpZXMgdG8gQUxMIGVuZHBvaW50IGNhbGxzKSDilIDilIAgLS0+CiAgICAgICAgPGxhYmVsIGNsYXNzPSJjZmctbGJsIj5CYXNlIFVSTCA8c3BhbiBzdHlsZT0iY29sb3I6dmFyKC0tcm9zZTIpO2ZvbnQtc2l6ZTo4cHgiPiogYWxsIGVuZHBvaW50cyB1c2UgdGhpczwvc3Bhbj48L2xhYmVsPgogICAgICAgIDxpbnB1dCBjbGFzcz0iY2ZnLWluIiBpZD0iY2YtZXJ3aW4tdXJsIiBwbGFjZWhvbGRlcj0iaHR0cHM6Ly9scHZyYTk3YTAxMDQud2VsbHNmYXJnby5uZXQ6NDUxMi9lcndpbkRJU3VpdGUiLz4KICAgICAgICA8bGFiZWwgY2xhc3M9ImNmZy1sYmwiPkFQSSBLZXkgPHNwYW4gc3R5bGU9ImNvbG9yOnZhcigtLXJvc2UyKTtmb250LXNpemU6OHB4Ij4qIHJlcXVpcmVkIGZvciBsaXZlIGNhbGxzPC9zcGFuPjwvbGFiZWw+CiAgICAgICAgPGlucHV0IGNsYXNzPSJjZmctaW4iIGlkPSJjZi1lcndpbi1rZXkiIHR5cGU9InBhc3N3b3JkIiBwbGFjZWhvbGRlcj0ieW91ci1hcGkta2V5LXZhbHVlIi8+CgogICAgICAgIDwhLS0g4pSA4pSAIERlZmF1bHQgcGFyYW1zIChhdXRvLWZpbGxlZCBmb3IgZW5kcG9pbnRzIHRoYXQgYWNjZXB0IHRoZW0pIOKUgOKUgCAtLT4KICAgICAgICA8ZGl2IHN0eWxlPSJtYXJnaW46OHB4IDAgNHB4O3BhZGRpbmc6NXB4IDdweDtiYWNrZ3JvdW5kOnJnYmEoMCwyMDEsMTc3LC4wNik7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDAsMjAxLDE3NywuMTUpO2JvcmRlci1yYWRpdXM6NHB4O2ZvbnQtc2l6ZTo5cHg7Y29sb3I6dmFyKC0tdGVhbDIpO2ZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlIj4KICAgICAgICAgICYjeDIxMzk7IERlZmF1bHQgcGFyYW1zIGJlbG93IGFyZSBzZW50IGF1dG9tYXRpY2FsbHkgdG8gYW55IGVuZHBvaW50IHRoYXQgYWNjZXB0cyB0aGVtLiBPdmVycmlkZSBwZXItZW5kcG9pbnQgaW4gdGhlIEVSV0lOIEFQSXMgdGFiLgogICAgICAgIDwvZGl2PgoKICAgICAgICA8bGFiZWwgY2xhc3M9ImNmZy1sYmwiPnN5c3RlbU5hbWUgPHNwYW4gc3R5bGU9ImNvbG9yOnZhcigtLXQzKTtmb250LXNpemU6OHB4Ij51c2VkIGJ5OiBkYXRhbGluZWFnZS90YWJsZSwgL3N5c3RlbSwgL29iamVjdHMsIC9tYXBwaW5ncy4uLjwvc3Bhbj48L2xhYmVsPgogICAgICAgIDxpbnB1dCBjbGFzcz0iY2ZnLWluIiBpZD0iY2YtZXJ3aW4tc3lzdGVtIiBwbGFjZWhvbGRlcj0iRk5JU0giLz4KCiAgICAgICAgPGxhYmVsIGNsYXNzPSJjZmctbGJsIj5lbnZpcm9ubWVudE5hbWUgPHNwYW4gc3R5bGU9ImNvbG9yOnZhcigtLXQzKTtmb250LXNpemU6OHB4Ij51c2VkIGJ5OiBkYXRhbGluZWFnZS90YWJsZSwgL2NvbHVtbiwgL29iamVjdHMuLi48L3NwYW4+PC9sYWJlbD4KICAgICAgICA8aW5wdXQgY2xhc3M9ImNmZy1pbiIgaWQ9ImNmLWVyd2luLWVudiIgcGxhY2Vob2xkZXI9IkNEU19EQ0JURElTUFVURSIvPgoKICAgICAgICA8bGFiZWwgY2xhc3M9ImNmZy1sYmwiPmxpbmVhZ2VUeXBlIDxzcGFuIHN0eWxlPSJjb2xvcjp2YXIoLS10Myk7Zm9udC1zaXplOjhweCI+dXNlZCBieTogZGF0YWxpbmVhZ2UvdGFibGUsIC9jb2x1bW4sIC9zeXN0ZW08L3NwYW4+PC9sYWJlbD4KICAgICAgICA8c2VsZWN0IGNsYXNzPSJjZmctaW4iIGlkPSJjZi1lcndpbi1saW50eXBlIiBzdHlsZT0iY3Vyc29yOnBvaW50ZXIiPgogICAgICAgICAgPG9wdGlvbiB2YWx1ZT0iRFVBTCIgc2VsZWN0ZWQ+RFVBTCAoc291cmNlICsgdGFyZ2V0KTwvb3B0aW9uPgogICAgICAgICAgPG9wdGlvbiB2YWx1ZT0iU09VUkNFIj5TT1VSQ0Ugb25seTwvb3B0aW9uPgogICAgICAgICAgPG9wdGlvbiB2YWx1ZT0iVEFSR0VUIj5UQVJHRVQgb25seTwvb3B0aW9uPgogICAgICAgIDwvc2VsZWN0PgoKICAgICAgICA8bGFiZWwgY2xhc3M9ImNmZy1sYmwiPnByb2plY3RJZHMgPHNwYW4gc3R5bGU9ImNvbG9yOnZhcigtLXQzKTtmb250LXNpemU6OHB4Ij51c2VkIGJ5OiBkYXRhbGluZWFnZS90YWJsZSwgL3N5c3RlbSwgL3RhYmxlcy4uLjwvc3Bhbj48L2xhYmVsPgogICAgICAgIDxpbnB1dCBjbGFzcz0iY2ZnLWluIiBpZD0iY2YtZXJ3aW4tcHJvamVjdCIgcGxhY2Vob2xkZXI9ImxlYXZlIGJsYW5rIGZvciBhbGwgcHJvamVjdHMiLz4KCiAgICAgICAgPCEtLSDilIDilIAgUGVyLWVuZHBvaW50IG92ZXJyaWRlcyBzZWN0aW9uIOKUgOKUgCAtLT4KICAgICAgICA8ZGl2IHN0eWxlPSJtYXJnaW46MTBweCAwIDVweDtmb250LXNpemU6OXB4O2ZvbnQtd2VpZ2h0OjcwMDtjb2xvcjp2YXIoLS1nb2xkMik7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2U7bGV0dGVyLXNwYWNpbmc6LjVweCI+UEVSLUVORFBPSU5UIE9WRVJSSURFUzwvZGl2PgogICAgICAgIDxkaXYgc3R5bGU9ImZvbnQtc2l6ZTo5cHg7Y29sb3I6dmFyKC0tdDIpO21hcmdpbi1ib3R0b206N3B4O2xpbmUtaGVpZ2h0OjEuNSI+CiAgICAgICAgICBUaGVzZSBvdmVycmlkZSBkZWZhdWx0cyBmb3Igc3BlY2lmaWMgZW5kcG9pbnQgdHlwZXMgb25seS4KICAgICAgICA8L2Rpdj4KCiAgICAgICAgPCEtLSBkYXRhbGluZWFnZS90YWJsZSBvdmVycmlkZSAtLT4KICAgICAgICA8ZGV0YWlscyBzdHlsZT0ibWFyZ2luLWJvdHRvbTo1cHgiPgogICAgICAgICAgPHN1bW1hcnkgc3R5bGU9ImZvbnQtc2l6ZTo5cHg7Y29sb3I6dmFyKC0tdGVhbDIpO2N1cnNvcjpwb2ludGVyO3BhZGRpbmc6NHB4IDdweDtiYWNrZ3JvdW5kOnZhcigtLXBhbmVsKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJkcik7Ym9yZGVyLXJhZGl1czo0cHg7bGlzdC1zdHlsZTpub25lIj4KICAgICAgICAgICAgJiN4MjVCNjsgL2FwaS9kYXRhbGluZWFnZS90YWJsZSAmYW1wOyAvc3lzdGVtIG92ZXJyaWRlcwogICAgICAgICAgPC9zdW1tYXJ5PgogICAgICAgICAgPGRpdiBzdHlsZT0icGFkZGluZzo3cHggMCAzcHgiPgogICAgICAgICAgICA8bGFiZWwgY2xhc3M9ImNmZy1sYmwiPnRhYmxlTmFtZSA8c3BhbiBzdHlsZT0iY29sb3I6dmFyKC0tdDMpO2ZvbnQtc2l6ZTo4cHgiPihsZWF2ZSBibGFuayDihpIgdXNlIHNlYXJjaCBxdWVyeSk8L3NwYW4+PC9sYWJlbD4KICAgICAgICAgICAgPGlucHV0IGNsYXNzPSJjZmctaW4iIGlkPSJjZi1lcC10YWJsZW5hbWUiIHBsYWNlaG9sZGVyPSJlLmcuIFRfRENCTU9fRElTUF9PUklHX0FDQ1QiLz4KICAgICAgICAgIDwvZGl2PgogICAgICAgIDwvZGV0YWlscz4KCiAgICAgICAgPCEtLSBkYXRhbGluZWFnZS9jb2x1bW4gb3ZlcnJpZGUgLS0+CiAgICAgICAgPGRldGFpbHMgc3R5bGU9Im1hcmdpbi1ib3R0b206NXB4Ij4KICAgICAgICAgIDxzdW1tYXJ5IHN0eWxlPSJmb250LXNpemU6OXB4O2NvbG9yOnZhcigtLXRlYWwyKTtjdXJzb3I6cG9pbnRlcjtwYWRkaW5nOjRweCA3cHg7YmFja2dyb3VuZDp2YXIoLS1wYW5lbCk7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1iZHIpO2JvcmRlci1yYWRpdXM6NHB4O2xpc3Qtc3R5bGU6bm9uZSI+CiAgICAgICAgICAgICYjeDI1QjY7IC9hcGkvZGF0YWxpbmVhZ2UvY29sdW1uIG92ZXJyaWRlcwogICAgICAgICAgPC9zdW1tYXJ5PgogICAgICAgICAgPGRpdiBzdHlsZT0icGFkZGluZzo3cHggMCAzcHgiPgogICAgICAgICAgICA8bGFiZWwgY2xhc3M9ImNmZy1sYmwiPmNvbHVtbk5hbWUgPHNwYW4gc3R5bGU9ImNvbG9yOnZhcigtLXQzKTtmb250LXNpemU6OHB4Ij4obGVhdmUgYmxhbmsg4oaSIHVzZSBzZWFyY2ggcXVlcnkpPC9zcGFuPjwvbGFiZWw+CiAgICAgICAgICAgIDxpbnB1dCBjbGFzcz0iY2ZnLWluIiBpZD0iY2YtZXAtY29sbmFtZSIgcGxhY2Vob2xkZXI9ImUuZy4gQUNDVF9OVU0iLz4KICAgICAgICAgIDwvZGl2PgogICAgICAgIDwvZGV0YWlscz4KCiAgICAgICAgPCEtLSBvYmplY3RzL3NlYXJjaCBvdmVycmlkZSAtLT4KICAgICAgICA8ZGV0YWlscyBzdHlsZT0ibWFyZ2luLWJvdHRvbTo1cHgiPgogICAgICAgICAgPHN1bW1hcnkgc3R5bGU9ImZvbnQtc2l6ZTo5cHg7Y29sb3I6dmFyKC0tdGVhbDIpO2N1cnNvcjpwb2ludGVyO3BhZGRpbmc6NHB4IDdweDtiYWNrZ3JvdW5kOnZhcigtLXBhbmVsKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJkcik7Ym9yZGVyLXJhZGl1czo0cHg7bGlzdC1zdHlsZTpub25lIj4KICAgICAgICAgICAgJiN4MjVCNjsgL2FwaS9vYmplY3RzL3NlYXJjaCBvdmVycmlkZXMKICAgICAgICAgIDwvc3VtbWFyeT4KICAgICAgICAgIDxkaXYgc3R5bGU9InBhZGRpbmc6N3B4IDAgM3B4Ij4KICAgICAgICAgICAgPGxhYmVsIGNsYXNzPSJjZmctbGJsIj50eXBlIDxzcGFuIHN0eWxlPSJjb2xvcjp2YXIoLS10Myk7Zm9udC1zaXplOjhweCI+KFRBQkxFIC8gQ09MVU1OIC8gTUFQUElORyk8L3NwYW4+PC9sYWJlbD4KICAgICAgICAgICAgPHNlbGVjdCBjbGFzcz0iY2ZnLWluIiBpZD0iY2YtZXAtb2JqdHlwZSIgc3R5bGU9ImN1cnNvcjpwb2ludGVyIj4KICAgICAgICAgICAgICA8b3B0aW9uIHZhbHVlPSJUQUJMRSIgc2VsZWN0ZWQ+VEFCTEU8L29wdGlvbj4KICAgICAgICAgICAgICA8b3B0aW9uIHZhbHVlPSJDT0xVTU4iPkNPTFVNTjwvb3B0aW9uPgogICAgICAgICAgICAgIDxvcHRpb24gdmFsdWU9Ik1BUFBJTkciPk1BUFBJTkc8L29wdGlvbj4KICAgICAgICAgICAgPC9zZWxlY3Q+CiAgICAgICAgICA8L2Rpdj4KICAgICAgICA8L2RldGFpbHM+CgogICAgICAgIDwhLS0gbWFwcGluZ3Mgb3ZlcnJpZGUgLS0+CiAgICAgICAgPGRldGFpbHMgc3R5bGU9Im1hcmdpbi1ib3R0b206NXB4Ij4KICAgICAgICAgIDxzdW1tYXJ5IHN0eWxlPSJmb250LXNpemU6OXB4O2NvbG9yOnZhcigtLXRlYWwyKTtjdXJzb3I6cG9pbnRlcjtwYWRkaW5nOjRweCA3cHg7YmFja2dyb3VuZDp2YXIoLS1wYW5lbCk7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1iZHIpO2JvcmRlci1yYWRpdXM6NHB4O2xpc3Qtc3R5bGU6bm9uZSI+CiAgICAgICAgICAgICYjeDI1QjY7IC9hcGkvbWFwcGluZ3Mgb3ZlcnJpZGVzCiAgICAgICAgICA8L3N1bW1hcnk+CiAgICAgICAgICA8ZGl2IHN0eWxlPSJwYWRkaW5nOjdweCAwIDNweCI+CiAgICAgICAgICAgIDxsYWJlbCBjbGFzcz0iY2ZnLWxibCI+TWFwcGluZyBuYW1lIGZpbHRlciA8c3BhbiBzdHlsZT0iY29sb3I6dmFyKC0tdDMpO2ZvbnQtc2l6ZTo4cHgiPihvcHRpb25hbCk8L3NwYW4+PC9sYWJlbD4KICAgICAgICAgICAgPGlucHV0IGNsYXNzPSJjZmctaW4iIGlkPSJjZi1lcC1tYXBuYW1lIiBwbGFjZWhvbGRlcj0iZS5nLiBNQVBfRElTUFVURV9MT0FEIi8+CiAgICAgICAgICA8L2Rpdj4KICAgICAgICA8L2RldGFpbHM+CgogICAgICAgIDwhLS0gZ2xvc3Nhcnkgb3ZlcnJpZGUgLS0+CiAgICAgICAgPGRldGFpbHMgc3R5bGU9Im1hcmdpbi1ib3R0b206NXB4Ij4KICAgICAgICAgIDxzdW1tYXJ5IHN0eWxlPSJmb250LXNpemU6OXB4O2NvbG9yOnZhcigtLXRlYWwyKTtjdXJzb3I6cG9pbnRlcjtwYWRkaW5nOjRweCA3cHg7YmFja2dyb3VuZDp2YXIoLS1wYW5lbCk7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1iZHIpO2JvcmRlci1yYWRpdXM6NHB4O2xpc3Qtc3R5bGU6bm9uZSI+CiAgICAgICAgICAgICYjeDI1QjY7IC9hcGkvZ2xvc3NhcnkvdGVybXMgb3ZlcnJpZGVzCiAgICAgICAgICA8L3N1bW1hcnk+CiAgICAgICAgICA8ZGl2IHN0eWxlPSJwYWRkaW5nOjdweCAwIDNweCI+CiAgICAgICAgICAgIDxsYWJlbCBjbGFzcz0iY2ZnLWxibCI+Q2F0ZWdvcnkgZmlsdGVyIDxzcGFuIHN0eWxlPSJjb2xvcjp2YXIoLS10Myk7Zm9udC1zaXplOjhweCI+KG9wdGlvbmFsKTwvc3Bhbj48L2xhYmVsPgogICAgICAgICAgICA8aW5wdXQgY2xhc3M9ImNmZy1pbiIgaWQ9ImNmLWVwLWdsb3NzY2F0IiBwbGFjZWhvbGRlcj0iZS5nLiBGaW5hbmNlLCBDdXN0b21lciIvPgogICAgICAgICAgPC9kaXY+CiAgICAgICAgPC9kZXRhaWxzPgoKICAgICAgICA8bGFiZWwgY2xhc3M9ImNmZy1sYmwiIHN0eWxlPSJtYXJnaW4tdG9wOjhweCI+TG9jYWwgTWV0YWRhdGEgRGlyIChmYWxsYmFjayk8L2xhYmVsPgogICAgICAgIDxpbnB1dCBjbGFzcz0iY2ZnLWluIiBpZD0iY2YtbWV0YSIgcGxhY2Vob2xkZXI9Ii4vbWV0YWRhdGFfZmlsZXMiLz4KICAgICAgICA8ZGl2IGNsYXNzPSJjZmctc2VjIj4mI3gxRjRENjsgRGF0YSBEaWN0aW9uYXJ5PC9kaXY+CiAgICAgICAgPGxhYmVsIGNsYXNzPSJjZmctbGJsIj5Mb2NhbCBEaWN0aW9uYXJ5IERpcjwvbGFiZWw+CiAgICAgICAgPGlucHV0IGNsYXNzPSJjZmctaW4iIGlkPSJjZi1kaWN0IiBwbGFjZWhvbGRlcj0iLi9kaWN0aW9uYXJ5X2ZpbGVzIi8+CiAgICAgICAgPGRpdiBjbGFzcz0iY2ZnLXNlYyI+JiN4MUY0QkI7IENvZGUgUmVwbzwvZGl2PgogICAgICAgIDxsYWJlbCBjbGFzcz0iY2ZnLWxibCI+R2l0SHViIFRva2VuPC9sYWJlbD4KICAgICAgICA8aW5wdXQgY2xhc3M9ImNmZy1pbiIgaWQ9ImNmLWdoLXRva2VuIiB0eXBlPSJwYXNzd29yZCIgcGxhY2Vob2xkZXI9ImdocF94eHh4eCIvPgogICAgICAgIDxsYWJlbCBjbGFzcz0iY2ZnLWxibCI+R2l0SHViIFJlcG8gKG93bmVyL3JlcG8pPC9sYWJlbD4KICAgICAgICA8aW5wdXQgY2xhc3M9ImNmZy1pbiIgaWQ9ImNmLWdoLXJlcG8iIHBsYWNlaG9sZGVyPSJteS1vcmcvZGF0YS1wbGF0Zm9ybSIvPgogICAgICAgIDxsYWJlbCBjbGFzcz0iY2ZnLWxibCI+TG9jYWwgQ29kZSBEaXI8L2xhYmVsPgogICAgICAgIDxpbnB1dCBjbGFzcz0iY2ZnLWluIiBpZD0iY2YtbG9jYWwiIHBsYWNlaG9sZGVyPSIuL2NvZGVfcmVwb19zYW1wbGVzIi8+CiAgICAgICAgPGRpdiBjbGFzcz0iY2ZnLXNlYyI+JiN4MUY5MTY7IFRhY2h5b24gTExNPC9kaXY+CiAgICAgICAgPGxhYmVsIGNsYXNzPSJjZmctbGJsIj5UYWNoeW9uIEFQSSBVUkw8L2xhYmVsPgogICAgICAgIDxpbnB1dCBjbGFzcz0iY2ZnLWluIiBpZD0iY2YtdGFjaHlvbi11cmwiIHBsYWNlaG9sZGVyPSJodHRwOi8vdGFjaHlvbi1hcGkvdjEvY2hhdCIvPgogICAgICAgIDxsYWJlbCBjbGFzcz0iY2ZnLWxibCI+VGFjaHlvbiBBUEkgS2V5PC9sYWJlbD4KICAgICAgICA8aW5wdXQgY2xhc3M9ImNmZy1pbiIgaWQ9ImNmLXRhY2h5b24ta2V5IiB0eXBlPSJwYXNzd29yZCIgcGxhY2Vob2xkZXI9ImtleS4uLiIvPgogICAgICA8L2Rpdj4KICAgIDwvZGV0YWlscz4KICA8L2Rpdj4KCiAgPGRpdiBjbGFzcz0ic2Itc2VjIiBpZD0icmVjZW50LXNlYyIgc3R5bGU9ImRpc3BsYXk6bm9uZSI+CiAgICA8ZGl2IGNsYXNzPSJzYi1sYmwiPiYjeDFGNTUzOyBSZWNlbnQ8L2Rpdj4KICAgIDxkaXYgaWQ9InJlY2VudC1saXN0Ij48L2Rpdj4KICA8L2Rpdj4KCjwvYXNpZGU+Cgo8IS0tIOKVkOKVkOKVkCBNQUlOIOKVkOKVkOKVkCAtLT4KPGRpdiBjbGFzcz0ibWFpbiI+CgogIDwhLS0gU3RpY2t5IG11bHRpLWZpZWxkIGFjdGlvbiBiYXIgLSBzaG93biBhZnRlciBtdWx0aS1zZWFyY2ggLS0+CiAgPGRpdiBpZD0ibXVsdGktYWN0aW9uLWJhciI+CiAgICA8c3BhbiBjbGFzcz0ibWFiLXRpdGxlIj4mI3gxRjUwMDsgTVVMVEktRklFTEQgUkVTVUxUUzwvc3Bhbj4KICAgIDxidXR0b24gY2xhc3M9Im1hYi1saW5lYWdlIiBvbmNsaWNrPSJzaG93VGFiKCdsaW5lYWdlJyk7c2V0VGltZW91dChmdW5jdGlvbigpe3NldExldmVsKCd0YWJsZScpO30sODApIj4KICAgICAgJiN4MUY1Nzg7IFZpZXcgQ29tYmluZWQgTGluZWFnZSBHcmFwaAogICAgPC9idXR0b24+CiAgICA8YnV0dG9uIGNsYXNzPSJtYWItbWV0YSIgb25jbGljaz0ic2hvd1RhYignbWV0YWRhdGEnKSI+CiAgICAgICYjeDFGNENCOyBWaWV3IEFsbCBNZXRhZGF0YQogICAgPC9idXR0b24+CiAgICA8YnV0dG9uIGNsYXNzPSJtYWItc3JjIiBvbmNsaWNrPSJzaG93VGFiKCdzb3VyY2VzJykiPgogICAgICAmI3gxRjRFMTsgVmlldyBTb3VyY2VzCiAgICA8L2J1dHRvbj4KICAgIDxzcGFuIGlkPSJtYWItc3RhdHMiIHN0eWxlPSJtYXJnaW4tbGVmdDphdXRvO2ZvbnQtc2l6ZToxMHB4O2NvbG9yOnZhcigtLXQyKTtmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZSI+PC9zcGFuPgogIDwvZGl2PgoKICA8ZGl2IGNsYXNzPSJ0b3AtYmFyIiBpZD0idG9wLWJhciIgc3R5bGU9ImRpc3BsYXk6bm9uZSI+CiAgICA8ZGl2IGNsYXNzPSJyZXMtYmFyIj4KICAgICAgPGRpdiBjbGFzcz0icmVzLXRpdGxlIiBpZD0icmVzLXRpdGxlIj5SZXN1bHRzPC9kaXY+CiAgICAgIDxkaXYgaWQ9InN0YXQtcGlsbHMiIHN0eWxlPSJkaXNwbGF5OmZsZXg7Z2FwOjVweDtmbGV4LXdyYXA6d3JhcCI+PC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9ImV4cC1yb3ciIGlkPSJleHAtcm93IiBzdHlsZT0iZGlzcGxheTpub25lIj4KICAgICAgICA8YnV0dG9uIGNsYXNzPSJleHAtYnRuIGV4cC14bCIgb25jbGljaz0iZG9FeHBvcnRFeGNlbCgpIj4mI3gxRjRDQTsgRXhjZWw8L2J1dHRvbj4KICAgICAgICA8YnV0dG9uIGNsYXNzPSJleHAtYnRuIGV4cC1qcyIgb25jbGljaz0iZG9FeHBvcnRKU09OKCkiPiYjeDFGNEU2OyBKU09OPC9idXR0b24+CiAgICAgIDwvZGl2PgogICAgPC9kaXY+CiAgICA8ZGl2IGNsYXNzPSJ0YWJzIj4KICAgICAgPGJ1dHRvbiBjbGFzcz0idGFiLWJ0biBvbiIgaWQ9InRiLWxpbmVhZ2UiICBvbmNsaWNrPSJzaG93VGFiKCdsaW5lYWdlJykiPiYjeDFGNTc4OyBMaW5lYWdlPC9idXR0b24+CiAgICAgIDxidXR0b24gY2xhc3M9InRhYi1idG4iICAgIGlkPSJ0Yi1tZXRhZGF0YSIgb25jbGljaz0ic2hvd1RhYignbWV0YWRhdGEnKSI+JiN4MUY0Q0I7IE1ldGFkYXRhPC9idXR0b24+CiAgICAgIDxidXR0b24gY2xhc3M9InRhYi1idG4iICAgIGlkPSJ0Yi1zb3VyY2VzIiAgb25jbGljaz0ic2hvd1RhYignc291cmNlcycpIj4mI3gxRjRFMTsgU291cmNlczwvYnV0dG9uPgogICAgICA8YnV0dG9uIGNsYXNzPSJ0YWItYnRuIiAgICBpZD0idGItbXVsdGkiICAgIG9uY2xpY2s9InNob3dUYWIoJ211bHRpJykiPiYjeDFGNTAwOyBNdWx0aS1GaWVsZDwvYnV0dG9uPgogICAgICA8YnV0dG9uIGNsYXNzPSJ0YWItYnRuIiAgICBpZD0idGItZXJ3aW4iICAgIG9uY2xpY2s9InNob3dUYWIoJ2Vyd2luJykiPiYjeDFGM0RCOyBFUldJTiBBUElzPC9idXR0b24+CiAgICA8L2Rpdj4KICA8L2Rpdj4KCiAgPGRpdiBjbGFzcz0iY29udGVudCIgaWQ9ImNvbnRlbnQiPgoKICAgIDwhLS0gRW1wdHkgc3RhdGUgLS0+CiAgICA8ZGl2IGNsYXNzPSJlbXB0eSIgaWQ9InN0LWVtcHR5Ij4KICAgICAgPGRpdiBjbGFzcz0iZW1wdHktaWNvbiI+JiN4MUY1MTc7PC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9ImVtcHR5LXRpdGxlIj5EYXRhTGluZWFnZSBJUTwvZGl2PgogICAgICA8ZGl2IGNsYXNzPSJlbXB0eS1kZXNjIj5EaXNjb3ZlciBjb21wbGV0ZSBlbmQtdG8tZW5kIGRhdGEgbGluZWFnZS4gRVJXSU4gREksIERhdGEgRGljdGlvbmFyeSwgQ29kZSBSZXBvcyAmYW1wOyBUYWNoeW9uIEFJICZtZGFzaDsgYWxsIGluIG9uZSBwbGF0Zm9ybS48L2Rpdj4KICAgICAgPGRpdiBjbGFzcz0iY3RhLXJvdyI+CiAgICAgICAgPGRpdiBjbGFzcz0iY2hpcCIgb25jbGljaz0ic2V0USgnQ1VTVE9NRVInKTtkb1NlYXJjaCgpIj4mI3gyNUI2OyBSdW4gQ1VTVE9NRVI8L2Rpdj4KICAgICAgICA8ZGl2IGNsYXNzPSJjaGlwIiBvbmNsaWNrPSJzZXRRKCdTQUxFUy5SRVZFTlVFX0FNVCcpO2RvU2VhcmNoKCkiPiYjeDI1QjY7IFJ1biBTQUxFUy5SRVZFTlVFX0FNVDwvZGl2PgogICAgICAgIDxkaXYgY2xhc3M9ImNoaXAiIG9uY2xpY2s9InNldFEoJ0VDTl9QUklNQVJZX0lEJyk7ZG9TZWFyY2goKSI+JiN4MjVCNjsgUnVuIEVDTl9QUklNQVJZX0lEPC9kaXY+CiAgICAgIDwvZGl2PgogICAgPC9kaXY+CgogICAgPCEtLSBMb2FkaW5nIHN0YXRlIC0tPgogICAgPGRpdiBjbGFzcz0ibG9hZGluZyIgaWQ9InN0LWxvYWRpbmciPgogICAgICA8ZGl2IGNsYXNzPSJyaW5nIj48L2Rpdj4KICAgICAgPGRpdiBzdHlsZT0iZm9udC1mYW1pbHk6J0JlYmFzIE5ldWUnLHNhbnMtc2VyaWY7Zm9udC1zaXplOjE3cHg7bGV0dGVyLXNwYWNpbmc6MnB4O2JhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDkwZGVnLHZhcigtLWdvbGQyKSx2YXIoLS10ZWFsMikpOy13ZWJraXQtYmFja2dyb3VuZC1jbGlwOnRleHQ7LXdlYmtpdC10ZXh0LWZpbGwtY29sb3I6dHJhbnNwYXJlbnQiPkRpc2NvdmVyaW5nIExpbmVhZ2UuLi48L2Rpdj4KICAgICAgPGRpdiBjbGFzcz0ibGQtc3RlcHMiPgogICAgICAgIDxkaXYgY2xhc3M9ImxkLXN0ZXAiIGlkPSJsZC1lcndpbiI+JiN4MUYzREI7IEVSV0lOIERJICZtZGFzaDsgQVBJICZyYXJyOyBsb2NhbCBmYWxsYmFjazwvZGl2PgogICAgICAgIDxkaXYgY2xhc3M9ImxkLXN0ZXAiIGlkPSJsZC1kaWN0Ij4mI3gxRjRENjsgRGF0YSBEaWN0aW9uYXJ5ICZtZGFzaDsgQVBJICZyYXJyOyBsb2NhbCBmYWxsYmFjazwvZGl2PgogICAgICAgIDxkaXYgY2xhc3M9ImxkLXN0ZXAiIGlkPSJsZC1jb2RlIj4mI3gxRjRCQjsgQ29kZSBSZXBvc2l0b3J5ICZtZGFzaDsgR2l0SHViICZyYXJyOyBsb2NhbCBzY2FuPC9kaXY+CiAgICAgICAgPGRpdiBjbGFzcz0ibGQtc3RlcCIgaWQ9ImxkLXRhY2h5b24iPiYjeDFGOTE2OyBUYWNoeW9uIEFJIGVucmljaG1lbnQ8L2Rpdj4KICAgICAgICA8ZGl2IGNsYXNzPSJsZC1zdGVwIiBpZD0ibGQtZ3JhcGgiPiYjeDFGNTc4OyBCdWlsZGluZyBsaW5lYWdlIGdyYXBoPC9kaXY+CiAgICAgIDwvZGl2PgogICAgPC9kaXY+CgogICAgPCEtLSDilIDilIAgVEFDSFlPTiBBSSDigJQgVFJJR0dFUiBCVVRUT04gKGFsd2F5cyB2aXNpYmxlIGFmdGVyIHNlYXJjaCkg4pSA4pSAIC0tPgogICAgPGRpdiBpZD0iYWktdHJpZ2dlci1yb3ciIHN0eWxlPSJkaXNwbGF5Om5vbmU7bWFyZ2luLWJvdHRvbToxNHB4Ij4KICAgICAgPGJ1dHRvbiBpZD0iYWktdHJpZ2dlci1idG4iIG9uY2xpY2s9ImdlbmVyYXRlQUlTdW1tYXJ5KCkiCiAgICAgICAgc3R5bGU9ImRpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjEwcHg7d2lkdGg6MTAwJTtwYWRkaW5nOjEzcHggMThweDsKICAgICAgICAgIGJhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDEzNWRlZyxyZ2JhKDE1NSwxMDksMjU1LC4xMikscmdiYSgwLDIwMSwxNzcsLjA4KSk7CiAgICAgICAgICBib3JkZXI6MXB4IHNvbGlkIHJnYmEoMTU1LDEwOSwyNTUsLjI4KTtib3JkZXItcmFkaXVzOnZhcigtLXIpO2N1cnNvcjpwb2ludGVyOwogICAgICAgICAgdHJhbnNpdGlvbjphbGwgLjI1czt0ZXh0LWFsaWduOmxlZnQ7cG9zaXRpb246cmVsYXRpdmU7b3ZlcmZsb3c6aGlkZGVuIj4KICAgICAgICA8IS0tIEFuaW1hdGVkIHNoaW1tZXIgbGluZSAtLT4KICAgICAgICA8ZGl2IHN0eWxlPSJwb3NpdGlvbjphYnNvbHV0ZTt0b3A6MDtsZWZ0Oi0xMDAlO3dpZHRoOjEwMCU7aGVpZ2h0OjJweDsKICAgICAgICAgIGJhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDkwZGVnLHRyYW5zcGFyZW50LHJnYmEoMTU1LDEwOSwyNTUsLjYpLHRyYW5zcGFyZW50KTsKICAgICAgICAgIGFuaW1hdGlvbjpzaGltbWVyIDIuNXMgZWFzZS1pbi1vdXQgaW5maW5pdGUiPjwvZGl2PgogICAgICAgIDwhLS0gSWNvbiAtLT4KICAgICAgICA8ZGl2IHN0eWxlPSJ3aWR0aDozNnB4O2hlaWdodDozNnB4O2JvcmRlci1yYWRpdXM6OXB4O2ZsZXgtc2hyaW5rOjA7CiAgICAgICAgICBiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCgxMzVkZWcscmdiYSgxNTUsMTA5LDI1NSwuMjUpLHJnYmEoMCwyMDEsMTc3LC4yKSk7CiAgICAgICAgICBib3JkZXI6MXB4IHNvbGlkIHJnYmEoMTU1LDEwOSwyNTUsLjM1KTtkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyOwogICAgICAgICAganVzdGlmeS1jb250ZW50OmNlbnRlcjtmb250LXNpemU6MTZweCI+JiN4MUY5MTY7PC9kaXY+CiAgICAgICAgPCEtLSBUZXh0IC0tPgogICAgICAgIDxkaXYgc3R5bGU9ImZsZXg6MSI+CiAgICAgICAgICA8ZGl2IHN0eWxlPSJmb250LWZhbWlseTonQmViYXMgTmV1ZScsc2Fucy1zZXJpZjtmb250LXNpemU6MTVweDtsZXR0ZXItc3BhY2luZzoxLjVweDsKICAgICAgICAgICAgYmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoOTBkZWcsdmFyKC0tdmlvMiksdmFyKC0tdGVhbDIpKTsKICAgICAgICAgICAgLXdlYmtpdC1iYWNrZ3JvdW5kLWNsaXA6dGV4dDstd2Via2l0LXRleHQtZmlsbC1jb2xvcjp0cmFuc3BhcmVudCI+CiAgICAgICAgICAgIEdlbmVyYXRlIEFJIFByb2Nlc3MgU3VtbWFyeQogICAgICAgICAgPC9kaXY+CiAgICAgICAgICA8ZGl2IHN0eWxlPSJmb250LXNpemU6MTBweDtjb2xvcjp2YXIoLS10Mik7bWFyZ2luLXRvcDoycHg7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2UiPgogICAgICAgICAgICBUYWNoeW9uIExMTSDigKIgU291cmNlIOKGkiBFVEwg4oaSIFRhcmdldCDigKIgUmlzayBhbmFseXNpcyDigKIgRnVsbCBkYXRhIGZsb3cKICAgICAgICAgIDwvZGl2PgogICAgICAgIDwvZGl2PgogICAgICAgIDwhLS0gQXJyb3cgLS0+CiAgICAgICAgPGRpdiBzdHlsZT0iZm9udC1zaXplOjE4cHg7Y29sb3I6dmFyKC0tdmlvMik7b3BhY2l0eTouNzt0cmFuc2l0aW9uOnRyYW5zZm9ybSAuMnMiIGlkPSJhaS1idG4tYXJyb3ciPiYjeDIxOTI7PC9kaXY+CiAgICAgIDwvYnV0dG9uPgogICAgPC9kaXY+CgogICAgPCEtLSDilIDilIAgVEFDSFlPTiBBSSBTVU1NQVJZIFBBTkVMIChzaG93biBvbmx5IGFmdGVyIGNsaWNraW5nIEdlbmVyYXRlKSDilIDilIAgLS0+CiAgICA8ZGl2IGlkPSJhaS1zdW1tYXJ5LWJveCIgc3R5bGU9ImRpc3BsYXk6bm9uZTttYXJnaW4tYm90dG9tOjE0cHgiPgoKICAgICAgPCEtLSBIZWFkZXIgYmFyIC0tPgogICAgICA8ZGl2IHN0eWxlPSJkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDoxMHB4O3BhZGRpbmc6MTRweCAxOHB4IDA7CiAgICAgICAgYmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoMTM1ZGVnLHJnYmEoMTU1LDEwOSwyNTUsLjEpLHJnYmEoMCwyMDEsMTc3LC4wNikpOwogICAgICAgIGJvcmRlcjoxcHggc29saWQgcmdiYSgxNTUsMTA5LDI1NSwuMjUpO2JvcmRlci1yYWRpdXM6dmFyKC0tcikgdmFyKC0tcikgMCAwOwogICAgICAgIGJvcmRlci1ib3R0b206bm9uZTtwb3NpdGlvbjpyZWxhdGl2ZTtvdmVyZmxvdzpoaWRkZW4iPgogICAgICAgIDwhLS0gVG9wIGFjY2VudCBsaW5lIC0tPgogICAgICAgIDxkaXYgc3R5bGU9InBvc2l0aW9uOmFic29sdXRlO3RvcDowO2xlZnQ6MDtyaWdodDowO2hlaWdodDoycHg7CiAgICAgICAgICBiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCg5MGRlZyx2YXIoLS12aW8pLHZhcigtLXRlYWwpLHZhcigtLXNreSkpIj48L2Rpdj4KICAgICAgICA8IS0tIEljb24gZ2xvdyBjaXJjbGUgLS0+CiAgICAgICAgPGRpdiBzdHlsZT0id2lkdGg6MzhweDtoZWlnaHQ6MzhweDtib3JkZXItcmFkaXVzOjEwcHg7ZmxleC1zaHJpbms6MDsKICAgICAgICAgIGJhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDEzNWRlZyxyZ2JhKDE1NSwxMDksMjU1LC4zKSxyZ2JhKDAsMjAxLDE3NywuMikpOwogICAgICAgICAgYm9yZGVyOjFweCBzb2xpZCByZ2JhKDE1NSwxMDksMjU1LC40KTtkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyOwogICAgICAgICAganVzdGlmeS1jb250ZW50OmNlbnRlcjtmb250LXNpemU6MThweDsKICAgICAgICAgIGJveC1zaGFkb3c6MCAwIDIwcHggcmdiYSgxNTUsMTA5LDI1NSwuMjUpIj4mI3gxRjkxNjs8L2Rpdj4KICAgICAgICA8ZGl2IHN0eWxlPSJmbGV4OjEiPgogICAgICAgICAgPGRpdiBzdHlsZT0iZm9udC1mYW1pbHk6J0JlYmFzIE5ldWUnLHNhbnMtc2VyaWY7Zm9udC1zaXplOjE2cHg7bGV0dGVyLXNwYWNpbmc6MS41cHg7CiAgICAgICAgICAgIGJhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDkwZGVnLHZhcigtLXZpbzIpLHZhcigtLXRlYWwyKSk7CiAgICAgICAgICAgIC13ZWJraXQtYmFja2dyb3VuZC1jbGlwOnRleHQ7LXdlYmtpdC10ZXh0LWZpbGwtY29sb3I6dHJhbnNwYXJlbnQiPgogICAgICAgICAgICBUYWNoeW9uIEFJIOKAlCBQcm9jZXNzIFN1bW1hcnkKICAgICAgICAgIDwvZGl2PgogICAgICAgICAgPGRpdiBzdHlsZT0iZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6NnB4O21hcmdpbi10b3A6MnB4Ij4KICAgICAgICAgICAgPHNwYW4gaWQ9ImFpLXN1bS1iYWRnZSIgc3R5bGU9ImZvbnQtc2l6ZTo4cHg7Zm9udC13ZWlnaHQ6NzAwO3BhZGRpbmc6MnB4IDhweDsKICAgICAgICAgICAgICBib3JkZXItcmFkaXVzOjEwcHg7YmFja2dyb3VuZDpyZ2JhKDE1NSwxMDksMjU1LC4xNSk7Y29sb3I6dmFyKC0tdmlvMik7CiAgICAgICAgICAgICAgYm9yZGVyOjFweCBzb2xpZCByZ2JhKDE1NSwxMDksMjU1LC4yNSk7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2UiPgogICAgICAgICAgICAgIFNUUlVDVFVSQUwgU1VNTUFSWQogICAgICAgICAgICA8L3NwYW4+CiAgICAgICAgICAgIDxzcGFuIGlkPSJhaS1zdW0tc3RhdHVzIiBzdHlsZT0iZm9udC1zaXplOjlweDtjb2xvcjp2YXIoLS10Mik7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2UiPjwvc3Bhbj4KICAgICAgICAgIDwvZGl2PgogICAgICAgIDwvZGl2PgogICAgICAgIDxidXR0b24gb25jbGljaz0iY2xvc2VBSVN1bW1hcnkoKSIKICAgICAgICAgIHN0eWxlPSJ3aWR0aDoyNnB4O2hlaWdodDoyNnB4O2JvcmRlci1yYWRpdXM6NnB4O2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyMik7CiAgICAgICAgICAgIGJhY2tncm91bmQ6dmFyKC0tcGFuZWwpO2NvbG9yOnZhcigtLXQxKTtjdXJzb3I6cG9pbnRlcjtmb250LXNpemU6MTJweDsKICAgICAgICAgICAgZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtqdXN0aWZ5LWNvbnRlbnQ6Y2VudGVyO3RyYW5zaXRpb246YWxsIC4xNXM7ZmxleC1zaHJpbms6MCI+JiN4MjcxNTs8L2J1dHRvbj4KICAgICAgPC9kaXY+CgogICAgICA8IS0tIDQtY2FyZCBncmlkIC0tPgogICAgICA8ZGl2IHN0eWxlPSJkaXNwbGF5OmdyaWQ7Z3JpZC10ZW1wbGF0ZS1jb2x1bW5zOjFmciAxZnI7Z2FwOjFweDsKICAgICAgICBiYWNrZ3JvdW5kOnJnYmEoMTU1LDEwOSwyNTUsLjE1KTtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMTU1LDEwOSwyNTUsLjIpO2JvcmRlci10b3A6bm9uZSI+CgogICAgICAgIDxkaXYgc3R5bGU9ImJhY2tncm91bmQ6dmFyKC0tYmcyKTtwYWRkaW5nOjEzcHggMTVweCI+CiAgICAgICAgICA8ZGl2IHN0eWxlPSJmb250LXNpemU6OHB4O2ZvbnQtd2VpZ2h0OjcwMDtjb2xvcjp2YXIoLS10Mik7dGV4dC10cmFuc2Zvcm06dXBwZXJjYXNlOwogICAgICAgICAgICBsZXR0ZXItc3BhY2luZzouN3B4O21hcmdpbi1ib3R0b206OHB4O2ZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlOwogICAgICAgICAgICBkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo1cHgiPgogICAgICAgICAgICA8c3BhbiBzdHlsZT0id2lkdGg6NnB4O2hlaWdodDo2cHg7Ym9yZGVyLXJhZGl1czo1MCU7YmFja2dyb3VuZDp2YXIoLS1saW1lKTtkaXNwbGF5OmlubGluZS1ibG9jayI+PC9zcGFuPgogICAgICAgICAgICBTb3VyY2UgU3lzdGVtcwogICAgICAgICAgPC9kaXY+CiAgICAgICAgICA8ZGl2IGlkPSJhaS1zdW0tc291cmNlcyIgc3R5bGU9ImZvbnQtc2l6ZToxMXB4O2NvbG9yOnZhcigtLXQwKTtsaW5lLWhlaWdodDoxLjYiPuKAlDwvZGl2PgogICAgICAgIDwvZGl2PgoKICAgICAgICA8ZGl2IHN0eWxlPSJiYWNrZ3JvdW5kOnZhcigtLWJnMik7cGFkZGluZzoxM3B4IDE1cHgiPgogICAgICAgICAgPGRpdiBzdHlsZT0iZm9udC1zaXplOjhweDtmb250LXdlaWdodDo3MDA7Y29sb3I6dmFyKC0tdDIpO3RleHQtdHJhbnNmb3JtOnVwcGVyY2FzZTsKICAgICAgICAgICAgbGV0dGVyLXNwYWNpbmc6LjdweDttYXJnaW4tYm90dG9tOjhweDtmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZTsKICAgICAgICAgICAgZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6NXB4Ij4KICAgICAgICAgICAgPHNwYW4gc3R5bGU9IndpZHRoOjZweDtoZWlnaHQ6NnB4O2JvcmRlci1yYWRpdXM6NTAlO2JhY2tncm91bmQ6dmFyKC0tc2t5KTtkaXNwbGF5OmlubGluZS1ibG9jayI+PC9zcGFuPgogICAgICAgICAgICBUYXJnZXQgU3lzdGVtcwogICAgICAgICAgPC9kaXY+CiAgICAgICAgICA8ZGl2IGlkPSJhaS1zdW0tdGFyZ2V0cyIgc3R5bGU9ImZvbnQtc2l6ZToxMXB4O2NvbG9yOnZhcigtLXQwKTtsaW5lLWhlaWdodDoxLjYiPuKAlDwvZGl2PgogICAgICAgIDwvZGl2PgoKICAgICAgICA8ZGl2IHN0eWxlPSJiYWNrZ3JvdW5kOnZhcigtLWJnMik7cGFkZGluZzoxM3B4IDE1cHgiPgogICAgICAgICAgPGRpdiBzdHlsZT0iZm9udC1zaXplOjhweDtmb250LXdlaWdodDo3MDA7Y29sb3I6dmFyKC0tdDIpO3RleHQtdHJhbnNmb3JtOnVwcGVyY2FzZTsKICAgICAgICAgICAgbGV0dGVyLXNwYWNpbmc6LjdweDttYXJnaW4tYm90dG9tOjhweDtmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZTsKICAgICAgICAgICAgZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6NXB4Ij4KICAgICAgICAgICAgPHNwYW4gc3R5bGU9IndpZHRoOjZweDtoZWlnaHQ6NnB4O2JvcmRlci1yYWRpdXM6NTAlO2JhY2tncm91bmQ6dmFyKC0tZ29sZCk7ZGlzcGxheTppbmxpbmUtYmxvY2siPjwvc3Bhbj4KICAgICAgICAgICAgRVRMIC8gVHJhbnNmb3JtCiAgICAgICAgICA8L2Rpdj4KICAgICAgICAgIDxkaXYgaWQ9ImFpLXN1bS10cmFuc2Zvcm0iIHN0eWxlPSJmb250LXNpemU6MTFweDtjb2xvcjp2YXIoLS10MCk7bGluZS1oZWlnaHQ6MS42Ij7igJQ8L2Rpdj4KICAgICAgICA8L2Rpdj4KCiAgICAgICAgPGRpdiBzdHlsZT0iYmFja2dyb3VuZDp2YXIoLS1iZzIpO3BhZGRpbmc6MTNweCAxNXB4Ij4KICAgICAgICAgIDxkaXYgc3R5bGU9ImZvbnQtc2l6ZTo4cHg7Zm9udC13ZWlnaHQ6NzAwO2NvbG9yOnZhcigtLXQyKTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7CiAgICAgICAgICAgIGxldHRlci1zcGFjaW5nOi43cHg7bWFyZ2luLWJvdHRvbTo4cHg7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2U7CiAgICAgICAgICAgIGRpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjVweCI+CiAgICAgICAgICAgIDxzcGFuIHN0eWxlPSJ3aWR0aDo2cHg7aGVpZ2h0OjZweDtib3JkZXItcmFkaXVzOjUwJTtiYWNrZ3JvdW5kOnZhcigtLXJvc2UpO2Rpc3BsYXk6aW5saW5lLWJsb2NrIj48L3NwYW4+CiAgICAgICAgICAgIFJpc2sgJmFtcDsgQ2xhc3NpZmljYXRpb24KICAgICAgICAgIDwvZGl2PgogICAgICAgICAgPGRpdiBpZD0iYWktc3VtLXJpc2siIHN0eWxlPSJmb250LXNpemU6MTFweDtjb2xvcjp2YXIoLS10MCk7bGluZS1oZWlnaHQ6MS42Ij7igJQ8L2Rpdj4KICAgICAgICA8L2Rpdj4KICAgICAgPC9kaXY+CgogICAgICA8IS0tIEluc2lnaHQgbmFycmF0aXZlIC0tPgogICAgICA8ZGl2IHN0eWxlPSJiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCgxMzVkZWcscmdiYSgxNTUsMTA5LDI1NSwuMDYpLHJnYmEoMCwyMDEsMTc3LC4wNCkpOwogICAgICAgIGJvcmRlcjoxcHggc29saWQgcmdiYSgxNTUsMTA5LDI1NSwuMTgpO2JvcmRlci10b3A6bm9uZTtwYWRkaW5nOjE0cHggMTZweDsKICAgICAgICBib3JkZXItcmFkaXVzOjAgMCB2YXIoLS1yKSB2YXIoLS1yKSI+CiAgICAgICAgPGRpdiBzdHlsZT0iZm9udC1zaXplOjhweDtmb250LXdlaWdodDo3MDA7Y29sb3I6dmFyKC0tdmlvMik7dGV4dC10cmFuc2Zvcm06dXBwZXJjYXNlOwogICAgICAgICAgbGV0dGVyLXNwYWNpbmc6LjdweDttYXJnaW4tYm90dG9tOjdweDtmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZSI+CiAgICAgICAgICAmI3gyNzI4OyBBSSBJbnNpZ2h0CiAgICAgICAgPC9kaXY+CiAgICAgICAgPGRpdiBpZD0iYWktc3VtLWluc2lnaHQiIHN0eWxlPSJmb250LXNpemU6MTJweDtjb2xvcjp2YXIoLS10MSk7bGluZS1oZWlnaHQ6MS43O2ZvbnQtc3R5bGU6aXRhbGljIj48L2Rpdj4KCiAgICAgICAgPCEtLSBEYXRhIGZsb3cgc3RyaXAgLS0+CiAgICAgICAgPGRpdiBpZD0iYWktc3VtLWZsb3ciIHN0eWxlPSJtYXJnaW4tdG9wOjEycHg7ZGlzcGxheTpub25lO3BhZGRpbmctdG9wOjExcHg7CiAgICAgICAgICBib3JkZXItdG9wOjFweCBzb2xpZCByZ2JhKDE1NSwxMDksMjU1LC4xMikiPgogICAgICAgICAgPGRpdiBzdHlsZT0iZm9udC1zaXplOjhweDtmb250LXdlaWdodDo3MDA7Y29sb3I6dmFyKC0tdDIpO3RleHQtdHJhbnNmb3JtOnVwcGVyY2FzZTsKICAgICAgICAgICAgbGV0dGVyLXNwYWNpbmc6LjdweDttYXJnaW4tYm90dG9tOjdweDtmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZSI+CiAgICAgICAgICAgIENvbXBsZXRlIERhdGEgRmxvdwogICAgICAgICAgPC9kaXY+CiAgICAgICAgICA8ZGl2IGlkPSJhaS1zdW0tZmxvdy1jb250ZW50IiBzdHlsZT0iZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtmbGV4LXdyYXA6d3JhcDtnYXA6NnB4Ij48L2Rpdj4KICAgICAgICA8L2Rpdj4KCiAgICAgICAgPCEtLSBUYWNoeW9uIENUQSBpZiBub3QgY29uZmlndXJlZCAtLT4KICAgICAgICA8ZGl2IGlkPSJhaS1zdW0tY3RhIiBzdHlsZT0iZGlzcGxheTpub25lO21hcmdpbi10b3A6MTBweDtwYWRkaW5nOjhweCAxMXB4OwogICAgICAgICAgYmFja2dyb3VuZDpyZ2JhKDE1NSwxMDksMjU1LC4wNik7Ym9yZGVyLXJhZGl1czo1cHg7Ym9yZGVyOjFweCBkYXNoZWQgcmdiYSgxNTUsMTA5LDI1NSwuMikiPgogICAgICAgICAgPGRpdiBzdHlsZT0iZm9udC1zaXplOjEwcHg7Y29sb3I6dmFyKC0tdmlvMikiPgogICAgICAgICAgICAmI3gxRjkxNjsgQ29uZmlndXJlIFRhY2h5b24gQUkgVVJMIGluIENvbmZpZ3VyYXRpb24g4oaSIFRhY2h5b24gTExNIGZvciByaWNoZXIsIGNvbnRleHQtYXdhcmUgaW5zaWdodHMuCiAgICAgICAgICA8L2Rpdj4KICAgICAgICA8L2Rpdj4KICAgICAgPC9kaXY+CiAgICA8L2Rpdj4KCiAgICA8IS0tIExpbmVhZ2UgcGFuZSAtLT4KICAgIDxkaXYgaWQ9InBhbmUtbGluZWFnZSIgc3R5bGU9ImRpc3BsYXk6bm9uZSI+CiAgICAgIDxkaXYgY2xhc3M9ImN0cmwtcm93Ij4KICAgICAgICA8ZGl2IGNsYXNzPSJsdi1iYXIiPgogICAgICAgICAgPGJ1dHRvbiBjbGFzcz0ibHYtYnRuIG9uIiBpZD0ibHYtc3lzdGVtIiBvbmNsaWNrPSJzZXRMZXZlbCgnc3lzdGVtJykiPiYjeDFGM0UyOyBTeXN0ZW08L2J1dHRvbj4KICAgICAgICAgIDxidXR0b24gY2xhc3M9Imx2LWJ0biIgICAgaWQ9Imx2LXRhYmxlIiAgb25jbGljaz0ic2V0TGV2ZWwoJ3RhYmxlJykiPiYjeDFGNUMzOyBUYWJsZTwvYnV0dG9uPgogICAgICAgICAgPGJ1dHRvbiBjbGFzcz0ibHYtYnRuIiAgICBpZD0ibHYtZmllbGQiICBvbmNsaWNrPSJzZXRMZXZlbCgnZmllbGQnKSI+JiN4MUY1MzI7IEZpZWxkPC9idXR0b24+CiAgICAgICAgPC9kaXY+CiAgICAgICAgPGRpdiBjbGFzcz0iYmMiIGlkPSJiYyI+PC9kaXY+CiAgICAgICAgPGRpdiBzdHlsZT0iZmxleDoxIj48L2Rpdj4KICAgICAgICA8c3BhbiBzdHlsZT0iZm9udC1zaXplOjlweDtjb2xvcjp2YXIoLS10Mik7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2UiPlZpZXc6PC9zcGFuPgogICAgICAgIDxkaXYgY2xhc3M9InZpZXctdG9nIj4KICAgICAgICAgIDxidXR0b24gY2xhc3M9InZ0LWJ0biBvbiIgaWQ9InZ0LWVyZCIgICBvbmNsaWNrPSJzZXRWaWV3KCdlcmQnKSI+JiN4MjVBNjsgRVJEIEZsb3c8L2J1dHRvbj4KICAgICAgICAgIDxidXR0b24gY2xhc3M9InZ0LWJ0biIgICAgaWQ9InZ0LWZvcmNlIiBvbmNsaWNrPSJzZXRWaWV3KCdmb3JjZScpIj4mI3gyNUM5OyBGb3JjZTwvYnV0dG9uPgogICAgICAgIDwvZGl2PgogICAgICA8L2Rpdj4KICAgICAgPGRpdiBjbGFzcz0iZ2JveCI+CiAgICAgICAgPGRpdiBjbGFzcz0iZy1oZHIiPgogICAgICAgICAgPGRpdj48ZGl2IGNsYXNzPSJnLXRpdGxlIiBpZD0iZy10aXRsZSI+TGluZWFnZSBHcmFwaDwvZGl2PjxkaXYgY2xhc3M9Imctc3ViIiBpZD0iZy1zdWIiPjwvZGl2PjwvZGl2PgogICAgICAgICAgPGRpdiBjbGFzcz0iZy10b29scyI+CiAgICAgICAgICAgIDxkaXYgY2xhc3M9ImctYnRuIiBvbmNsaWNrPSJnWm9vbUluKCkiIHRpdGxlPSJab29tIEluIj4rPC9kaXY+CiAgICAgICAgICAgIDxkaXYgY2xhc3M9ImctYnRuIiBvbmNsaWNrPSJnWm9vbU91dCgpIiB0aXRsZT0iWm9vbSBPdXQiPiZtaW51czs8L2Rpdj4KICAgICAgICAgICAgPGRpdiBjbGFzcz0iZy1idG4iIG9uY2xpY2s9ImdGaXQoKSIgdGl0bGU9IkZpdCI+JiN4MjI5Rjs8L2Rpdj4KICAgICAgICAgICAgPGRpdiBjbGFzcz0iZy1idG4iIG9uY2xpY2s9ImdGdWxsc2NyZWVuKCkiIHRpdGxlPSJGdWxsc2NyZWVuIj4mI3gyNkY2OzwvZGl2PgogICAgICAgICAgPC9kaXY+CiAgICAgICAgPC9kaXY+CiAgICAgICAgPGRpdiBpZD0iZXJkLXZpZXciPgogICAgICAgICAgPGRpdiBpZD0iZXJkLXBhbiIgc3R5bGU9IndpZHRoOjEwMCU7aGVpZ2h0OjEwMCU7cG9zaXRpb246YWJzb2x1dGUiPgogICAgICAgICAgICA8ZGl2IGlkPSJlcmQtd29ybGQiIHN0eWxlPSJwb3NpdGlvbjphYnNvbHV0ZTt0cmFuc2Zvcm0tb3JpZ2luOjAgMCI+CiAgICAgICAgICAgICAgPGRpdiBpZD0iZXJkLWhvc3QiPjwvZGl2PgogICAgICAgICAgICAgIDxzdmcgaWQ9ImVyZC1zdmciPjwvc3ZnPgogICAgICAgICAgICA8L2Rpdj4KICAgICAgICAgIDwvZGl2PgogICAgICAgIDwvZGl2PgogICAgICAgIDxkaXYgaWQ9ImZvcmNlLXZpZXciIHN0eWxlPSJkaXNwbGF5Om5vbmU7aGVpZ2h0OjQ5MHB4Ij4KICAgICAgICAgIDxzdmcgaWQ9ImZvcmNlLXN2ZyIgc3R5bGU9IndpZHRoOjEwMCU7aGVpZ2h0OjQ5MHB4Ij48L3N2Zz4KICAgICAgICA8L2Rpdj4KICAgICAgICA8ZGl2IGNsYXNzPSJnLWxlZ2VuZCIgaWQ9ImctbGVnZW5kIj48L2Rpdj4KICAgICAgPC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9Imx2LWNhcmRzIj4KICAgICAgICA8ZGl2IGNsYXNzPSJsdi1jYXJkIiBvbmNsaWNrPSJzZXRMZXZlbCgnc3lzdGVtJykiPgogICAgICAgICAgPGRpdiBjbGFzcz0ibHYtY2FyZC1pY29uIj4mI3gxRjNFMjs8L2Rpdj4KICAgICAgICAgIDxkaXYgY2xhc3M9Imx2LWNhcmQtbmFtZSI+U3lzdGVtIExldmVsPC9kaXY+CiAgICAgICAgICA8ZGl2IGNsYXNzPSJsdi1jYXJkLWRlc2MiPkNyb3NzLXN5c3RlbSBsaW5lYWdlIGZyb20gc291cmNlIHBsYXRmb3JtcyB0byBkYXRhIHdhcmVob3VzZTwvZGl2PgogICAgICAgIDwvZGl2PgogICAgICAgIDxkaXYgY2xhc3M9Imx2LWNhcmQiIG9uY2xpY2s9InNldExldmVsKCd0YWJsZScpIj4KICAgICAgICAgIDxkaXYgY2xhc3M9Imx2LWNhcmQtaWNvbiI+JiN4MUY1QzM7PC9kaXY+CiAgICAgICAgICA8ZGl2IGNsYXNzPSJsdi1jYXJkLW5hbWUiPlRhYmxlIExldmVsPC9kaXY+CiAgICAgICAgICA8ZGl2IGNsYXNzPSJsdi1jYXJkLWRlc2MiPlRhYmxlLXRvLXRhYmxlIGZsb3cgd2l0aCBFVEwgdHJhbnNmb3JtYXRpb25zPC9kaXY+CiAgICAgICAgPC9kaXY+CiAgICAgICAgPGRpdiBjbGFzcz0ibHYtY2FyZCIgb25jbGljaz0ic2V0TGV2ZWwoJ2ZpZWxkJykiPgogICAgICAgICAgPGRpdiBjbGFzcz0ibHYtY2FyZC1pY29uIj4mI3gxRjUzMjs8L2Rpdj4KICAgICAgICAgIDxkaXYgY2xhc3M9Imx2LWNhcmQtbmFtZSI+RmllbGQgTGV2ZWw8L2Rpdj4KICAgICAgICAgIDxkaXYgY2xhc3M9Imx2LWNhcmQtZGVzYyI+Q29sdW1uLXRvLWNvbHVtbiBtYXBwaW5nIHdpdGggYnVzaW5lc3MgcnVsZXM8L2Rpdj4KICAgICAgICA8L2Rpdj4KICAgICAgPC9kaXY+CiAgICA8L2Rpdj4KCiAgICA8IS0tIE1ldGFkYXRhIHBhbmUgLS0+CiAgICA8ZGl2IGlkPSJwYW5lLW1ldGFkYXRhIiBzdHlsZT0iZGlzcGxheTpub25lIj4KICAgICAgPGRpdiBzdHlsZT0iZm9udC1zaXplOjEwcHg7Y29sb3I6dmFyKC0tdDIpO21hcmdpbi1ib3R0b206OHB4O2ZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlIiBpZD0ibWV0YS1jb3VudCI+PC9kaXY+CiAgICAgIDxkaXYgY2xhc3M9Im1ldGEtZ3JpZCI+CiAgICAgICAgPGRpdj48ZGl2IGNsYXNzPSJtZXRhLXRibCIgaWQ9Im1ldGEtdGJsIj48L2Rpdj48L2Rpdj4KICAgICAgICA8ZGl2IGNsYXNzPSJmZGV0YWlsIiBpZD0iZmRldGFpbCI+CiAgICAgICAgICA8ZGl2IHN0eWxlPSJwYWRkaW5nOjM4cHg7dGV4dC1hbGlnbjpjZW50ZXI7Y29sb3I6dmFyKC0tdDIpIj4KICAgICAgICAgICAgPGRpdiBzdHlsZT0iZm9udC1zaXplOjMwcHg7bWFyZ2luLWJvdHRvbToxMHB4O2ZpbHRlcjpkcm9wLXNoYWRvdygwIDAgMTBweCByZ2JhKDI0NSwxNjYsMzUsLjM1KSkiPiYjeDI2MUQ7PC9kaXY+CiAgICAgICAgICAgIDxkaXYgc3R5bGU9ImZvbnQtc2l6ZToxMnB4O2ZvbnQtd2VpZ2h0OjYwMCI+U2VsZWN0IGEgZmllbGQgdG8gdmlldyBkZXRhaWxzPC9kaXY+CiAgICAgICAgICAgIDxkaXYgc3R5bGU9ImZvbnQtc2l6ZToxMHB4O2NvbG9yOnZhcigtLXQzKTttYXJnaW4tdG9wOjVweCI+UmlzayAmYnVsbDsgTG9naWMgJmJ1bGw7IExpbmVhZ2UgJmJ1bGw7IEltcGFjdCAmYnVsbDsgUERGPC9kaXY+CiAgICAgICAgICA8L2Rpdj4KICAgICAgICA8L2Rpdj4KICAgICAgPC9kaXY+CiAgICA8L2Rpdj4KCiAgICA8IS0tIFNvdXJjZXMgcGFuZSAtLT4KICAgIDxkaXYgaWQ9InBhbmUtc291cmNlcyIgc3R5bGU9ImRpc3BsYXk6bm9uZSI+CiAgICAgIDxkaXYgY2xhc3M9InNyYy1jYXJkcyIgaWQ9InNyYy1jYXJkcyI+PC9kaXY+CiAgICA8L2Rpdj4KCiAgICA8IS0tIEVSV0lOIEFQSSBFeHBsb3JlciBwYW5lIC0tPgogICAgPGRpdiBpZD0icGFuZS1lcndpbiIgc3R5bGU9ImRpc3BsYXk6bm9uZSI+CgogICAgICA8IS0tIEhlYWRlciBiYXIgLS0+CiAgICAgIDxkaXYgc3R5bGU9ImRpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjEwcHg7bWFyZ2luLWJvdHRvbToxNHB4O2ZsZXgtd3JhcDp3cmFwIj4KICAgICAgICA8ZGl2IHN0eWxlPSJmb250LWZhbWlseTonQmViYXMgTmV1ZScsc2Fucy1zZXJpZjtmb250LXNpemU6MThweDtsZXR0ZXItc3BhY2luZzoxcHg7CiAgICAgICAgICBiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCg5MGRlZyx2YXIoLS1nb2xkMiksdmFyKC0tdGVhbDIpKTsKICAgICAgICAgIC13ZWJraXQtYmFja2dyb3VuZC1jbGlwOnRleHQ7LXdlYmtpdC10ZXh0LWZpbGwtY29sb3I6dHJhbnNwYXJlbnQiPgogICAgICAgICAgRVJXSU4gREkgQVBJIEV4cGxvcmVyCiAgICAgICAgPC9kaXY+CiAgICAgICAgPHNwYW4gaWQ9ImVyd2luLWVwLWNvdW50IiBzdHlsZT0iZm9udC1zaXplOjEwcHg7Y29sb3I6dmFyKC0tdDIpOwogICAgICAgICAgZm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2UiPjwvc3Bhbj4KICAgICAgICA8YnV0dG9uIG9uY2xpY2s9ImxvYWRFcndpbkVuZHBvaW50cygpIgogICAgICAgICAgc3R5bGU9InBhZGRpbmc6NXB4IDEycHg7YmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoMTM1ZGVnLHZhcigtLWdvbGQzKSx2YXIoLS10ZWFsMykpOwogICAgICAgICAgICBib3JkZXI6bm9uZTtib3JkZXItcmFkaXVzOjZweDtjb2xvcjojZmZmO2ZvbnQtc2l6ZToxMXB4O2ZvbnQtd2VpZ2h0OjcwMDtjdXJzb3I6cG9pbnRlciI+CiAgICAgICAgICAmI3gyN0YzOyBSZWZyZXNoIEVuZHBvaW50cwogICAgICAgIDwvYnV0dG9uPgogICAgICAgIDxidXR0b24gb25jbGljaz0ibG9hZEVyd2luU3dhZ2dlcigpIgogICAgICAgICAgc3R5bGU9InBhZGRpbmc6NXB4IDEycHg7YmFja2dyb3VuZDp2YXIoLS1wYW5lbCk7Ym9yZGVyOjFweCBzb2xpZCB2YXIoLS1iZHIyKTsKICAgICAgICAgICAgYm9yZGVyLXJhZGl1czo2cHg7Y29sb3I6dmFyKC0tdDEpO2ZvbnQtc2l6ZToxMXB4O2ZvbnQtd2VpZ2h0OjcwMDtjdXJzb3I6cG9pbnRlciI+CiAgICAgICAgICAmI3gxRjRDQjsgVmlldyBTd2FnZ2VyIFNwZWMKICAgICAgICA8L2J1dHRvbj4KICAgICAgPC9kaXY+CgogICAgICA8IS0tIFF1aWNrIGNhbGwgYnVpbGRlciAtLT4KICAgICAgPGRpdiBzdHlsZT0iYmFja2dyb3VuZDp2YXIoLS1iZzIpO2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyMik7Ym9yZGVyLXJhZGl1czp2YXIoLS1yKTsKICAgICAgICBwYWRkaW5nOjE0cHg7bWFyZ2luLWJvdHRvbToxNHB4O2JveC1zaGFkb3c6dmFyKC0tc2gyKSI+CiAgICAgICAgPGRpdiBzdHlsZT0iZm9udC1zaXplOjEwcHg7Zm9udC13ZWlnaHQ6NzAwO2NvbG9yOnZhcigtLWdvbGQyKTtsZXR0ZXItc3BhY2luZzouNXB4OwogICAgICAgICAgZm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2U7bWFyZ2luLWJvdHRvbToxMHB4Ij4KICAgICAgICAgICYjeDI2QTE7IFFVSUNLIEVORFBPSU5UIENBTExFUgogICAgICAgIDwvZGl2PgogICAgICAgIDxkaXYgc3R5bGU9ImRpc3BsYXk6Z3JpZDtncmlkLXRlbXBsYXRlLWNvbHVtbnM6MWZyIDFmcjtnYXA6OHB4O21hcmdpbi1ib3R0b206OHB4Ij4KICAgICAgICAgIDxkaXY+CiAgICAgICAgICAgIDxsYWJlbCBzdHlsZT0iZm9udC1zaXplOjlweDtjb2xvcjp2YXIoLS10Mik7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2U7ZGlzcGxheTpibG9jazttYXJnaW4tYm90dG9tOjNweCI+RW5kcG9pbnQgUGF0aDwvbGFiZWw+CiAgICAgICAgICAgIDxzZWxlY3QgaWQ9ImVyd2luLWVwLXNlbGVjdCIgY2xhc3M9ImNmZy1pbiIgc3R5bGU9IndpZHRoOjEwMCU7Y3Vyc29yOnBvaW50ZXIiCiAgICAgICAgICAgICAgb25jaGFuZ2U9Im9uRXJ3aW5FcFNlbGVjdCh0aGlzLnZhbHVlKSI+CiAgICAgICAgICAgICAgPG9wdGlvbiB2YWx1ZT0iIj4tLSBTZWxlY3QgZW5kcG9pbnQgLS08L29wdGlvbj4KICAgICAgICAgICAgPC9zZWxlY3Q+CiAgICAgICAgICA8L2Rpdj4KICAgICAgICAgIDxkaXY+CiAgICAgICAgICAgIDxsYWJlbCBzdHlsZT0iZm9udC1zaXplOjlweDtjb2xvcjp2YXIoLS10Mik7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2U7ZGlzcGxheTpibG9jazttYXJnaW4tYm90dG9tOjNweCI+RGVzY3JpcHRpb248L2xhYmVsPgogICAgICAgICAgICA8ZGl2IGlkPSJlcndpbi1lcC1kZXNjIiBzdHlsZT0iZm9udC1zaXplOjEwcHg7Y29sb3I6dmFyKC0tdGVhbDIpOwogICAgICAgICAgICAgIGJhY2tncm91bmQ6dmFyKC0tcGFuZWwpO2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyKTtib3JkZXItcmFkaXVzOnZhcigtLXIzKTsKICAgICAgICAgICAgICBwYWRkaW5nOjVweCA4cHg7aGVpZ2h0OjI2cHg7b3ZlcmZsb3c6aGlkZGVuO2ZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlIj4KICAgICAgICAgICAgICBTZWxlY3QgYW4gZW5kcG9pbnQgYWJvdmUKICAgICAgICAgICAgPC9kaXY+CiAgICAgICAgICA8L2Rpdj4KICAgICAgICA8L2Rpdj4KICAgICAgICA8ZGl2IGlkPSJlcndpbi1wYXJhbS1idWlsZGVyIiBzdHlsZT0ibWFyZ2luLWJvdHRvbTo4cHgiPjwvZGl2PgogICAgICAgIDxkaXYgc3R5bGU9ImRpc3BsYXk6ZmxleDtnYXA6N3B4Ij4KICAgICAgICAgIDxidXR0b24gb25jbGljaz0iY2FsbEVyd2luRW5kcG9pbnQoKSIKICAgICAgICAgICAgc3R5bGU9ImZsZXg6MTtoZWlnaHQ6MzBweDtiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCgxMzVkZWcsdmFyKC0tZ29sZDMpLHZhcigtLXRlYWwzKSk7CiAgICAgICAgICAgICAgYm9yZGVyOm5vbmU7Ym9yZGVyLXJhZGl1czo2cHg7Y29sb3I6I2ZmZjtmb250LXNpemU6MTJweDtmb250LXdlaWdodDo3MDA7Y3Vyc29yOnBvaW50ZXIiPgogICAgICAgICAgICAmI3gyNUI2OyBDYWxsIEVuZHBvaW50CiAgICAgICAgICA8L2J1dHRvbj4KICAgICAgICAgIDxidXR0b24gb25jbGljaz0iY2FsbEFuZFNlYXJjaCgpIgogICAgICAgICAgICBzdHlsZT0iZmxleDoxO2hlaWdodDozMHB4O2JhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDEzNWRlZyx2YXIoLS10ZWFsMyksdmFyKC0tc2t5KSk7CiAgICAgICAgICAgICAgYm9yZGVyOm5vbmU7Ym9yZGVyLXJhZGl1czo2cHg7Y29sb3I6I2ZmZjtmb250LXNpemU6MTJweDtmb250LXdlaWdodDo3MDA7Y3Vyc29yOnBvaW50ZXIiPgogICAgICAgICAgICAmI3gxRjU3ODsgQ2FsbCArIFNob3cgTGluZWFnZQogICAgICAgICAgPC9idXR0b24+CiAgICAgICAgPC9kaXY+CiAgICAgIDwvZGl2PgoKICAgICAgPCEtLSBSZXN1bHRzIC0tPgogICAgICA8ZGl2IGlkPSJlcndpbi1jYWxsLXJlc3VsdCIgc3R5bGU9ImRpc3BsYXk6bm9uZTtiYWNrZ3JvdW5kOnZhcigtLWJnMik7CiAgICAgICAgYm9yZGVyOjFweCBzb2xpZCB2YXIoLS1iZHIyKTtib3JkZXItcmFkaXVzOnZhcigtLXIpO3BhZGRpbmc6MTRweDsKICAgICAgICBtYXJnaW4tYm90dG9tOjE0cHg7Ym94LXNoYWRvdzp2YXIoLS1zaDIpIj4KICAgICAgICA8ZGl2IHN0eWxlPSJkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo4cHg7bWFyZ2luLWJvdHRvbTo4cHgiPgogICAgICAgICAgPGRpdiBpZD0iZXJ3aW4tcmVzdWx0LXN0YXR1cyIgc3R5bGU9ImZvbnQtc2l6ZToxMHB4O2ZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlIj48L2Rpdj4KICAgICAgICAgIDxkaXYgaWQ9ImVyd2luLXJlc3VsdC1lbmRwb2ludCIgc3R5bGU9ImZvbnQtc2l6ZTo5cHg7Y29sb3I6dmFyKC0tdDIpO2ZvbnQtZmFtaWx5OidKZXRCcmFpbnMgTW9ubycsbW9ub3NwYWNlIj48L2Rpdj4KICAgICAgICA8L2Rpdj4KICAgICAgICA8cHJlIGlkPSJlcndpbi1yZXN1bHQtanNvbiIgc3R5bGU9ImJhY2tncm91bmQ6cmdiYSgwLDAsMCwuNDUpO2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyMik7CiAgICAgICAgICBib3JkZXItcmFkaXVzOjRweDtwYWRkaW5nOjEwcHg7Zm9udC1mYW1pbHk6J0pldEJyYWlucyBNb25vJyxtb25vc3BhY2U7Zm9udC1zaXplOjlweDsKICAgICAgICAgIGNvbG9yOnZhcigtLXRlYWwyKTtvdmVyZmxvdzphdXRvO21heC1oZWlnaHQ6MzIwcHg7bGluZS1oZWlnaHQ6MS41OwogICAgICAgICAgYm9yZGVyLWxlZnQ6MnB4IHNvbGlkIHZhcigtLWdvbGQpO3doaXRlLXNwYWNlOnByZS13cmFwO3dvcmQtYnJlYWs6YnJlYWstYWxsIj48L3ByZT4KICAgICAgPC9kaXY+CgogICAgICA8IS0tIEVuZHBvaW50IGNhdGFsb2cgLS0+CiAgICAgIDxkaXYgc3R5bGU9ImZvbnQtc2l6ZToxMHB4O2ZvbnQtd2VpZ2h0OjcwMDtjb2xvcjp2YXIoLS10Mik7bGV0dGVyLXNwYWNpbmc6LjVweDsKICAgICAgICBmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZTttYXJnaW4tYm90dG9tOjhweCI+CiAgICAgICAgQUxMIEFWQUlMQUJMRSBFTkRQT0lOVFMKICAgICAgPC9kaXY+CiAgICAgIDxkaXYgaWQ9ImVyd2luLWVwLWxpc3QiPjwvZGl2PgoKICAgIDwvZGl2PgoKICAgIDwhLS0gTXVsdGkgcGFuZSAtLT4KICAgIDxkaXYgaWQ9InBhbmUtbXVsdGkiIHN0eWxlPSJkaXNwbGF5Om5vbmUiPgoKICAgICAgPCEtLSBTVEFUSUMgQUNUSU9OIEJBUiDigJQgYWx3YXlzIHZpc2libGUsIG5vIEpTIG5lZWRlZCB0byByZW5kZXIgLS0+CiAgICAgIDxkaXYgc3R5bGU9ImRpc3BsYXk6ZmxleDtnYXA6OHB4O21hcmdpbi1ib3R0b206MTRweDtmbGV4LXdyYXA6d3JhcDtwYWRkaW5nOjEycHggMDtib3JkZXItYm90dG9tOjJweCBzb2xpZCByZ2JhKDI0NSwxNjYsMzUsLjI1KSI+CiAgICAgICAgPGJ1dHRvbiBvbmNsaWNrPSJzaG93VGFiKCdsaW5lYWdlJyk7c2V0VGltZW91dChmdW5jdGlvbigpe3NldExldmVsKCd0YWJsZScpO30sODApIgogICAgICAgICAgc3R5bGU9ImRpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjdweDtwYWRkaW5nOjEwcHggMThweDsKICAgICAgICAgICAgYmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoMTM1ZGVnLCNCODc4MEEsIzAwOEY3RSk7CiAgICAgICAgICAgIGJvcmRlcjpub25lO2JvcmRlci1yYWRpdXM6N3B4O2NvbG9yOiNmZmY7Zm9udC1zaXplOjEzcHg7Zm9udC13ZWlnaHQ6NzAwOwogICAgICAgICAgICBjdXJzb3I6cG9pbnRlcjtib3gtc2hhZG93OjAgMCAxOHB4IHJnYmEoMjQ1LDE2NiwzNSwuMzUpOwogICAgICAgICAgICBsZXR0ZXItc3BhY2luZzouM3B4O3RyYW5zaXRpb246YWxsIC4ycyIKICAgICAgICAgIG9ubW91c2VvdmVyPSJ0aGlzLnN0eWxlLnRyYW5zZm9ybT0ndHJhbnNsYXRlWSgtMnB4KSciCiAgICAgICAgICBvbm1vdXNlb3V0PSJ0aGlzLnN0eWxlLnRyYW5zZm9ybT0ndHJhbnNsYXRlWSgwKSciPgogICAgICAgICAgJiN4MUY1Nzg7IFZpZXcgQ29tYmluZWQgTGluZWFnZSBHcmFwaAogICAgICAgIDwvYnV0dG9uPgogICAgICAgIDxidXR0b24gb25jbGljaz0ic2hvd1RhYignbWV0YWRhdGEnKSIKICAgICAgICAgIHN0eWxlPSJkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo3cHg7cGFkZGluZzoxMHB4IDE4cHg7CiAgICAgICAgICAgIGJhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDEzNWRlZywjMDA4RjdFLCMxZDZmYTYpOwogICAgICAgICAgICBib3JkZXI6bm9uZTtib3JkZXItcmFkaXVzOjdweDtjb2xvcjojZmZmO2ZvbnQtc2l6ZToxM3B4O2ZvbnQtd2VpZ2h0OjcwMDsKICAgICAgICAgICAgY3Vyc29yOnBvaW50ZXI7Ym94LXNoYWRvdzowIDAgMThweCByZ2JhKDAsMjAxLDE3NywuMyk7CiAgICAgICAgICAgIGxldHRlci1zcGFjaW5nOi4zcHg7dHJhbnNpdGlvbjphbGwgLjJzIgogICAgICAgICAgb25tb3VzZW92ZXI9InRoaXMuc3R5bGUudHJhbnNmb3JtPSd0cmFuc2xhdGVZKC0ycHgpJyIKICAgICAgICAgIG9ubW91c2VvdXQ9InRoaXMuc3R5bGUudHJhbnNmb3JtPSd0cmFuc2xhdGVZKDApJyI+CiAgICAgICAgICAmI3gxRjRDQjsgVmlldyBBbGwgTWV0YWRhdGEKICAgICAgICA8L2J1dHRvbj4KICAgICAgICA8YnV0dG9uIG9uY2xpY2s9InNob3dUYWIoJ3NvdXJjZXMnKSIKICAgICAgICAgIHN0eWxlPSJkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo3cHg7cGFkZGluZzoxMHB4IDE4cHg7CiAgICAgICAgICAgIGJhY2tncm91bmQ6cmdiYSgyNTUsMjU1LDI1NSwuMDcpOwogICAgICAgICAgICBib3JkZXI6MXB4IHNvbGlkIHJnYmEoMjU1LDI1NSwyNTUsLjE1KTtib3JkZXItcmFkaXVzOjdweDtjb2xvcjojRjJGOEZGOwogICAgICAgICAgICBmb250LXNpemU6MTNweDtmb250LXdlaWdodDo3MDA7Y3Vyc29yOnBvaW50ZXI7CiAgICAgICAgICAgIGxldHRlci1zcGFjaW5nOi4zcHg7dHJhbnNpdGlvbjphbGwgLjJzIgogICAgICAgICAgb25tb3VzZW92ZXI9InRoaXMuc3R5bGUudHJhbnNmb3JtPSd0cmFuc2xhdGVZKC0ycHgpJyIKICAgICAgICAgIG9ubW91c2VvdXQ9InRoaXMuc3R5bGUudHJhbnNmb3JtPSd0cmFuc2xhdGVZKDApJyI+CiAgICAgICAgICAmI3gxRjRFMTsgVmlldyBTb3VyY2VzCiAgICAgICAgPC9idXR0b24+CiAgICAgICAgPHNwYW4gaWQ9Im11bHRpLXN1bW1hcnkiIHN0eWxlPSJtYXJnaW4tbGVmdDphdXRvO2ZvbnQtc2l6ZToxMHB4O2NvbG9yOiM0RTZFOEE7CiAgICAgICAgICBmb250LWZhbWlseTonSmV0QnJhaW5zIE1vbm8nLG1vbm9zcGFjZTtkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyIj48L3NwYW4+CiAgICAgIDwvZGl2PgoKICAgICAgPGRpdiBpZD0ibXVsdGktcmVzdWx0cyI+PC9kaXY+CgogICAgICA8IS0tIEJPVFRPTSBBQ1RJT04gQkFSIOKAlCBhbHdheXMgdmlzaWJsZSB3aGVuIHNjcm9sbGVkIGRvd24gLS0+CiAgICAgIDxkaXYgaWQ9Im11bHRpLWJvdHRvbS1iYXIiIHN0eWxlPSJkaXNwbGF5Om5vbmU7Z2FwOjhweDttYXJnaW4tdG9wOjE0cHg7ZmxleC13cmFwOndyYXA7CiAgICAgICAgcGFkZGluZy10b3A6MTJweDtib3JkZXItdG9wOjJweCBzb2xpZCByZ2JhKDI0NSwxNjYsMzUsLjI1KSI+CiAgICAgICAgPGJ1dHRvbiBvbmNsaWNrPSJzaG93VGFiKCdsaW5lYWdlJyk7c2V0VGltZW91dChmdW5jdGlvbigpe3NldExldmVsKCd0YWJsZScpO30sODApIgogICAgICAgICAgc3R5bGU9ImRpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjdweDtwYWRkaW5nOjEwcHggMThweDsKICAgICAgICAgICAgYmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoMTM1ZGVnLCNCODc4MEEsIzAwOEY3RSk7CiAgICAgICAgICAgIGJvcmRlcjpub25lO2JvcmRlci1yYWRpdXM6N3B4O2NvbG9yOiNmZmY7Zm9udC1zaXplOjEzcHg7Zm9udC13ZWlnaHQ6NzAwOwogICAgICAgICAgICBjdXJzb3I6cG9pbnRlcjtib3gtc2hhZG93OjAgMCAxOHB4IHJnYmEoMjQ1LDE2NiwzNSwuMzUpO2xldHRlci1zcGFjaW5nOi4zcHgiPgogICAgICAgICAgJiN4MUY1Nzg7IFZpZXcgQ29tYmluZWQgTGluZWFnZSBHcmFwaAogICAgICAgIDwvYnV0dG9uPgogICAgICAgIDxidXR0b24gb25jbGljaz0ic2hvd1RhYignbWV0YWRhdGEnKSIKICAgICAgICAgIHN0eWxlPSJkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo3cHg7cGFkZGluZzoxMHB4IDE4cHg7CiAgICAgICAgICAgIGJhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDEzNWRlZywjMDA4RjdFLCMxZDZmYTYpOwogICAgICAgICAgICBib3JkZXI6bm9uZTtib3JkZXItcmFkaXVzOjdweDtjb2xvcjojZmZmO2ZvbnQtc2l6ZToxM3B4O2ZvbnQtd2VpZ2h0OjcwMDsKICAgICAgICAgICAgY3Vyc29yOnBvaW50ZXI7Ym94LXNoYWRvdzowIDAgMThweCByZ2JhKDAsMjAxLDE3NywuMyk7bGV0dGVyLXNwYWNpbmc6LjNweCI+CiAgICAgICAgICAmI3gxRjRDQjsgVmlldyBBbGwgTWV0YWRhdGEKICAgICAgICA8L2J1dHRvbj4KICAgICAgPC9kaXY+CgogICAgPC9kaXY+CgogIDwvZGl2PjwhLS0gY29udGVudCAtLT4KPC9kaXY+PCEtLSBtYWluIC0tPgo8L2Rpdj48IS0tIGFwcCAtLT4KCjwhLS0gTW9kYWwgLS0+CjxkaXYgY2xhc3M9Im92ZXJsYXkiIGlkPSJvdmVybGF5IiBvbmNsaWNrPSJjbG9zZU92ZXJsYXkoZXZlbnQpIj4KICA8ZGl2IGNsYXNzPSJtb2RhbCI+CiAgICA8ZGl2IGNsYXNzPSJtb2RhbC1oZHIiPgogICAgICA8ZGl2IGNsYXNzPSJtb2RhbC10aXRsZSIgaWQ9Im1vZGFsLXRpdGxlIj5EZXRhaWxzPC9kaXY+CiAgICAgIDxidXR0b24gY2xhc3M9Im1vZGFsLXgiIG9uY2xpY2s9ImNsb3NlTW9kYWwoKSI+JiN4MjcxNTs8L2J1dHRvbj4KICAgIDwvZGl2PgogICAgPGRpdiBjbGFzcz0ibW9kYWwtYm9keSIgaWQ9Im1vZGFsLWJvZHkiPjwvZGl2PgogICAgPGRpdiBjbGFzcz0ibW9kYWwtZm9vdCI+PGJ1dHRvbiBjbGFzcz0iYnRuLWNsb3NlIiBvbmNsaWNrPSJjbG9zZU1vZGFsKCkiPkNsb3NlPC9idXR0b24+PC9kaXY+CiAgPC9kaXY+CjwvZGl2PgoKPGRpdiBjbGFzcz0idG9hc3QiIGlkPSJ0b2FzdCI+PC9kaXY+Cgo8c2NyaXB0PgondXNlIHN0cmljdCc7CgovLyDilIDilIAgU1RBVEUg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmNvbnN0IEFQSSA9ICcnOwpsZXQgRCA9IG51bGw7CmxldCBjdXJMZXZlbCA9ICdzeXN0ZW0nOwpsZXQgY3VyVmlldyAgPSAnZXJkJzsKbGV0IHNlbEZpZWxkID0gbnVsbDsKbGV0IGhpc3QgPSBbXTsKbGV0IHNyY3MgPSB7ZXJ3aW46dHJ1ZSwgZGljdDp0cnVlLCBjb2RlOnRydWUsIHRhY2h5b246dHJ1ZX07CmxldCBlcmRTY2FsZSA9IDEsIGVyZFggPSAyMCwgZXJkWSA9IDIwLCBlcmREcmFnZ2luZyA9IGZhbHNlLCBlcmREcmFnU3RhcnQgPSB7eDowLHk6MH07CmxldCBub2RlUG9zaXRpb25zID0ge307CmxldCBmU3ZnID0gbnVsbCwgZlpvb20gPSBudWxsLCBmRyA9IG51bGw7CgovLyDilIDilIAgSEVMUEVSUyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZnVuY3Rpb24gZXNjKHMpeyByZXR1cm4gKHN8fCcnKS5yZXBsYWNlKC8mL2csJyZhbXA7JykucmVwbGFjZSgvPC9nLCcmbHQ7JykucmVwbGFjZSgvPi9nLCcmZ3Q7Jyk7IH0KZnVuY3Rpb24gZ3YoaWQpeyBjb25zdCBlbD1kb2N1bWVudC5nZXRFbGVtZW50QnlJZChpZCk7IHJldHVybiBlbCA/IGVsLnZhbHVlIDogJyc7IH0KZnVuY3Rpb24gc3YoaWQsdil7IGNvbnN0IGVsPWRvY3VtZW50LmdldEVsZW1lbnRCeUlkKGlkKTsgaWYoZWwpIGVsLnZhbHVlPXY7IH0KZnVuY3Rpb24gZWwoaWQpeyByZXR1cm4gZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoaWQpOyB9CmZ1bmN0aW9uIHNob3coaWQpeyBjb25zdCBlPWVsKGlkKTsgaWYoZSkgZS5zdHlsZS5kaXNwbGF5PSdibG9jayc7IH0KZnVuY3Rpb24gaGlkZShpZCl7IGNvbnN0IGU9ZWwoaWQpOyBpZihlKSBlLnN0eWxlLmRpc3BsYXk9J25vbmUnOyB9CgpmdW5jdGlvbiB0b2FzdChtc2csIHR5cGUpewogIGNvbnN0IHQ9ZWwoJ3RvYXN0Jyk7CiAgY29uc3QgYmM9e2luZm86J3ZhcigtLXNreSknLHN1Y2Nlc3M6J3ZhcigtLWxpbWUpJyxlcnJvcjondmFyKC0tcm9zZSknLHdhcm46J3ZhcigtLWdvbGQpJ307CiAgdC5zdHlsZS5ib3JkZXJMZWZ0Q29sb3IgPSBiY1t0eXBlXXx8YmMuaW5mbzsKICB0LnRleHRDb250ZW50ID0gbXNnOwogIHQuY2xhc3NMaXN0LmFkZCgnc2hvdycpOwogIHNldFRpbWVvdXQoKCk9PnQuY2xhc3NMaXN0LnJlbW92ZSgnc2hvdycpLCAzMjAwKTsKfQoKZnVuY3Rpb24gZnIoayx2KXsKICByZXR1cm4gJzxkaXYgY2xhc3M9ImZkLXJvdyI+PHNwYW4gY2xhc3M9ImZkLWsiPicraysnPC9zcGFuPjxzcGFuIGNsYXNzPSJmZC12Ij4nK3YrJzwvc3Bhbj48L2Rpdj4nOwp9CgpmdW5jdGlvbiBjb25mQmFkZ2UoYyl7CiAgY29uc3QgY2xzID0gYz09PSdQSUktUmVzdHJpY3RlZCc/J2NvbmYtcGlpLXInIDogYz09PSdQSUktQ29uZmlkZW50aWFsJz8nY29uZi1waWktYycgOgogICAgICAgICAgICAgIGM9PT0nQ29uZmlkZW50aWFsJz8nY29uZi1jb25mJyAgICA6IGM9PT0nUHVibGljJz8nY29uZi1wdWInIDogJ2NvbmYtaW50JzsKICByZXR1cm4gJzxzcGFuIGNsYXNzPSJjb25mLXBpbGwgJytjbHMrJyI+Jytlc2MoYykrJzwvc3Bhbj4nOwp9CgovLyDilIDilIAgSU5JVCDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKKGFzeW5jIGZ1bmN0aW9uIGluaXQoKXsKICB0cnl7CiAgICBjb25zdCByID0gYXdhaXQgZmV0Y2goQVBJKycvYXBpL2hlYWx0aCcpOwogICAgaWYoci5vayl7CiAgICAgIGNvbnN0IGQgPSBhd2FpdCByLmpzb24oKTsKICAgICAgZWwoJ2FwaS1zdGF0dXMnKS50ZXh0Q29udGVudCA9ICd2JytkLnZlcnNpb24rJyBDb25uZWN0ZWQgXHUyNzEzJzsKICAgIH0gZWxzZSB7CiAgICAgIGVsKCdhcGktc3RhdHVzJykudGV4dENvbnRlbnQgPSAnTW9jayBNb2RlJzsKICAgIH0KICB9IGNhdGNoKGUpewogICAgZWwoJ2FwaS1zdGF0dXMnKS50ZXh0Q29udGVudCA9ICdNb2NrIE1vZGUnOwogIH0KICB0cnl7CiAgICBjb25zdCByID0gYXdhaXQgZmV0Y2goQVBJKycvYXBpL2NvbmZpZycpOwogICAgaWYoci5vayl7CiAgICAgIGNvbnN0IGMgPSBhd2FpdCByLmpzb24oKTsKICAgICAgaWYoYy5MT0NBTF9NRVRBX0RJUiAmJiBjLkxPQ0FMX01FVEFfRElSIT09Jy4vbWV0YWRhdGFfZmlsZXMnKSBzdignY2YtbWV0YScsIGMuTE9DQUxfTUVUQV9ESVIpOwogICAgICBpZihjLkxPQ0FMX0RJQ1RfRElSICYmIGMuTE9DQUxfRElDVF9ESVIhPT0nLi9kaWN0aW9uYXJ5X2ZpbGVzJykgc3YoJ2NmLWRpY3QnLCBjLkxPQ0FMX0RJQ1RfRElSKTsKICAgICAgaWYoYy5MT0NBTF9DT0RFX0RJUiAmJiBjLkxPQ0FMX0NPREVfRElSIT09Jy4vY29kZV9yZXBvX3NhbXBsZXMnKSBzdignY2YtbG9jYWwnLCBjLkxPQ0FMX0NPREVfRElSKTsKICAgICAgaWYoYy5HSVRIVUJfUkVQTykgc3YoJ2NmLWdoLXJlcG8nLCBjLkdJVEhVQl9SRVBPKTsKICAgICAgaWYoYy5UQUNIWU9OX1VSTCkgc3YoJ2NmLXRhY2h5b24tdXJsJywgYy5UQUNIWU9OX1VSTCk7CiAgICAgIC8vIFByZS1maWxsIEVSV0lOIGZpZWxkcyBmcm9tIHNlcnZlciBjb25maWcgKHNldCBpbiBTVEFSVF9BUFAuYmF0KQogICAgICBpZihjLkVSV0lOX0JBU0VfVVJMICYmIGMuRVJXSU5fQkFTRV9VUkwhPT0naHR0cDovL2Vyd2luLWRpLWhvc3QvZXJ3aW4tZGkvYXBpJykKICAgICAgICBzdignY2YtZXJ3aW4tdXJsJywgYy5FUldJTl9CQVNFX1VSTCk7CiAgICAgIGlmKGMuRVJXSU5fU1lTVEVNKSAgc3YoJ2NmLWVyd2luLXN5c3RlbScsICBjLkVSV0lOX1NZU1RFTSk7CiAgICAgIGlmKGMuRVJXSU5fRU5WKSAgICAgc3YoJ2NmLWVyd2luLWVudicsICAgICBjLkVSV0lOX0VOVik7CiAgICAgIGlmKGMuRVJXSU5fUFJPSkVDVCkgc3YoJ2NmLWVyd2luLXByb2plY3QnLCBjLkVSV0lOX1BST0pFQ1QpOwogICAgfQogIH0gY2F0Y2goZSl7fQp9KSgpOwoKZWwoJ3EnKS5hZGRFdmVudExpc3RlbmVyKCdrZXlkb3duJywgZnVuY3Rpb24oZSl7IGlmKGUua2V5PT09J0VudGVyJykgZG9VbmlmaWVkU2VhcmNoKCk7IH0pOwoKZnVuY3Rpb24gc2V0UShxKXsgZWwoJ3EnKS52YWx1ZT1xOyBlbCgncScpLmZvY3VzKCk7IH0KCmZ1bmN0aW9uIHF1aWNrU2VhcmNoKHEpewogIGVsKCdxJykudmFsdWU9cTsKICAvLyBoaWRlIG11bHRpLXRhIGlmIHZpc2libGUKICB2YXIgdGE9ZWwoJ211bHRpLXRhJyk7IGlmKHRhKSB0YS5zdHlsZS5kaXNwbGF5PSdub25lJzsKICBkb1VuaWZpZWRTZWFyY2goKTsKfQoKdmFyIF9tdWx0aU1vZGUgPSBmYWxzZTsKZnVuY3Rpb24gdG9nZ2xlTXVsdGlNb2RlKCl7CiAgX211bHRpTW9kZSA9ICFfbXVsdGlNb2RlOwogIHZhciB0YSAgPSBlbCgnbXVsdGktdGEnKTsKICB2YXIgaW5wID0gZWwoJ3EnKTsKICB2YXIgYnRuID0gZWwoJ211bHRpLXRvZ2dsZS1idG4nKTsKICBpZihfbXVsdGlNb2RlKXsKICAgIHRhLnN0eWxlLmRpc3BsYXkgID0gJ2Jsb2NrJzsKICAgIGlucC5zdHlsZS5kaXNwbGF5ID0gJ25vbmUnOwogICAgYnRuLnN0eWxlLmJhY2tncm91bmQgPSAnbGluZWFyLWdyYWRpZW50KDEzNWRlZyx2YXIoLS10ZWFsMyksdmFyKC0tc2t5KSknOwogICAgYnRuLnN0eWxlLmNvbG9yID0gJyNmZmYnOwogICAgYnRuLnN0eWxlLmJvcmRlckNvbG9yID0gJ3RyYW5zcGFyZW50JzsKICAgIHRhLmZvY3VzKCk7CiAgfSBlbHNlIHsKICAgIHRhLnN0eWxlLmRpc3BsYXkgID0gJ25vbmUnOwogICAgaW5wLnN0eWxlLmRpc3BsYXkgPSAnYmxvY2snOwogICAgYnRuLnN0eWxlLmJhY2tncm91bmQgPSAndmFyKC0tcGFuZWwpJzsKICAgIGJ0bi5zdHlsZS5jb2xvciA9ICd2YXIoLS10MSknOwogICAgYnRuLnN0eWxlLmJvcmRlckNvbG9yID0gJ3ZhcigtLWJkcjIpJzsKICAgIGlucC5mb2N1cygpOwogIH0KfQoKLy8gVW5pZmllZCBlbnRyeSBwb2ludCDigJQgZGV0ZWN0cyBzaW5nbGUgdnMgbXVsdGkgYXV0b21hdGljYWxseQphc3luYyBmdW5jdGlvbiBkb1VuaWZpZWRTZWFyY2goKXsKICBpZihfbXVsdGlNb2RlKXsKICAgIC8vIE11bHRpLWZpZWxkIG1vZGUKICAgIGRvTXVsdGlTZWFyY2goKTsKICAgIHJldHVybjsKICB9CiAgdmFyIHJhdyA9IGVsKCdxJykudmFsdWUudHJpbSgpOwogIGlmKCFyYXcpeyB0b2FzdCgnRW50ZXIgYSB0YWJsZSwgZmllbGQgb3IgY29tbWEtc2VwYXJhdGVkIGxpc3QnLCd3YXJuJyk7IHJldHVybjsgfQogIC8vIERldGVjdCBpZiBjb21tYS1zZXBhcmF0ZWQgb3IgbXVsdGktd29yZCDigJQgdHJlYXQgYXMgbXVsdGktZmllbGQKICB2YXIgZmllbGRzID0gcmF3LnNwbGl0KC9bXG5ccixdKy8pLm1hcChmdW5jdGlvbihmKXtyZXR1cm4gZi50cmltKCk7fSkuZmlsdGVyKEJvb2xlYW4pOwogIGlmKGZpZWxkcy5sZW5ndGggPiAxKXsKICAgIC8vIEF1dG8tZmlsbCBtdWx0aSB0ZXh0YXJlYSBhbmQgcnVuIG11bHRpIHNlYXJjaAogICAgZWwoJ211bHRpLXRhJykudmFsdWUgPSBmaWVsZHMuam9pbignXG4nKTsKICAgIGRvTXVsdGlTZWFyY2goKTsKICB9IGVsc2UgewogICAgZG9TZWFyY2goKTsKICB9Cn0KCmZ1bmN0aW9uIHRvZ2dsZVNyYyhzKXsKICBzcmNzW3NdID0gIXNyY3Nbc107CiAgZWwoJ3RvZy0nK3MpLmNsYXNzTGlzdC50b2dnbGUoJ29uJywgc3Jjc1tzXSk7Cn0KCi8vIOKUgOKUgCBTRUFSQ0gg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmFzeW5jIGZ1bmN0aW9uIGRvU2VhcmNoKCl7CiAgY29uc3QgcSA9IGVsKCdxJykudmFsdWUudHJpbSgpOwogIGlmKCFxKXsgdG9hc3QoJ0VudGVyIGEgc2VhcmNoIHRlcm0nLCd3YXJuJyk7IHJldHVybjsgfQoKICAvLyBIaWRlIGFsbCBzdGF0ZSBwYW5lcwogIGhpZGUoJ3N0LWVtcHR5Jyk7CiAgaGlkZUFsbFBhbmVzKCk7CiAgc2hvdygnc3QtbG9hZGluZycpOwogIGhpZGUoJ3RvcC1iYXInKTsKICB2YXIgYWliPWVsKCdhaS1zdW1tYXJ5LWJveCcpOyBpZihhaWIpIGFpYi5zdHlsZS5kaXNwbGF5PSdub25lJzsKICB2YXIgYXRyPWVsKCdhaS10cmlnZ2VyLXJvdycpOyBpZihhdHIpIGF0ci5zdHlsZS5kaXNwbGF5PSdub25lJzsKICBlbCgnZ28tYnRuJykuZGlzYWJsZWQgPSB0cnVlOwogIGVsKCdtdWx0aS1hY3Rpb24tYmFyJykuc3R5bGUuZGlzcGxheSA9ICdub25lJzsKICBhbmltTG9hZFN0ZXBzKCk7CgogIHRyeXsKICAgIGNvbnN0IGJvZHkgPSB7CiAgICAgIHF1ZXJ5OiAgICAgICAgICBxLAogICAgICB1c2VfZXJ3aW46ICAgICAgc3Jjcy5lcndpbiwKICAgICAgdXNlX2RhdGFfZGljdDogIHNyY3MuZGljdCwKICAgICAgdXNlX2NvZGVfcmVwbzogIHNyY3MuY29kZSwKICAgICAgdXNlX3RhY2h5b246ICAgIHNyY3MudGFjaHlvbiwKICAgICAgLy8gRVJXSU4gREkg4oCUIHNoYXJlZCBjb25uZWN0aW9uICh1c2VkIGJ5IEFMTCBlbmRwb2ludHMpCiAgICAgIGVyd2luX2Jhc2VfdXJsOiAgICAgZ3YoJ2NmLWVyd2luLXVybCcpICAgICAgfHwgbnVsbCwKICAgICAgZXJ3aW5fYXBpX2tleTogICAgICBndignY2YtZXJ3aW4ta2V5JykgICAgICB8fCBudWxsLAogICAgICAvLyBFUldJTiBESSDigJQgZGVmYXVsdCBwYXJhbXMgKGF1dG8tZmlsbGVkIGZvciBhbnkgZW5kcG9pbnQgdGhhdCBhY2NlcHRzIHRoZW0pCiAgICAgIGVyd2luX3N5c3RlbTogICAgICAgZ3YoJ2NmLWVyd2luLXN5c3RlbScpICAgfHwgbnVsbCwKICAgICAgZXJ3aW5fZW52OiAgICAgICAgICBndignY2YtZXJ3aW4tZW52JykgICAgICAgfHwgbnVsbCwKICAgICAgZXJ3aW5fcHJvamVjdDogICAgICBndignY2YtZXJ3aW4tcHJvamVjdCcpICB8fCBudWxsLAogICAgICBlcndpbl9saW5lYWdlX3R5cGU6IGd2KCdjZi1lcndpbi1saW50eXBlJykgIHx8ICdEVUFMJywKICAgICAgLy8gRVJXSU4gREkg4oCUIHBlci1lbmRwb2ludCBvdmVycmlkZXMKICAgICAgZXJ3aW5fZXBfdGFibGVuYW1lOiBndignY2YtZXAtdGFibGVuYW1lJykgICB8fCBudWxsLAogICAgICBlcndpbl9lcF9jb2xuYW1lOiAgIGd2KCdjZi1lcC1jb2xuYW1lJykgICAgIHx8IG51bGwsCiAgICAgIGVyd2luX2VwX29ianR5cGU6ICAgZ3YoJ2NmLWVwLW9ianR5cGUnKSAgICAgfHwgJ1RBQkxFJywKICAgICAgZXJ3aW5fZXBfbWFwbmFtZTogICBndignY2YtZXAtbWFwbmFtZScpICAgICB8fCBudWxsLAogICAgICBlcndpbl9lcF9nbG9zc2NhdDogIGd2KCdjZi1lcC1nbG9zc2NhdCcpICAgIHx8IG51bGwsCiAgICAgIC8vIE90aGVyIHNvdXJjZSBwYXJhbXMKICAgICAgZ2l0aHViX3Rva2VuOiAgIGd2KCdjZi1naC10b2tlbicpICAgIHx8IG51bGwsCiAgICAgIGdpdGh1Yl9yZXBvOiAgICBndignY2YtZ2gtcmVwbycpICAgICB8fCBudWxsLAogICAgICBsb2NhbF9jb2RlX2RpcjogZ3YoJ2NmLWxvY2FsJykgICAgICAgfHwgbnVsbCwKICAgICAgbG9jYWxfbWV0YV9kaXI6IGd2KCdjZi1tZXRhJykgICAgICAgIHx8IG51bGwsCiAgICAgIGxvY2FsX2RpY3RfZGlyOiBndignY2YtZGljdCcpICAgICAgICB8fCBudWxsLAogICAgICB0YWNoeW9uX3VybDogICAgZ3YoJ2NmLXRhY2h5b24tdXJsJykgfHwgbnVsbCwKICAgICAgdGFjaHlvbl9rZXk6ICAgIGd2KCdjZi10YWNoeW9uLWtleScpIHx8IG51bGwsCiAgICB9OwogICAgY29uc3QgciA9IGF3YWl0IGZldGNoKEFQSSsnL2FwaS9zZWFyY2gnLCB7CiAgICAgIG1ldGhvZDonUE9TVCcsCiAgICAgIGhlYWRlcnM6eydDb250ZW50LVR5cGUnOidhcHBsaWNhdGlvbi9qc29uJ30sCiAgICAgIGJvZHk6IEpTT04uc3RyaW5naWZ5KGJvZHkpCiAgICB9KTsKICAgIGlmKCFyLm9rKSB0aHJvdyBuZXcgRXJyb3IoYXdhaXQgci50ZXh0KCkpOwogICAgRCA9IGF3YWl0IHIuanNvbigpOwogICAgaWYoIWhpc3QuaW5jbHVkZXMocSkpeyBoaXN0LnVuc2hpZnQocSk7IGhpc3Q9aGlzdC5zbGljZSgwLDYpOyByZW5kZXJSZWNlbnQoKTsgfQogICAgcmVuZGVyQWxsKEQpOwogIH0gY2F0Y2goZSl7CiAgICB0b2FzdCgnRXJyb3I6ICcrZS5tZXNzYWdlLCAnZXJyb3InKTsKICAgIHNob3coJ3N0LWVtcHR5Jyk7CiAgfSBmaW5hbGx5ewogICAgaGlkZSgnc3QtbG9hZGluZycpOwogICAgZWwoJ2dvLWJ0bicpLmRpc2FibGVkID0gZmFsc2U7CiAgfQp9CgpmdW5jdGlvbiBhbmltTG9hZFN0ZXBzKCl7CiAgWydlcndpbicsJ2RpY3QnLCdjb2RlJywndGFjaHlvbicsJ2dyYXBoJ10uZm9yRWFjaChmdW5jdGlvbihzLCBpKXsKICAgIGNvbnN0IGUgPSBlbCgnbGQtJytzKTsKICAgIGlmKCFlKSByZXR1cm47CiAgICBlLmNsYXNzTGlzdC5yZW1vdmUoJ2FjdGl2ZScsJ2RvbmUnKTsKICAgIHNldFRpbWVvdXQoZnVuY3Rpb24oKXsgZS5jbGFzc0xpc3QuYWRkKCdhY3RpdmUnKTsgfSwgaSozOTApOwogICAgc2V0VGltZW91dChmdW5jdGlvbigpeyBlLmNsYXNzTGlzdC5yZW1vdmUoJ2FjdGl2ZScpOyBlLmNsYXNzTGlzdC5hZGQoJ2RvbmUnKTsgfSwgaSozOTArNTgwKTsKICB9KTsKfQoKZnVuY3Rpb24gdXBkYXRlU3JjUGlsbHMoc291cmNlc1VzZWQpewogIHZhciBtYXAgPSB7J0VSV0lOIERJJzonZXJ3aW4nLCdEYXRhIERpY3Rpb25hcnknOidkaWN0JywnQ29kZSBSZXBvJzonY29kZScsJ1RhY2h5b24gQUknOid0YWNoeW9uJ307CiAgT2JqZWN0LmtleXMobWFwKS5mb3JFYWNoKGZ1bmN0aW9uKG5hbWUpewogICAgdmFyIGtleSA9IG1hcFtuYW1lXTsKICAgIHZhciBwaWxsID0gZWwoJ3BpbGwtJytrZXkpOwogICAgdmFyIGRvdCAgPSBlbCgnc2RvdC0nK2tleSk7CiAgICBpZighcGlsbCkgcmV0dXJuOwogICAgdmFyIG1hdGNoID0gc291cmNlc1VzZWQuZmluZChmdW5jdGlvbihzKXsgcmV0dXJuIHMuaW5jbHVkZXMobmFtZSk7IH0pfHwnJzsKICAgIHBpbGwuY2xhc3NOYW1lPSdzcmMtcGlsbCc7CiAgICBpZihtYXRjaC5pbmNsdWRlcygnTElWRScpfHxtYXRjaC5pbmNsdWRlcygnR0lUSFVCJykpewogICAgICBwaWxsLmNsYXNzTGlzdC5hZGQoJ2xpdmUnKTsKICAgIH0gZWxzZSBpZihtYXRjaC5pbmNsdWRlcygnTE9DQUwnKXx8bWF0Y2guaW5jbHVkZXMoJ0ZJTEUnKXx8bWF0Y2guaW5jbHVkZXMoJ0VYQ0VMJykpewogICAgICBwaWxsLmNsYXNzTGlzdC5hZGQoJ2ZpbGUnKTsKICAgIH0gZWxzZSBpZihtYXRjaC5pbmNsdWRlcygnTExNJyl8fG1hdGNoLmluY2x1ZGVzKCdBSScpKXsKICAgICAgcGlsbC5jbGFzc0xpc3QuYWRkKCdhY3RpdmUnKTsKICAgIH0gZWxzZSB7CiAgICAgIHBpbGwuY2xhc3NMaXN0LmFkZCgnbW9jaycpOwogICAgfQogIH0pOwp9CgovLyBTdG9yZSBsYXN0IEQgZm9yIHdoZW4gdXNlciBjbGlja3MgR2VuZXJhdGUKdmFyIF9sYXN0U2VhcmNoRGF0YSA9IG51bGw7CgpmdW5jdGlvbiByZW5kZXJBSVN1bW1hcnlCb3goZCl7CiAgX2xhc3RTZWFyY2hEYXRhID0gZDsKICB2YXIgYm94ID0gZWwoJ2FpLXN1bW1hcnktYm94Jyk7CiAgdmFyIHRyaWcgPSBlbCgnYWktdHJpZ2dlci1yb3cnKTsKICBpZighYm94IHx8ICF0cmlnKSByZXR1cm47CgogIC8vIEhpZGUgdGhlIGV4cGFuZGVkIGJveCDigJQgb25seSBzaG93IHRyaWdnZXIgYnV0dG9uCiAgYm94LnN0eWxlLmRpc3BsYXkgPSAnbm9uZSc7CgogIC8vIFNob3cgdHJpZ2dlciByb3cgaWYgdGhlcmUgaXMgZ3JhcGggZGF0YQogIHZhciBub2RlcyA9IGQubm9kZXN8fFtdOwogIGlmKG5vZGVzLmxlbmd0aCA9PT0gMCl7IHRyaWcuc3R5bGUuZGlzcGxheT0nbm9uZSc7IHJldHVybjsgfQogIHRyaWcuc3R5bGUuZGlzcGxheSA9ICdibG9jayc7CgogIC8vIFJlc2V0IGJ1dHRvbiBzdGF0ZQogIHZhciBidG4gPSBlbCgnYWktdHJpZ2dlci1idG4nKTsKICBpZihidG4pewogICAgYnRuLmRpc2FibGVkID0gZmFsc2U7CiAgICBidG4uc3R5bGUub3BhY2l0eSA9ICcxJzsKICAgIGJ0bi5xdWVyeVNlbGVjdG9yKCdkaXY6bGFzdC1jaGlsZCcpLmlubmVySFRNTCA9ICcmI3gyMTkyOyc7CiAgfQp9CgpmdW5jdGlvbiBjbG9zZUFJU3VtbWFyeSgpewogIHZhciBib3ggPSBlbCgnYWktc3VtbWFyeS1ib3gnKTsKICB2YXIgdHJpZyA9IGVsKCdhaS10cmlnZ2VyLXJvdycpOwogIGlmKGJveCkgYm94LnN0eWxlLmRpc3BsYXk9J25vbmUnOwogIGlmKHRyaWcpIHRyaWcuc3R5bGUuZGlzcGxheT0nYmxvY2snOwp9Cgphc3luYyBmdW5jdGlvbiBnZW5lcmF0ZUFJU3VtbWFyeSgpewogIHZhciBkID0gX2xhc3RTZWFyY2hEYXRhOwogIGlmKCFkKSByZXR1cm47CgogIHZhciBidG4gPSBlbCgnYWktdHJpZ2dlci1idG4nKTsKICAvLyBTaG93IGxvYWRpbmcgc3RhdGUgb24gYnV0dG9uCiAgaWYoYnRuKXsKICAgIGJ0bi5kaXNhYmxlZCA9IHRydWU7CiAgICBidG4uc3R5bGUub3BhY2l0eSA9ICcuNyc7CiAgICBidG4ucXVlcnlTZWxlY3RvcignZGl2Omxhc3QtY2hpbGQnKS5pbm5lckhUTUwgPSAnPHNwYW4gc3R5bGU9ImZvbnQtc2l6ZToxNHB4O2FuaW1hdGlvbjpzcGluIC42cyBsaW5lYXIgaW5maW5pdGU7ZGlzcGxheTppbmxpbmUtYmxvY2siPiYjeDIxQkI7PC9zcGFuPic7CiAgfQoKICAvLyBJZiBUYWNoeW9uIGNvbmZpZ3VyZWQg4oCUIGNhbGwgdGhlIEFQSSBmb3IgcmVhbCBBSSBzdW1tYXJ5CiAgdmFyIHRhY2h5b25VcmwgPSBndignY2YtdGFjaHlvbi11cmwnKTsKICBpZih0YWNoeW9uVXJsICYmIGQubm9kZXMgJiYgZC5ub2Rlcy5sZW5ndGgpewogICAgdHJ5ewogICAgICB2YXIgciA9IGF3YWl0IGZldGNoKCcvYXBpL3NlYXJjaCcsIHsKICAgICAgICBtZXRob2Q6J1BPU1QnLCBoZWFkZXJzOnsnQ29udGVudC1UeXBlJzonYXBwbGljYXRpb24vanNvbid9LAogICAgICAgIGJvZHk6IEpTT04uc3RyaW5naWZ5KHsKICAgICAgICAgIHF1ZXJ5OiBkLnF1ZXJ5LCB1c2VfdGFjaHlvbjogdHJ1ZSwKICAgICAgICAgIHVzZV9lcndpbjogZmFsc2UsIHVzZV9kYXRhX2RpY3Q6IGZhbHNlLCB1c2VfY29kZV9yZXBvOiBmYWxzZSwKICAgICAgICAgIHRhY2h5b25fdXJsOiB0YWNoeW9uVXJsLCB0YWNoeW9uX2tleTogZ3YoJ2NmLXRhY2h5b24ta2V5JykKICAgICAgICB9KQogICAgICB9KTsKICAgICAgaWYoci5vayl7CiAgICAgICAgdmFyIGZyZXNoID0gYXdhaXQgci5qc29uKCk7CiAgICAgICAgaWYoZnJlc2guYWlfaW5zaWdodHMpIGQgPSBPYmplY3QuYXNzaWduKHt9LCBkLCB7YWlfaW5zaWdodHM6IGZyZXNoLmFpX2luc2lnaHRzfSk7CiAgICAgIH0KICAgIH0gY2F0Y2goZSl7fQogIH0KCiAgX2ZpbGxBSVN1bW1hcnlCb3goZCk7Cn0KCmZ1bmN0aW9uIF9maWxsQUlTdW1tYXJ5Qm94KGQpewogIHZhciBib3ggID0gZWwoJ2FpLXN1bW1hcnktYm94Jyk7CiAgdmFyIHRyaWcgPSBlbCgnYWktdHJpZ2dlci1yb3cnKTsKICBpZighYm94KSByZXR1cm47CgogIC8vIEJ1aWxkIHN0cnVjdHVyZWQgc3VtbWFyeSBmcm9tIGdyYXBoIGRhdGEKICB2YXIgbm9kZXMgICA9IGQubm9kZXN8fFtdOwogIHZhciBtZXRhICAgID0gZC5tZXRhZGF0YXx8W107CiAgdmFyIHNvdXJjZXMgPSBbLi4ubmV3IFNldChub2Rlcy5maWx0ZXIoZnVuY3Rpb24obil7cmV0dXJuIG4udHlwZT09PSdzb3VyY2VfdGFibGUnO30pLm1hcChmdW5jdGlvbihuKXtyZXR1cm4gbi5zeXN0ZW18fG4ubGFiZWw7fSkpXTsKICB2YXIgdGFyZ2V0cyA9IFsuLi5uZXcgU2V0KG5vZGVzLmZpbHRlcihmdW5jdGlvbihuKXtyZXR1cm4gbi50eXBlPT09J3RhcmdldF90YWJsZSc7fSkubWFwKGZ1bmN0aW9uKG4pe3JldHVybiBuLnN5c3RlbXx8bi5sYWJlbDt9KSldOwogIHZhciB0cmZzICAgID0gbm9kZXMuZmlsdGVyKGZ1bmN0aW9uKG4pe3JldHVybiBuLnR5cGU9PT0ndHJhbnNmb3JtYXRpb24nO30pLm1hcChmdW5jdGlvbihuKXtyZXR1cm4gbi5sYWJlbDt9KTsKICB2YXIgaGlnaFJpc2s9IG1ldGEuZmlsdGVyKGZ1bmN0aW9uKG0pe3JldHVybiBtLnJpc2tfbGV2ZWw9PT0nSGlnaCc7fSkubGVuZ3RoOwogIHZhciBwaWkgICAgID0gbWV0YS5maWx0ZXIoZnVuY3Rpb24obSl7cmV0dXJuIChtLmNvbmZpZGVudGlhbGl0eXx8JycpLmluY2x1ZGVzKCdQSUknKTt9KS5sZW5ndGg7CgogIGlmKHNvdXJjZXMubGVuZ3RoPT09MCAmJiB0YXJnZXRzLmxlbmd0aD09PTApIHsgaWYoYm94KSBib3guc3R5bGUuZGlzcGxheT0nbm9uZSc7IHJldHVybjsgfQoKICAvLyBTaG93IHN1bW1hcnkgYm94LCBoaWRlIHRyaWdnZXIKICBpZihib3gpIGJveC5zdHlsZS5kaXNwbGF5PSdibG9jayc7CiAgaWYodHJpZykgdHJpZy5zdHlsZS5kaXNwbGF5PSdub25lJzsKCiAgLy8gRmlsbCBzdHJ1Y3R1cmVkIGNhcmRzCiAgZWwoJ2FpLXN1bS1zb3VyY2VzJykuaW5uZXJIVE1MICAgPSBzb3VyY2VzLmxlbmd0aCA/IHNvdXJjZXMubWFwKGZ1bmN0aW9uKHMpe3JldHVybiAnPHNwYW4gY2xhc3M9ImR0eXBlIiBzdHlsZT0iZm9udC1zaXplOjlweCI+Jytlc2MocykrJzwvc3Bhbj4nO30pLmpvaW4oJyAnKSA6ICfigJQnOwogIGVsKCdhaS1zdW0tdGFyZ2V0cycpLmlubmVySFRNTCAgID0gdGFyZ2V0cy5sZW5ndGggPyB0YXJnZXRzLm1hcChmdW5jdGlvbihzKXtyZXR1cm4gJzxzcGFuIGNsYXNzPSJkdHlwZSIgc3R5bGU9ImZvbnQtc2l6ZTo5cHg7Ym9yZGVyLWNvbG9yOnJnYmEoNTksMTU4LDI1NSwuMik7Y29sb3I6dmFyKC0tc2t5MikiPicrZXNjKHMpKyc8L3NwYW4+Jzt9KS5qb2luKCcgJykgOiAn4oCUJzsKICBlbCgnYWktc3VtLXRyYW5zZm9ybScpLmlubmVySFRNTCA9IHRyZnMubGVuZ3RoID8gdHJmcy5zbGljZSgwLDIpLm1hcChmdW5jdGlvbih0KXtyZXR1cm4gJzxzcGFuIGNsYXNzPSJkdHlwZSIgc3R5bGU9ImZvbnQtc2l6ZTo5cHg7Ym9yZGVyLWNvbG9yOnJnYmEoMjQ1LDE2NiwzNSwuMik7Y29sb3I6dmFyKC0tZ29sZDIpIj4nK2VzYyh0LnJlcGxhY2UoJ1xuJywnICcpLnNsaWNlKDAsMzApKSsnPC9zcGFuPic7fSkuam9pbignICcpIDogJzxzcGFuIHN0eWxlPSJjb2xvcjp2YXIoLS10MykiPkRpcmVjdCBtYXBwaW5nPC9zcGFuPic7CiAgZWwoJ2FpLXN1bS1yaXNrJykuaW5uZXJIVE1MICAgICAgPSAnPHNwYW4gY2xhc3M9InJiIHJiLUgiIHN0eWxlPSJmb250LXNpemU6OXB4Ij4nK2hpZ2hSaXNrKycgSGlnaDwvc3Bhbj4gJm5ic3A7PHNwYW4gc3R5bGU9ImZvbnQtc2l6ZTo5cHg7Y29sb3I6dmFyKC0tdDIpIj4nK3BpaSsnIFBJSSBmaWVsZHM8L3NwYW4+JzsKCiAgLy8gRGF0YSBmbG93IHN0cmlwCiAgdmFyIGZsb3dDb250ZW50ID0gZWwoJ2FpLXN1bS1mbG93LWNvbnRlbnQnKTsKICBpZihmbG93Q29udGVudCAmJiBub2Rlcy5sZW5ndGg+MCl7CiAgICB2YXIgZmxvd05vZGVzID0gW107CiAgICBub2Rlcy5mb3JFYWNoKGZ1bmN0aW9uKG4pewogICAgICBpZihuLnR5cGU9PT0nc291cmNlX3RhYmxlJykgZmxvd05vZGVzLnVuc2hpZnQobik7CiAgICAgIGVsc2UgaWYobi50eXBlPT09J3RhcmdldF90YWJsZScpIGZsb3dOb2Rlcy5wdXNoKG4pOwogICAgICBlbHNlIGZsb3dOb2Rlcy5zcGxpY2UoTWF0aC5mbG9vcihmbG93Tm9kZXMubGVuZ3RoLzIpLDAsbik7CiAgICB9KTsKICAgIGZsb3dDb250ZW50LmlubmVySFRNTCA9IGZsb3dOb2Rlcy5zbGljZSgwLDYpLm1hcChmdW5jdGlvbihuLGkpewogICAgICB2YXIgY2xzID0gbi50eXBlPT09J3NvdXJjZV90YWJsZSc/J2xmLXNyYyc6bi50eXBlPT09J3RhcmdldF90YWJsZSc/J2xmLXRndCc6Jyc7CiAgICAgIHZhciBhcnIgPSBpPE1hdGgubWluKGZsb3dOb2Rlcy5sZW5ndGgsNiktMT8nPHNwYW4gY2xhc3M9ImxmLWFyciI+JnJhcnI7PC9zcGFuPic6Jyc7CiAgICAgIHJldHVybiAnPHNwYW4gY2xhc3M9ImxmLW4gJytjbHMrJyIgc3R5bGU9ImZvbnQtc2l6ZTo5cHgiPicrZXNjKChuLmxhYmVsfHwnJykucmVwbGFjZSgnXG4nLCcgJykuc2xpY2UoMCwyMikpKyc8L3NwYW4+JythcnI7CiAgICB9KS5qb2luKCcnKTsKICAgIGVsKCdhaS1zdW0tZmxvdycpLnN0eWxlLmRpc3BsYXk9J2Jsb2NrJzsKICB9CgogIC8vIElmIFRhY2h5b24gcmVzcG9uZGVkIOKAlCBzaG93IHJpY2hlciBpbnNpZ2h0OyBlbHNlIHNob3cgc3RydWN0dXJhbCBzdW1tYXJ5CiAgaWYoZC5haV9pbnNpZ2h0cyl7CiAgICBlbCgnYWktc3VtLWluc2lnaHQnKS50ZXh0Q29udGVudCA9IGQuYWlfaW5zaWdodHM7CiAgICBlbCgnYWktc3VtLXN0YXR1cycpLnRleHRDb250ZW50ICA9ICcmI3gxRjdFMjsgVGFjaHlvbiBBSSBMaXZlJzsKICAgIGVsKCdhaS1zdW0tYmFkZ2UnKS50ZXh0Q29udGVudCAgID0gJ1RBQ0hZT04gQUknOwogIH0gZWxzZSB7CiAgICAvLyBBdXRvLWdlbmVyYXRlZCBzdHJ1Y3R1cmFsIHN1bW1hcnkgZnJvbSBncmFwaAogICAgdmFyIHN1bW1hcnkgPSAnRGF0YSBmbG93cyBmcm9tICc7CiAgICBpZihzb3VyY2VzLmxlbmd0aCkgc3VtbWFyeSArPSBzb3VyY2VzLnNsaWNlKDAsMikuam9pbignICYgJyk7CiAgICBlbHNlIHN1bW1hcnkgKz0gJ3NvdXJjZSBzeXN0ZW1zJzsKICAgIGlmKHRyZnMubGVuZ3RoKSBzdW1tYXJ5ICs9ICcgdGhyb3VnaCAnICsgdHJmcy5zbGljZSgwLDEpLmpvaW4oJywgJyk7CiAgICBpZih0YXJnZXRzLmxlbmd0aCkgc3VtbWFyeSArPSAnIGludG8gJyArIHRhcmdldHMuc2xpY2UoMCwyKS5qb2luKCcgJiAnKTsKICAgIHN1bW1hcnkgKz0gJy4gRm91bmQgJyttZXRhLmxlbmd0aCsnIGZpZWxkJysobWV0YS5sZW5ndGghPT0xPydzJzonJykrJyBhY3Jvc3MgJytub2Rlcy5sZW5ndGgrJyBvYmplY3RzJzsKICAgIGlmKGhpZ2hSaXNrPjApIHN1bW1hcnkgKz0gJywgaW5jbHVkaW5nICcraGlnaFJpc2srJyBoaWdoLXJpc2sgZmllbGQnKyhoaWdoUmlzaz4xPydzJzonJyk7CiAgICBpZihwaWk+MCkgc3VtbWFyeSArPSAnIGFuZCAnK3BpaSsnIFBJSS1jbGFzc2lmaWVkIGZpZWxkJysocGlpPjE/J3MnOicnKTsKICAgIHN1bW1hcnkgKz0gJy4nOwogICAgdmFyIGluc0VsID0gZWwoJ2FpLXN1bS1pbnNpZ2h0Jyk7CiAgICBpZihpbnNFbCl7IGluc0VsLnRleHRDb250ZW50ID0gc3VtbWFyeTsgaW5zRWwuY2xhc3NMaXN0LmFkZCgnYWktdHlwaW5nJyk7IH0KICAgIHZhciBzdGF0RWwgPSBlbCgnYWktc3VtLXN0YXR1cycpOwogICAgaWYoc3RhdEVsKSBzdGF0RWwudGV4dENvbnRlbnQgPSAnQXV0by1nZW5lcmF0ZWQgKFRhY2h5b24gbm90IGNvbmZpZ3VyZWQpJzsKICAgIHZhciBiYWRnZUVsID0gZWwoJ2FpLXN1bS1iYWRnZScpOwogICAgaWYoYmFkZ2VFbCkgYmFkZ2VFbC50ZXh0Q29udGVudCA9ICdTVFJVQ1RVUkFMIFNVTU1BUlknOwogICAgdmFyIGN0YUVsID0gZWwoJ2FpLXN1bS1jdGEnKTsKICAgIGlmKGN0YUVsKSBjdGFFbC5zdHlsZS5kaXNwbGF5ID0gZ3YoJ2NmLXRhY2h5b24tdXJsJykgPyAnbm9uZScgOiAnYmxvY2snOwogIH0KICAvLyBSZXNldCBidXR0b24KICB2YXIgYnRuID0gZWwoJ2FpLXRyaWdnZXItYnRuJyk7CiAgaWYoYnRuKXsgYnRuLmRpc2FibGVkPWZhbHNlOyBidG4uc3R5bGUub3BhY2l0eT0nMSc7IH0KfQoKZnVuY3Rpb24gcmVuZGVyUmVjZW50KCl7CiAgZWwoJ3JlY2VudC1zZWMnKS5zdHlsZS5kaXNwbGF5ID0gJ2Jsb2NrJzsKICBlbCgncmVjZW50LWxpc3QnKS5pbm5lckhUTUwgPSBoaXN0Lm1hcChmdW5jdGlvbihoKXsKICAgIHJldHVybiAnPGRpdiBjbGFzcz0icmVjZW50LWl0ZW0iIG9uY2xpY2s9InNldFEoXCcnK2VzYyhoKSsnXCcpO2RvU2VhcmNoKCkiPicrZXNjKGgpKyc8L2Rpdj4nOwogIH0pLmpvaW4oJycpOwp9CgovLyDilIDilIAgUkVOREVSIEFMTCDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZnVuY3Rpb24gcmVuZGVyQWxsKGQpewogIGVsKCd0b3AtYmFyJykuc3R5bGUuZGlzcGxheSA9ICdibG9jayc7CiAgZWwoJ3Jlcy10aXRsZScpLnRleHRDb250ZW50ID0gJ0xpbmVhZ2U6ICInK2QucXVlcnkrJyInOwogIGVsKCdleHAtcm93Jykuc3R5bGUuZGlzcGxheSA9ICdmbGV4JzsKCiAgLy8gVXBkYXRlIHNpZGViYXIgc291cmNlIHBpbGxzCiAgdXBkYXRlU3JjUGlsbHMoZC5zb3VyY2VzX3VzZWR8fFtdKTsKCiAgLy8gUmVuZGVyIEFJIHN1bW1hcnkgYm94CiAgcmVuZGVyQUlTdW1tYXJ5Qm94KGQpOwoKICAvLyBTdGF0dXMgcGlsbHMKICB2YXIgU0MgPSB7ZXJ3aW46JyM3RkQ5NjInLGRpY3Q6JyM5QjZERkYnLGNvZGU6JyMzQjlFRkYnLHRhY2h5b246JyNGNUE2MjMnfTsKICB2YXIgU04gPSB7J0VSV0lOIERJJzonZXJ3aW4nLCdEYXRhIERpY3Rpb25hcnknOidkaWN0JywnQ29kZSBSZXBvJzonY29kZScsJ1RhY2h5b24gQUknOid0YWNoeW9uJ307CiAgdmFyIHBpbGxzID0gZC5zb3VyY2VzX3VzZWQubWFwKGZ1bmN0aW9uKHMpewogICAgdmFyIGsgPSBPYmplY3Qua2V5cyhTTikuZmluZChmdW5jdGlvbih4KXsgcmV0dXJuIHMuaW5jbHVkZXMoeCk7IH0pOwogICAgdmFyIGNvbG9yID0gayA/IChTQ1tTTltrXV18fCcjNEU2RThBJykgOiAnIzRFNkU4QSc7CiAgICB2YXIgaXNMaXZlID0gcy5pbmNsdWRlcygnTElWRScpOwogICAgdmFyIGlzRmlsZSA9IHMuaW5jbHVkZXMoJ0xPQ0FMJyl8fHMuaW5jbHVkZXMoJ0dJVEhVQicpfHxzLmluY2x1ZGVzKCdGSUxFJyk7CiAgICB2YXIgY2xzID0gaXNMaXZlPydzcC1saXZlJzppc0ZpbGU/J3NwLWZpbGUnOidzcC1tb2NrJzsKICAgIHZhciBkb3RDbHMgPSBpc0xpdmU/J2xpdmUnOmlzRmlsZT8nZmlsZSc6J21vY2snOwogICAgdmFyIGRrID0gayA/IFNOW2tdIDogbnVsbDsKICAgIGlmKGRrKXsKICAgICAgdmFyIGRvdCA9IGVsKCdkb3QtJytkayk7CiAgICAgIGlmKGRvdCkgZG90LmNsYXNzTmFtZSA9ICdzdC1kb3QgJysoZGs9PT0ndGFjaHlvbic/J2FpJzpkb3RDbHMpOwogICAgfQogICAgcmV0dXJuICc8ZGl2IGNsYXNzPSJzcGlsbCAnK2NscysnIj48ZGl2IGNsYXNzPSJzZG90IiBzdHlsZT0iYmFja2dyb3VuZDonK2NvbG9yKyciPjwvZGl2PicrZXNjKHMpKyc8L2Rpdj4nOwogIH0pLmpvaW4oJycpOwoKICAvLyBTT1IgY291bnQKICB2YXIgc29yU2V0ID0gbmV3IFNldCgpOwogIGQubWV0YWRhdGEuZm9yRWFjaChmdW5jdGlvbihtKXsgKG0uc29yc3x8W10pLmZvckVhY2goZnVuY3Rpb24ocyl7IHNvclNldC5hZGQocy5zb3IpOyB9KTsgfSk7CiAgaWYoc29yU2V0LnNpemU+MCl7CiAgICBwaWxscyArPSAnPGRpdiBjbGFzcz0ic3BpbGwiIHN0eWxlPSJib3JkZXItY29sb3I6cmdiYSgyNDUsMTY2LDM1LC4zKTtjb2xvcjp2YXIoLS1nb2xkMikiPjxkaXYgY2xhc3M9InNkb3QiIHN0eWxlPSJiYWNrZ3JvdW5kOnZhcigtLWdvbGQpIj48L2Rpdj5cdTI2MDUgJytzb3JTZXQuc2l6ZSsnIFNPUicrKHNvclNldC5zaXplIT09MT8ncyc6JycpKyc8L2Rpdj4nOwogIH0KICBwaWxscyArPSAnPGRpdiBjbGFzcz0ic3BpbGwiPjxkaXYgY2xhc3M9InNkb3QiIHN0eWxlPSJiYWNrZ3JvdW5kOnZhcigtLXQyKSI+PC9kaXY+XHUyM0YxICcrZC5zZWFyY2hfdGltZV9tcysnbXM8L2Rpdj4nOwogIHBpbGxzICs9ICc8ZGl2IGNsYXNzPSJzcGlsbCI+PGRpdiBjbGFzcz0ic2RvdCIgc3R5bGU9ImJhY2tncm91bmQ6dmFyKC0tdDIpIj48L2Rpdj5cdUQ4M0RcdUREMTcgJytkLm5vZGVzLmxlbmd0aCsnbiAmYnVsbDsgJytkLmVkZ2VzLmxlbmd0aCsnZTwvZGl2Pic7CiAgZWwoJ3N0YXQtcGlsbHMnKS5pbm5lckhUTUwgPSBwaWxsczsKCiAgaWYoZC5haV9pbnNpZ2h0cyl7CiAgICBlbCgnYWktc2VjJykuc3R5bGUuZGlzcGxheSA9ICdibG9jayc7CiAgICBlbCgnYWktdGV4dCcpLnRleHRDb250ZW50ID0gZC5haV9pbnNpZ2h0czsKICB9CgogIHJlbmRlck1ldGFkYXRhKGQubWV0YWRhdGEpOwogIHJlbmRlclNvdXJjZXMoZCk7CiAgLy8gU2hvdyBsaW5lYWdlIHBhbmUgZmlyc3QsIFRIRU4gcmVuZGVyIGdyYXBoCiAgc2hvd1RhYignbGluZWFnZScpOwogIHNldFRpbWVvdXQoZnVuY3Rpb24oKXsKICAgIHNldExldmVsKCdzeXN0ZW0nKTsKICB9LCA4MCk7CiAgLy8gTWV0YWRhdGEgdGFiIGlzIHByZS1yZW5kZXJlZCDigJQgdXNlciBjYW4gc3dpdGNoIHdpdGhvdXQgcmUtcmVuZGVyCn0KCi8vIOKUgOKUgCBUQUJTIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAp2YXIgQUxMX1RBQlMgPSBbJ2xpbmVhZ2UnLCdtZXRhZGF0YScsJ3NvdXJjZXMnLCdtdWx0aScsJ2Vyd2luJ107CgpmdW5jdGlvbiBzaG93VGFiKHQpewogIEFMTF9UQUJTLmZvckVhY2goZnVuY3Rpb24oeCl7CiAgICB2YXIgcGFuZSA9IGVsKCdwYW5lLScreCk7CiAgICB2YXIgYnRuICA9IGVsKCd0Yi0nK3gpOwogICAgaWYocGFuZSkgcGFuZS5zdHlsZS5kaXNwbGF5ID0geD09PXQgPyAnYmxvY2snOidub25lJzsKICAgIGlmKGJ0bikgIGJ0bi5jbGFzc0xpc3QudG9nZ2xlKCdvbicsIHg9PT10KTsKICB9KTsKICBpZih0PT09J2xpbmVhZ2UnICYmIEQpIHNldFRpbWVvdXQocmVuZGVyR3JhcGgsIDYwKTsKICBpZih0PT09J2Vyd2luJyAmJiBlcndpbkVuZHBvaW50cy5sZW5ndGg9PT0wKSBsb2FkRXJ3aW5FbmRwb2ludHMoKTsKfQoKZnVuY3Rpb24gaGlkZUFsbFBhbmVzKCl7CiAgQUxMX1RBQlMuZm9yRWFjaChmdW5jdGlvbih4KXsKICAgIHZhciBwID0gZWwoJ3BhbmUtJyt4KTsgaWYocCkgcC5zdHlsZS5kaXNwbGF5PSdub25lJzsKICB9KTsKfQoKLy8g4pSA4pSAIExFVkVMUyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZnVuY3Rpb24gc2V0TGV2ZWwobHYpewogIGN1ckxldmVsID0gbHY7CiAgWydzeXN0ZW0nLCd0YWJsZScsJ2ZpZWxkJ10uZm9yRWFjaChmdW5jdGlvbih4KXsKICAgIHZhciBiID0gZWwoJ2x2LScreCk7IGlmKGIpIGIuY2xhc3NMaXN0LnRvZ2dsZSgnb24nLCB4PT09bHYpOwogIH0pOwogIHVwZGF0ZUJyZWFkY3J1bWIobHYpOwogIGlmKEQpIHJlbmRlckdyYXBoKCk7Cn0KCmZ1bmN0aW9uIHNldFZpZXcodil7CiAgY3VyVmlldyA9IHY7CiAgZWwoJ3Z0LWVyZCcpLmNsYXNzTGlzdC50b2dnbGUoJ29uJywgdj09PSdlcmQnKTsKICBlbCgndnQtZm9yY2UnKS5jbGFzc0xpc3QudG9nZ2xlKCdvbicsIHY9PT0nZm9yY2UnKTsKICBlbCgnZXJkLXZpZXcnKS5zdHlsZS5kaXNwbGF5ICAgPSB2PT09J2VyZCcgICA/ICdibG9jayc6J25vbmUnOwogIGVsKCdmb3JjZS12aWV3Jykuc3R5bGUuZGlzcGxheSA9IHY9PT0nZm9yY2UnID8gJ2Jsb2NrJzonbm9uZSc7CiAgcmVuZGVyR3JhcGgoKTsKfQoKZnVuY3Rpb24gdXBkYXRlQnJlYWRjcnVtYihsdil7CiAgaWYoIUQpIHJldHVybjsKICB2YXIgcSAgID0gRC5xdWVyeS50b1VwcGVyQ2FzZSgpOwogIHZhciB0YmwgPSBxLmluY2x1ZGVzKCcuJyk/cS5zcGxpdCgnLicpWzBdOnE7CiAgdmFyIHBhcnRzID0gWwogICAge2w6J0FsbCBTeXN0ZW1zJywgdjonc3lzdGVtJ30sCiAgICB7bDp0YmwrJyBUYWJsZXMnLCB2Oid0YWJsZSd9CiAgXTsKICBpZihsdj09PSdmaWVsZCcpIHBhcnRzLnB1c2goe2w6J0ZpZWxkcycsdjonZmllbGQnfSk7CiAgZWwoJ2JjJykuaW5uZXJIVE1MID0gcGFydHMubWFwKGZ1bmN0aW9uKHAsaSl7CiAgICB2YXIgc2VwID0gaT4wID8gJzxzcGFuIGNsYXNzPSJiYy1zZXAiPiZyc2FxdW87PC9zcGFuPicgOiAnJzsKICAgIGlmKHAudj09PWx2KSByZXR1cm4gc2VwKyc8c3BhbiBjbGFzcz0iYmMtY3VyIj4nK2VzYyhwLmwpKyc8L3NwYW4+JzsKICAgIHJldHVybiBzZXArJzxzcGFuIGNsYXNzPSJiYy1pdGVtIiBvbmNsaWNrPSJzZXRMZXZlbChcJycrcC52KydcJykiPicrZXNjKHAubCkrJzwvc3Bhbj4nOwogIH0pLmpvaW4oJycpOwp9CgovLyDilIDilIAgR1JBUEggREFUQSDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZnVuY3Rpb24gYnVpbGRHcmFwaERhdGEoZCwgbHYpewogIHZhciBxICAgPSBkLnF1ZXJ5LnRvVXBwZXJDYXNlKCk7CiAgdmFyIHRibCA9IHEuaW5jbHVkZXMoJy4nKT9xLnNwbGl0KCcuJylbMF06cTsKICB2YXIgZmxkID0gcS5pbmNsdWRlcygnLicpP3Euc3BsaXQoJy4nKVsxXTpudWxsOwoKICBpZihsdj09PSdzeXN0ZW0nKXsKICAgIHZhciBzeXNNYXAgPSB7fTsKICAgIHZhciBUTSA9IHtzb3VyY2VfdGFibGU6J3N5c19zcmMnLHRhcmdldF90YWJsZTonc3lzX3RndCcsdHJhbnNmb3JtYXRpb246J3N5c190cmYnLGNvZGVfcmVwbzonc3lzX2NvZGUnfTsKICAgIGQubm9kZXMuZm9yRWFjaChmdW5jdGlvbihuKXsKICAgICAgdmFyIHNuID0gbi5zeXN0ZW18fChuLnR5cGU9PT0nY29kZV9yZXBvJz8nQ29kZSBSZXBvc2l0b3J5JzonVW5rbm93bicpOwogICAgICBpZighc3lzTWFwW3NuXSkgc3lzTWFwW3NuXSA9IHtpZDonU1lTXycrc24ucmVwbGFjZSgvXFcvZywnXycpLGxhYmVsOnNuLHR5cGU6VE1bbi50eXBlXXx8J3N5c19zcmMnLGZpZWxkczpbXX07CiAgICB9KTsKICAgIHZhciBub2RlcyA9IE9iamVjdC52YWx1ZXMoc3lzTWFwKTsKICAgIHZhciBlZGdlcyA9IFtdOwogICAgdmFyIHNlZW4gID0ge307CiAgICBkLmVkZ2VzLmZvckVhY2goZnVuY3Rpb24oZSl7CiAgICAgIHZhciBzbiA9IGQubm9kZXMuZmluZChmdW5jdGlvbihuKXsgcmV0dXJuIG4uaWQ9PT1lLnNvdXJjZTsgfSk7CiAgICAgIHZhciB0biA9IGQubm9kZXMuZmluZChmdW5jdGlvbihuKXsgcmV0dXJuIG4uaWQ9PT1lLnRhcmdldDsgfSk7CiAgICAgIGlmKCFzbnx8IXRuKSByZXR1cm47CiAgICAgIHZhciBzayA9ICdTWVNfJysoc24uc3lzdGVtfHwoc24udHlwZT09PSdjb2RlX3JlcG8nPydDb2RlIFJlcG9zaXRvcnknOidVbmtub3duJykpLnJlcGxhY2UoL1xXL2csJ18nKTsKICAgICAgdmFyIHRrID0gJ1NZU18nKyh0bi5zeXN0ZW18fCh0bi50eXBlPT09J2NvZGVfcmVwbyc/J0NvZGUgUmVwb3NpdG9yeSc6J1Vua25vd24nKSkucmVwbGFjZSgvXFcvZywnXycpOwogICAgICBpZihzaz09PXRrKSByZXR1cm47CiAgICAgIHZhciBrID0gc2srJz4nK3RrOwogICAgICBpZighc2VlbltrXSl7IHNlZW5ba109MTsgZWRnZXMucHVzaCh7aWQ6ayxzb3VyY2U6c2ssdGFyZ2V0OnRrLGxhYmVsOmUubGFiZWx8fCdmbG93J30pOyB9CiAgICB9KTsKICAgIHJldHVybiB7bm9kZXM6bm9kZXMsIGVkZ2VzOmVkZ2VzfTsKICB9CgogIGlmKGx2PT09J3RhYmxlJyl7CiAgICByZXR1cm4gewogICAgICBub2RlczogZC5ub2Rlcy5tYXAoZnVuY3Rpb24obil7IHJldHVybiBPYmplY3QuYXNzaWduKHt9LG4se2xhYmVsOm4ubGFiZWwucmVwbGFjZSgvXG4uKi8sJycpfSk7IH0pLAogICAgICBlZGdlczogZC5lZGdlcwogICAgfTsKICB9CgogIGlmKGx2PT09J2ZpZWxkJyl7CiAgICB2YXIgbm9kZXM9W10sZWRnZXM9W107CiAgICB2YXIgc049ZC5ub2Rlcy5maW5kKGZ1bmN0aW9uKG4pe3JldHVybiBuLnR5cGU9PT0nc291cmNlX3RhYmxlJzt9KTsKICAgIHZhciB0Tj1kLm5vZGVzLmZpbmQoZnVuY3Rpb24obil7cmV0dXJuIG4udHlwZT09PSd0YXJnZXRfdGFibGUnO30pOwogICAgdmFyIHJOPWQubm9kZXMuZmluZChmdW5jdGlvbihuKXtyZXR1cm4gbi50eXBlPT09J3RyYW5zZm9ybWF0aW9uJzt9KTsKICAgIHZhciBjTj1kLm5vZGVzLmZpbmQoZnVuY3Rpb24obil7cmV0dXJuIG4udHlwZT09PSdjb2RlX3JlcG8nO30pOwogICAgaWYoc04pKHNOLmZpZWxkc3x8W10pLmZvckVhY2goZnVuY3Rpb24oZil7CiAgICAgIG5vZGVzLnB1c2goe2lkOidGU18nK2YubmFtZSxsYWJlbDpmLm5hbWUsdHlwZTonc3JjX2ZpZWxkJyxmaWVsZE5hbWU6Zi5uYW1lLAogICAgICAgIHRhYmxlTmFtZTpzTi5sYWJlbCxkYXRhVHlwZTpmLnR5cGUsZGVzY3JpcHRpb246Zi5kZXNjcmlwdGlvbixyaXNrX2xldmVsOmYucmlza3x8J0xvdycsCiAgICAgICAgaGlnaGxpZ2h0ZWQ6ISEoZmxkJiZmLm5hbWU9PT1mbGQpLGZpZWxkczpbXX0pOwogICAgfSk7CiAgICBpZihyTikgbm9kZXMucHVzaChPYmplY3QuYXNzaWduKHt9LHJOLHtsYWJlbDonVHJhbnNmb3JtJyxmaWVsZHM6W119KSk7CiAgICBpZih0TikodE4uZmllbGRzfHxbXSkuZm9yRWFjaChmdW5jdGlvbihmKXsKICAgICAgbm9kZXMucHVzaCh7aWQ6J0ZUXycrZi5uYW1lLGxhYmVsOmYubmFtZSx0eXBlOid0Z3RfZmllbGQnLGZpZWxkTmFtZTpmLm5hbWUsCiAgICAgICAgdGFibGVOYW1lOnROLmxhYmVsLGRhdGFUeXBlOmYudHlwZSxkZXNjcmlwdGlvbjpmLmRlc2NyaXB0aW9uLHJpc2tfbGV2ZWw6Zi5yaXNrfHwnTG93JywKICAgICAgICBoaWdobGlnaHRlZDohIShmbGQmJmYubmFtZT09PWZsZCksZmllbGRzOltdfSk7CiAgICB9KTsKICAgIGlmKGNOKSBub2Rlcy5wdXNoKE9iamVjdC5hc3NpZ24oe30sY04se2xhYmVsOidDb2RlIFJlcG8nLGZpZWxkczpbXX0pKTsKICAgIGlmKHNOJiZyTikoc04uZmllbGRzfHxbXSkuZm9yRWFjaChmdW5jdGlvbihmKXsKICAgICAgZWRnZXMucHVzaCh7aWQ6J0VGU18nK2YubmFtZSxzb3VyY2U6J0ZTXycrZi5uYW1lLHRhcmdldDpyTi5pZCxsYWJlbDonXHUyMTkyJ30pOwogICAgfSk7CiAgICBpZihyTiYmdE4pKHROLmZpZWxkc3x8W10pLmZvckVhY2goZnVuY3Rpb24oZil7CiAgICAgIGVkZ2VzLnB1c2goe2lkOidFRlRfJytmLm5hbWUsc291cmNlOnJOLmlkLHRhcmdldDonRlRfJytmLm5hbWUsbGFiZWw6J1x1MjE5Mid9KTsKICAgIH0pOwogICAgaWYoY04mJnJOKSBlZGdlcy5wdXNoKHtpZDonRUMnLHNvdXJjZTpjTi5pZCx0YXJnZXQ6ck4uaWQsbGFiZWw6J2ltcGwnfSk7CiAgICByZXR1cm4ge25vZGVzOm5vZGVzLGVkZ2VzOmVkZ2VzfTsKICB9CgogIHJldHVybiB7bm9kZXM6ZC5ub2RlcywgZWRnZXM6ZC5lZGdlc307Cn0KCmZ1bmN0aW9uIHJlbmRlckdyYXBoKCl7CiAgaWYoIUQpIHJldHVybjsKICB2YXIgdGl0bGVzID0gewogICAgc3lzdGVtOicmI3gxRjNFMjsgU1lTVEVNIExFVkVMJywKICAgIHRhYmxlOicmI3gxRjVDMzsgVEFCTEUgTEVWRUwnLAogICAgZmllbGQ6JyYjeDFGNTMyOyBGSUVMRCBMRVZFTCcKICB9OwogIHZhciBzdWJzID0gewogICAgc3lzdGVtOidDcm9zcy1zeXN0ZW0gbGluZWFnZTogc291cmNlIHBsYXRmb3JtcyBcdTIxOTIgRVRMIFx1MjE5MiBkYXRhIHdhcmVob3VzZScsCiAgICB0YWJsZTonVGFibGUtdG8tdGFibGUgZGF0YSBmbG93IHdpdGggdHJhbnNmb3JtYXRpb24gbG9naWMnLAogICAgZmllbGQ6J0NvbHVtbi10by1jb2x1bW4gbWFwcGluZyB3aXRoIGJ1c2luZXNzIHJ1bGVzJwogIH07CiAgZWwoJ2ctdGl0bGUnKS5pbm5lckhUTUwgPSAodGl0bGVzW2N1ckxldmVsXXx8JycpICsgJyAmbWRhc2g7ICcgKyAoY3VyVmlldz09PSdlcmQnPydFUkQgRmxvdyc6J0ZvcmNlIEdyYXBoJyk7CiAgZWwoJ2ctc3ViJykudGV4dENvbnRlbnQgPSBzdWJzW2N1ckxldmVsXXx8Jyc7CiAgdmFyIGdkID0gYnVpbGRHcmFwaERhdGEoRCwgY3VyTGV2ZWwpOwogIGlmKGN1clZpZXc9PT0nZXJkJykgZHJhd0VSRChnZC5ub2RlcywgZ2QuZWRnZXMpOwogIGVsc2UgZHJhd0ZvcmNlKGdkLm5vZGVzLCBnZC5lZGdlcyk7CiAgZHJhd0xlZ2VuZCgpOwp9CgovLyDilIDilIAgRVJEIEdSQVBIIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApmdW5jdGlvbiBlcmRDbGFzcyh0KXsKICBpZih0PT09J3NvdXJjZV90YWJsZSd8fHQ9PT0nc3JjX2ZpZWxkJ3x8dD09PSdzeXNfc3JjJykgcmV0dXJuICdlYy1zcmMnOwogIGlmKHQ9PT0ndGFyZ2V0X3RhYmxlJ3x8dD09PSd0Z3RfZmllbGQnfHx0PT09J3N5c190Z3QnKSByZXR1cm4gJ2VjLXRndCc7CiAgaWYodD09PSd0cmFuc2Zvcm1hdGlvbid8fHQ9PT0nc3lzX3RyZicpIHJldHVybiAnZWMtdHJmJzsKICBpZih0PT09J2NvZGVfcmVwbyd8fHQ9PT0nc3lzX2NvZGUnKSByZXR1cm4gJ2VjLWNvZGUnOwogIHJldHVybiAnZWMtc3JjJzsKfQpmdW5jdGlvbiBlcmRDb2xvcih0KXsKICBpZighdCkgcmV0dXJuICcjNEU2RThBJzsKICBpZih0LmluZGV4T2YoJ3NyYycpPi0xfHx0PT09J3NvdXJjZV90YWJsZScpIHJldHVybiAnIzdGRDk2Mic7CiAgaWYodC5pbmRleE9mKCd0Z3QnKT4tMXx8dD09PSd0YXJnZXRfdGFibGUnKSByZXR1cm4gJyMzQjlFRkYnOwogIGlmKHQuaW5kZXhPZigndHJmJyk+LTF8fHQ9PT0ndHJhbnNmb3JtYXRpb24nKSByZXR1cm4gJyNGNUE2MjMnOwogIGlmKHQuaW5kZXhPZignY29kZScpPi0xKSByZXR1cm4gJyMwMEM5QjEnOwogIHJldHVybiAnIzRFNkU4QSc7Cn0KZnVuY3Rpb24gZXJkQmFkZ2UodCl7CiAgdmFyIG09e3NvdXJjZV90YWJsZTonU1JDJyx0YXJnZXRfdGFibGU6J1RHVCcsdHJhbnNmb3JtYXRpb246J0VUTCcsY29kZV9yZXBvOidDT0RFJywKICAgIHNyY19maWVsZDonU1JDJyx0Z3RfZmllbGQ6J1RHVCcsc3lzX3NyYzonU1lTJyxzeXNfdGd0OidTWVMnLHN5c190cmY6J0VUTCcsc3lzX2NvZGU6J0NPREUnfTsKICByZXR1cm4gbVt0XXx8J09CSic7Cn0KZnVuY3Rpb24gZXJkSWNvbih0KXsKICB2YXIgbT17c291cmNlX3RhYmxlOicmI3gxRjVDNDsnLHRhcmdldF90YWJsZTonJiN4MUYzRUQ7Jyx0cmFuc2Zvcm1hdGlvbjonJiN4MjY5OTsnLGNvZGVfcmVwbzonJiN4MUY0QkI7JywKICAgIHNyY19maWVsZDonJiN4MUY0RTU7Jyx0Z3RfZmllbGQ6JyYjeDFGNEU0Oycsc3lzX3NyYzonJiN4MUYzRTI7JyxzeXNfdGd0OicmI3gxRjNFRDsnLAogICAgc3lzX3RyZjonJiN4MjY5OTsnLHN5c19jb2RlOicmI3gxRjRCQjsnfTsKICByZXR1cm4gbVt0XXx8JyYjeDFGNEU2Oyc7Cn0KCmZ1bmN0aW9uIGRyYXdFUkQobm9kZXMsIGVkZ2VzKXsKICB2YXIgaG9zdCA9IGVsKCdlcmQtaG9zdCcpOwogIHZhciBzdmcgID0gZWwoJ2VyZC1zdmcnKTsKICBob3N0LmlubmVySFRNTCA9ICcnOwogIHdoaWxlKHN2Zy5maXJzdENoaWxkKSBzdmcucmVtb3ZlQ2hpbGQoc3ZnLmZpcnN0Q2hpbGQpOwogIHN2Zy5zZXRBdHRyaWJ1dGUoJ3dpZHRoJywnMzAwMCcpOwogIHN2Zy5zZXRBdHRyaWJ1dGUoJ2hlaWdodCcsJzIwMDAnKTsKCiAgLy8gQXJyb3cgbWFya2VyIGRlZnMKICB2YXIgbnMgID0gJ2h0dHA6Ly93d3cudzMub3JnLzIwMDAvc3ZnJzsKICB2YXIgZGVmID0gZG9jdW1lbnQuY3JlYXRlRWxlbWVudE5TKG5zLCdkZWZzJyk7CiAgW1snZycsJyM3RkQ5NjInXSxbJ2InLCcjM0I5RUZGJ10sWydhJywnI0Y1QTYyMyddLFsnYycsJyMwMEM5QjEnXV0uZm9yRWFjaChmdW5jdGlvbihrYyl7CiAgICB2YXIgbSA9IGRvY3VtZW50LmNyZWF0ZUVsZW1lbnROUyhucywnbWFya2VyJyk7CiAgICBtLnNldEF0dHJpYnV0ZSgnaWQnLCdlbScra2NbMF0pOyBtLnNldEF0dHJpYnV0ZSgndmlld0JveCcsJzAgLTQgOCA4Jyk7CiAgICBtLnNldEF0dHJpYnV0ZSgncmVmWCcsJzcnKTsgbS5zZXRBdHRyaWJ1dGUoJ3JlZlknLCcwJyk7CiAgICBtLnNldEF0dHJpYnV0ZSgnbWFya2VyV2lkdGgnLCc1Jyk7IG0uc2V0QXR0cmlidXRlKCdtYXJrZXJIZWlnaHQnLCc1Jyk7IG0uc2V0QXR0cmlidXRlKCdvcmllbnQnLCdhdXRvJyk7CiAgICB2YXIgcCA9IGRvY3VtZW50LmNyZWF0ZUVsZW1lbnROUyhucywncGF0aCcpOwogICAgcC5zZXRBdHRyaWJ1dGUoJ2QnLCdNMCwtNEw4LDBMMCw0Jyk7IHAuc2V0QXR0cmlidXRlKCdmaWxsJyxrY1sxXSk7IHAuc2V0QXR0cmlidXRlKCdvcGFjaXR5JywnLjknKTsKICAgIG0uYXBwZW5kQ2hpbGQocCk7IGRlZi5hcHBlbmRDaGlsZChtKTsKICB9KTsKICBzdmcuYXBwZW5kQ2hpbGQoZGVmKTsKCiAgLy8gQ29sdW1uIGxheW91dAogIHZhciBjb2xPZiA9IHtzb3VyY2VfdGFibGU6MCxzcmNfZmllbGQ6MCxzeXNfc3JjOjAsdHJhbnNmb3JtYXRpb246MSxzeXNfdHJmOjEsCiAgICAgICAgICAgICAgIHRhcmdldF90YWJsZToyLHRndF9maWVsZDoyLHN5c190Z3Q6Mixjb2RlX3JlcG86MyxzeXNfY29kZTozfTsKICB2YXIgY29scyAgPSBbW10sW10sW10sW11dOwogIG5vZGVzLmZvckVhY2goZnVuY3Rpb24obil7IGNvbHNbKGNvbE9mW24udHlwZV0hPT11bmRlZmluZWQ/Y29sT2Zbbi50eXBlXTowKV0ucHVzaChuKTsgfSk7CgogIHZhciBDVz0xOTIsIENHPTEzMCwgUlM9MjgsIFJHPTE2LCBGSD0yNCwgSEg9MzQ7CiAgdmFyIENYPVszNiwgMzYrQ1crQ0csIDM2KzIqKENXK0NHKSwgMzYrMyooQ1crQ0cpXTsKICBub2RlUG9zaXRpb25zID0ge307CgogIGNvbHMuZm9yRWFjaChmdW5jdGlvbihjb2wsIGNpKXsKICAgIHZhciB5ID0gUlM7CiAgICBjb2wuZm9yRWFjaChmdW5jdGlvbihuKXsKICAgICAgdmFyIGZsID0gbi5maWVsZHN8fFtdOwogICAgICB2YXIgaCAgPSBISCArIE1hdGgubWF4KDEsIGZsLmxlbmd0aCkqRkg7CiAgICAgIG5vZGVQb3NpdGlvbnNbbi5pZF0gPSB7eDpDWFtjaV0sIHk6eSwgdzpDVywgaDpoLCBjb2w6Y2l9OwogICAgICB5ICs9IGggKyBSRzsKICAgIH0pOwogIH0pOwoKICB2YXIgaGlGaWVsZCA9IEQgJiYgRC5xdWVyeSAmJiBELnF1ZXJ5LmluZGV4T2YoJy4nKT4tMSA/IEQucXVlcnkudG9VcHBlckNhc2UoKS5zcGxpdCgnLicpWzFdIDogbnVsbDsKCiAgbm9kZXMuZm9yRWFjaChmdW5jdGlvbihuKXsKICAgIHZhciBwb3MgPSBub2RlUG9zaXRpb25zW24uaWRdOyBpZighcG9zKSByZXR1cm47CiAgICB2YXIgY2xzICAgPSBlcmRDbGFzcyhuLnR5cGUpOwogICAgdmFyIGNvbG9yID0gZXJkQ29sb3Iobi50eXBlKTsKICAgIHZhciBiYWRnZSA9IGVyZEJhZGdlKG4udHlwZSk7CiAgICB2YXIgaWNvbiAgPSBlcmRJY29uKG4udHlwZSk7CiAgICB2YXIgZmllbGRzID0gbi5maWVsZHN8fChuLmZpZWxkTmFtZT9be25hbWU6bi5maWVsZE5hbWV8fG4ubGFiZWwsdHlwZTpuLmRhdGFUeXBlfHwnJyxkZXNjcmlwdGlvbjpuLmRlc2NyaXB0aW9ufHwnJ31dOltdKTsKCiAgICB2YXIgZmllbGRzSHRtbCA9ICcnOwogICAgaWYoZmllbGRzLmxlbmd0aCl7CiAgICAgIGZpZWxkc0h0bWwgPSBmaWVsZHMubWFwKGZ1bmN0aW9uKGYpewogICAgICAgIHZhciBoaSA9ICEhKGhpRmllbGQgJiYgZi5uYW1lPT09aGlGaWVsZCk7CiAgICAgICAgdmFyIHJjID0ge0hpZ2g6JyNGRjRGNkQnLE1lZGl1bTonI0Y1QTYyMycsTG93OicjN0ZEOTYyJ31bZi5yaXNrfHwnTG93J118fCcjNEU2RThBJzsKICAgICAgICByZXR1cm4gJzxkaXYgY2xhc3M9ImVyZC1maWVsZCcrKGhpPycgaGknOicnKSsnIiBvbmNsaWNrPSJvbkVyZEZpZWxkKFwnJytlc2Mobi5pZCkrJ1wnLFwnJytlc2MoZi5uYW1lKSsnXCcpIiB0aXRsZT0iJytlc2MoZi5kZXNjcmlwdGlvbnx8Zi5uYW1lKSsnIj4nKwogICAgICAgICAgJzxkaXYgY2xhc3M9ImVyZC1hbmMiIHN0eWxlPSJib3JkZXItY29sb3I6Jytjb2xvcisnNTA7YmFja2dyb3VuZDonK2NvbG9yKycxNSI+PC9kaXY+JysKICAgICAgICAgICc8ZGl2IGNsYXNzPSJlcmQtZm4iJysoaGk/JyBzdHlsZT0iY29sb3I6I0Y1QTYyMyInOicnKSsnPicrZXNjKGYubmFtZSkrJzwvZGl2PicrCiAgICAgICAgICAnPGRpdiBjbGFzcz0iZXJkLWZ0Ij4nK2VzYyhmLnR5cGV8fCcnKSsnPC9kaXY+JysKICAgICAgICAgICc8ZGl2IGNsYXNzPSJlcmQtZnJpc2siIHN0eWxlPSJiYWNrZ3JvdW5kOicrcmMrJyI+PC9kaXY+JysKICAgICAgICAnPC9kaXY+JzsKICAgICAgfSkuam9pbignJyk7CiAgICB9IGVsc2UgewogICAgICBmaWVsZHNIdG1sID0gJzxkaXYgY2xhc3M9ImVyZC1maWVsZCIgb25jbGljaz0ib25FcmROb2RlKFwnJytlc2Mobi5pZCkrJ1wnKSI+JysKICAgICAgICAnPGRpdiBjbGFzcz0iZXJkLWFuYyIgc3R5bGU9ImJvcmRlci1jb2xvcjonK2NvbG9yKyc1MDtiYWNrZ3JvdW5kOicrY29sb3IrJzE1Ij48L2Rpdj4nKwogICAgICAgICc8ZGl2IGNsYXNzPSJlcmQtZm4iIHN0eWxlPSJjb2xvcjonK2NvbG9yKyciPicrZXNjKG4ubGFiZWx8fG4uaWQpKyc8L2Rpdj4nKwogICAgICAgICc8L2Rpdj4nOwogICAgfQoKICAgIHZhciBjYXJkID0gZG9jdW1lbnQuY3JlYXRlRWxlbWVudCgnZGl2Jyk7CiAgICBjYXJkLmNsYXNzTmFtZSA9ICdlcmQtY2FyZCAnK2NsczsKICAgIGNhcmQuaWQgPSAnZWMtJytuLmlkOwogICAgY2FyZC5zdHlsZS5jc3NUZXh0ID0gJ2xlZnQ6Jytwb3MueCsncHg7dG9wOicrcG9zLnkrJ3B4O3dpZHRoOicrcG9zLncrJ3B4Oyc7CiAgICBjYXJkLmlubmVySFRNTCA9CiAgICAgICc8ZGl2IGNsYXNzPSJlcmQtaGRyIj4nKwogICAgICAgICc8ZGl2IGNsYXNzPSJlcmQtZHJhZyI+PHNwYW4gY2xhc3M9ImVyZC1pY28iPicraWNvbisnPC9zcGFuPicrCiAgICAgICAgICAnPHNwYW4gY2xhc3M9ImVyZC1uYW1lIiB0aXRsZT0iJytlc2Mobi5sYWJlbHx8bi5pZCkrJyI+Jytlc2Mobi5sYWJlbHx8bi5pZCkrJzwvc3Bhbj48L2Rpdj4nKwogICAgICAgICc8c3BhbiBjbGFzcz0iZXJkLWJhZGdlIj4nK2JhZGdlKyc8L3NwYW4+JysKICAgICAgICAoZmllbGRzLmxlbmd0aCA/ICc8c3BhbiBjbGFzcz0iZXJkLWZjIj4nK2ZpZWxkcy5sZW5ndGgrJ2Y8L3NwYW4+JyA6ICcnKSsKICAgICAgJzwvZGl2PicrZmllbGRzSHRtbDsKCiAgICAvLyBEcmFnIG9uIGhlYWRlcgogICAgdmFyIGRyYWc9ZmFsc2UsIG94PTAsIG95PTA7CiAgICB2YXIgZHJhZ0hhbmRsZSA9IGNhcmQucXVlcnlTZWxlY3RvcignLmVyZC1kcmFnJyk7CiAgICBkcmFnSGFuZGxlLnN0eWxlLmN1cnNvciA9ICdtb3ZlJzsKICAgIGRyYWdIYW5kbGUuYWRkRXZlbnRMaXN0ZW5lcignbW91c2Vkb3duJywgZnVuY3Rpb24oZSl7CiAgICAgIGRyYWc9dHJ1ZTsgb3g9ZS5jbGllbnRYLW5vZGVQb3NpdGlvbnNbbi5pZF0ueDsgb3k9ZS5jbGllbnRZLW5vZGVQb3NpdGlvbnNbbi5pZF0ueTsKICAgICAgZS5wcmV2ZW50RGVmYXVsdCgpOyBlLnN0b3BQcm9wYWdhdGlvbigpOwogICAgfSk7CiAgICBkb2N1bWVudC5hZGRFdmVudExpc3RlbmVyKCdtb3VzZW1vdmUnLCBmdW5jdGlvbihlKXsKICAgICAgaWYoIWRyYWcpIHJldHVybjsKICAgICAgbm9kZVBvc2l0aW9uc1tuLmlkXS54ID0gZS5jbGllbnRYLW94OyBub2RlUG9zaXRpb25zW24uaWRdLnkgPSBlLmNsaWVudFktb3k7CiAgICAgIGNhcmQuc3R5bGUubGVmdCA9IG5vZGVQb3NpdGlvbnNbbi5pZF0ueCsncHgnOyBjYXJkLnN0eWxlLnRvcCA9IG5vZGVQb3NpdGlvbnNbbi5pZF0ueSsncHgnOwogICAgICBkcmF3Q29ubmVjdG9ycyhub2RlcywgZWRnZXMpOwogICAgfSk7CiAgICBkb2N1bWVudC5hZGRFdmVudExpc3RlbmVyKCdtb3VzZXVwJywgZnVuY3Rpb24oKXsgZHJhZz1mYWxzZTsgfSk7CgogICAgaG9zdC5hcHBlbmRDaGlsZChjYXJkKTsKICB9KTsKCiAgc2V0VGltZW91dChmdW5jdGlvbigpeyBkcmF3Q29ubmVjdG9ycyhub2RlcyxlZGdlcyk7IH0sIDUwKTsKICBzZXR1cEVSRFBhbigpOwogIGVyZFNjYWxlPTE7IGVyZFg9MjA7IGVyZFk9MjA7CiAgZWwoJ2VyZC13b3JsZCcpLnN0eWxlLnRyYW5zZm9ybSA9ICd0cmFuc2xhdGUoJytlcmRYKydweCwnK2VyZFkrJ3B4KSBzY2FsZSgnK2VyZFNjYWxlKycpJzsKfQoKZnVuY3Rpb24gZHJhd0Nvbm5lY3RvcnMobm9kZXMsIGVkZ2VzKXsKICB2YXIgc3ZnID0gZWwoJ2VyZC1zdmcnKTsKICB2YXIgbnMgID0gJ2h0dHA6Ly93d3cudzMub3JnLzIwMDAvc3ZnJzsKICAvLyBSZW1vdmUgYWxsIG5vbi1kZWZzIGNoaWxkcmVuCiAgdmFyIHRvUmVtb3ZlID0gW107CiAgZm9yKHZhciBpPTA7aTxzdmcuY2hpbGROb2Rlcy5sZW5ndGg7aSsrKXsKICAgIGlmKHN2Zy5jaGlsZE5vZGVzW2ldLm5vZGVOYW1lIT09J2RlZnMnKSB0b1JlbW92ZS5wdXNoKHN2Zy5jaGlsZE5vZGVzW2ldKTsKICB9CiAgdG9SZW1vdmUuZm9yRWFjaChmdW5jdGlvbihjKXsgc3ZnLnJlbW92ZUNoaWxkKGMpOyB9KTsKCiAgZWRnZXMuZm9yRWFjaChmdW5jdGlvbihlKXsKICAgIHZhciBzcCA9IG5vZGVQb3NpdGlvbnNbZS5zb3VyY2VdLCB0cCA9IG5vZGVQb3NpdGlvbnNbZS50YXJnZXRdOwogICAgaWYoIXNwfHwhdHApIHJldHVybjsKICAgIHZhciBzbiA9IG5vZGVzLmZpbmQoZnVuY3Rpb24obil7IHJldHVybiBuLmlkPT09ZS5zb3VyY2U7IH0pOwogICAgdmFyIGNvbCA9IGVyZENvbG9yKHNuID8gc24udHlwZSA6ICcnKTsKICAgIHZhciBtayA9IChzbiYmKHNuLnR5cGUuaW5kZXhPZignc3JjJyk+LTF8fHNuLnR5cGU9PT0nc291cmNlX3RhYmxlJykpPydnJzoKICAgICAgICAgICAgIChzbiYmKHNuLnR5cGUuaW5kZXhPZigndGd0Jyk+LTF8fHNuLnR5cGU9PT0ndGFyZ2V0X3RhYmxlJykpPydiJzoKICAgICAgICAgICAgIChzbiYmKHNuLnR5cGUuaW5kZXhPZigndHJmJyk+LTF8fHNuLnR5cGU9PT0ndHJhbnNmb3JtYXRpb24nKSk/J2EnOidjJzsKICAgIHZhciB4MT1zcC54K3NwLncsIHkxPXNwLnkrKHNwLmh8fDYwKS8yOwogICAgdmFyIHgyPXRwLngsICAgICAgeTI9dHAueSsodHAuaHx8NjApLzI7CiAgICB2YXIgbXg9KHgxK3gyKS8yOwogICAgdmFyIHBhdGggPSBkb2N1bWVudC5jcmVhdGVFbGVtZW50TlMobnMsJ3BhdGgnKTsKICAgIHBhdGguc2V0QXR0cmlidXRlKCdkJywnTScreDErJywnK3kxKycgQycrbXgrJywnK3kxKycgJytteCsnLCcreTIrJyAnK3gyKycsJyt5Mik7CiAgICBwYXRoLnNldEF0dHJpYnV0ZSgnZmlsbCcsJ25vbmUnKTsgcGF0aC5zZXRBdHRyaWJ1dGUoJ3N0cm9rZScsY29sKTsKICAgIHBhdGguc2V0QXR0cmlidXRlKCdzdHJva2Utd2lkdGgnLCcxLjgnKTsgcGF0aC5zZXRBdHRyaWJ1dGUoJ3N0cm9rZS1vcGFjaXR5JywnLjUnKTsKICAgIHBhdGguc2V0QXR0cmlidXRlKCdtYXJrZXItZW5kJywndXJsKCNlbScrbWsrJyknKTsKICAgIHN2Zy5hcHBlbmRDaGlsZChwYXRoKTsKICB9KTsKfQoKZnVuY3Rpb24gc2V0dXBFUkRQYW4oKXsKICB2YXIgcGFuICAgPSBlbCgnZXJkLXBhbicpOwogIHZhciB3b3JsZCA9IGVsKCdlcmQtd29ybGQnKTsKICBwYW4ub25tb3VzZWRvd24gPSBmdW5jdGlvbihlKXsKICAgIGlmKGUudGFyZ2V0LmNsb3Nlc3QoJy5lcmQtZHJhZycpKSByZXR1cm47CiAgICBlcmREcmFnZ2luZz10cnVlOyBlcmREcmFnU3RhcnQ9e3g6ZS5jbGllbnRYLWVyZFgsIHk6ZS5jbGllbnRZLWVyZFl9OwogICAgcGFuLnN0eWxlLmN1cnNvcj0nZ3JhYmJpbmcnOwogIH07CiAgZG9jdW1lbnQuYWRkRXZlbnRMaXN0ZW5lcignbW91c2Vtb3ZlJywgZnVuY3Rpb24oZSl7CiAgICBpZighZXJkRHJhZ2dpbmcpIHJldHVybjsKICAgIGVyZFg9ZS5jbGllbnRYLWVyZERyYWdTdGFydC54OyBlcmRZPWUuY2xpZW50WS1lcmREcmFnU3RhcnQueTsKICAgIHdvcmxkLnN0eWxlLnRyYW5zZm9ybT0ndHJhbnNsYXRlKCcrZXJkWCsncHgsJytlcmRZKydweCkgc2NhbGUoJytlcmRTY2FsZSsnKSc7CiAgfSk7CiAgZG9jdW1lbnQuYWRkRXZlbnRMaXN0ZW5lcignbW91c2V1cCcsIGZ1bmN0aW9uKCl7CiAgICBpZihlcmREcmFnZ2luZyl7IGVyZERyYWdnaW5nPWZhbHNlOyBlbCgnZXJkLXBhbicpLnN0eWxlLmN1cnNvcj0nJzsgfQogIH0pOwogIHBhbi5vbndoZWVsID0gZnVuY3Rpb24oZSl7CiAgICBlLnByZXZlbnREZWZhdWx0KCk7CiAgICB2YXIgZiA9IGUuZGVsdGFZPDAgPyAxLjEyIDogMC45OwogICAgZXJkU2NhbGUgPSBNYXRoLm1heCgwLjE1LCBNYXRoLm1pbig0LCBlcmRTY2FsZSpmKSk7CiAgICB3b3JsZC5zdHlsZS50cmFuc2Zvcm09J3RyYW5zbGF0ZSgnK2VyZFgrJ3B4LCcrZXJkWSsncHgpIHNjYWxlKCcrZXJkU2NhbGUrJyknOwogIH07Cn0KCmZ1bmN0aW9uIGFwcGx5RVJEVHJhbnNmb3JtKCl7CiAgZWwoJ2VyZC13b3JsZCcpLnN0eWxlLnRyYW5zZm9ybT0ndHJhbnNsYXRlKCcrZXJkWCsncHgsJytlcmRZKydweCkgc2NhbGUoJytlcmRTY2FsZSsnKSc7Cn0KCmZ1bmN0aW9uIG9uRXJkRmllbGQobmlkLCBmbmFtZSl7CiAgdmFyIG0gPSBEICYmIEQubWV0YWRhdGEgPyBELm1ldGFkYXRhLmZpbmQoZnVuY3Rpb24oeCl7IHJldHVybiB4LmZpZWxkX25hbWU9PT1mbmFtZTsgfSkgOiBudWxsOwogIGlmKG0peyBzaG93VGFiKCdtZXRhZGF0YScpOyBzZWxlY3RGaWVsZChtKTsgfQogIGVsc2UgdG9hc3QoJ0ZpZWxkOiAnK2ZuYW1lLCAnaW5mbycpOwp9CmZ1bmN0aW9uIG9uRXJkTm9kZShuaWQpewogIHZhciBuID0gRCAmJiBELm5vZGVzID8gRC5ub2Rlcy5maW5kKGZ1bmN0aW9uKHgpeyByZXR1cm4geC5pZD09PW5pZDsgfSkgOiBudWxsOwogIGlmKG4pIG9wZW5Ob2RlTW9kYWwobik7Cn0KCi8vIOKUgOKUgCBGT1JDRSBHUkFQSCDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZnVuY3Rpb24gZHJhd0ZvcmNlKG5vZGVzLCBlZGdlcyl7CiAgdmFyIGNvbnQgPSBlbCgnZm9yY2Utc3ZnJyk7CiAgdmFyIFcgPSBjb250LnBhcmVudEVsZW1lbnQuY2xpZW50V2lkdGh8fDg0MCwgSD00OTA7CiAgZDMuc2VsZWN0KCcjZm9yY2Utc3ZnJykuc2VsZWN0QWxsKCcqJykucmVtb3ZlKCk7CiAgZlN2ZyA9IGQzLnNlbGVjdCgnI2ZvcmNlLXN2ZycpLmF0dHIoJ3ZpZXdCb3gnLCcwIDAgJytXKycgJytIKS5hdHRyKCd3aWR0aCcsVykuYXR0cignaGVpZ2h0JyxIKTsKICB2YXIgZGVmcyA9IGZTdmcuYXBwZW5kKCdkZWZzJyk7CiAgdmFyIEFDPXtnOicjN0ZEOTYyJyxiOicjM0I5RUZGJyxhOicjRjVBNjIzJyxjOicjMDBDOUIxJyxyOicjRkY0RjZEJyxkZjonIzFFMzQ1MCd9OwogIE9iamVjdC5lbnRyaWVzKEFDKS5mb3JFYWNoKGZ1bmN0aW9uKGt2KXsKICAgIGRlZnMuYXBwZW5kKCdtYXJrZXInKS5hdHRyKCdpZCcsJ2ZhJytrdlswXSkuYXR0cigndmlld0JveCcsJzAgLTQgOCA4JykuYXR0cigncmVmWCcsJzI4JykKICAgICAgLmF0dHIoJ3JlZlknLCcwJykuYXR0cignbWFya2VyV2lkdGgnLCc1JykuYXR0cignbWFya2VySGVpZ2h0JywnNScpLmF0dHIoJ29yaWVudCcsJ2F1dG8nKQogICAgICAuYXBwZW5kKCdwYXRoJykuYXR0cignZCcsJ00wLC00TDgsMEwwLDQnKS5hdHRyKCdmaWxsJyxrdlsxXSkuYXR0cignb3BhY2l0eScsJy44Jyk7CiAgfSk7CiAgZGVmcy5hcHBlbmQoJ2ZpbHRlcicpLmF0dHIoJ2lkJywnZ2xvdycpLmh0bWwoJzxmZUdhdXNzaWFuQmx1ciBzdGREZXZpYXRpb249IjMiIHJlc3VsdD0iYiIvPjxmZU1lcmdlPjxmZU1lcmdlTm9kZSBpbj0iYiIvPjxmZU1lcmdlTm9kZSBpbj0iU291cmNlR3JhcGhpYyIvPjwvZmVNZXJnZT4nKTsKICBkZWZzLmFwcGVuZCgnZmlsdGVyJykuYXR0cignaWQnLCdnbG93SCcpLmh0bWwoJzxmZUdhdXNzaWFuQmx1ciBzdGREZXZpYXRpb249IjciIHJlc3VsdD0iYiIvPjxmZU1lcmdlPjxmZU1lcmdlTm9kZSBpbj0iYiIvPjxmZU1lcmdlTm9kZSBpbj0iU291cmNlR3JhcGhpYyIvPjwvZmVNZXJnZT4nKTsKICBmRyA9IGZTdmcuYXBwZW5kKCdnJyk7CiAgZlpvb20gPSBkMy56b29tKCkuc2NhbGVFeHRlbnQoWy4xLDVdKS5vbignem9vbScsIGZ1bmN0aW9uKGV2KXsgZkcuYXR0cigndHJhbnNmb3JtJyxldi50cmFuc2Zvcm0pOyB9KTsKICBmU3ZnLmNhbGwoZlpvb20pOwoKICB2YXIgTlM9ewogICAgc291cmNlX3RhYmxlOntyOjMyLGZpbGw6JyMwODE4MEEnLHN0cm9rZTonIzdGRDk2Micsc3c6MixhazonZyd9LAogICAgdGFyZ2V0X3RhYmxlOntyOjMyLGZpbGw6JyMwODBEMTgnLHN0cm9rZTonIzNCOUVGRicsc3c6MixhazonYid9LAogICAgdHJhbnNmb3JtYXRpb246e3I6MjYsZmlsbDonIzE4MTMwQScsc3Ryb2tlOicjRjVBNjIzJyxzdzoyLGFrOidhJ30sCiAgICBjb2RlX3JlcG86e3I6MjIsZmlsbDonIzA0MTQxMicsc3Ryb2tlOicjMDBDOUIxJyxzdzoxLjUsYWs6J2MnfSwKICAgIHNyY19maWVsZDp7cjoxNyxmaWxsOicjMDgxODBBJyxzdHJva2U6JyM3RkQ5NjInLHN3OjEuNSxhazonZyd9LAogICAgdGd0X2ZpZWxkOntyOjE3LGZpbGw6JyMwODBEMTgnLHN0cm9rZTonIzNCOUVGRicsc3c6MS41LGFrOidiJ30sCiAgICBzeXNfc3JjOntyOjM2LGZpbGw6JyMwODE4MEEnLHN0cm9rZTonIzdGRDk2Micsc3c6Mi41LGFrOidnJ30sCiAgICBzeXNfdGd0OntyOjM2LGZpbGw6JyMwODBEMTgnLHN0cm9rZTonIzNCOUVGRicsc3c6Mi41LGFrOidiJ30sCiAgICBzeXNfdHJmOntyOjI4LGZpbGw6JyMxODEzMEEnLHN0cm9rZTonI0Y1QTYyMycsc3c6MixhazonYSd9LAogICAgc3lzX2NvZGU6e3I6MjIsZmlsbDonIzA0MTQxMicsc3Ryb2tlOicjMDBDOUIxJyxzdzoxLjUsYWs6J2MnfSwKICB9OwogIGZ1bmN0aW9uIGdzKHQpeyByZXR1cm4gTlNbdF18fHtyOjIyLGZpbGw6JyMwODBFMTgnLHN0cm9rZTonIzFFMzQ1MCcsc3c6MS41LGFrOidkZid9OyB9CgogIHZhciBsaW5rID0gZkcuYXBwZW5kKCdnJykuc2VsZWN0QWxsKCdwYXRoJykuZGF0YShlZGdlcykuZW50ZXIoKS5hcHBlbmQoJ3BhdGgnKQogICAgLmF0dHIoJ2ZpbGwnLCdub25lJykKICAgIC5hdHRyKCdzdHJva2UnLGZ1bmN0aW9uKGQpeyB2YXIgcz1ub2Rlcy5maW5kKGZ1bmN0aW9uKG4pe3JldHVybiBuLmlkPT09ZC5zb3VyY2U7fSk7IHJldHVybiBBQ1tncyhzJiZzLnR5cGUpLmFrXXx8QUMuZGY7IH0pCiAgICAuYXR0cignc3Ryb2tlLXdpZHRoJywxLjYpLmF0dHIoJ3N0cm9rZS1vcGFjaXR5JywuNDUpCiAgICAuYXR0cignbWFya2VyLWVuZCcsZnVuY3Rpb24oZCl7IHZhciBzPW5vZGVzLmZpbmQoZnVuY3Rpb24obil7cmV0dXJuIG4uaWQ9PT1kLnNvdXJjZTt9KTsgcmV0dXJuICd1cmwoI2ZhJysoZ3MocyYmcy50eXBlKS5ha3x8J2RmJykrJyknOyB9KTsKCiAgdmFyIG5vZGUgPSBmRy5hcHBlbmQoJ2cnKS5zZWxlY3RBbGwoJ2cnKS5kYXRhKG5vZGVzKS5lbnRlcigpLmFwcGVuZCgnZycpCiAgICAuc3R5bGUoJ2N1cnNvcicsJ3BvaW50ZXInKQogICAgLm9uKCdjbGljaycsIGZ1bmN0aW9uKGV2LGQpeyBvbkZvcmNlTm9kZUNsaWNrKGQpOyB9KQogICAgLmNhbGwoZDMuZHJhZygpCiAgICAgIC5vbignc3RhcnQnLGZ1bmN0aW9uKGV2LGQpeyBpZighZXYuYWN0aXZlKSBzaW0uYWxwaGFUYXJnZXQoLjMpLnJlc3RhcnQoKTsgZC5meD1kLng7IGQuZnk9ZC55OyB9KQogICAgICAub24oJ2RyYWcnLCBmdW5jdGlvbihldixkKXsgZC5meD1ldi54OyBkLmZ5PWV2Lnk7IH0pCiAgICAgIC5vbignZW5kJywgIGZ1bmN0aW9uKGV2LGQpeyBpZighZXYuYWN0aXZlKSBzaW0uYWxwaGFUYXJnZXQoMCk7IGQuZng9bnVsbDsgZC5meT1udWxsOyB9KSk7CgogIG5vZGUuYXBwZW5kKCdjaXJjbGUnKQogICAgLmF0dHIoJ3InLGZ1bmN0aW9uKGQpe3JldHVybiBncyhkLnR5cGUpLnI7fSkKICAgIC5hdHRyKCdmaWxsJyxmdW5jdGlvbihkKXtyZXR1cm4gZ3MoZC50eXBlKS5maWxsO30pCiAgICAuYXR0cignc3Ryb2tlJyxmdW5jdGlvbihkKXtyZXR1cm4gZ3MoZC50eXBlKS5zdHJva2U7fSkKICAgIC5hdHRyKCdzdHJva2Utd2lkdGgnLGZ1bmN0aW9uKGQpe3JldHVybiBncyhkLnR5cGUpLnN3O30pCiAgICAuYXR0cignZmlsdGVyJywndXJsKCNnbG93KScpCiAgICAub24oJ21vdXNlb3ZlcicsZnVuY3Rpb24oKXsgZDMuc2VsZWN0KHRoaXMpLmF0dHIoJ2ZpbHRlcicsJ3VybCgjZ2xvd0gpJyk7IH0pCiAgICAub24oJ21vdXNlb3V0JywgZnVuY3Rpb24oKXsgZDMuc2VsZWN0KHRoaXMpLmF0dHIoJ2ZpbHRlcicsJ3VybCgjZ2xvdyknKTsgfSk7CgogIG5vZGUuZWFjaChmdW5jdGlvbihkKXsKICAgIHZhciByID0gZ3MoZC50eXBlKS5yOwogICAgKGQubGFiZWx8fCcnKS5zcGxpdCgnXG4nKS5mb3JFYWNoKGZ1bmN0aW9uKGxuLGkpewogICAgICBkMy5zZWxlY3QodGhpcykuYXBwZW5kKCd0ZXh0JykuYXR0cigndGV4dC1hbmNob3InLCdtaWRkbGUnKS5hdHRyKCdkb21pbmFudC1iYXNlbGluZScsJ21pZGRsZScpCiAgICAgICAgLmF0dHIoJ3knLHIrMTArKGkqMTEpKS5hdHRyKCdmb250LXNpemUnLCc5cHgnKS5hdHRyKCdmaWxsJywnIzRFNkU4QScpCiAgICAgICAgLmF0dHIoJ2ZvbnQtZmFtaWx5JywnSmV0QnJhaW5zIE1vbm8sbW9ub3NwYWNlJykuYXR0cigncG9pbnRlci1ldmVudHMnLCdub25lJykKICAgICAgICAudGV4dChsbi5sZW5ndGg+MTQ/bG4uc2xpY2UoMCwxMikrJ1x1MjAyNic6bG4pOwogICAgfS5iaW5kKHRoaXMpKTsKICB9KTsKCiAgLy8gVG9vbHRpcAogIG5vZGUub24oJ21vdXNlb3Zlci50dCcsZnVuY3Rpb24oZXYsZCl7CiAgICBkb2N1bWVudC5nZXRFbGVtZW50QnlJZCgnZy10dCcpICYmIGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdnLXR0JykucmVtb3ZlKCk7CiAgICB2YXIgdHQ9ZG9jdW1lbnQuY3JlYXRlRWxlbWVudCgnZGl2Jyk7IHR0LmlkPSdnLXR0JzsKICAgIHZhciBjPWdzKGQudHlwZSkuc3Ryb2tlOwogICAgdHQuc3R5bGUuY3NzVGV4dD0ncG9zaXRpb246Zml4ZWQ7YmFja2dyb3VuZDojMDYwOTBmO2JvcmRlcjoxcHggc29saWQgJytjKyc0MDtib3JkZXItbGVmdDoycHggc29saWQgJytjKyc7Ym9yZGVyLXJhZGl1czo2cHg7cGFkZGluZzo3cHggMTFweDtmb250LXNpemU6MTBweDtjb2xvcjojRjJGOEZGO3BvaW50ZXItZXZlbnRzOm5vbmU7ei1pbmRleDozMDA7bWF4LXdpZHRoOjIwMHB4O2xpbmUtaGVpZ2h0OjEuNTtib3gtc2hhZG93OjAgNnB4IDI0cHggcmdiYSgwLDAsMCwuOCknOwogICAgdHQuaW5uZXJIVE1MPSc8YiBzdHlsZT0iY29sb3I6JytjKyciPicrKGQubGFiZWx8fGQuaWQpLnJlcGxhY2UoJ1xuJywnICcpKyc8L2I+PGJyPjxzcGFuIHN0eWxlPSJmb250LXNpemU6OHB4O2NvbG9yOiM0RTZFOEEiPicrKGQudHlwZXx8JycpLnJlcGxhY2UoL18vZywnICcpKyc8L3NwYW4+JysoZC5zeXN0ZW0/Jzxicj48c3BhbiBzdHlsZT0iZm9udC1zaXplOjhweDtjb2xvcjojNEU2RThBIj4nK2Quc3lzdGVtKyc8L3NwYW4+JzonJyk7CiAgICBkb2N1bWVudC5ib2R5LmFwcGVuZENoaWxkKHR0KTsKICAgIG1vdmVUVChldik7CiAgfSkub24oJ21vdXNlbW92ZS50dCcsbW92ZVRUKS5vbignbW91c2VvdXQudHQnLGZ1bmN0aW9uKCl7IHZhciB0PWRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdnLXR0Jyk7IGlmKHQpIHQucmVtb3ZlKCk7IH0pOwoKICBmdW5jdGlvbiBtb3ZlVFQoZXYpewogICAgdmFyIHQ9ZG9jdW1lbnQuZ2V0RWxlbWVudEJ5SWQoJ2ctdHQnKTsKICAgIGlmKHQpeyB0LnN0eWxlLmxlZnQ9KGV2LmNsaWVudFgrMTIpKydweCc7IHQuc3R5bGUudG9wPShldi5jbGllbnRZLTYpKydweCc7IH0KICB9CgogIHZhciBzaW0gPSBkMy5mb3JjZVNpbXVsYXRpb24obm9kZXMpCiAgICAuZm9yY2UoJ2xpbmsnLCBkMy5mb3JjZUxpbmsoZWRnZXMpLmlkKGZ1bmN0aW9uKGQpe3JldHVybiBkLmlkO30pLmRpc3RhbmNlKGN1ckxldmVsPT09J3N5c3RlbSc/MTcwOjEyMCkuc3RyZW5ndGgoLjYpKQogICAgLmZvcmNlKCdjaGFyZ2UnLCBkMy5mb3JjZU1hbnlCb2R5KCkuc3RyZW5ndGgoLTQ2MCkpCiAgICAuZm9yY2UoJ2NlbnRlcicsIGQzLmZvcmNlQ2VudGVyKFcvMiwgSC8yKSkKICAgIC5mb3JjZSgnY29sbGlkZScsIGQzLmZvcmNlQ29sbGlkZShmdW5jdGlvbihkKXtyZXR1cm4gZ3MoZC50eXBlKS5yKzEzO30pKQogICAgLm9uKCd0aWNrJywgZnVuY3Rpb24oKXsKICAgICAgbGluay5hdHRyKCdkJyxmdW5jdGlvbihkKXsKICAgICAgICB2YXIgZHg9ZC50YXJnZXQueC1kLnNvdXJjZS54LCBkeT1kLnRhcmdldC55LWQuc291cmNlLnk7CiAgICAgICAgdmFyIGRyPU1hdGguc3FydChkeCpkeCtkeSpkeSkqMS4yOwogICAgICAgIHJldHVybiAnTScrZC5zb3VyY2UueCsnLCcrZC5zb3VyY2UueSsnIEEnK2RyKycsJytkcisnIDAgMCwxICcrZC50YXJnZXQueCsnLCcrZC50YXJnZXQueTsKICAgICAgfSk7CiAgICAgIG5vZGUuYXR0cigndHJhbnNmb3JtJyxmdW5jdGlvbihkKXtyZXR1cm4gJ3RyYW5zbGF0ZSgnK2QueCsnLCcrZC55KycpJzt9KTsKICAgIH0pOwoKICBpZihjdXJMZXZlbD09PSd0YWJsZSd8fGN1ckxldmVsPT09J3N5c3RlbScpewogICAgbm9kZXMuZm9yRWFjaChmdW5jdGlvbihuKXsKICAgICAgaWYobi50eXBlPT09J3NvdXJjZV90YWJsZSd8fG4udHlwZT09PSdzeXNfc3JjJyl7bi5meD1XKi4xNjtuLmZ5PUgqLjQ0O30KICAgICAgZWxzZSBpZihuLnR5cGU9PT0ndGFyZ2V0X3RhYmxlJ3x8bi50eXBlPT09J3N5c190Z3QnKXtuLmZ4PVcqLjc4O24uZnk9SCouNDQ7fQogICAgICBlbHNlIGlmKG4udHlwZT09PSd0cmFuc2Zvcm1hdGlvbid8fG4udHlwZT09PSdzeXNfdHJmJyl7bi5meD1XKi40NjtuLmZ5PUgqLjM3O30KICAgICAgZWxzZSBpZihuLnR5cGU9PT0nY29kZV9yZXBvJ3x8bi50eXBlPT09J3N5c19jb2RlJyl7bi5meD1XKi40NjtuLmZ5PUgqLjcwO30KICAgIH0pOwogICAgc2V0VGltZW91dChmdW5jdGlvbigpewogICAgICBub2Rlcy5mb3JFYWNoKGZ1bmN0aW9uKG4peyBpZihuLnR5cGUhPT0nY29kZV9yZXBvJyYmbi50eXBlIT09J3N5c19jb2RlJyl7bi5meD1udWxsO24uZnk9bnVsbDt9IH0pOwogICAgICBzaW0uYWxwaGEoLjI1KS5yZXN0YXJ0KCk7CiAgICB9LCAxODAwKTsKICB9Cn0KCmZ1bmN0aW9uIG9uRm9yY2VOb2RlQ2xpY2soZCl7CiAgdmFyIGd0dCA9IGRvY3VtZW50LmdldEVsZW1lbnRCeUlkKCdnLXR0Jyk7IGlmKGd0dCkgZ3R0LnJlbW92ZSgpOwogIGlmKGN1ckxldmVsPT09J3N5c3RlbScpeyBzZXRMZXZlbCgndGFibGUnKTsgdG9hc3QoJ0RyaWxsZWQgaW50byBUYWJsZSBMZXZlbCcsJ2luZm8nKTsgfQogIGVsc2UgaWYoY3VyTGV2ZWw9PT0ndGFibGUnKXsgc2V0TGV2ZWwoJ2ZpZWxkJyk7IHRvYXN0KCdEcmlsbGVkIGludG8gRmllbGQgTGV2ZWwnLCdpbmZvJyk7IH0KICBlbHNlewogICAgdmFyIG0gPSBEICYmIEQubWV0YWRhdGEgPyBELm1ldGFkYXRhLmZpbmQoZnVuY3Rpb24oeCl7IHJldHVybiB4LmZpZWxkX25hbWU9PT0oZC5maWVsZE5hbWV8fGQuZmllbGQpOyB9KSA6IG51bGw7CiAgICBpZihtKXsgc2hvd1RhYignbWV0YWRhdGEnKTsgc2VsZWN0RmllbGQobSk7IH0gZWxzZSBvcGVuTm9kZU1vZGFsKGQpOwogIH0KfQoKLy8g4pSA4pSAIEdSQVBIIENPTlRST0xTIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgApmdW5jdGlvbiBnWm9vbUluKCl7CiAgaWYoY3VyVmlldz09PSdlcmQnKXsgZXJkU2NhbGU9TWF0aC5taW4oNCxlcmRTY2FsZSoxLjI4KTsgYXBwbHlFUkRUcmFuc2Zvcm0oKTsgfQogIGVsc2UgaWYoZlN2ZyYmZlpvb20pIGZTdmcudHJhbnNpdGlvbigpLmNhbGwoZlpvb20uc2NhbGVCeSwxLjQpOwp9CmZ1bmN0aW9uIGdab29tT3V0KCl7CiAgaWYoY3VyVmlldz09PSdlcmQnKXsgZXJkU2NhbGU9TWF0aC5tYXgoLjE1LGVyZFNjYWxlKi43OCk7IGFwcGx5RVJEVHJhbnNmb3JtKCk7IH0KICBlbHNlIGlmKGZTdmcmJmZab29tKSBmU3ZnLnRyYW5zaXRpb24oKS5jYWxsKGZab29tLnNjYWxlQnksLjcyKTsKfQpmdW5jdGlvbiBnRml0KCl7CiAgaWYoY3VyVmlldz09PSdlcmQnKXsgZXJkU2NhbGU9MTsgZXJkWD0yMDsgZXJkWT0yMDsgYXBwbHlFUkRUcmFuc2Zvcm0oKTsgfQogIGVsc2UgaWYoZlN2ZyYmZlpvb20pIGZTdmcudHJhbnNpdGlvbigpLmNhbGwoZlpvb20udHJhbnNmb3JtLCBkMy56b29tSWRlbnRpdHkpOwp9CmZ1bmN0aW9uIGdGdWxsc2NyZWVuKCl7CiAgdmFyIGMgPSBkb2N1bWVudC5xdWVyeVNlbGVjdG9yKCcuZ2JveCcpOwogIGlmKCFkb2N1bWVudC5mdWxsc2NyZWVuRWxlbWVudCkgYy5yZXF1ZXN0RnVsbHNjcmVlbigpLmNhdGNoKGZ1bmN0aW9uKCl7fSk7CiAgZWxzZSBkb2N1bWVudC5leGl0RnVsbHNjcmVlbigpOwp9CgpmdW5jdGlvbiBkcmF3TGVnZW5kKCl7CiAgdmFyIGl0ZW1zID0gWwogICAge2M6JyM3RkQ5NjInLGw6J1NvdXJjZSd9LHtjOicjRjVBNjIzJyxsOidUcmFuc2Zvcm0vRVRMJ30sCiAgICB7YzonIzNCOUVGRicsbDonVGFyZ2V0J30se2M6JyMwMEM5QjEnLGw6J0NvZGUgUmVwbyd9CiAgXTsKICBlbCgnZy1sZWdlbmQnKS5pbm5lckhUTUwgPSBpdGVtcy5tYXAoZnVuY3Rpb24oaSl7CiAgICByZXR1cm4gJzxkaXYgY2xhc3M9ImxlZy1pIj48ZGl2IGNsYXNzPSJsZWctZCIgc3R5bGU9ImJhY2tncm91bmQ6JytpLmMrJyI+PC9kaXY+JytpLmwrJzwvZGl2Pic7CiAgfSkuam9pbignJyk7Cn0KCi8vIOKUgOKUgCBNRVRBREFUQSDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZnVuY3Rpb24gcmVuZGVyTWV0YWRhdGEobWV0YSl7CiAgZWwoJ21ldGEtY291bnQnKS50ZXh0Q29udGVudCA9IG1ldGEubGVuZ3RoKycgZmllbGQnKyhtZXRhLmxlbmd0aCE9PTE/J3MnOicnKSsnIGRpc2NvdmVyZWQgXHUwMEI3IENvbmZpZGVudGlhbGl0eSBcdTAwQjcgU09SIFx1MDBCNyBFcXVpdmFsZW50IEZpZWxkcyBcdTAwQjcgTGluZWFnZSBQYXRoJzsKICBpZighbWV0YS5sZW5ndGgpewogICAgZWwoJ21ldGEtdGJsJykuaW5uZXJIVE1MID0gJzxkaXYgc3R5bGU9InBhZGRpbmc6NDBweDt0ZXh0LWFsaWduOmNlbnRlcjtjb2xvcjp2YXIoLS10MikiPjxkaXYgc3R5bGU9ImZvbnQtc2l6ZToyOHB4O21hcmdpbi1ib3R0b206OHB4Ij4mI3gxRjRDQjs8L2Rpdj48ZGl2Pk5vIG1ldGFkYXRhIGZvdW5kPC9kaXY+PC9kaXY+JzsKICAgIHJldHVybjsKICB9CiAgdmFyIGNvbHMgPSAnMTQwcHggODBweCA1NnB4IDY0cHggMWZyIDgwcHgnOwogIHZhciByb3dzID0gbWV0YS5tYXAoZnVuY3Rpb24obSxpKXsKICAgIHJldHVybiAnPGRpdiBjbGFzcz0ibWV0YS1yb3ciIGlkPSJtci0nK2krJyIgc3R5bGU9ImdyaWQtdGVtcGxhdGUtY29sdW1uczonK2NvbHMrJyIgb25jbGljaz0ic2VsZWN0RmllbGQoRC5tZXRhZGF0YVsnK2krJ10pIj4nKwogICAgICAnPGRpdiBzdHlsZT0iZm9udC1mYW1pbHk6XCdKZXRCcmFpbnMgTW9ub1wnLG1vbm9zcGFjZTtmb250LXNpemU6MTBweDtmb250LXdlaWdodDo2MDA7YmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoOTBkZWcsdmFyKC0tZ29sZDIpLHZhcigtLXRlYWwyKSk7LXdlYmtpdC1iYWNrZ3JvdW5kLWNsaXA6dGV4dDstd2Via2l0LXRleHQtZmlsbC1jb2xvcjp0cmFuc3BhcmVudCI+JysobS5pc19waz8nJiN4MUY1MTE7ICc6JycpK2VzYyhtLmZpZWxkX25hbWUpKyc8L2Rpdj4nKwogICAgICAnPGRpdj48c3BhbiBjbGFzcz0iZHR5cGUiIHN0eWxlPSJmb250LXNpemU6OHB4Ij4nK2VzYyhtLmRhdGFfdHlwZXx8J1ZBUkNIQVInKSsnPC9zcGFuPjwvZGl2PicrCiAgICAgICc8ZGl2PjxzcGFuIGNsYXNzPSJyYiByYi0nKyhtLnJpc2tfbGV2ZWx8fCdMJylbMF0rJyIgc3R5bGU9ImZvbnQtc2l6ZTo4cHgiPicrZXNjKG0ucmlza19sZXZlbHx8J0xvdycpKyc8L3NwYW4+PC9kaXY+JysKICAgICAgJzxkaXY+Jytjb25mQmFkZ2UobS5jb25maWRlbnRpYWxpdHl8fCdJbnRlcm5hbCcpKyc8L2Rpdj4nKwogICAgICAnPGRpdiBzdHlsZT0iZm9udC1zaXplOjEwcHg7Y29sb3I6dmFyKC0tdDIpO292ZXJmbG93OmhpZGRlbjt0ZXh0LW92ZXJmbG93OmVsbGlwc2lzO3doaXRlLXNwYWNlOm5vd3JhcCI+Jytlc2MobS5kZXNjcmlwdGlvbnx8J1x1MjAxNCcpKyc8L2Rpdj4nKwogICAgICAnPGRpdj48c3BhbiBjbGFzcz0iaW5nLXBpbGwgJysoKG0uaW5nZXN0aW9uX3BhdHRlcm58fCcnKT09PSdSZWFsLVRpbWUgU3RyZWFtaW5nJz8naW5nLXJ0JzonaW5nLWJhdGNoJykrJyIgc3R5bGU9ImZvbnQtc2l6ZTo4cHgiPicrKCgobS5pbmdlc3Rpb25fcGF0dGVybnx8JycpPT09J1JlYWwtVGltZSBTdHJlYW1pbmcnKT8nUlQnOidCdGNoJykrJzwvc3Bhbj48L2Rpdj4nKwogICAgJzwvZGl2Pic7CiAgfSkuam9pbignJyk7CiAgZWwoJ21ldGEtdGJsJykuaW5uZXJIVE1MID0KICAgICc8ZGl2IGNsYXNzPSJtZXRhLWhkciIgc3R5bGU9ImdyaWQtdGVtcGxhdGUtY29sdW1uczonK2NvbHMrJyI+JysKICAgICAgJzxkaXY+RmllbGQ8L2Rpdj48ZGl2PlR5cGU8L2Rpdj48ZGl2PlJpc2s8L2Rpdj48ZGl2PkNvbmZpZC48L2Rpdj48ZGl2PkRlc2NyaXB0aW9uPC9kaXY+PGRpdj5Jbmdlc3Rpb248L2Rpdj4nKwogICAgJzwvZGl2Picrcm93czsKfQoKZnVuY3Rpb24gc2VsZWN0RmllbGQobSl7CiAgc2VsRmllbGQgPSBtOwogIGRvY3VtZW50LnF1ZXJ5U2VsZWN0b3JBbGwoJy5tZXRhLXJvdycpLmZvckVhY2goZnVuY3Rpb24ocil7IHIuY2xhc3NMaXN0LnJlbW92ZSgnc2VsJyk7IH0pOwogIHZhciBpID0gRC5tZXRhZGF0YS5pbmRleE9mKG0pOwogIHZhciByb3cgPSBlbCgnbXItJytpKTsgaWYocm93KSByb3cuY2xhc3NMaXN0LmFkZCgnc2VsJyk7CgogIHZhciBzb3JIdG1sID0gJyc7CiAgaWYobS5zb3JzICYmIG0uc29ycy5sZW5ndGgpewogICAgc29ySHRtbCA9ICc8ZGl2IGNsYXNzPSJmZC1zZWMiPicrCiAgICAgICc8ZGl2IGNsYXNzPSJmZC1zZWMtdCI+JiN4MUY0Q0I7IFN5c3RlbXMgb2YgUmVjb3JkIChTT1IpPC9kaXY+JysKICAgICAgJzxkaXYgc3R5bGU9ImZvbnQtc2l6ZTo5cHg7Y29sb3I6dmFyKC0tdDIpO21hcmdpbi1ib3R0b206NXB4Ij5DbGljayB0byBzZXQgYXV0aG9yaXRhdGl2ZSBTT1IuPC9kaXY+JysKICAgICAgbS5zb3JzLm1hcChmdW5jdGlvbihzKXsKICAgICAgICByZXR1cm4gJzxkaXYgY2xhc3M9InNvci1jYXJkICcrKHMuaXNfYXV0aG9yaXRhdGl2ZT8nYXV0aCc6JycpKyciIG9uY2xpY2s9ImRvU2V0U09SKFwnJytlc2MobS5maWVsZF9uYW1lKSsnXCcsXCcnK2VzYyhzLnNvcikrJ1wnKSI+JysKICAgICAgICAgICc8ZGl2PjxkaXYgY2xhc3M9InNvci1uYW1lIj4nK2VzYyhzLnNvcikrJzwvZGl2PjxkaXYgY2xhc3M9InNvci1zeXMiPicrZXNjKHMudGFibGV8fHMuc3lzdGVtfHwnJykrJzwvZGl2PjwvZGl2PicrCiAgICAgICAgICAocy5pc19hdXRob3JpdGF0aXZlPyc8c3BhbiBjbGFzcz0ic29yLWF1dGgiPlx1MjYwNSBBVVRIPC9zcGFuPic6JycpKwogICAgICAgICc8L2Rpdj4nOwogICAgICB9KS5qb2luKCcnKSsKICAgICc8L2Rpdj4nOwogIH0KCiAgdmFyIGVxdWl2SHRtbCA9ICcnOwogIGlmKG0uZXF1aXZhbGVudF9maWVsZHMgJiYgbS5lcXVpdmFsZW50X2ZpZWxkcy5sZW5ndGgpewogICAgZXF1aXZIdG1sID0gJzxkaXYgY2xhc3M9ImZkLXNlYyI+JysKICAgICAgJzxkaXYgY2xhc3M9ImZkLXNlYy10Ij4mI3gxRjUxNzsgRXF1aXZhbGVudCBGaWVsZHMgQWNyb3NzIFNPUnM8L2Rpdj4nKwogICAgICAnPGRpdiBzdHlsZT0iZm9udC1zaXplOjlweDtjb2xvcjp2YXIoLS10Mik7bWFyZ2luLWJvdHRvbTo1cHgiPlNhbWUgZmllbGQsIGRpZmZlcmVudCBuYW1lcyBhY3Jvc3Mgc3lzdGVtcy48L2Rpdj4nKwogICAgICAnPGRpdj4nK20uZXF1aXZhbGVudF9maWVsZHMubWFwKGZ1bmN0aW9uKGUpewogICAgICAgIHJldHVybiAnPHNwYW4gY2xhc3M9ImVxLWl0ZW0gJysoZS5hbGlhc190eXBlPT09J2tub3duX2VxdWl2YWxlbnQnPydlcS1leGFjdCc6JycpKyciIHRpdGxlPSInKyhlLmFsaWFzX3R5cGU9PT0na25vd25fZXF1aXZhbGVudCc/J0tub3duIGFsaWFzJzonRnV6enkgbWF0Y2ggKCcrZS5zaW1pbGFyaXR5KycpJykrJyIgb25jbGljaz0ic2V0UShcJycrZXNjKGUuZmllbGQpKydcJyk7ZG9TZWFyY2goKSI+Jytlc2MoZS5maWVsZCkrJzwvc3Bhbj4nOwogICAgICB9KS5qb2luKCcnKSsnPC9kaXY+JysKICAgICc8L2Rpdj4nOwogIH0KCiAgdmFyIGhvcHNIdG1sID0gJyc7CiAgaWYobS5saW5lYWdlX2hvcHMgJiYgbS5saW5lYWdlX2hvcHMubGVuZ3RoKXsKICAgIGhvcHNIdG1sID0gJzxkaXYgY2xhc3M9ImZkLXNlYyI+JysKICAgICAgJzxkaXYgY2xhc3M9ImZkLXNlYy10Ij4mI3gxRjUwMzsgRnVsbCBMaW5lYWdlIFBhdGggKCcrbS5saW5lYWdlX2hvcHMubGVuZ3RoKycgaG9wcyk8L2Rpdj4nKwogICAgICBtLmxpbmVhZ2VfaG9wcy5tYXAoZnVuY3Rpb24oaCxpKXsKICAgICAgICByZXR1cm4gJzxkaXYgY2xhc3M9ImhvcC1yb3ciPicrCiAgICAgICAgICAnPGRpdiBjbGFzcz0iaG9wLW4iPicraC5ob3ArJzwvZGl2PicrCiAgICAgICAgICAnPGRpdiBjbGFzcz0iaG9wLWxibCI+Jytlc2MoaC5zeXN0ZW18fGgudGFibGV8fCcnKSsnPC9kaXY+JysKICAgICAgICAgICc8ZGl2IGNsYXNzPSJob3AtdHlwZSI+Jytlc2MoaC50eXBlPyhoLnR5cGUucmVwbGFjZSgvXy9nLCcgJykpOicnKSsnIDwvZGl2PicrCiAgICAgICAgICAoaTxtLmxpbmVhZ2VfaG9wcy5sZW5ndGgtMT8nPHNwYW4gc3R5bGU9ImNvbG9yOnZhcigtLXQzKTtmb250LXNpemU6MTBweCI+JnJhcnI7PC9zcGFuPic6JycpKwogICAgICAgICc8L2Rpdj4nOwogICAgICB9KS5qb2luKCcnKSsKICAgICc8L2Rpdj4nOwogIH0KCiAgZWwoJ2ZkZXRhaWwnKS5pbm5lckhUTUwgPQogICAgJzxkaXYgY2xhc3M9ImZkLWhlYWQiPicrCiAgICAgICc8ZGl2IGNsYXNzPSJmZC1uYW1lIj4nK2VzYyhtLmZpZWxkX25hbWUpKyc8L2Rpdj4nKwogICAgICAnPGRpdiBjbGFzcz0iZmQtdGJsIj4nK2VzYyhtLnRhYmxlX25hbWUpKyc8L2Rpdj4nKwogICAgICAnPGRpdiBzdHlsZT0ibWFyZ2luLXRvcDo2cHg7ZGlzcGxheTpmbGV4O2dhcDo0cHg7ZmxleC13cmFwOndyYXA7YWxpZ24taXRlbXM6Y2VudGVyIj4nKwogICAgICAgICc8c3BhbiBjbGFzcz0iZHR5cGUiPicrZXNjKG0uZGF0YV90eXBlfHwnVkFSQ0hBUicpKyc8L3NwYW4+JysKICAgICAgICAnPHNwYW4gY2xhc3M9InJiIHJiLScrKG0ucmlza19sZXZlbHx8J0wnKVswXSsnIj4nK2VzYyhtLnJpc2tfbGV2ZWx8fCdMb3cnKSsnPC9zcGFuPicrCiAgICAgICAgKG0uaXNfcGs/JzxzcGFuIGNsYXNzPSJyYiIgc3R5bGU9ImJhY2tncm91bmQ6cmdiYSgyNDUsMTY2LDM1LC4xKTtjb2xvcjp2YXIoLS1nb2xkMik7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDI0NSwxNjYsMzUsLjI1KSI+JiN4MUY1MTE7IFBLPC9zcGFuPic6JycpKwogICAgICAgICghbS5pc19udWxsYWJsZT8nPHNwYW4gY2xhc3M9InJiIiBzdHlsZT0iYmFja2dyb3VuZDpyZ2JhKDI1NSw3OSwxMDksLjA4KTtjb2xvcjp2YXIoLS1yb3NlMik7Ym9yZGVyOjFweCBzb2xpZCByZ2JhKDI1NSw3OSwxMDksLjIpIj5OT1QgTlVMTDwvc3Bhbj4nOicnKSsKICAgICAgICAobS50YWdzfHxbXSkubWFwKGZ1bmN0aW9uKHQpe3JldHVybiAnPHNwYW4gY2xhc3M9InRhZyI+Jytlc2ModCkrJzwvc3Bhbj4nO30pLmpvaW4oJycpKwogICAgICAnPC9kaXY+JysKICAgICc8L2Rpdj4nKwogICAgJzxkaXYgY2xhc3M9ImZkLWJvZHkiPicrCiAgICAgICc8ZGl2IGNsYXNzPSJmZC1zZWMiPicrCiAgICAgICAgJzxkaXYgY2xhc3M9ImZkLXNlYy10Ij4mI3gxRjRDQjsgT3ZlcnZpZXc8L2Rpdj4nKwogICAgICAgIGZyKCdEZXNjcmlwdGlvbicsZXNjKG0uZGVzY3JpcHRpb258fCdcdTIwMTQnKSkrCiAgICAgICAgZnIoJ1NvdXJjZSBTeXN0ZW0nLGVzYyhtLnNvdXJjZV9zeXN0ZW18fCdcdTIwMTQnKSkrCiAgICAgICAgZnIoJ093bmVyJyxlc2MobS5vd25lcnx8J1x1MjAxNCcpKSsKICAgICAgICBmcignVXBkYXRlZCcsZXNjKG0ubGFzdF91cGRhdGVkfHwnXHUyMDE0JykpKwogICAgICAnPC9kaXY+JysKICAgICAgJzxkaXYgY2xhc3M9ImZkLXNlYyI+JysKICAgICAgICAnPGRpdiBjbGFzcz0iZmQtc2VjLXQiPiYjeDI2QTE7IENvbmZpZGVudGlhbGl0eSAmYW1wOyBJbmdlc3Rpb248L2Rpdj4nKwogICAgICAgICc8ZGl2IHN0eWxlPSJkaXNwbGF5OmZsZXg7Z2FwOjZweDtmbGV4LXdyYXA6d3JhcDttYXJnaW4tYm90dG9tOjRweCI+JysKICAgICAgICAgIGNvbmZCYWRnZShtLmNvbmZpZGVudGlhbGl0eXx8J0ludGVybmFsJykrCiAgICAgICAgICAnPHNwYW4gY2xhc3M9ImluZy1waWxsICcrKChtLmluZ2VzdGlvbl9wYXR0ZXJufHwnJyk9PT0nUmVhbC1UaW1lIFN0cmVhbWluZyc/J2luZy1ydCc6J2luZy1iYXRjaCcpKyciPicrZXNjKG0uaW5nZXN0aW9uX3BhdHRlcm58fCdCYXRjaCcpKyc8L3NwYW4+JysKICAgICAgICAnPC9kaXY+JysKICAgICAgJzwvZGl2PicrCiAgICAgICc8ZGl2IGNsYXNzPSJmZC1zZWMiPicrCiAgICAgICAgJzxkaXYgY2xhc3M9ImZkLXNlYy10Ij4mI3gyNjk5OyBCdXNpbmVzcyBMb2dpYzwvZGl2PicrCiAgICAgICAgJzxkaXYgY2xhc3M9ImZkLWNvZGUiPicrZXNjKG0uYnVzaW5lc3NfbG9naWN8fCdObyBsb2dpYyBkZWZpbmVkJykrJzwvZGl2PicrCiAgICAgICc8L2Rpdj4nKwogICAgICAobS5idXNpbmVzc19ydWxlcyYmbS5idXNpbmVzc19ydWxlcy5sZW5ndGg/CiAgICAgICAgJzxkaXYgY2xhc3M9ImZkLXNlYyI+PGRpdiBjbGFzcz0iZmQtc2VjLXQiPiYjeDFGNENGOyBCdXNpbmVzcyBSdWxlczwvZGl2PicrCiAgICAgICAgbS5idXNpbmVzc19ydWxlcy5tYXAoZnVuY3Rpb24ocil7CiAgICAgICAgICByZXR1cm4gJzxkaXYgc3R5bGU9ImZvbnQtc2l6ZToxMHB4O2NvbG9yOnZhcigtLXQxKTttYXJnaW4tYm90dG9tOjRweDtkaXNwbGF5OmZsZXg7Z2FwOjVweCI+PHNwYW4gc3R5bGU9ImNvbG9yOnZhcigtLXRlYWwpIj4mcnNhcXVvOzwvc3Bhbj4nK2VzYyhyKSsnPC9kaXY+JzsKICAgICAgICB9KS5qb2luKCcnKSsnPC9kaXY+JzonJykrCiAgICAgICc8ZGl2IGNsYXNzPSJmZC1zZWMiPicrCiAgICAgICAgJzxkaXYgY2xhc3M9ImZkLXNlYy10Ij4mI3gxRjUxNzsgTGluZWFnZSBGbG93PC9kaXY+JysKICAgICAgICAnPGRpdiBjbGFzcz0ibGluZWFnZS1mbG93Ij4nKwogICAgICAgICAgKG0uc291cmNlc3x8W10pLm1hcChmdW5jdGlvbihzKXtyZXR1cm4gJzxzcGFuIGNsYXNzPSJsZi1uIGxmLXNyYyI+JiN4MUY0RTU7ICcrZXNjKHMpKyc8L3NwYW4+Jzt9KS5qb2luKCcnKSsKICAgICAgICAgICc8c3BhbiBjbGFzcz0ibGYtYXJyIj4mcmFycjs8L3NwYW4+JysKICAgICAgICAgICc8c3BhbiBjbGFzcz0ibGYtbiIgc3R5bGU9ImJvcmRlci1jb2xvcjpyZ2JhKDI0NSwxNjYsMzUsLjMpO2NvbG9yOnZhcigtLWdvbGQyKSI+JiN4MjFCQTsgVHJhbnNmb3JtPC9zcGFuPicrCiAgICAgICAgICAnPHNwYW4gY2xhc3M9ImxmLWFyciI+JnJhcnI7PC9zcGFuPicrCiAgICAgICAgICAobS50YXJnZXRzfHxbXSkubWFwKGZ1bmN0aW9uKHMpe3JldHVybiAnPHNwYW4gY2xhc3M9ImxmLW4gbGYtdGd0Ij4mI3gxRjRFNDsgJytlc2MocykrJzwvc3Bhbj4nO30pLmpvaW4oJycpKwogICAgICAgICc8L2Rpdj4nKwogICAgICAnPC9kaXY+JysKICAgICAgJzxkaXYgY2xhc3M9ImZkLXNlYyI+JysKICAgICAgICAnPGRpdiBjbGFzcz0iZmQtc2VjLXQiPiYjeDI2QTA7IEltcGFjdDwvZGl2PicrCiAgICAgICAgZnIoJ0ltcGFjdCcsZXNjKG0uaW1wYWN0fHwnTm8gaW1wYWN0IGFzc2Vzc21lbnQnKSkrCiAgICAgICc8L2Rpdj4nKwogICAgICBzb3JIdG1sICsgZXF1aXZIdG1sICsgaG9wc0h0bWwgKwogICAgICAnPGRpdiBjbGFzcz0iZmQtc2VjIj4nKwogICAgICAgICc8ZGl2IGNsYXNzPSJmZC1zZWMtdCI+JiN4MUY2RTE7IFJpc2sgTWF0cml4PC9kaXY+JysKICAgICAgICAnPGRpdiBjbGFzcz0icmlzay10YmwiPicrCiAgICAgICAgICBbe2w6J0hpZ2gnLGk6JyYjeDFGNTM0OycsZDonQnJlYWtpbmcgY2hhbmdlIFx1MjAxNCBkb3duc3RyZWFtIGFmZmVjdGVkJyxhOidEYXRhIFN0ZXdhcmQgKyBDQUIgYXBwcm92YWwnfSwKICAgICAgICAgICB7bDonTWVkaXVtJyxpOicmI3gxRjdFMTsnLGQ6J01vZGVyYXRlIGltcGFjdCBcdTIwMTQgdmVyaWZ5IGRlcGVuZGVuY2llcycsYTonUGVlciByZXZpZXcgKyBVQVQnfSwKICAgICAgICAgICB7bDonTG93JyxpOicmI3gxRjdFMjsnLGQ6J01pbmltYWwgcmlzaycsYTonU3RhbmRhcmQgUFIgcmV2aWV3J31dLm1hcChmdW5jdGlvbihyKXsKICAgICAgICAgICAgcmV0dXJuICc8ZGl2IGNsYXNzPSJyaXNrLXJvdyIgc3R5bGU9IicrKHIubD09PW0ucmlza19sZXZlbD8nYmFja2dyb3VuZDpyZ2JhKDI0NSwxNjYsMzUsLjA1KSc6JycpKyciPicrJzxkaXYgY2xhc3M9InJpc2stbCI+JytyLmkrJyA8c3BhbiBjbGFzcz0icmIgcmItJytyLmxbMF0rJyI+JytyLmwrJzwvc3Bhbj48L2Rpdj4nKyc8ZGl2IGNsYXNzPSJyaXNrLXIiPjxkaXYgY2xhc3M9InJpc2stcmQiPicrci5kKyc8L2Rpdj48ZGl2IGNsYXNzPSJyaXNrLXJhIj4nK3IuYSsnPC9kaXY+PC9kaXY+PC9kaXY+JzsKICAgICAgICAgIH0pLmpvaW4oJycpKwogICAgICAgICc8L2Rpdj4nKwogICAgICAnPC9kaXY+JysKICAgICc8L2Rpdj4nKwogICAgJzxkaXYgY2xhc3M9ImZkLWZvb3QiPjxidXR0b24gY2xhc3M9InBkZi1idG4iIG9uY2xpY2s9ImRvRG93bmxvYWRQREYoKSI+JiN4MUY0QzQ7IERPV05MT0FEIFBERiBSRVBPUlQ8L2J1dHRvbj48L2Rpdj4nOwp9CgovLyDilIDilIAgU09VUkNFUyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZnVuY3Rpb24gcmVuZGVyU291cmNlcyhkYXRhKXsKICB2YXIgc2QgPSBkYXRhLnNvdXJjZV9kZXRhaWxzfHx7fTsKICB2YXIgY2FyZHMgPSBbXTsKCiAgLy8gRVJXSU4KICBpZihzcmNzLmVyd2luKXsKICAgIHZhciBzdGF0ID0gKHNkLmVyd2luJiZzZC5lcndpbi5zdGF0dXMpfHwnbW9jayc7CiAgICB2YXIgc3RhdExhYmVscz17bGl2ZTonJiN4MUY3RTI7IExJVkUgQVBJJyxsb2NhbF9leGNlbDonJiN4MUY3RTE7IExPQ0FMIEVYQ0VMJyxsb2NhbF9qc29uOicmI3gxRjdFMTsgTE9DQUwgSlNPTicsbW9jazonJiN4MjZBQTsgTU9DSyd9OwogICAgdmFyIHN0YXRDb2xvcnM9e2xpdmU6J3ZhcigtLWxpbWUyKScsbG9jYWxfZXhjZWw6J3ZhcigtLWdvbGQyKScsbG9jYWxfanNvbjondmFyKC0tZ29sZDIpJyxtb2NrOid2YXIoLS10MiknfTsKICAgIHZhciBlbiA9IGRhdGEubm9kZXMuZmlsdGVyKGZ1bmN0aW9uKG4pe3JldHVybiBuLnR5cGUhPT0nY29kZV9yZXBvJzt9KTsKICAgIHZhciBzb3JMaXN0ID0gW107CiAgICBkYXRhLm1ldGFkYXRhLmZvckVhY2goZnVuY3Rpb24obSl7KG0uc29yc3x8W10pLmZvckVhY2goZnVuY3Rpb24ocyl7aWYoc29yTGlzdC5pbmRleE9mKHMuc29yKTwwKSBzb3JMaXN0LnB1c2gocy5zb3IpO30pO30pOwogICAgY2FyZHMucHVzaCgnPGRpdiBjbGFzcz0ic3JjLWNhcmQiPicrCiAgICAgICc8ZGl2IGNsYXNzPSJzYy1oZHIiPicrCiAgICAgICAgJzxkaXYgY2xhc3M9InNjLWljbyIgc3R5bGU9ImJhY2tncm91bmQ6cmdiYSgxMjcsMjE3LDk4LC4xKSI+JiN4MUYzREI7PC9kaXY+JysKICAgICAgICAnPGRpdj48ZGl2IGNsYXNzPSJzYy1uYW1lIj5FUldJTiBESTwvZGl2PjxkaXYgY2xhc3M9InNjLXN0YXQiIHN0eWxlPSJjb2xvcjonKyhzdGF0Q29sb3JzW3N0YXRdfHwndmFyKC0tdDIpJykrJyI+Jysoc3RhdExhYmVsc1tzdGF0XXx8c3RhdCkrJyAmYnVsbDsgJytlbi5sZW5ndGgrJyBvYmplY3RzPC9kaXY+PC9kaXY+JysKICAgICAgICAoKHNkLmVyd2luJiZzZC5lcndpbi5zb3VyY2VfZmlsZSk/JzxzcGFuIGNsYXNzPSJzYy1iYWRnZSBiLWZpbGUiIHRpdGxlPSInK2VzYyhzZC5lcndpbi5zb3VyY2VfZmlsZSkrJyI+JiN4MUY0QzE7IEZJTEU8L3NwYW4+JzonPHNwYW4gY2xhc3M9InNjLWJhZGdlIGItbGl2ZSI+RVJXSU48L3NwYW4+JykrCiAgICAgICc8L2Rpdj4nKwogICAgICAoKHNkLmVyd2luJiZzZC5lcndpbi5zb3VyY2VfZmlsZSk/JzxkaXYgc3R5bGU9ImZvbnQtc2l6ZTo5cHg7Y29sb3I6dmFyKC0tdDIpO21hcmdpbi1ib3R0b206OHB4O3dvcmQtYnJlYWs6YnJlYWstYWxsIj4mI3gxRjRDMjsgJytlc2Moc2QuZXJ3aW4uc291cmNlX2ZpbGUpKyc8L2Rpdj4nOicnKSsKICAgICAgZW4uc2xpY2UoMCw0KS5tYXAoZnVuY3Rpb24obil7CiAgICAgICAgcmV0dXJuICc8ZGl2IHN0eWxlPSJtYXJnaW4tYm90dG9tOjZweDtwYWRkaW5nOjZweCA4cHg7YmFja2dyb3VuZDpyZ2JhKDAsMCwwLC4zKTtib3JkZXItcmFkaXVzOjRweDtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJkcikiPicrCiAgICAgICAgICAnPGRpdiBzdHlsZT0iZm9udC1zaXplOjEwcHg7Zm9udC13ZWlnaHQ6NjAwIj4nK2VzYyhuLmxhYmVsKSsnPC9kaXY+JysKICAgICAgICAgICc8ZGl2IHN0eWxlPSJmb250LXNpemU6OXB4O2NvbG9yOnZhcigtLXQyKTttYXJnaW4tdG9wOjJweCI+Jytlc2Mobi5kYXRhYmFzZXx8J1x1MjAxNCcpKycgJmJ1bGw7ICcrZXNjKG4uc2NoZW1hX25hbWV8fCdcdTIwMTQnKSsnICZidWxsOyAnK2VzYyhuLnN5c3RlbXx8J1x1MjAxNCcpKyc8L2Rpdj4nKwogICAgICAgICAgKG4uZmllbGRzJiZuLmZpZWxkcy5sZW5ndGg/JzxkaXYgc3R5bGU9Im1hcmdpbi10b3A6NHB4O2Rpc3BsYXk6ZmxleDtmbGV4LXdyYXA6d3JhcDtnYXA6MnB4Ij4nK24uZmllbGRzLm1hcChmdW5jdGlvbihmKXtyZXR1cm4gJzxzcGFuIGNsYXNzPSJkdHlwZSIgc3R5bGU9ImZvbnQtc2l6ZTo4cHgiPicrZXNjKGYubmFtZSkrJzwvc3Bhbj4nO30pLmpvaW4oJycpKyc8L2Rpdj4nOicnKSsKICAgICAgICAnPC9kaXY+JzsKICAgICAgfSkuam9pbignJykrCiAgICAgIChzb3JMaXN0Lmxlbmd0aD8nPGRpdiBzdHlsZT0ibWFyZ2luLXRvcDo4cHg7cGFkZGluZy10b3A6NnB4O2JvcmRlci10b3A6MXB4IHNvbGlkIHZhcigtLWJkcikiPjxkaXYgc3R5bGU9ImZvbnQtc2l6ZTo5cHg7Y29sb3I6dmFyKC0tdDIpO21hcmdpbi1ib3R0b206NHB4O2ZvbnQtZmFtaWx5OlwnSmV0QnJhaW5zIE1vbm9cJyxtb25vc3BhY2UiPktub3duIFNPUnMgaW4gbGluZWFnZTo8L2Rpdj48ZGl2IHN0eWxlPSJkaXNwbGF5OmZsZXg7Z2FwOjRweDtmbGV4LXdyYXA6d3JhcCI+Jytzb3JMaXN0LnNsaWNlKDAsOCkubWFwKGZ1bmN0aW9uKHMpe3JldHVybiAnPHNwYW4gY2xhc3M9ImR0eXBlIiBzdHlsZT0iY29sb3I6dmFyKC0tZ29sZDIpO2JvcmRlci1jb2xvcjpyZ2JhKDI0NSwxNjYsMzUsLjI1KSI+Jytlc2MocykrJzwvc3Bhbj4nO30pLmpvaW4oJycpKyc8L2Rpdj48L2Rpdj4nOicnKSsKICAgICc8L2Rpdj4nKTsKICB9CgogIC8vIERhdGEgRGljdGlvbmFyeQogIGlmKHNyY3MuZGljdCl7CiAgICB2YXIgZHN0YXQgPSAoc2QuZGljdCYmc2QuZGljdC5zdGF0dXMpfHwnbW9jayc7CiAgICB2YXIgZFN0YXRMYWJlbHM9e2xpdmU6JyYjeDFGN0UyOyBMSVZFIEFQSScsbG9jYWxfZXhjZWw6JyYjeDFGN0UxOyBMT0NBTCBFWENFTCcsbG9jYWxfanNvbjonJiN4MUY3RTE7IExPQ0FMIEpTT04nLGxvY2FsX2ZpbGU6JyYjeDFGN0UxOyBMT0NBTCBGSUxFJyxtb2NrOicmI3gyNkFBOyBNT0NLJ307CiAgICB2YXIgZFN0YXRDb2xvcnM9e2xpdmU6J3ZhcigtLXZpbzIpJyxsb2NhbF9leGNlbDondmFyKC0tZ29sZDIpJyxsb2NhbF9qc29uOid2YXIoLS1nb2xkMiknLGxvY2FsX2ZpbGU6J3ZhcigtLWdvbGQyKScsbW9jazondmFyKC0tdDIpJ307CiAgICBjYXJkcy5wdXNoKCc8ZGl2IGNsYXNzPSJzcmMtY2FyZCI+JysKICAgICAgJzxkaXYgY2xhc3M9InNjLWhkciI+JysKICAgICAgICAnPGRpdiBjbGFzcz0ic2MtaWNvIiBzdHlsZT0iYmFja2dyb3VuZDpyZ2JhKDE1NSwxMDksMjU1LC4xKSI+JiN4MUY0RDY7PC9kaXY+JysKICAgICAgICAnPGRpdj48ZGl2IGNsYXNzPSJzYy1uYW1lIj5EYXRhIERpY3Rpb25hcnk8L2Rpdj48ZGl2IGNsYXNzPSJzYy1zdGF0IiBzdHlsZT0iY29sb3I6JysoZFN0YXRDb2xvcnNbZHN0YXRdfHwndmFyKC0tdDIpJykrJyI+JysoZFN0YXRMYWJlbHNbZHN0YXRdfHxkc3RhdCkrJyAmYnVsbDsgJytkYXRhLm1ldGFkYXRhLmxlbmd0aCsnIGZpZWxkczwvZGl2PjwvZGl2PicrCiAgICAgICAgKChzZC5kaWN0JiZzZC5kaWN0LnNvdXJjZV9maWxlKT8nPHNwYW4gY2xhc3M9InNjLWJhZGdlIGItZmlsZSI+JiN4MUY0QzE7IEZJTEU8L3NwYW4+JzonPHNwYW4gY2xhc3M9InNjLWJhZGdlIGItYWkiPkRJQ1Q8L3NwYW4+JykrCiAgICAgICc8L2Rpdj4nKwogICAgICAoKHNkLmRpY3QmJnNkLmRpY3Quc291cmNlX2ZpbGUpPyc8ZGl2IHN0eWxlPSJmb250LXNpemU6OXB4O2NvbG9yOnZhcigtLXQyKTttYXJnaW4tYm90dG9tOjdweDt3b3JkLWJyZWFrOmJyZWFrLWFsbCI+JiN4MUY0QzI7ICcrZXNjKHNkLmRpY3Quc291cmNlX2ZpbGUpKyc8L2Rpdj4nOicnKSsKICAgICAgZGF0YS5tZXRhZGF0YS5zbGljZSgwLDUpLm1hcChmdW5jdGlvbihtKXsKICAgICAgICByZXR1cm4gJzxkaXYgc3R5bGU9Im1hcmdpbi1ib3R0b206NXB4O3BhZGRpbmc6NXB4IDdweDtiYWNrZ3JvdW5kOnJnYmEoMCwwLDAsLjMpO2JvcmRlci1yYWRpdXM6NHB4O2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyKSI+JysKICAgICAgICAgICc8ZGl2IHN0eWxlPSJkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo0cHg7bWFyZ2luLWJvdHRvbToycHgiPicrCiAgICAgICAgICAgICc8c3BhbiBzdHlsZT0iZm9udC1mYW1pbHk6XCdKZXRCcmFpbnMgTW9ub1wnLG1vbm9zcGFjZTtmb250LXNpemU6MTBweDtmb250LXdlaWdodDo2MDA7YmFja2dyb3VuZDpsaW5lYXItZ3JhZGllbnQoOTBkZWcsdmFyKC0tZ29sZDIpLHZhcigtLXRlYWwyKSk7LXdlYmtpdC1iYWNrZ3JvdW5kLWNsaXA6dGV4dDstd2Via2l0LXRleHQtZmlsbC1jb2xvcjp0cmFuc3BhcmVudCI+Jytlc2MobS5maWVsZF9uYW1lKSsnPC9zcGFuPicrCiAgICAgICAgICAgICc8c3BhbiBjbGFzcz0iZHR5cGUiIHN0eWxlPSJmb250LXNpemU6OHB4Ij4nK2VzYyhtLmRhdGFfdHlwZXx8JycpKyc8L3NwYW4+JysKICAgICAgICAgICAgJzxzcGFuIGNsYXNzPSJyYiByYi0nKyhtLnJpc2tfbGV2ZWx8fCdMJylbMF0rJyIgc3R5bGU9ImZvbnQtc2l6ZTo4cHgiPicrZXNjKG0ucmlza19sZXZlbHx8J0xvdycpKyc8L3NwYW4+JysKICAgICAgICAgICc8L2Rpdj4nKwogICAgICAgICAgJzxkaXYgc3R5bGU9ImZvbnQtc2l6ZTo5cHg7Y29sb3I6dmFyKC0tdDIpIj4nK2VzYyhtLmRlc2NyaXB0aW9ufHwnXHUyMDE0JykrJzwvZGl2PicrCiAgICAgICAgJzwvZGl2Pic7CiAgICAgIH0pLmpvaW4oJycpKwogICAgJzwvZGl2PicpOwogIH0KCiAgLy8gQ29kZSBSZXBvCiAgaWYoc3Jjcy5jb2RlKXsKICAgIHZhciBjbm9kZSA9IGRhdGEubm9kZXMuZmluZChmdW5jdGlvbihuKXtyZXR1cm4gbi50eXBlPT09J2NvZGVfcmVwbyc7fSk7CiAgICB2YXIgY3JlcyAgPSAoY25vZGUmJmNub2RlLmNvZGVfcmVzdWx0cyl8fFtdOwogICAgdmFyIGNzdGF0ID0gKHNkLmNvZGUmJnNkLmNvZGUuc3RhdHVzKXx8J21vY2snOwogICAgdmFyIGNMYWJlbD17Z2l0aHViOicmI3gxRjQxOTsgR0lUSFVCJyxsb2NhbDonJiN4MUY0QzE7IExPQ0FMJyxtb2NrOicmI3gyNkFBOyBNT0NLJ31bY3N0YXRdfHxjc3RhdC50b1VwcGVyQ2FzZSgpOwogICAgdmFyIGNDb2xvcj17Z2l0aHViOid2YXIoLS1saW1lMiknLGxvY2FsOid2YXIoLS1za3kyKScsbW9jazondmFyKC0tdDIpJ31bY3N0YXRdfHwndmFyKC0tdDIpJzsKICAgIGNhcmRzLnB1c2goJzxkaXYgY2xhc3M9InNyYy1jYXJkIj4nKwogICAgICAnPGRpdiBjbGFzcz0ic2MtaGRyIj4nKwogICAgICAgICc8ZGl2IGNsYXNzPSJzYy1pY28iIHN0eWxlPSJiYWNrZ3JvdW5kOnJnYmEoNTksMTU4LDI1NSwuMSkiPiYjeDFGNEJCOzwvZGl2PicrCiAgICAgICAgJzxkaXY+PGRpdiBjbGFzcz0ic2MtbmFtZSI+Q29kZSBSZXBvc2l0b3J5PC9kaXY+PGRpdiBjbGFzcz0ic2Mtc3RhdCIgc3R5bGU9ImNvbG9yOicrY0NvbG9yKyciPicrY0xhYmVsKycgJmJ1bGw7ICcrY3Jlcy5sZW5ndGgrJyBzY3JpcHRzPC9kaXY+PC9kaXY+JysKICAgICAgICAnPHNwYW4gY2xhc3M9InNjLWJhZGdlICcrKGNzdGF0PT09J2dpdGh1Yic/J2ItbGl2ZSc6Y3N0YXQ9PT0nbG9jYWwnPydiLWZpbGUnOidiLW1vY2snKSsnIj4nK2NzdGF0LnRvVXBwZXJDYXNlKCkrJzwvc3Bhbj4nKwogICAgICAnPC9kaXY+JysKICAgICAgKGNyZXMubGVuZ3RoP2NyZXMubWFwKGZ1bmN0aW9uKGMpewogICAgICAgIHJldHVybiAnPGRpdiBzdHlsZT0ibWFyZ2luLWJvdHRvbTo5cHgiPicrCiAgICAgICAgICAnPGRpdiBzdHlsZT0iZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6NXB4O21hcmdpbi1ib3R0b206M3B4Ij4nKwogICAgICAgICAgICAnPHNwYW4gY2xhc3M9ImxhbmctcGlsbCBscC0nK2VzYyhjLmxhbmd1YWdlKSsnIj4nK2VzYyhjLmxhbmd1YWdlKSsnPC9zcGFuPicrCiAgICAgICAgICAgICc8c3BhbiBzdHlsZT0iZm9udC1zaXplOjlweDtjb2xvcjp2YXIoLS10MSkiPicrZXNjKGMuZmlsZSkrJzwvc3Bhbj4nKwogICAgICAgICAgICAoYy51cmw/JzxhIGhyZWY9IicrZXNjKGMudXJsKSsnIiB0YXJnZXQ9Il9ibGFuayIgc3R5bGU9ImZvbnQtc2l6ZTo4cHg7Y29sb3I6dmFyKC0tc2t5Mik7bWFyZ2luLWxlZnQ6YXV0byI+R2l0SHViICZuZWFycjs8L2E+JzonJykrCiAgICAgICAgICAnPC9kaXY+JysKICAgICAgICAgICc8ZGl2IGNsYXNzPSJjb2RlLWJsayI+Jytlc2MoYy5idXNpbmVzc19sb2dpYykrJzwvZGl2PicrCiAgICAgICAgICAnPGRpdiBzdHlsZT0iZm9udC1zaXplOjhweDtjb2xvcjp2YXIoLS10MykiPiYjeDFGNEMxOyAnK2VzYyhjLnBhdGgpKyc8L2Rpdj4nKwogICAgICAgICAgKGMuYWlfc3VtbWFyeT8nPGRpdiBjbGFzcz0iYWktaW5zaWdodC1jYXJkIj48ZGl2IGNsYXNzPSJhaS1pbnNpZ2h0LXRleHQiPicrZXNjKGMuYWlfc3VtbWFyeSkrJzwvZGl2PjwvZGl2Pic6JycpKwogICAgICAgICc8L2Rpdj4nOwogICAgICB9KS5qb2luKCcnKToKICAgICAgJzxkaXYgc3R5bGU9ImZvbnQtc2l6ZToxMXB4O2NvbG9yOnZhcigtLXQzKSI+Tm8gc2NyaXB0cyBmb3VuZDwvZGl2PicpKwogICAgJzwvZGl2PicpOwogIH0KCiAgLy8gVGFjaHlvbiBBSQogIGlmKHNyY3MudGFjaHlvbiAmJiBkYXRhLmFpX2luc2lnaHRzKXsKICAgIGNhcmRzLnB1c2goJzxkaXYgY2xhc3M9InNyYy1jYXJkIiBzdHlsZT0iYm9yZGVyLWNvbG9yOnJnYmEoMTU1LDEwOSwyNTUsLjMpIj4nKwogICAgICAnPGRpdiBjbGFzcz0ic2MtaGRyIj4nKwogICAgICAgICc8ZGl2IGNsYXNzPSJzYy1pY28iIHN0eWxlPSJiYWNrZ3JvdW5kOnJnYmEoMTU1LDEwOSwyNTUsLjEpIj4mI3gxRjkxNjs8L2Rpdj4nKwogICAgICAgICc8ZGl2PjxkaXYgY2xhc3M9InNjLW5hbWUiPlRhY2h5b24gQUk8L2Rpdj48ZGl2IGNsYXNzPSJzYy1zdGF0IiBzdHlsZT0iY29sb3I6dmFyKC0tdmlvMikiPiYjeDFGN0UzOyBMTE0gRU5SSUNITUVOVDwvZGl2PjwvZGl2PicrCiAgICAgICAgJzxzcGFuIGNsYXNzPSJzYy1iYWRnZSBiLWFpIj5MTE08L3NwYW4+JysKICAgICAgJzwvZGl2PicrCiAgICAgICc8ZGl2IGNsYXNzPSJhaS1pbnNpZ2h0LWNhcmQiPjxkaXYgY2xhc3M9ImFpLWluc2lnaHQtdGV4dCI+Jytlc2MoZGF0YS5haV9pbnNpZ2h0cykrJzwvZGl2PjwvZGl2PicrCiAgICAnPC9kaXY+Jyk7CiAgfQoKICBlbCgnc3JjLWNhcmRzJykuaW5uZXJIVE1MID0gY2FyZHMuam9pbignJyk7Cn0KCi8vIOKUgOKUgCBNT0RBTCDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKZnVuY3Rpb24gb3Blbk5vZGVNb2RhbChkKXsKICBlbCgnbW9kYWwtdGl0bGUnKS50ZXh0Q29udGVudCA9IGQubGFiZWx8fGQuaWQ7CiAgZWwoJ21vZGFsLWJvZHknKS5pbm5lckhUTUwgPQogICAgZnIoJ1R5cGUnLCBlc2MoKGQudHlwZXx8JycpLnJlcGxhY2UoL18vZywnICcpKSkrCiAgICBmcignU3lzdGVtJywgZXNjKGQuc3lzdGVtfHwnXHUyMDE0JykpKwogICAgZnIoJ0RhdGFiYXNlJywgZXNjKGQuZGF0YWJhc2V8fCdcdTIwMTQnKSkrCiAgICBmcignUmlzaycsICc8c3BhbiBjbGFzcz0icmIgcmItJysoZC5yaXNrX2xldmVsfHwnTCcpWzBdKyciPicrZXNjKGQucmlza19sZXZlbHx8J0xvdycpKyc8L3NwYW4+JykrCiAgICAoZC5kZXNjcmlwdGlvbj9mcignRGVzY3JpcHRpb24nLGVzYyhkLmRlc2NyaXB0aW9uKSk6JycpKwogICAgKGQuY29kZV9yZXN1bHRzPwogICAgICAnPGRpdiBzdHlsZT0ibWFyZ2luLXRvcDoxMXB4Ij48ZGl2IGNsYXNzPSJmZC1zZWMtdCI+Q29kZSBSZWZlcmVuY2VzPC9kaXY+JysKICAgICAgZC5jb2RlX3Jlc3VsdHMubWFwKGZ1bmN0aW9uKGMpewogICAgICAgIHJldHVybiAnPGRpdiBzdHlsZT0ibWFyZ2luLWJvdHRvbTo3cHgiPjxzcGFuIGNsYXNzPSJsYW5nLXBpbGwgbHAtJytlc2MoYy5sYW5ndWFnZSkrJyI+Jytlc2MoYy5sYW5ndWFnZSkrJzwvc3Bhbj48ZGl2IGNsYXNzPSJjb2RlLWJsayI+Jytlc2MoYy5idXNpbmVzc19sb2dpYykrJzwvZGl2PjwvZGl2Pic7CiAgICAgIH0pLmpvaW4oJycpKyc8L2Rpdj4nOicnKTsKICBlbCgnb3ZlcmxheScpLmNsYXNzTGlzdC5hZGQoJ29wZW4nKTsKfQpmdW5jdGlvbiBjbG9zZU1vZGFsKCl7IGVsKCdvdmVybGF5JykuY2xhc3NMaXN0LnJlbW92ZSgnb3BlbicpOyB9CmZ1bmN0aW9uIGNsb3NlT3ZlcmxheShlKXsgaWYoZS50YXJnZXQ9PT1lbCgnb3ZlcmxheScpKSBjbG9zZU1vZGFsKCk7IH0KCi8vIOKUgOKUgCBTT1IgU0VMRUNUIOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgOKUgAphc3luYyBmdW5jdGlvbiBkb1NldFNPUihmaWVsZE5hbWUsIHNvcil7CiAgdHJ5ewogICAgdmFyIHIgPSBhd2FpdCBmZXRjaChBUEkrJy9hcGkvc29yLXNlbGVjdCcsewogICAgICBtZXRob2Q6J1BPU1QnLCBoZWFkZXJzOnsnQ29udGVudC1UeXBlJzonYXBwbGljYXRpb24vanNvbid9LAogICAgICBib2R5OiBKU09OLnN0cmluZ2lmeSh7ZmllbGRfbmFtZTpmaWVsZE5hbWUsIHByZWZlcnJlZF9zb3I6c29yfSkKICAgIH0pOwogICAgaWYoIXIub2spIHRocm93IG5ldyBFcnJvcihhd2FpdCByLnRleHQoKSk7CiAgICB0b2FzdCgnXHUyNjA1ICcrZmllbGROYW1lKycgXHUyMTkyICcrc29yKycgc2V0IGFzIGF1dGhvcml0YXRpdmUgU09SJywgJ3N1Y2Nlc3MnKTsKICAgIC8vIFVwZGF0ZSBpbi1tZW1vcnkgZGF0YSBhbmQgcmUtcmVuZGVyCiAgICBpZihzZWxGaWVsZCAmJiBzZWxGaWVsZC5maWVsZF9uYW1lPT09ZmllbGROYW1lICYmIHNlbEZpZWxkLnNvcnMpewogICAgICBzZWxGaWVsZC5zb3JzLmZvckVhY2goZnVuY3Rpb24ocyl7IHMuaXNfYXV0aG9yaXRhdGl2ZSA9IHMuc29yLnRvVXBwZXJDYXNlKCk9PT1zb3IudG9VcHBlckNhc2UoKTsgfSk7CiAgICAgIHNlbEZpZWxkLnByZWZlcnJlZF9zb3IgPSBzb3I7CiAgICAgIHNlbGVjdEZpZWxkKHNlbEZpZWxkKTsKICAgIH0KICB9IGNhdGNoKGUpeyB0b2FzdCgnU09SIGVycm9yOiAnK2UubWVzc2FnZSwnZXJyb3InKTsgfQp9CgovLyDilIDilIAgRVhQT1JUUyDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKYXN5bmMgZnVuY3Rpb24gZG9FeHBvcnRFeGNlbCgpewogIGlmKCFEKXsgdG9hc3QoJ1J1biBhIHNlYXJjaCBmaXJzdCcsJ3dhcm4nKTsgcmV0dXJuOyB9CiAgdG9hc3QoJyYjeDFGNENBOyBHZW5lcmF0aW5nIEV4Y2VsLi4uJywnaW5mbycpOwogIHRyeXsKICAgIHZhciByID0gYXdhaXQgZmV0Y2goQVBJKycvYXBpL2V4cG9ydC9leGNlbCcsewogICAgICBtZXRob2Q6J1BPU1QnLCBoZWFkZXJzOnsnQ29udGVudC1UeXBlJzonYXBwbGljYXRpb24vanNvbid9LAogICAgICBib2R5OiBKU09OLnN0cmluZ2lmeSh7cXVlcnk6RC5xdWVyeSxtZXRhZGF0YTpELm1ldGFkYXRhLG5vZGVzOkQubm9kZXMsZWRnZXM6RC5lZGdlcyxzb3VyY2VzX3VzZWQ6RC5zb3VyY2VzX3VzZWQsYWlfaW5zaWdodHM6RC5haV9pbnNpZ2h0c3x8bnVsbH0pCiAgICB9KTsKICAgIGlmKCFyLm9rKSB0aHJvdyBuZXcgRXJyb3IoYXdhaXQgci50ZXh0KCkpOwogICAgdmFyIGJsb2IgPSBhd2FpdCByLmJsb2IoKTsKICAgIHZhciB1cmwgID0gVVJMLmNyZWF0ZU9iamVjdFVSTChibG9iKTsKICAgIHZhciBhICAgID0gZG9jdW1lbnQuY3JlYXRlRWxlbWVudCgnYScpOwogICAgYS5ocmVmID0gdXJsOyBhLmRvd25sb2FkID0gJ0RhdGFMaW5lYWdlXycrRC5xdWVyeSsnXycrbmV3IERhdGUoKS50b0lTT1N0cmluZygpLnNsaWNlKDAsMTApKycueGxzeCc7CiAgICBhLmNsaWNrKCk7IFVSTC5yZXZva2VPYmplY3RVUkwodXJsKTsKICAgIHRvYXN0KCdcdTI3MDUgRXhjZWwgZG93bmxvYWRlZCEnLCdzdWNjZXNzJyk7CiAgfSBjYXRjaChlKXsgdG9hc3QoJ0V4Y2VsIGVycm9yOiAnK2UubWVzc2FnZSwnZXJyb3InKTsgfQp9Cgphc3luYyBmdW5jdGlvbiBkb0V4cG9ydEpTT04oKXsKICBpZighRCl7IHRvYXN0KCdSdW4gYSBzZWFyY2ggZmlyc3QnLCd3YXJuJyk7IHJldHVybjsgfQogIHRvYXN0KCcmI3gxRjRFNjsgR2VuZXJhdGluZyBKU09OLi4uJywnaW5mbycpOwogIHRyeXsKICAgIHZhciByID0gYXdhaXQgZmV0Y2goQVBJKycvYXBpL2V4cG9ydC9qc29uJyx7CiAgICAgIG1ldGhvZDonUE9TVCcsIGhlYWRlcnM6eydDb250ZW50LVR5cGUnOidhcHBsaWNhdGlvbi9qc29uJ30sCiAgICAgIGJvZHk6IEpTT04uc3RyaW5naWZ5KHtxdWVyeTpELnF1ZXJ5LG1ldGFkYXRhOkQubWV0YWRhdGEsbm9kZXM6RC5ub2RlcyxlZGdlczpELmVkZ2VzLHNvdXJjZXNfdXNlZDpELnNvdXJjZXNfdXNlZCxhaV9pbnNpZ2h0czpELmFpX2luc2lnaHRzfHxudWxsfSkKICAgIH0pOwogICAgaWYoIXIub2spIHRocm93IG5ldyBFcnJvcihhd2FpdCByLnRleHQoKSk7CiAgICB2YXIgYmxvYiA9IGF3YWl0IHIuYmxvYigpOwogICAgdmFyIHVybCAgPSBVUkwuY3JlYXRlT2JqZWN0VVJMKGJsb2IpOwogICAgdmFyIGEgICAgPSBkb2N1bWVudC5jcmVhdGVFbGVtZW50KCdhJyk7CiAgICBhLmhyZWYgPSB1cmw7IGEuZG93bmxvYWQgPSAnRGF0YUxpbmVhZ2VfJytELnF1ZXJ5KydfJytuZXcgRGF0ZSgpLnRvSVNPU3RyaW5nKCkuc2xpY2UoMCwxMCkrJy5qc29uJzsKICAgIGEuY2xpY2soKTsgVVJMLnJldm9rZU9iamVjdFVSTCh1cmwpOwogICAgdG9hc3QoJ1x1MjcwNSBKU09OIGRvd25sb2FkZWQhJywnc3VjY2VzcycpOwogIH0gY2F0Y2goZSl7IHRvYXN0KCdKU09OIGVycm9yOiAnK2UubWVzc2FnZSwnZXJyb3InKTsgfQp9Cgphc3luYyBmdW5jdGlvbiBkb0Rvd25sb2FkUERGKCl7CiAgaWYoIXNlbEZpZWxkKXsgdG9hc3QoJ1NlbGVjdCBhIGZpZWxkIGZpcnN0Jywnd2FybicpOyByZXR1cm47IH0KICB0b2FzdCgnJiN4MUY0QzQ7IEdlbmVyYXRpbmcgUERGLi4uJywnaW5mbycpOwogIHRyeXsKICAgIHZhciBwYXlsb2FkID0gSlNPTi5wYXJzZShKU09OLnN0cmluZ2lmeShzZWxGaWVsZCkpOwogICAgcGF5bG9hZC5haV9pbnNpZ2h0cyA9IChEJiZELmFpX2luc2lnaHRzKXx8Jyc7CiAgICB2YXIgciA9IGF3YWl0IGZldGNoKEFQSSsnL2FwaS9nZW5lcmF0ZS1yZXBvcnQnLHsKICAgICAgbWV0aG9kOidQT1NUJywgaGVhZGVyczp7J0NvbnRlbnQtVHlwZSc6J2FwcGxpY2F0aW9uL2pzb24nfSwKICAgICAgYm9keTogSlNPTi5zdHJpbmdpZnkocGF5bG9hZCkKICAgIH0pOwogICAgaWYoIXIub2spIHRocm93IG5ldyBFcnJvcihhd2FpdCByLnRleHQoKSk7CiAgICB2YXIgYmxvYiA9IGF3YWl0IHIuYmxvYigpOwogICAgdmFyIHVybCAgPSBVUkwuY3JlYXRlT2JqZWN0VVJMKGJsb2IpOwogICAgdmFyIGEgICAgPSBkb2N1bWVudC5jcmVhdGVFbGVtZW50KCdhJyk7CiAgICBhLmhyZWYgPSB1cmw7IGEuZG93bmxvYWQgPSAnbGluZWFnZV8nK3NlbEZpZWxkLmZpZWxkX25hbWUrJ18nK25ldyBEYXRlKCkudG9JU09TdHJpbmcoKS5zbGljZSgwLDEwKSsnLnBkZic7CiAgICBhLmNsaWNrKCk7IFVSTC5yZXZva2VPYmplY3RVUkwodXJsKTsKICAgIHRvYXN0KCdcdTI3MDUgUERGIGRvd25sb2FkZWQhJywnc3VjY2VzcycpOwogIH0gY2F0Y2goZSl7IHRvYXN0KCdQREYgZXJyb3I6ICcrZS5tZXNzYWdlLCdlcnJvcicpOyB9Cn0KCi8vIOKUgOKUgCBNVUxUSS1GSUVMRCBTRUFSQ0gg4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSA4pSACmFzeW5jIGZ1bmN0aW9uIGRvTXVsdGlTZWFyY2goKXsKICB2YXIgcmF3ID0gZWwoJ211bHRpLXRhJykudmFsdWUudHJpbSgpOwogIGlmKCFyYXcpeyB0b2FzdCgnRW50ZXIgYXQgbGVhc3Qgb25lIGZpZWxkIG5hbWUnLCd3YXJuJyk7IHJldHVybjsgfQogIHZhciBmaWVsZHMgPSByYXcuc3BsaXQoL1tcblxyLF0rLykubWFwKGZ1bmN0aW9uKGYpe3JldHVybiBmLnRyaW0oKTt9KS5maWx0ZXIoQm9vbGVhbik7CiAgaWYoIWZpZWxkcy5sZW5ndGgpeyB0b2FzdCgnTm8gdmFsaWQgZmllbGQgbmFtZXMnLCd3YXJuJyk7IHJldHVybjsgfQoKICBlbCgndG9wLWJhcicpLnN0eWxlLmRpc3BsYXkgPSAnYmxvY2snOwogIGVsKCdleHAtcm93Jykuc3R5bGUuZGlzcGxheSA9ICdub25lJzsKICBoaWRlKCdzdC1lbXB0eScpOwogIGhpZGVBbGxQYW5lcygpOwogIHNob3dUYWIoJ211bHRpJyk7CiAgZWwoJ211bHRpLWFjdGlvbi1iYXInKS5zdHlsZS5kaXNwbGF5ID0gJ2ZsZXgnOwogIGVsKCdtYWItc3RhdHMnKS50ZXh0Q29udGVudCA9IGZpZWxkcy5sZW5ndGggKyAnIGZpZWxkcyc7CiAgZWwoJ3Jlcy10aXRsZScpLnRleHRDb250ZW50ID0gJ011bHRpLUZpZWxkOiAnK2ZpZWxkcy5sZW5ndGgrJyBmaWVsZHMnOwogIGVsKCdtdWx0aS1zdW1tYXJ5JykudGV4dENvbnRlbnQgPSAnU2VhcmNoaW5nICcrZmllbGRzLmxlbmd0aCsnIGZpZWxkcy4uLic7CiAgZWwoJ211bHRpLXJlc3VsdHMnKS5pbm5lckhUTUwgPSAnPGRpdiBzdHlsZT0iY29sb3I6dmFyKC0tdDIpO2ZvbnQtc2l6ZToxMXB4O3BhZGRpbmc6MjBweCAwIj5cdTIzRjMgUnVubmluZyBtdWx0aS1maWVsZCBzZWFyY2guLi48L2Rpdj4nOwoKICB0cnl7CiAgICB2YXIgYm9keSA9IHsKICAgICAgZmllbGRzOiAgICAgICAgIGZpZWxkcywKICAgICAgdXNlX2Vyd2luOiAgICAgIHNyY3MuZXJ3aW4sCiAgICAgIHVzZV9kYXRhX2RpY3Q6ICBzcmNzLmRpY3QsCiAgICAgIHVzZV9jb2RlX3JlcG86ICBzcmNzLmNvZGUsCiAgICAgIHVzZV90YWNoeW9uOiAgICBzcmNzLnRhY2h5b24sCiAgICAgIGVyd2luX2Jhc2VfdXJsOiAgICAgZ3YoJ2NmLWVyd2luLXVybCcpICAgICAgfHwgbnVsbCwKICAgICAgZXJ3aW5fYXBpX2tleTogICAgICBndignY2YtZXJ3aW4ta2V5JykgICAgICB8fCBudWxsLAogICAgICBlcndpbl9zeXN0ZW06ICAgICAgIGd2KCdjZi1lcndpbi1zeXN0ZW0nKSAgIHx8IG51bGwsCiAgICAgIGVyd2luX2VudjogICAgICAgICAgZ3YoJ2NmLWVyd2luLWVudicpICAgICAgIHx8IG51bGwsCiAgICAgIGVyd2luX3Byb2plY3Q6ICAgICAgZ3YoJ2NmLWVyd2luLXByb2plY3QnKSAgfHwgbnVsbCwKICAgICAgZXJ3aW5fbGluZWFnZV90eXBlOiBndignY2YtZXJ3aW4tbGludHlwZScpICB8fCAnRFVBTCcsCiAgICAgIGVyd2luX2VwX3RhYmxlbmFtZTogZ3YoJ2NmLWVwLXRhYmxlbmFtZScpICAgfHwgbnVsbCwKICAgICAgZXJ3aW5fZXBfY29sbmFtZTogICBndignY2YtZXAtY29sbmFtZScpICAgICB8fCBudWxsLAogICAgICBlcndpbl9lcF9vYmp0eXBlOiAgIGd2KCdjZi1lcC1vYmp0eXBlJykgICAgIHx8ICdUQUJMRScsCiAgICAgIGVyd2luX2VwX21hcG5hbWU6ICAgZ3YoJ2NmLWVwLW1hcG5hbWUnKSAgICAgfHwgbnVsbCwKICAgICAgZXJ3aW5fZXBfZ2xvc3NjYXQ6ICBndignY2YtZXAtZ2xvc3NjYXQnKSAgICB8fCBudWxsLAogICAgICBsb2NhbF9tZXRhX2RpcjogZ3YoJ2NmLW1ldGEnKSAgICAgICAgfHwgbnVsbCwKICAgICAgbG9jYWxfZGljdF9kaXI6IGd2KCdjZi1kaWN0JykgICAgICAgIHx8IG51bGwsCiAgICAgIGxvY2FsX2NvZGVfZGlyOiBndignY2YtbG9jYWwnKSAgICAgICB8fCBudWxsLAogICAgICB0YWNoeW9uX3VybDogICAgZ3YoJ2NmLXRhY2h5b24tdXJsJykgfHwgbnVsbCwKICAgICAgdGFjaHlvbl9rZXk6ICAgIGd2KCdjZi10YWNoeW9uLWtleScpIHx8IG51bGwsCiAgICB9OwogICAgdmFyIHIgPSBhd2FpdCBmZXRjaChBUEkrJy9hcGkvbXVsdGktc2VhcmNoJyx7CiAgICAgIG1ldGhvZDonUE9TVCcsIGhlYWRlcnM6eydDb250ZW50LVR5cGUnOidhcHBsaWNhdGlvbi9qc29uJ30sCiAgICAgIGJvZHk6IEpTT04uc3RyaW5naWZ5KGJvZHkpCiAgICB9KTsKICAgIGlmKCFyLm9rKSB0aHJvdyBuZXcgRXJyb3IoYXdhaXQgci50ZXh0KCkpOwogICAgdmFyIGQgPSBhd2FpdCByLmpzb24oKTsKCiAgICBlbCgnbWFiLXN0YXRzJykudGV4dENvbnRlbnQgPSBkLnN1bW1hcnkgKyAnIMK3ICcgKyBkLnNlYXJjaF90aW1lX21zICsgJ21zJzsKICAgIC8vIEJ1aWxkIGNvbWJpbmVkIEQgc28gTGluZWFnZS9NZXRhZGF0YS9Tb3VyY2VzIHRhYnMgd29yawogICAgRCA9IHsKICAgICAgcXVlcnk6IGZpZWxkcy5qb2luKCcsICcpLAogICAgICBub2RlczogZC5jb21iaW5lZF9ncmFwaCA/IGQuY29tYmluZWRfZ3JhcGgubm9kZXMgOiBbXSwKICAgICAgZWRnZXM6IGQuY29tYmluZWRfZ3JhcGggPyBkLmNvbWJpbmVkX2dyYXBoLmVkZ2VzIDogW10sCiAgICAgIG1ldGFkYXRhOiBkLmNvbWJpbmVkX21ldGFkYXRhIHx8IFtdLAogICAgICBzb3VyY2VzX3VzZWQ6IFsnTXVsdGktRmllbGQgU2VhcmNoIFsnK2ZpZWxkcy5sZW5ndGgrJyBmaWVsZHNdJ10sCiAgICAgIHNvdXJjZV9kZXRhaWxzOiB7fSwKICAgICAgc2VhcmNoX3RpbWVfbXM6IGQuc2VhcmNoX3RpbWVfbXMsCiAgICAgIGFpX2luc2lnaHRzOiBudWxsCiAgICB9OwoKICAgIC8vIFNob3cgZXhwb3J0IGJ1dHRvbnMgYW5kIHVwZGF0ZSByZXN1bHQgYmFyCiAgICBlbCgnZXhwLXJvdycpLnN0eWxlLmRpc3BsYXkgPSAnZmxleCc7CiAgICBlbCgnc3RhdC1waWxscycpLmlubmVySFRNTCA9CiAgICAgICc8ZGl2IGNsYXNzPSJzcGlsbCBzcC1saXZlIj48ZGl2IGNsYXNzPSJzZG90IiBzdHlsZT0iYmFja2dyb3VuZDp2YXIoLS10ZWFsKSI+PC9kaXY+JyArIGZpZWxkcy5sZW5ndGggKyAnIGZpZWxkcyBzZWFyY2hlZDwvZGl2PicgKwogICAgICAnPGRpdiBjbGFzcz0ic3BpbGwiPjxkaXYgY2xhc3M9InNkb3QiIHN0eWxlPSJiYWNrZ3JvdW5kOnZhcigtLXQyKSI+PC9kaXY+JyArIEQubm9kZXMubGVuZ3RoICsgJyBub2RlczwvZGl2PicgKwogICAgICAnPGRpdiBjbGFzcz0ic3BpbGwiPjxkaXYgY2xhc3M9InNkb3QiIHN0eWxlPSJiYWNrZ3JvdW5kOnZhcigtLXQyKSI+PC9kaXY+XHUyM0YxICcgKyBkLnNlYXJjaF90aW1lX21zICsgJ21zPC9kaXY+JzsKCiAgICByZW5kZXJNZXRhZGF0YShELm1ldGFkYXRhKTsKICAgIHJlbmRlclNvdXJjZXMoRCk7CiAgICByZW5kZXJNdWx0aVJlc3VsdHMoZCk7CgogIH0gY2F0Y2goZSl7CiAgICBlbCgnbXVsdGktcmVzdWx0cycpLmlubmVySFRNTCA9ICc8ZGl2IHN0eWxlPSJjb2xvcjp2YXIoLS1yb3NlMik7Zm9udC1zaXplOjExcHg7cGFkZGluZzoyMHB4Ij5cdTI3NEMgRXJyb3I6ICcrZXNjKGUubWVzc2FnZSkrJzwvZGl2Pic7CiAgICB0b2FzdCgnTXVsdGktc2VhcmNoIGVycm9yOiAnK2UubWVzc2FnZSwnZXJyb3InKTsKICB9Cn0KCmZ1bmN0aW9uIHJlbmRlck11bHRpUmVzdWx0cyhkKXsKICAgIGVsKCdtdWx0aS1zdW1tYXJ5JykudGV4dENvbnRlbnQgPSAnXHUyNzA1ICcgKyBkLnN1bW1hcnkgKyAnIFx1MDBCNyAnICsgZC5zZWFyY2hfdGltZV9tcyArICdtcyc7CgogIC8vIFJlbmRlciBkZWRpY2F0ZWQgYWN0aW9uIGJhciBhYm92ZSByZXN1bHRzCiAgZWwoJ211bHRpLXJlc3VsdHMnKS5pbm5lckhUTUwgPQogICAgJzxkaXYgc3R5bGU9ImRpc3BsYXk6ZmxleDtnYXA6OHB4O21hcmdpbi1ib3R0b206MTRweDtmbGV4LXdyYXA6d3JhcCI+JyArCiAgICAgICc8YnV0dG9uIG9uY2xpY2s9InNob3dUYWIoXCdsaW5lYWdlXCcpO3NldFRpbWVvdXQoZnVuY3Rpb24oKXtzZXRMZXZlbChcJ3RhYmxlXCcpO30sODApIiAnICsKICAgICAgICAnc3R5bGU9ImRpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjZweDtwYWRkaW5nOjlweCAxNnB4OycgKwogICAgICAgICdiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCgxMzVkZWcsdmFyKC0tZ29sZDMpLHZhcigtLXRlYWwzKSk7JyArCiAgICAgICAgJ2JvcmRlcjpub25lO2JvcmRlci1yYWRpdXM6N3B4O2NvbG9yOiNmZmY7Zm9udC1zaXplOjEycHg7Zm9udC13ZWlnaHQ6NzAwOycgKwogICAgICAgICdjdXJzb3I6cG9pbnRlcjtib3gtc2hhZG93OjAgMCAxNnB4IHJnYmEoMjQ1LDE2NiwzNSwuMyk7dHJhbnNpdGlvbjphbGwgLjJzIj4nICsKICAgICAgICAnXHVEODNEXHVERDc4IFZpZXcgQ29tYmluZWQgTGluZWFnZSBHcmFwaDwvYnV0dG9uPicgKwogICAgICAnPGJ1dHRvbiBvbmNsaWNrPSJzaG93VGFiKFwnbWV0YWRhdGFcJykiICcgKwogICAgICAgICdzdHlsZT0iZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6NnB4O3BhZGRpbmc6OXB4IDE2cHg7JyArCiAgICAgICAgJ2JhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDEzNWRlZyx2YXIoLS10ZWFsMyksdmFyKC0tc2t5KSk7JyArCiAgICAgICAgJ2JvcmRlcjpub25lO2JvcmRlci1yYWRpdXM6N3B4O2NvbG9yOiNmZmY7Zm9udC1zaXplOjEycHg7Zm9udC13ZWlnaHQ6NzAwOycgKwogICAgICAgICdjdXJzb3I6cG9pbnRlcjtib3gtc2hhZG93OjAgMCAxNnB4IHJnYmEoMCwyMDEsMTc3LC4yNSk7dHJhbnNpdGlvbjphbGwgLjJzIj4nICsKICAgICAgICAnXHVEODNEXHVEQ0NCIFZpZXcgQWxsIE1ldGFkYXRhPC9idXR0b24+JyArCiAgICAgICc8YnV0dG9uIG9uY2xpY2s9InNob3dUYWIoXCdzb3VyY2VzXCcpIiAnICsKICAgICAgICAnc3R5bGU9ImRpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjZweDtwYWRkaW5nOjlweCAxNnB4OycgKwogICAgICAgICdiYWNrZ3JvdW5kOnZhcigtLXBhbmVsKTtib3JkZXI6MXB4IHNvbGlkIHZhcigtLWJkcjIpOycgKwogICAgICAgICdib3JkZXItcmFkaXVzOjdweDtjb2xvcjp2YXIoLS10MCk7Zm9udC1zaXplOjEycHg7Zm9udC13ZWlnaHQ6NzAwO2N1cnNvcjpwb2ludGVyIj4nICsKICAgICAgICAnXHVEODNEXHVEQ0UxIFZpZXcgU291cmNlczwvYnV0dG9uPicgKwogICAgJzwvZGl2Pic7CgogIHZhciBodG1sID0gZC5maWVsZHNfc2VhcmNoZWQubWFwKGZ1bmN0aW9uKGZpZWxkKXsKICAgIHZhciByZXMgPSBkLnJlc3VsdHNfYnlfZmllbGRbZmllbGRdOwogICAgaWYocmVzICYmIHJlcy5lcnJvcil7CiAgICAgIHJldHVybiAnPGRpdiBjbGFzcz0ibXJlcy1jYXJkIj48ZGl2IGNsYXNzPSJtcmVzLWhkciI+Jytlc2MoZmllbGQpKyc8L2Rpdj48ZGl2IHN0eWxlPSJjb2xvcjp2YXIoLS1yb3NlMik7Zm9udC1zaXplOjEwcHgiPlx1Mjc0QyAnK2VzYyhyZXMuZXJyb3IpKyc8L2Rpdj48L2Rpdj4nOwogICAgfQogICAgaWYoIXJlcykgcmV0dXJuICcnOwogICAgdmFyIG1ldGEgICA9IHJlcy5tZXRhZGF0YXx8W107CiAgICB2YXIgZmlyc3QgID0gbWV0YVswXXx8e307CiAgICB2YXIgY29uZiAgID0gZmlyc3QuY29uZmlkZW50aWFsaXR5fHwnSW50ZXJuYWwnOwogICAgdmFyIGVxdWl2cyA9IGZpcnN0LmVxdWl2YWxlbnRfZmllbGRzfHxbXTsKICAgIHZhciBzb3JzICAgPSBmaXJzdC5zb3JzfHxbXTsKICAgIHZhciBob3BzICAgPSBmaXJzdC5saW5lYWdlX2hvcHN8fFtdOwogICAgdmFyIGluZyAgICA9IGZpcnN0LmluZ2VzdGlvbl9wYXR0ZXJufHwnQmF0Y2gnOwogICAgdmFyIGVyd2luU3RhdCA9IHJlcy5lcndpbl9zdGF0dXN8fCdtb2NrJzsKCiAgICByZXR1cm4gJzxkaXYgY2xhc3M9Im1yZXMtY2FyZCI+JysKICAgICAgJzxkaXYgc3R5bGU9ImRpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjdweDttYXJnaW4tYm90dG9tOjlweDtmbGV4LXdyYXA6d3JhcCI+JysKICAgICAgICAnPGRpdiBjbGFzcz0ibXJlcy1oZHIiIHN0eWxlPSJtYXJnaW4tYm90dG9tOjAiPicrZXNjKGZpZWxkKSsnPC9kaXY+JysKICAgICAgICBjb25mQmFkZ2UoY29uZikrCiAgICAgICAgJzxzcGFuIGNsYXNzPSJpbmctcGlsbCAnKyhpbmc9PT0nUmVhbC1UaW1lIFN0cmVhbWluZyc/J2luZy1ydCc6J2luZy1iYXRjaCcpKyciPicrZXNjKGluZykrJzwvc3Bhbj4nKwogICAgICAgICc8c3BhbiBzdHlsZT0ibWFyZ2luLWxlZnQ6YXV0bztmb250LXNpemU6OXB4O2NvbG9yOnZhcigtLXQyKTtmb250LWZhbWlseTpcJ0pldEJyYWlucyBNb25vXCcsbW9ub3NwYWNlIj4nK2VzYyhlcndpblN0YXQpKyc8L3NwYW4+JysKICAgICAgJzwvZGl2PicrCiAgICAgICc8ZGl2IHN0eWxlPSJkaXNwbGF5OmdyaWQ7Z3JpZC10ZW1wbGF0ZS1jb2x1bW5zOjFmciAxZnI7Z2FwOjlweCI+JysKICAgICAgICAnPGRpdj4nKwogICAgICAgICAgJzxkaXYgc3R5bGU9ImZvbnQtc2l6ZTo5cHg7Zm9udC13ZWlnaHQ6NzAwO2NvbG9yOnZhcigtLXQyKTt0ZXh0LXRyYW5zZm9ybTp1cHBlcmNhc2U7bGV0dGVyLXNwYWNpbmc6LjVweDttYXJnaW4tYm90dG9tOjVweDtmb250LWZhbWlseTpcJ0pldEJyYWlucyBNb25vXCcsbW9ub3NwYWNlIj5TeXN0ZW1zIG9mIFJlY29yZDwvZGl2PicrCiAgICAgICAgICBzb3JzLnNsaWNlKDAsNCkubWFwKGZ1bmN0aW9uKHMpewogICAgICAgICAgICByZXR1cm4gJzxkaXYgY2xhc3M9InNvci1jYXJkICcrKHMuaXNfYXV0aG9yaXRhdGl2ZT8nYXV0aCc6JycpKyciIG9uY2xpY2s9ImRvU2V0U09SKFwnJytlc2MoZmllbGQpKydcJyxcJycrZXNjKHMuc29yKSsnXCcpIiBzdHlsZT0icGFkZGluZzo0cHggN3B4O21hcmdpbi1ib3R0b206M3B4Ij4nKwogICAgICAgICAgICAgICc8ZGl2IGNsYXNzPSJzb3ItbmFtZSIgc3R5bGU9ImZvbnQtc2l6ZTo5cHgiPicrZXNjKHMuc29yKSsnPC9kaXY+JysKICAgICAgICAgICAgICAocy5pc19hdXRob3JpdGF0aXZlPyc8c3BhbiBjbGFzcz0ic29yLWF1dGgiIHN0eWxlPSJmb250LXNpemU6N3B4Ij5cdTI2MDU8L3NwYW4+JzonJykrCiAgICAgICAgICAgICc8L2Rpdj4nOwogICAgICAgICAgfSkuam9pbignJykrCiAgICAgICAgJzwvZGl2PicrCiAgICAgICAgJzxkaXY+JysKICAgICAgICAgICc8ZGl2IHN0eWxlPSJmb250LXNpemU6OXB4O2ZvbnQtd2VpZ2h0OjcwMDtjb2xvcjp2YXIoLS10Mik7dGV4dC10cmFuc2Zvcm06dXBwZXJjYXNlO2xldHRlci1zcGFjaW5nOi41cHg7bWFyZ2luLWJvdHRvbTo1cHg7Zm9udC1mYW1pbHk6XCdKZXRCcmFpbnMgTW9ub1wnLG1vbm9zcGFjZSI+RXF1aXZhbGVudCBGaWVsZHM8L2Rpdj4nKwogICAgICAgICAgKGVxdWl2cy5sZW5ndGg/ZXF1aXZzLnNsaWNlKDAsNSkubWFwKGZ1bmN0aW9uKGUpewogICAgICAgICAgICByZXR1cm4gJzxzcGFuIGNsYXNzPSJlcS1pdGVtICcrKGUuYWxpYXNfdHlwZT09PSdrbm93bl9lcXVpdmFsZW50Jz8nZXEtZXhhY3QnOicnKSsnIiBzdHlsZT0iZm9udC1zaXplOjhweCIgdGl0bGU9IicrZXNjKGUuYWxpYXNfdHlwZSkrJyIgb25jbGljaz0ic2V0UShcJycrZXNjKGUuZmllbGQpKydcJyk7ZG9TZWFyY2goKSI+Jytlc2MoZS5maWVsZCkrJzwvc3Bhbj4nOwogICAgICAgICAgfSkuam9pbignJyk6JzxzcGFuIHN0eWxlPSJjb2xvcjp2YXIoLS10Myk7Zm9udC1zaXplOjlweCI+Tm9uZSBmb3VuZDwvc3Bhbj4nKSsKICAgICAgICAnPC9kaXY+JysKICAgICAgJzwvZGl2PicrCiAgICAgIChob3BzLmxlbmd0aD8KICAgICAgICAnPGRpdiBzdHlsZT0ibWFyZ2luLXRvcDo4cHg7cGFkZGluZy10b3A6N3B4O2JvcmRlci10b3A6MXB4IHNvbGlkIHZhcigtLWJkcikiPicrCiAgICAgICAgICAnPGRpdiBzdHlsZT0iZm9udC1zaXplOjlweDtmb250LXdlaWdodDo3MDA7Y29sb3I6dmFyKC0tdDIpO3RleHQtdHJhbnNmb3JtOnVwcGVyY2FzZTtsZXR0ZXItc3BhY2luZzouNXB4O21hcmdpbi1ib3R0b206NXB4O2ZvbnQtZmFtaWx5OlwnSmV0QnJhaW5zIE1vbm9cJyxtb25vc3BhY2UiPkxpbmVhZ2UgUGF0aCAoJytob3BzLmxlbmd0aCsnIGhvcHMpPC9kaXY+JysKICAgICAgICAgICc8ZGl2IHN0eWxlPSJkaXNwbGF5OmZsZXg7YWxpZ24taXRlbXM6Y2VudGVyO2dhcDo0cHg7ZmxleC13cmFwOndyYXAiPicrCiAgICAgICAgICAgIGhvcHMubWFwKGZ1bmN0aW9uKGgsaSl7CiAgICAgICAgICAgICAgcmV0dXJuICc8c3BhbiBzdHlsZT0iZm9udC1zaXplOjlweDtmb250LWZhbWlseTpcJ0pldEJyYWlucyBNb25vXCcsbW9ub3NwYWNlO2NvbG9yOnZhcigtLXQxKSI+Jytlc2MoaC5zeXN0ZW18fGgudGFibGV8fCcnKSsnPC9zcGFuPicrKGk8aG9wcy5sZW5ndGgtMT8nPHNwYW4gc3R5bGU9ImNvbG9yOnZhcigtLXQzKSI+JnJhcnI7PC9zcGFuPic6JycpOwogICAgICAgICAgICB9KS5qb2luKCcnKSsKICAgICAgICAgICc8L2Rpdj4nKwogICAgICAgICc8L2Rpdj4nOicnKSArCiAgICAgICc8ZGl2IHN0eWxlPSJtYXJnaW4tdG9wOjlweDtkaXNwbGF5OmZsZXg7Z2FwOjZweCI+JysKICAgICAgICAnPGJ1dHRvbiBvbmNsaWNrPSJzZXRRKFwnJytlc2MoZmllbGQpKydcJyk7ZG9TZWFyY2goKSIgJysKICAgICAgICAnc3R5bGU9ImZsZXg6MTtoZWlnaHQ6MjZweDtiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCgxMzVkZWcsdmFyKC0tZ29sZDMpLHZhcigtLXRlYWwzKSk7Ym9yZGVyOm5vbmU7Ym9yZGVyLXJhZGl1czo0cHg7Y29sb3I6I2ZmZjtmb250LXNpemU6MTBweDtmb250LXdlaWdodDo3MDA7Y3Vyc29yOnBvaW50ZXIiPicrCiAgICAgICAgJ1x1MjZBMSBTZWFyY2ggVGhpcyBGaWVsZDwvYnV0dG9uPicrCiAgICAgICAgJzxidXR0b24gb25jbGljaz0ic2V0UShcJycrZXNjKGZpZWxkKSsnXCcpO2RvU2VhcmNoKCk7c2V0VGltZW91dChmdW5jdGlvbigpe3Nob3dUYWIoXCdsaW5lYWdlXCcpO30sODAwKSIgJysKICAgICAgICAnc3R5bGU9ImZsZXg6MTtoZWlnaHQ6MjZweDtiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCgxMzVkZWcsdmFyKC0tdGVhbDMpLHZhcigtLXNreSkpO2JvcmRlcjpub25lO2JvcmRlci1yYWRpdXM6NHB4O2NvbG9yOiNmZmY7Zm9udC1zaXplOjEwcHg7Zm9udC13ZWlnaHQ6NzAwO2N1cnNvcjpwb2ludGVyIj4nKwogICAgICAgICdcdUQ4M0RcdURENzggVmlldyBMaW5lYWdlPC9idXR0b24+JysKICAgICAgJzwvZGl2PicrCiAgICAgICc8L2Rpdj4nOwogIH0pLmpvaW4oJycpOwogIGVsKCdtdWx0aS1yZXN1bHRzJykuaW5uZXJIVE1MICs9IGh0bWw7CiAgZWwoJ211bHRpLWJvdHRvbS1iYXInKS5zdHlsZS5kaXNwbGF5ID0gJ2ZsZXgnOwogIC8vIEFkZCBib3R0b20gYWN0aW9uIGJhciB0b28gKGxlZ2FjeSkKICBlbCgnbXVsdGktcmVzdWx0cycpLmlubmVySFRNTCArPQogICAgJzxkaXYgc3R5bGU9ImRpc3BsYXk6ZmxleDtnYXA6OHB4O21hcmdpbi10b3A6MTRweDtmbGV4LXdyYXA6d3JhcDsnICsKICAgICdwYWRkaW5nLXRvcDoxMnB4O2JvcmRlci10b3A6MXB4IHNvbGlkIHZhcigtLWJkcjIpIj4nICsKICAgICAgJzxidXR0b24gb25jbGljaz0ic2hvd1RhYihcJ2xpbmVhZ2VcJyk7c2V0VGltZW91dChmdW5jdGlvbigpe3NldExldmVsKFwndGFibGVcJyk7fSw4MCkiICcgKwogICAgICAgICdzdHlsZT0iZGlzcGxheTpmbGV4O2FsaWduLWl0ZW1zOmNlbnRlcjtnYXA6NnB4O3BhZGRpbmc6OXB4IDE2cHg7JyArCiAgICAgICAgJ2JhY2tncm91bmQ6bGluZWFyLWdyYWRpZW50KDEzNWRlZyx2YXIoLS1nb2xkMyksdmFyKC0tdGVhbDMpKTsnICsKICAgICAgICAnYm9yZGVyOm5vbmU7Ym9yZGVyLXJhZGl1czo3cHg7Y29sb3I6I2ZmZjtmb250LXNpemU6MTJweDtmb250LXdlaWdodDo3MDA7JyArCiAgICAgICAgJ2N1cnNvcjpwb2ludGVyO2JveC1zaGFkb3c6MCAwIDE2cHggcmdiYSgyNDUsMTY2LDM1LC4zKSI+JyArCiAgICAgICAgJ1x1RDgzRFx1REQ3OCBWaWV3IENvbWJpbmVkIExpbmVhZ2UgR3JhcGg8L2J1dHRvbj4nICsKICAgICAgJzxidXR0b24gb25jbGljaz0ic2hvd1RhYihcJ21ldGFkYXRhXCcpIiAnICsKICAgICAgICAnc3R5bGU9ImRpc3BsYXk6ZmxleDthbGlnbi1pdGVtczpjZW50ZXI7Z2FwOjZweDtwYWRkaW5nOjlweCAxNnB4OycgKwogICAgICAgICdiYWNrZ3JvdW5kOmxpbmVhci1ncmFkaWVudCgxMzVkZWcsdmFyKC0tdGVhbDMpLHZhcigtLXNreSkpOycgKwogICAgICAgICdib3JkZXI6bm9uZTtib3JkZXItcmFkaXVzOjdweDtjb2xvcjojZmZmO2ZvbnQtc2l6ZToxMnB4O2ZvbnQtd2VpZ2h0OjcwMDsnICsKICAgICAgICAnY3Vyc29yOnBvaW50ZXI7Ym94LXNoYWRvdzowIDAgMTZweCByZ2JhKDAsMjAxLDE3NywuMjUpIj4nICsKICAgICAgICAnXHVEODNEXHVEQ0NCIFZpZXcgQWxsIE1ldGFkYXRhPC9idXR0b24+JyArCiAgICAnPC9kaXY+JzsKfQoKLy8g4pSA4pSAIFVQTE9BRCDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKLy8g4pSA4pSAIEVSV0lOIEFQSSBFWFBMT1JFUiDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIDilIAKdmFyIGVyd2luRW5kcG9pbnRzID0gW107Cgphc3luYyBmdW5jdGlvbiBsb2FkRXJ3aW5FbmRwb2ludHMoKSB7CiAgZWwoJ2Vyd2luLWVwLWNvdW50JykudGV4dENvbnRlbnQgPSAnTG9hZGluZy4uLic7CiAgZWwoJ2Vyd2luLWVwLWxpc3QnKS5pbm5lckhUTUwgPSAnPGRpdiBzdHlsZT0iY29sb3I6dmFyKC0tdDIpO2ZvbnQtc2l6ZToxMXB4O3BhZGRpbmc6MTZweCAwIj4mI3gyM0YzOyBEaXNjb3ZlcmluZyBlbmRwb2ludHMuLi48L2Rpdj4nOwogIHRyeSB7CiAgICB2YXIgciA9IGF3YWl0IGZldGNoKEFQSSArICcvYXBpL2Vyd2luL2VuZHBvaW50cycpOwogICAgaWYgKCFyLm9rKSB0aHJvdyBuZXcgRXJyb3IoYXdhaXQgci50ZXh0KCkpOwogICAgdmFyIGQgPSBhd2FpdCByLmpzb24oKTsKICAgIGVyd2luRW5kcG9pbnRzID0gZC5lbmRwb2ludHM7CiAgICBlbCgnZXJ3aW4tZXAtY291bnQnKS50ZXh0Q29udGVudCA9IGQudG90YWwgKyAnIGVuZHBvaW50cycgKyAoZC5zd2FnZ2VyX2Rpc2NvdmVyZWQgPyAnIChTd2FnZ2VyKScgOiAnIChidWlsdC1pbiByZWdpc3RyeSknKTsKICAgIHJlbmRlckVyd2luRW5kcG9pbnRzKGQpOwogICAgcG9wdWxhdGVFcndpblNlbGVjdChkLmVuZHBvaW50cyk7CiAgfSBjYXRjaChlKSB7CiAgICBlbCgnZXJ3aW4tZXAtY291bnQnKS50ZXh0Q29udGVudCA9ICdFcnJvcjogJyArIGUubWVzc2FnZTsKICAgIGVsKCdlcndpbi1lcC1saXN0JykuaW5uZXJIVE1MID0gJzxkaXYgc3R5bGU9ImNvbG9yOnZhcigtLXJvc2UyKTtmb250LXNpemU6MTFweDtwYWRkaW5nOjE2cHgiPicgKyBlc2MoZS5tZXNzYWdlKSArICc8L2Rpdj4nOwogIH0KfQoKZnVuY3Rpb24gcG9wdWxhdGVFcndpblNlbGVjdChlcHMpIHsKICB2YXIgc2VsID0gZWwoJ2Vyd2luLWVwLXNlbGVjdCcpOwogIHNlbC5pbm5lckhUTUwgPSAnPG9wdGlvbiB2YWx1ZT0iIj4tLSBTZWxlY3QgZW5kcG9pbnQgLS08L29wdGlvbj4nOwogIGVwcy5mb3JFYWNoKGZ1bmN0aW9uKGVwKSB7CiAgICB2YXIgb3B0ID0gZG9jdW1lbnQuY3JlYXRlRWxlbWVudCgnb3B0aW9uJyk7CiAgICBvcHQudmFsdWUgPSBlcC5wYXRoOwogICAgb3B0LnRleHRDb250ZW50ID0gZXAubWV0aG9kICsgJyAnICsgZXAucGF0aDsKICAgIHNlbC5hcHBlbmRDaGlsZChvcHQpOwogIH0pOwp9CgpmdW5jdGlvbiBvbkVyd2luRXBTZWxlY3QocGF0aCkgewogIHZhciBlcCA9IGVyd2luRW5kcG9pbnRzLmZpbmQoZnVuY3Rpb24oZSl7IHJldHVybiBlLnBhdGggPT09IHBhdGg7IH0pOwogIGlmICghZXApIHsgZWwoJ2Vyd2luLWVwLWRlc2MnKS50ZXh0Q29udGVudCA9ICdTZWxlY3QgYW4gZW5kcG9pbnQgYWJvdmUnOyBlbCgnZXJ3aW4tcGFyYW0tYnVpbGRlcicpLmlubmVySFRNTCA9ICcnOyByZXR1cm47IH0KICBlbCgnZXJ3aW4tZXAtZGVzYycpLnRleHRDb250ZW50ID0gZXAuZGVzY3JpcHRpb24gfHwgZXAucGF0aDsKICAvLyBCdWlsZCBwYXJhbSBpbnB1dHMKICB2YXIgcGIgPSBlbCgnZXJ3aW4tcGFyYW0tYnVpbGRlcicpOwogIGlmICghZXAucGFyYW1zIHx8ICFlcC5wYXJhbXMubGVuZ3RoKSB7IHBiLmlubmVySFRNTCA9ICc8ZGl2IHN0eWxlPSJmb250LXNpemU6MTBweDtjb2xvcjp2YXIoLS10MikiPk5vIHBhcmFtZXRlcnMgcmVxdWlyZWQ8L2Rpdj4nOyByZXR1cm47IH0KICB2YXIgZGVmYXVsdHMgPSB7CiAgICB0YWJsZU5hbWU6ICAgICAgIGVsKCdxJykgPyBlbCgncScpLnZhbHVlLnNwbGl0KCcuJykucG9wKCkudG9VcHBlckNhc2UoKSA6ICcnLAogICAgY29sdW1uTmFtZTogICAgICBlbCgncScpID8gKGVsKCdxJykudmFsdWUuc3BsaXQoJy4nKVsxXSB8fCAnJykgOiAnJywKICAgIHN5c3RlbU5hbWU6ICAgICAgZ3YoJ2NmLWVyd2luLXN5c3RlbScpLAogICAgZW52aXJvbm1lbnROYW1lOiBndignY2YtZXJ3aW4tZW52JyksCiAgICBsaW5lYWdlVHlwZTogICAgIGd2KCdjZi1lcndpbi1saW50eXBlJykgfHwgJ0RVQUwnLAogICAgcHJvamVjdElkczogICAgICBndignY2YtZXJ3aW4tcHJvamVjdCcpLAogICAgbmFtZTogICAgICAgICAgICBlbCgncScpID8gZWwoJ3EnKS52YWx1ZS5zcGxpdCgnLicpLnBvcCgpLnRvVXBwZXJDYXNlKCkgOiAnJywKICAgIHR5cGU6ICAgICAgICAgICAgJ1RBQkxFJywKICB9OwogIHZhciByb3dzID0gJyc7CiAgZXAucGFyYW1zLmZvckVhY2goZnVuY3Rpb24ocCkgewogICAgdmFyIHJlcSA9IGVwLnJlcXVpcmVkICYmIGVwLnJlcXVpcmVkLmluY2x1ZGVzKHApOwogICAgcm93cyArPSAnPGRpdiBzdHlsZT0iZGlzcGxheTpmbGV4O2dhcDo2cHg7YWxpZ24taXRlbXM6Y2VudGVyO21hcmdpbi1ib3R0b206NXB4Ij4nICsKICAgICAgJzxsYWJlbCBzdHlsZT0iZm9udC1zaXplOjlweDtjb2xvcjonKyhyZXE/J3ZhcigtLWdvbGQyKSc6J3ZhcigtLXQyKScpKwogICAgICAnO2ZvbnQtZmFtaWx5OkpldEJyYWlucyBNb25vLG1vbm9zcGFjZTttaW4td2lkdGg6MTQwcHg7ZmxleC1zaHJpbms6MCI+JyArCiAgICAgIGVzYyhwKSArIChyZXE/JyAqJzonJykgKyAnPC9sYWJlbD4nICsKICAgICAgJzxpbnB1dCBpZD0iZXJ3aW4tcC0nK2VzYyhwKSsnIiBjbGFzcz0iY2ZnLWluIiBzdHlsZT0iZmxleDoxO21hcmdpbi1ib3R0b206MCIgJyArCiAgICAgICd2YWx1ZT0iJytlc2MoZGVmYXVsdHNbcF18fCcnKSsnIiBwbGFjZWhvbGRlcj0iJytlc2MocCkrJyIvPicgKwogICAgICAnPC9kaXY+JzsKICB9KTsKICBwYi5pbm5lckhUTUwgPSAnPGRpdiBzdHlsZT0iZm9udC1zaXplOjlweDtjb2xvcjp2YXIoLS10Mik7bWFyZ2luLWJvdHRvbTo2cHg7Zm9udC1mYW1pbHk6SmV0QnJhaW5zIE1vbm8sbW9ub3NwYWNlIj4qID0gcmVxdWlyZWQ8L2Rpdj4nICsgcm93czsKfQoKYXN5bmMgZnVuY3Rpb24gY2FsbEVyd2luRW5kcG9pbnQoKSB7CiAgdmFyIHBhdGggPSBlbCgnZXJ3aW4tZXAtc2VsZWN0JykudmFsdWU7CiAgaWYgKCFwYXRoKSB7IHRvYXN0KCdTZWxlY3QgYW4gZW5kcG9pbnQgZmlyc3QnLCAnd2FybicpOyByZXR1cm47IH0KICB2YXIgZXAgPSBlcndpbkVuZHBvaW50cy5maW5kKGZ1bmN0aW9uKGUpeyByZXR1cm4gZS5wYXRoID09PSBwYXRoOyB9KTsKICBpZiAoIWVwKSByZXR1cm47CiAgLy8gQ29sbGVjdCBwYXJhbXMKICB2YXIgcGFyYW1zID0ge307CiAgKGVwLnBhcmFtc3x8W10pLmZvckVhY2goZnVuY3Rpb24ocCkgewogICAgdmFyIGlucCA9IGVsKCdlcndpbi1wLScgKyBwKTsKICAgIGlmIChpbnAgJiYgaW5wLnZhbHVlLnRyaW0oKSkgcGFyYW1zW3BdID0gaW5wLnZhbHVlLnRyaW0oKTsKICB9KTsKICBlbCgnZXJ3aW4tY2FsbC1yZXN1bHQnKS5zdHlsZS5kaXNwbGF5ID0gJ2Jsb2NrJzsKICBlbCgnZXJ3aW4tcmVzdWx0LXN0YXR1cycpLmlubmVySFRNTCA9ICc8c3BhbiBzdHlsZT0iY29sb3I6dmFyKC0tZ29sZDIpIj4mI3gyM0YzOyBDYWxsaW5nICcgKyBlc2MocGF0aCkgKyAnLi4uPC9zcGFuPic7CiAgZWwoJ2Vyd2luLXJlc3VsdC1qc29uJykudGV4dENvbnRlbnQgPSAnJzsKICB0cnkgewogICAgdmFyIHIgPSBhd2FpdCBmZXRjaChBUEkgKyAnL2FwaS9lcndpbi9jYWxsJywgewogICAgICBtZXRob2Q6ICdQT1NUJywgaGVhZGVyczogeydDb250ZW50LVR5cGUnOidhcHBsaWNhdGlvbi9qc29uJ30sCiAgICAgIGJvZHk6IEpTT04uc3RyaW5naWZ5KHsKICAgICAgICBwYXRoOiBwYXRoLCBtZXRob2Q6IGVwLm1ldGhvZCB8fCAnR0VUJywgcGFyYW1zOiBwYXJhbXMsCiAgICAgICAgYmFzZV91cmw6IGd2KCdjZi1lcndpbi11cmwnKSB8fCBudWxsLAogICAgICAgIGFwaV9rZXk6ICBndignY2YtZXJ3aW4ta2V5JykgfHwgbnVsbCwKICAgICAgfSkKICAgIH0pOwogICAgdmFyIGQgPSBhd2FpdCByLmpzb24oKTsKICAgIHZhciBzdGF0dXNDb2xvciA9IGQuc3RhdHVzX2NvZGUgPT09IDIwMCA/ICd2YXIoLS1saW1lMiknIDogJ3ZhcigtLXJvc2UyKSc7CiAgICBlbCgnZXJ3aW4tcmVzdWx0LXN0YXR1cycpLmlubmVySFRNTCA9ICc8c3BhbiBzdHlsZT0iY29sb3I6JytzdGF0dXNDb2xvcisnIj5IVFRQICcgKyBkLnN0YXR1c19jb2RlICsgJzwvc3Bhbj4nOwogICAgZWwoJ2Vyd2luLXJlc3VsdC1lbmRwb2ludCcpLnRleHRDb250ZW50ID0gZC5lbmRwb2ludDsKICAgIGVsKCdlcndpbi1yZXN1bHQtanNvbicpLnRleHRDb250ZW50ID0gSlNPTi5zdHJpbmdpZnkoZC5yZXNwb25zZSwgbnVsbCwgMik7CiAgICB0b2FzdCgoZC5zdGF0dXNfY29kZT09PTIwMCA/ICfinIUgJyA6ICfimqDvuI8gJykgKyAnSFRUUCAnICsgZC5zdGF0dXNfY29kZSArICcg4oCUICcgKyBwYXRoLCBkLnN0YXR1c19jb2RlPT09MjAwPydzdWNjZXNzJzond2FybicpOwogIH0gY2F0Y2goZSkgewogICAgZWwoJ2Vyd2luLXJlc3VsdC1zdGF0dXMnKS5pbm5lckhUTUwgPSAnPHNwYW4gc3R5bGU9ImNvbG9yOnZhcigtLXJvc2UyKSI+RXJyb3I6ICcgKyBlc2MoZS5tZXNzYWdlKSArICc8L3NwYW4+JzsKICAgIHRvYXN0KCdDYWxsIGZhaWxlZDogJyArIGUubWVzc2FnZSwgJ2Vycm9yJyk7CiAgfQp9Cgphc3luYyBmdW5jdGlvbiBjYWxsQW5kU2VhcmNoKCkgewogIC8vIENhbGwgZW5kcG9pbnQgZmlyc3QsIHRoZW4gdHJpZ2dlciBsaW5lYWdlIHNlYXJjaCB1c2luZyB0YWJsZU5hbWUgcGFyYW0KICB2YXIgdGFibGVJbnB1dCA9IGVsKCdlcndpbi1wLXRhYmxlTmFtZScpIHx8IGVsKCdlcndpbi1wLW5hbWUnKTsKICBpZiAodGFibGVJbnB1dCAmJiB0YWJsZUlucHV0LnZhbHVlKSB7CiAgICBlbCgncScpLnZhbHVlID0gdGFibGVJbnB1dC52YWx1ZTsKICB9CiAgYXdhaXQgY2FsbEVyd2luRW5kcG9pbnQoKTsKICBhd2FpdCBkb1NlYXJjaCgpOwogIHNob3dUYWIoJ2xpbmVhZ2UnKTsKfQoKYXN5bmMgZnVuY3Rpb24gbG9hZEVyd2luU3dhZ2dlcigpIHsKICBlbCgnZXJ3aW4tY2FsbC1yZXN1bHQnKS5zdHlsZS5kaXNwbGF5ID0gJ2Jsb2NrJzsKICBlbCgnZXJ3aW4tcmVzdWx0LXN0YXR1cycpLmlubmVySFRNTCA9ICc8c3BhbiBzdHlsZT0iY29sb3I6dmFyKC0tZ29sZDIpIj4mI3gyM0YzOyBGZXRjaGluZyBTd2FnZ2VyIHNwZWMuLi48L3NwYW4+JzsKICBlbCgnZXJ3aW4tcmVzdWx0LWpzb24nKS50ZXh0Q29udGVudCA9ICcnOwogIHRyeSB7CiAgICB2YXIgciA9IGF3YWl0IGZldGNoKEFQSSArICcvYXBpL2Vyd2luL3N3YWdnZXInKTsKICAgIGlmICghci5vaykgeyB2YXIgZSA9IGF3YWl0IHIuanNvbigpOyB0aHJvdyBuZXcgRXJyb3IoZS5kZXRhaWx8fHIuc3RhdHVzVGV4dCk7IH0KICAgIHZhciBkID0gYXdhaXQgci5qc29uKCk7CiAgICBlbCgnZXJ3aW4tcmVzdWx0LXN0YXR1cycpLmlubmVySFRNTCA9ICc8c3BhbiBzdHlsZT0iY29sb3I6dmFyKC0tbGltZTIpIj4mI3gyNzA1OyBTd2FnZ2VyIHNwZWMgbG9hZGVkPC9zcGFuPic7CiAgICBlbCgnZXJ3aW4tcmVzdWx0LWVuZHBvaW50JykudGV4dENvbnRlbnQgPSAnT3BlbkFQSSAnICsgKGQub3BlbmFwaXx8ZC5zd2FnZ2VyfHwnc3BlYycpOwogICAgdmFyIHBhdGhzID0gT2JqZWN0LmtleXMoZC5wYXRoc3x8e30pOwogICAgZWwoJ2Vyd2luLXJlc3VsdC1qc29uJykudGV4dENvbnRlbnQgPSAnUGF0aHMgZm91bmQ6ICcgKyBwYXRocy5sZW5ndGggKyAnXG5cbicgKyBKU09OLnN0cmluZ2lmeShkLmluZm98fHt9LCBudWxsLCAyKSArICdcblxuRW5kcG9pbnRzOlxuJyArIHBhdGhzLmpvaW4oJ1xuJyk7CiAgICB0b2FzdCgn4pyFIFN3YWdnZXIgc3BlYzogJyArIHBhdGhzLmxlbmd0aCArICcgcGF0aHMgZm91bmQnLCAnc3VjY2VzcycpOwogIH0gY2F0Y2goZSkgewogICAgZWwoJ2Vyd2luLXJlc3VsdC1zdGF0dXMnKS5pbm5lckhUTUwgPSAnPHNwYW4gc3R5bGU9ImNvbG9yOnZhcigtLXJvc2UyKSI+RXJyb3I6ICcgKyBlc2MoZS5tZXNzYWdlKSArICc8L3NwYW4+JzsKICAgIHRvYXN0KCdTd2FnZ2VyIGVycm9yOiAnICsgZS5tZXNzYWdlLCAnd2FybicpOwogIH0KfQoKZnVuY3Rpb24gcmVuZGVyRXJ3aW5FbmRwb2ludHMoZGF0YSkgewogIHZhciBncm91cHMgPSB7fTsKICBkYXRhLmVuZHBvaW50cy5mb3JFYWNoKGZ1bmN0aW9uKGVwKSB7CiAgICB2YXIgZ3JwID0gKGVwLnBhdGguc3BsaXQoJy8nKS5maWx0ZXIoQm9vbGVhbilbMV0pIHx8ICdvdGhlcic7CiAgICBpZiAoIWdyb3Vwc1tncnBdKSBncm91cHNbZ3JwXSA9IFtdOwogICAgZ3JvdXBzW2dycF0ucHVzaChlcCk7CiAgfSk7CiAgdmFyIGNvbnRhaW5lciA9IGRvY3VtZW50LmNyZWF0ZUVsZW1lbnQoJ2RpdicpOwogIE9iamVjdC5rZXlzKGdyb3Vwcykuc29ydCgpLmZvckVhY2goZnVuY3Rpb24oZ3JwKSB7CiAgICB2YXIgZ3JwRGl2ID0gZG9jdW1lbnQuY3JlYXRlRWxlbWVudCgnZGl2Jyk7CiAgICBncnBEaXYuc3R5bGUubWFyZ2luQm90dG9tID0gJzEycHgnOwogICAgdmFyIGdycEhkciA9IGRvY3VtZW50LmNyZWF0ZUVsZW1lbnQoJ2RpdicpOwogICAgZ3JwSGRyLnN0eWxlLmNzc1RleHQgPSAnZm9udC1zaXplOjEwcHg7Zm9udC13ZWlnaHQ6NzAwO2NvbG9yOnZhcigtLWdvbGQyKTtmb250LWZhbWlseTptb25vc3BhY2U7bWFyZ2luLWJvdHRvbTo2cHg7dGV4dC10cmFuc2Zvcm06dXBwZXJjYXNlJzsKICAgIGdycEhkci50ZXh0Q29udGVudCA9IGdycC50b1VwcGVyQ2FzZSgpICsgJyAoJyArIGdyb3Vwc1tncnBdLmxlbmd0aCArICcpJzsKICAgIGdycERpdi5hcHBlbmRDaGlsZChncnBIZHIpOwogICAgZ3JvdXBzW2dycF0uZm9yRWFjaChmdW5jdGlvbihlcCkgewogICAgICB2YXIgcm93ID0gZG9jdW1lbnQuY3JlYXRlRWxlbWVudCgnZGl2Jyk7CiAgICAgIHJvdy5zdHlsZS5jc3NUZXh0ID0gJ2JhY2tncm91bmQ6dmFyKC0tcGFuZWwpO2JvcmRlcjoxcHggc29saWQgdmFyKC0tYmRyKTtib3JkZXItcmFkaXVzOjVweDtwYWRkaW5nOjdweCAxMHB4O21hcmdpbi1ib3R0b206NHB4O2N1cnNvcjpwb2ludGVyJzsKICAgICAgcm93Lm9ubW91c2VvdmVyID0gZnVuY3Rpb24oKXsgdGhpcy5zdHlsZS5iYWNrZ3JvdW5kPSd2YXIoLS1wYW5lbDIpJzsgfTsKICAgICAgcm93Lm9ubW91c2VvdXQgID0gZnVuY3Rpb24oKXsgdGhpcy5zdHlsZS5iYWNrZ3JvdW5kPSd2YXIoLS1wYW5lbCknOyB9OwogICAgICB2YXIgZXBQYXRoID0gZXAucGF0aDsKICAgICAgcm93Lm9uY2xpY2sgPSBmdW5jdGlvbigpeyBlbCgnZXJ3aW4tZXAtc2VsZWN0JykudmFsdWU9ZXBQYXRoOyBvbkVyd2luRXBTZWxlY3QoZXBQYXRoKTsgfTsKICAgICAgdmFyIHRvcCA9IGRvY3VtZW50LmNyZWF0ZUVsZW1lbnQoJ2RpdicpOwogICAgICB0b3Auc3R5bGUuZGlzcGxheSA9ICdmbGV4JzsgdG9wLnN0eWxlLmdhcCA9ICc3cHgnOyB0b3Auc3R5bGUuYWxpZ25JdGVtcyA9ICdjZW50ZXInOwogICAgICB2YXIgYmFkZ2UgPSBkb2N1bWVudC5jcmVhdGVFbGVtZW50KCdzcGFuJyk7CiAgICAgIGJhZGdlLnN0eWxlLmNzc1RleHQgPSAnYmFja2dyb3VuZDpyZ2JhKDAsMjAxLDE3NywuMTIpO2NvbG9yOnZhcigtLXRlYWwyKTtib3JkZXI6MXB4IHNvbGlkIHJnYmEoMCwyMDEsMTc3LC4yKTtib3JkZXItcmFkaXVzOjNweDtwYWRkaW5nOjFweCA2cHg7Zm9udC1zaXplOjlweDtmb250LXdlaWdodDo3MDA7Zm9udC1mYW1pbHk6bW9ub3NwYWNlJzsKICAgICAgYmFkZ2UudGV4dENvbnRlbnQgPSBlcC5tZXRob2Q7CiAgICAgIHZhciBwdGggPSBkb2N1bWVudC5jcmVhdGVFbGVtZW50KCdzcGFuJyk7CiAgICAgIHB0aC5zdHlsZS5jc3NUZXh0ID0gJ2ZvbnQtZmFtaWx5Om1vbm9zcGFjZTtmb250LXNpemU6MTBweDtjb2xvcjp2YXIoLS10MCknOwogICAgICBwdGgudGV4dENvbnRlbnQgPSBlcC5wYXRoOwogICAgICB0b3AuYXBwZW5kQ2hpbGQoYmFkZ2UpOyB0b3AuYXBwZW5kQ2hpbGQocHRoKTsKICAgICAgaWYgKGVwLnJlcXVpcmVkICYmIGVwLnJlcXVpcmVkLmxlbmd0aCkgewogICAgICAgIHZhciByMiA9IGRvY3VtZW50LmNyZWF0ZUVsZW1lbnQoJ3NwYW4nKTsKICAgICAgICByMi5zdHlsZS5jc3NUZXh0ID0gJ2ZvbnQtc2l6ZTo4cHg7Y29sb3I6dmFyKC0tcm9zZTIpJzsKICAgICAgICByMi50ZXh0Q29udGVudCA9ICdyZXF1aXJlczogJyArIGVwLnJlcXVpcmVkLmpvaW4oJywgJyk7CiAgICAgICAgdG9wLmFwcGVuZENoaWxkKHIyKTsKICAgICAgfQogICAgICByb3cuYXBwZW5kQ2hpbGQodG9wKTsKICAgICAgaWYgKGVwLmRlc2NyaXB0aW9uKSB7CiAgICAgICAgdmFyIGQyID0gZG9jdW1lbnQuY3JlYXRlRWxlbWVudCgnZGl2Jyk7CiAgICAgICAgZDIuc3R5bGUuY3NzVGV4dCA9ICdmb250LXNpemU6OXB4O2NvbG9yOnZhcigtLXQyKTttYXJnaW4tdG9wOjJweCc7CiAgICAgICAgZDIudGV4dENvbnRlbnQgPSBlcC5kZXNjcmlwdGlvbjsKICAgICAgICByb3cuYXBwZW5kQ2hpbGQoZDIpOwogICAgICB9CiAgICAgIGlmIChlcC5wYXJhbXMgJiYgZXAucGFyYW1zLmxlbmd0aCkgewogICAgICAgIHZhciBwMiA9IGRvY3VtZW50LmNyZWF0ZUVsZW1lbnQoJ2RpdicpOwogICAgICAgIHAyLnN0eWxlLmNzc1RleHQgPSAnZm9udC1zaXplOjhweDtjb2xvcjp2YXIoLS10Myk7bWFyZ2luLXRvcDoycHg7Zm9udC1mYW1pbHk6bW9ub3NwYWNlJzsKICAgICAgICBwMi50ZXh0Q29udGVudCA9ICdwYXJhbXM6ICcgKyBlcC5wYXJhbXMuam9pbignLCAnKTsKICAgICAgICByb3cuYXBwZW5kQ2hpbGQocDIpOwogICAgICB9CiAgICAgIGdycERpdi5hcHBlbmRDaGlsZChyb3cpOwogICAgfSk7CiAgICBjb250YWluZXIuYXBwZW5kQ2hpbGQoZ3JwRGl2KTsKICB9KTsKICBlbCgnZXJ3aW4tZXAtbGlzdCcpLmlubmVySFRNTCA9ICcnOwogIGVsKCdlcndpbi1lcC1saXN0JykuYXBwZW5kQ2hpbGQoY29udGFpbmVyKTsKfQoKYXN5bmMgZnVuY3Rpb24gZG9VcGxvYWQoaW5wdXQpewogIHZhciBmaWxlID0gaW5wdXQuZmlsZXNbMF07IGlmKCFmaWxlKSByZXR1cm47CiAgdmFyIHN0YXR1c0VsID0gZWwoJ3VwbG9hZC1zdGF0dXMnKTsKICBzdGF0dXNFbC5zdHlsZS5jb2xvciA9ICd2YXIoLS1nb2xkMiknOwogIHN0YXR1c0VsLnRleHRDb250ZW50ID0gJ1x1MjNGMyBVcGxvYWRpbmcgJytmaWxlLm5hbWUrJy4uLic7CiAgdHJ5ewogICAgdmFyIGZkID0gbmV3IEZvcm1EYXRhKCk7CiAgICBmZC5hcHBlbmQoJ2ZpbGUnLCBmaWxlKTsKICAgIHZhciByID0gYXdhaXQgZmV0Y2goQVBJKycvYXBpL3VwbG9hZC1pbnB1dHMnLHttZXRob2Q6J1BPU1QnLCBib2R5OmZkfSk7CiAgICBpZighci5vaykgdGhyb3cgbmV3IEVycm9yKGF3YWl0IHIudGV4dCgpKTsKICAgIHZhciBkID0gYXdhaXQgci5qc29uKCk7CiAgICBzdGF0dXNFbC5zdHlsZS5jb2xvciA9ICd2YXIoLS1saW1lMiknOwogICAgaWYoZC5maWVsZHNfZGV0ZWN0ZWQgJiYgZC5maWVsZHNfZGV0ZWN0ZWQubGVuZ3RoKXsKICAgICAgc3RhdHVzRWwudGV4dENvbnRlbnQgPSAnXHUyNzA1ICcrZC50b3RhbF9maWVsZHNfZGV0ZWN0ZWQrJyBmaWVsZHMgZGV0ZWN0ZWQnOwogICAgICBlbCgnbXVsdGktdGEnKS52YWx1ZSA9IGQuZmllbGRzX2RldGVjdGVkLnNsaWNlKDAsMTApLmpvaW4oJ1xuJyk7CiAgICAgIHRvYXN0KCdcdTI3MDUgJytkLnRvdGFsX2ZpZWxkc19kZXRlY3RlZCsnIGZpZWxkcyBsb2FkZWQgaW50byBNdWx0aS1GaWVsZCcsJ3N1Y2Nlc3MnKTsKICAgIH0gZWxzZSBpZihkLmV4dHJhY3RlZF9maWxlcyl7CiAgICAgIHN0YXR1c0VsLnRleHRDb250ZW50ID0gJ1x1MjcwNSBFeHRyYWN0ZWQgJytkLmV4dHJhY3RlZF9maWxlcy5sZW5ndGgrJyBmaWxlcyc7CiAgICAgIHRvYXN0KCdcdTI3MDUgRXh0cmFjdGVkICcrZC5leHRyYWN0ZWRfZmlsZXMubGVuZ3RoKycgZmlsZXMnLCdzdWNjZXNzJyk7CiAgICB9IGVsc2UgewogICAgICBzdGF0dXNFbC50ZXh0Q29udGVudCA9ICdcdTI3MDUgRmlsZSB1cGxvYWRlZCc7CiAgICAgIHRvYXN0KCdcdTI3MDUgVXBsb2FkZWQgc3VjY2Vzc2Z1bGx5Jywnc3VjY2VzcycpOwogICAgfQogIH0gY2F0Y2goZSl7CiAgICBzdGF0dXNFbC5zdHlsZS5jb2xvciA9ICd2YXIoLS1yb3NlMiknOwogICAgc3RhdHVzRWwudGV4dENvbnRlbnQgPSAnXHUyNzRDICcrZS5tZXNzYWdlOwogICAgdG9hc3QoJ1VwbG9hZCBlcnJvcjogJytlLm1lc3NhZ2UsJ2Vycm9yJyk7CiAgfQogIGlucHV0LnZhbHVlID0gJyc7Cn0KPC9zY3JpcHQ+CjwvYm9keT4KPC9odG1sPgo="

@app.get("/")
async def ui():
    import base64
    from fastapi.responses import HTMLResponse
    html = base64.b64decode(_HTML_B64).decode("utf-8")
    # Inject a version marker so browser always knows it's fresh
    html = html.replace(
        '</title>',
        f'</title>\n<meta name="version" content="{datetime.now().isoformat()}">'
    )
    return HTMLResponse(
        content=html,
        headers={
            "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
            "Pragma": "no-cache",
            "Expires": "-1",
            "Surrogate-Control": "no-store",
            "Vary": "*"
        }
    )

if __name__=="__main__":
    import uvicorn
    uvicorn.run("main:app",host="0.0.0.0",port=8080,reload=True)
